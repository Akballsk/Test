import urllib.request
import urllib.parse
import json
import time
import os
import datetime
import threading
import re
import uuid
import random
import concurrent.futures
import shutil
import queue

TELEGRAM_TOKEN = "8763092816:AAEoOb-6ed_rVIbIpnSeZZXfNujTsYOHCPc"
ADMIN_ID = 7061092833
IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

# ============= API PROVIDERS =============
APIS = {
    "s1": {
        "name":     "Server 1 (SmsBower)",
        "short":    "S1",
        "icon":     "🔵",
        "emoji_id": "5440539497383087970",
        "key":      "O181ILPnweEEuO58M7AQYL61yqRlWjwd",
        "url":      "https://smsbower.page/stubs/handler_api.php",
    },
    "s2": {
        "name":     "Server 2 (HeroSMS)",
        "short":    "S2",
        "icon":     "🟢",
        "emoji_id": "5447203607294265305",
        "key":      "2b014c857fbc7ec81ff4b76A2872AA27",
        "url":      "https://hero-sms.com/stubs/handler_api.php",
    },
    "s3": {
        "name":     "Server 3 (SmsIndia)",
        "short":    "S3",
        "icon":     "🟡",
        "emoji_id": "5453902265922376865",
        "key":      "6cd48cbacd48e2e3c48eecd14614468723df",
        "url":      "https://api.smsindia.pro/stubs/handler_api.php",
        "operator": True,
    },
    "s4": {
        "name":     "Server 4 (UotpStore)",
        "short":    "S4",
        "icon":     "🟤",
        "emoji_id": "5453902265922376865",
        "key":      "4WXKaMC8KK6eDmkML3aUStti2riXTKQBGAcXNB9Q",
        "url":      "https://uotp.store/api/stubs/handler_api.php",
        "operator": True,
    },
    "s5": {
        "name":     "Server 5 (SmsProvider)",
        "short":    "S5",
        "icon":     "🟣",
        "emoji_id": "5453902265922376865",
        "key":      "XLOK1XSC7gHXXeB83TrEXYqnQpoYuDgnNdfCeePX5Ii0MrfiwyK2UFmuW1Pb",
        "url":      "https://smsprovider.in/stubs/handler_api.php",
        "operator": False,
    },
}

# ============= COUNTRY FLAGS & UTILS =============
COUNTRY_FLAGS = {
    "afghanistan": "🇦🇫", "albania": "🇦🇱", "algeria": "🇩🇿", "andorra": "🇦🇩", "angola": "🇦🇴",
    "antigua and barbuda": "🇦🇬", "argentina": "🇦🇷", "armenia": "🇦🇲", "australia": "🇦🇺",
    "austria": "🇦🇹", "azerbaijan": "🇦🇿", "bahamas": "🇧🇸", "bahrain": "🇧🇭",
    "bangladesh": "🇧🇩", "barbados": "🇧🇧", "belarus": "🇧🇾", "belgium": "🇧🇪", "belize": "🇧🇿",
    "benin": "🇧🇯", "bhutan": "🇧🇹", "bolivia": "🇧🇴", "bosnia and herzegovina": "🇧🇦",
    "botswana": "🇧🇼", "brazil": "🇧🇷", "brunei": "🇧🇳", "bulgaria": "🇧🇬",
    "burkina faso": "🇧🇫", "burundi": "🇧🇮", "cambodia": "🇰🇭", "cameroon": "🇨🇲",
    "canada": "🇨🇦", "cape verde": "🇨🇻", "central african republic": "🇨🇫", "chad": "🇹🇩",
    "chile": "🇨🇱", "china": "🇨🇳", "colombia": "🇨🇴", "comoros": "🇰🇲", "congo": "🇨🇬",
    "costa rica": "🇨🇷", "cote d'ivoire": "🇨🇮", "ivory coast": "🇨🇮",
    "croatia": "🇭🇷", "cuba": "🇨🇺", "cyprus": "🇨🇾", "czech republic": "🇨🇿",
    "denmark": "🇩🇰", "djibouti": "🇩🇯", "dominica": "🇩🇲", "dominican republic": "🇩🇴",
    "drc": "🇨🇩", "ecuador": "🇪🇨", "egypt": "🇪🇬", "el salvador": "🇸🇻",
    "equatorial guinea": "🇬🇶", "eritrea": "🇪🇷", "estonia": "🇪🇪", "eswatini": "🇸🇿",
    "ethiopia": "🇪🇹", "fiji": "🇫🇯", "finland": "🇫🇮", "france": "🇫🇷",
    "gabon": "🇬🇦", "gambia": "🇬🇲", "georgia": "🇬🇪", "germany": "🇩🇪", "ghana": "🇬🇭",
    "greece": "🇬🇷", "grenada": "🇬🇩", "guatemala": "🇬🇹", "guinea": "🇬🇳",
    "guinea bissau": "🇬🇼", "guyana": "🇬🇾", "haiti": "🇭🇹", "honduras": "🇭🇳",
    "hong kong": "🇭🇰", "hungary": "🇭🇺", "iceland": "🇮🇸", "india": "🇮🇳",
    "indonesia": "🇮🇩", "iran": "🇮🇷", "iraq": "🇮🇶", "ireland": "🇮🇪", "israel": "🇮🇱",
    "italy": "🇮🇹", "jamaica": "🇯🇲", "japan": "🇯🇵", "jordan": "🇯🇴", "kazakhstan": "🇰🇿",
    "kenya": "🇰🇪", "kiribati": "🇰🇮", "kosovo": "🇽🇰", "kuwait": "🇰🇼", "kyrgyzstan": "🇰🇬",
    "laos": "🇱🇦", "latvia": "🇱🇻", "lebanon": "🇱🇧", "lesotho": "🇱🇸", "liberia": "🇱🇷",
    "libya": "🇱🇾", "liechtenstein": "🇱🇮", "lithuania": "🇱🇹", "luxembourg": "🇱🇺",
    "macau": "🇲🇴", "madagascar": "🇲🇬", "malawi": "🇲🇼", "malaysia": "🇲🇾", "maldives": "🇲🇻",
    "mali": "🇲🇱", "malta": "🇲🇹", "marshall islands": "🇲🇭", "mauritania": "🇲🇷",
    "mauritius": "🇲🇺", "mexico": "🇲🇽", "micronesia": "🇫🇲", "moldova": "🇲🇩",
    "monaco": "🇲🇨", "mongolia": "🇲🇳", "montenegro": "🇲🇪", "morocco": "🇲🇦",
    "mozambique": "🇲🇿", "myanmar": "🇲🇲", "namibia": "🇳🇦", "nauru": "🇳🇷", "nepal": "🇳🇵",
    "netherlands": "🇳🇱", "new zealand": "🇳🇿", "nicaragua": "🇳🇮", "niger": "🇳🇪",
    "nigeria": "🇳🇬", "north korea": "🇰🇵", "north macedonia": "🇲🇰", "norway": "🇳🇴",
    "oman": "🇴🇲", "pakistan": "🇵🇰", "palau": "🇵🇼", "palestine": "🇵🇸", "panama": "🇵🇦",
    "papua new guinea": "🇵🇬", "paraguay": "🇵🇾", "peru": "🇵🇪", "philippines": "🇵🇭",
    "poland": "🇵🇱", "portugal": "🇵🇹", "qatar": "🇶🇦", "romania": "🇷🇴", "russia": "🇷🇺",
    "rwanda": "🇷🇼", "saudi arabia": "🇸🇦", "senegal": "🇸🇳", "serbia": "🇷🇸",
    "seychelles": "🇸🇨", "sierra leone": "🇸🇱", "singapore": "🇸🇬", "slovakia": "🇸🇰",
    "slovenia": "🇸🇮", "somalia": "🇸🇴", "south africa": "🇿🇦",
    "south korea": "🇰🇷", "south sudan": "🇸🇸", "spain": "🇪🇸", "sri lanka": "🇱🇰",
    "sudan": "🇸🇩", "suriname": "🇸🇷", "sweden": "🇸🇪", "switzerland": "🇨🇭", "syria": "🇸🇾",
    "taiwan": "🇹🇼", "tajikistan": "🇹🇯", "tanzania": "🇹🇿", "thailand": "🇹🇭",
    "timor leste": "🇹🇱", "togo": "🇹🇬", "tonga": "🇹🇴", "trinidad and tobago": "🇹🇹",
    "tunisia": "🇹🇳", "turkey": "🇹🇷", "turkmenistan": "🇹🇲", "tuvalu": "🇹🇻",
    "uganda": "🇺🇬", "ukraine": "🇺🇦", "uae": "🇦🇪", "united arab emirates": "🇦🇪",
    "united kingdom": "🇬🇧", "uk": "🇬🇧", "usa": "🇺🇸", "united states": "🇺🇸",
    "uruguay": "🇺🇾", "uzbekistan": "🇺🇿", "vanuatu": "🇻🇺", "venezuela": "🇻🇪",
    "vietnam": "🇻🇳", "yemen": "🇾🇪", "zambia": "🇿🇲", "zimbabwe": "🇿🇼",
    "anguilla": "🇦🇮", "aruba": "🇦🇼", "bermuda": "🇧🇲", "cayman islands": "🇰🇾",
    "greenland": "🇬🇱", "guam": "🇬🇺", "puerto rico": "🇵🇷",
}

# Build CC (ISO-3166-1 alpha-2) → canonical country name from COUNTRY_FLAGS
_CC_TO_NAME = {}
for _cn, _fl in COUNTRY_FLAGS.items():
    _ri = [ch for ch in _fl if 0x1F1E6 <= ord(ch) <= 0x1F1FF]
    if len(_ri) == 2:
        _cc2 = chr(ord(_ri[0]) - 0x1F1E6 + 65) + chr(ord(_ri[1]) - 0x1F1E6 + 65)
        if _cc2 not in _CC_TO_NAME:               # keep first / most common name
            _CC_TO_NAME[_cc2] = _cn.title()
# Manual overrides for clarity
_CC_TO_NAME.update({
    "CI": "Ivory Coast", "CD": "DRC", "KR": "South Korea",
    "KP": "North Korea", "GB": "United Kingdom", "US": "USA",
    "AE": "UAE", "TW": "Taiwan",
})

def _expand_country_name(text):
    """If text is a 2-letter ISO country code, expand to full country name; else return unchanged."""
    s = text.strip()
    if len(s) == 2 and s.isalpha():
        full = _CC_TO_NAME.get(s.upper())
        if full:
            return full
    return s


def get_country_flag(country_name):
    if not country_name:
        return "🌍"
    name = str(country_name).lower().strip()
    # strip emoji flag chars at start
    clean = name.lstrip("🇦🇧🇨🇩🇪🇫🇬🇭🇮🇯🇰🇱🇲🇳🇴🇵🇶🇷🇸🇹🇺🇻🇼🇽🇾🇿 ").strip()
    if clean in COUNTRY_FLAGS:
        return COUNTRY_FLAGS[clean]
    for country, flag in COUNTRY_FLAGS.items():
        if len(country) >= 4 and (country in clean or clean in country):
            return flag
    return "🌍"


PHONE_DIALCODES = {
    "1": "usa", "7": "russia", "20": "egypt", "27": "south africa",
    "30": "greece", "31": "netherlands", "32": "belgium", "33": "france",
    "34": "spain", "36": "hungary", "39": "italy", "40": "romania",
    "41": "switzerland", "43": "austria", "44": "united kingdom", "45": "denmark",
    "46": "sweden", "47": "norway", "48": "poland", "49": "germany",
    "51": "peru", "52": "mexico", "53": "cuba", "54": "argentina",
    "55": "brazil", "56": "chile", "57": "colombia", "58": "venezuela",
    "60": "malaysia", "61": "australia", "62": "indonesia", "63": "philippines",
    "64": "new zealand", "65": "singapore", "66": "thailand",
    "77": "kazakhstan", "81": "japan", "82": "south korea", "84": "vietnam",
    "86": "china", "90": "turkey", "91": "india", "92": "pakistan",
    "93": "afghanistan", "94": "sri lanka", "95": "myanmar", "98": "iran",
    "212": "morocco", "213": "algeria", "216": "tunisia", "218": "libya",
    "220": "gambia", "221": "senegal", "222": "mauritania", "223": "mali",
    "224": "guinea", "225": "ivory coast", "226": "burkina faso", "227": "niger",
    "228": "togo", "229": "benin", "230": "mauritius", "231": "liberia",
    "232": "sierra leone", "233": "ghana", "234": "nigeria", "235": "chad",
    "236": "central african republic", "237": "cameroon", "238": "cape verde",
    "240": "equatorial guinea", "241": "gabon", "242": "congo", "243": "drc",
    "244": "angola", "245": "guinea bissau", "248": "seychelles", "249": "sudan",
    "250": "rwanda", "251": "ethiopia", "252": "somalia", "253": "djibouti",
    "254": "kenya", "255": "tanzania", "256": "uganda", "257": "burundi",
    "258": "mozambique", "260": "zambia", "261": "madagascar", "263": "zimbabwe",
    "264": "namibia", "265": "malawi", "266": "lesotho", "267": "botswana",
    "268": "eswatini", "269": "comoros", "291": "eritrea",
    "297": "aruba", "299": "greenland",
    "350": "gibraltar", "351": "portugal", "352": "luxembourg", "353": "ireland",
    "354": "iceland", "355": "albania", "356": "malta", "357": "cyprus",
    "358": "finland", "359": "bulgaria", "370": "lithuania", "371": "latvia",
    "372": "estonia", "373": "moldova", "374": "armenia", "375": "belarus",
    "376": "andorra", "377": "monaco", "380": "ukraine", "381": "serbia",
    "382": "montenegro", "383": "kosovo", "385": "croatia", "386": "slovenia",
    "387": "bosnia and herzegovina", "389": "north macedonia",
    "420": "czech republic", "421": "slovakia", "423": "liechtenstein",
    "501": "belize", "502": "guatemala", "503": "el salvador", "504": "honduras",
    "505": "nicaragua", "506": "costa rica", "507": "panama", "509": "haiti",
    "591": "bolivia", "592": "guyana", "593": "ecuador", "595": "paraguay",
    "597": "suriname", "598": "uruguay",
    "670": "timor leste", "673": "brunei", "674": "nauru",
    "675": "papua new guinea", "676": "tonga", "678": "vanuatu", "679": "fiji",
    "680": "palau", "685": "samoa", "686": "kiribati", "688": "tuvalu",
    "691": "micronesia", "692": "marshall islands",
    "850": "north korea", "852": "hong kong", "853": "macau",
    "855": "cambodia", "856": "laos", "880": "bangladesh", "886": "taiwan",
    "960": "maldives", "961": "lebanon", "962": "jordan", "963": "syria",
    "964": "iraq", "965": "kuwait", "966": "saudi arabia", "967": "yemen",
    "968": "oman", "970": "palestine", "971": "uae", "972": "israel",
    "973": "bahrain", "974": "qatar", "975": "bhutan", "976": "mongolia",
    "977": "nepal", "992": "tajikistan", "993": "turkmenistan",
    "994": "azerbaijan", "995": "georgia", "996": "kyrgyzstan", "998": "uzbekistan",
}


def detect_country_from_phone(number):
    """Given a phone number, return (flag_emoji, country_name) by matching dialcode."""
    digits = str(number).lstrip("+").strip()
    for length in (3, 2, 1):
        code = digits[:length]
        country = PHONE_DIALCODES.get(code)
        if country:
            flag = COUNTRY_FLAGS.get(country, "🌍")
            return flag, country.title()
    return "🌍", ""


COUNTRY_FLAG_PREMIUM_IDS = {
    "🇦🇫": "5291937511591925566", "🇦🇱": "5294202819077756005", "🇩🇿": "5294048127240655242",
    "🇦🇸": "5291994273879709721", "🇦🇩": "5294215205763434181", "🇦🇴": "5294516785482062829",
    "🇦🇮": "5292186323342350940", "🇦🇬": "5294005972136647964", "🇦🇷": "5292208210495689627",
    "🇦🇲": "5291978717508164018", "🇦🇼": "5294007002928798927", "🇦🇺": "5294444247779399477",
    "🇦🇹": "5291975174160145850", "🇦🇿": "5294323533428579078", "🇧🇸": "5294031587321600012",
    "🇧🇭": "5294108398516720753", "🇧🇩": "5291824687096027834", "🇧🇧": "5294526187165471742",
    "🇧🇾": "5294134426018536120", "🇧🇪": "5291774466043435275", "🇧🇿": "5294171848068584842",
    "🇧🇯": "5293984969746566866", "🇧🇹": "5294121983498277263", "🇧🇴": "5294201479047957700",
    "🇧🇼": "5294026179957772585", "🇧🇷": "5291892229751723900", "🇧🇳": "5292098293692650297",
    "🇧🇬": "5294308947719640437", "🇧🇫": "5294153164960848949", "🇧🇮": "5294051631933967760",
    "🇰🇭": "5294225191562400452", "🇨🇲": "5291997306126626950", "🇨🇦": "5292290347450259214",
    "🇨🇻": "5292203503211535593", "🇨🇫": "5294210571493724819", "🇹🇩": "5291780728105753403",
    "🇨🇱": "5294231037012888049", "🇨🇳": "5294068833277990704", "🇨🇴": "5294010206974397371",
    "🇰🇲": "5294351381996521508", "🇨🇬": "5294035229453865597", "🇨🇰": "5292098684534675100",
    "🇨🇷": "5292063805105263554", "🇨🇮": "5293991322003200135", "🇭🇷": "5291999676948569127",
    "🇨🇺": "5291963947115631526", "🇨🇾": "5294062721539526918", "🇨🇿": "5294242852467923382",
    "🇩🇰": "5294531860817268837", "🇩🇯": "5294127214768468283", "🇩🇲": "5294485513825178032",
    "🇩🇴": "5294522197140857947", "🇪🇨": "5292083733753517221", "🇪🇬": "5293992082212409502",
    "🇸🇻": "5294337307388695687", "🇬🇶": "5292170045416297012", "🇪🇷": "5291922054004625949",
    "🇪🇪": "5291951143818123103", "🇪🇹": "5292245976143124155", "🇪🇺": "5291992809295861098",
    "🇬🇮": "5292055799286224027", "🇬🇲": "5294399820637688352", "🇬🇱": "5292014752283774878",
    "🇫🇮": "5294049961191690629", "🇫🇷": "5291817660529533837", "🇬🇦": "5294321325815389139",
    "🇬🇪": "5294349389131697267", "🇩🇪": "5292013274815028523", "🇬🇭": "5294347396266873249",
    "🇬🇷": "5291948395039054764", "🇬🇼": "5294409819321550432", "🇬🇹": "5294336633078831209",
    "🇬🇳": "5291892096607739008", "🇬🇾": "5292062692708736193", "🇭🇹": "5292045130587462814",
    "🇭🇳": "5291901034434682297", "🇭🇰": "5292166459118606932", "🇭🇺": "5294229581018975260",
    "🇮🇸": "5294354358408859664", "🇮🇳": "5291933173674957761", "🇮🇷": "5294220170745630736",
    "🇮🇶": "5294325010897327367", "🇮🇪": "5294471971793293647", "🇮🇱": "5294069056616289553",
    "🇮🇹": "5291826830284709120", "🇯🇲": "5294505107465982830", "🇯🇵": "5291799063321139445",
    "🇯🇴": "5291988613112814801", "🇰🇿": "5294227175837290463", "🇰🇪": "5292111852904416801",
    "🇰🇮": "5294538934628405146", "🇰🇵": "5294193812531333564", "🇰🇷": "5294408281723262763",
    "🇰🇼": "5292066437920218075", "🇰🇬": "5292091954320922577",
}


def _flag_premium_id(flag_char):
    """Return premium emoji ID for a flag emoji char, or empty string."""
    return COUNTRY_FLAG_PREMIUM_IDS.get(flag_char, "")


def _extract_prefix_digits(rng_val):
    """Extract leading digit prefix from a range like '22501XXX' → '22501'."""
    digits = ""
    for ch in str(rng_val).upper():
        if ch.isdigit():
            digits += ch
        else:
            break
    return digits


def _nsrv_gen_id(prefix):
    return prefix + "_" + uuid.uuid4().hex[:8]


def _nsrv_total_ranges(data):
    count = 0
    for srv in data.get("services_data", {}).values():
        for cnt in srv.get("countries", {}).values():
            count += len(cnt.get("ranges", {}))
    return count




# ============= PREMIUM EMOJIS =============
PREMIUM_EMOJIS = {
    # ── UI icons ──────────────────────────────────────────────────
    "🔵": "6282902529680347017",
    "❌": "5251420989144725682",
    "🔢": "5226513232549664618",
    "🔄": "5384466886857614542",
    "📸": "5258205968025525531",
    "📤": "5433614747381538714",
    "🚫": "5271476789904968419",
    "✅": "6199333289417773100",
    "📞": "6159019669815562830",
    "🔔": "6204010762206189094",
    "🛒": "6257905485010178374",
    "📱": "5251543627640894460",
    "🧾": "5204242830687494041",
    "💳": "6158821233736553062",
    "🌍": "5399898266265475100",
    "👑": "5253809746875542344",
    "💰": "6159157637050011672",
    "📨": "5406631276042002796",
    "🤖": "5258093637450866522",
    "👤": "5256143829672672750",
    "🔑": "5307843983102204243",
    "👋": "5472055112702629499",
    "📭": "5352896944496728039",
    "⚡": "6257850144356569423",
    "🔴": "5411225014148014586",
    "👇": "6061913594579128741",
    "🕐": "6278227444893749704",
    "🛑": "5240151566190656407",
    "📊": "6258267877170745815",
    "📩": "5319185587276619437",
    "😊": "6158872240768160988",
    "⏳": "5350427505805238170",
    "⚠️": "5420323339723881652",
    "🏠": "5321429707688801442",
    "➕": "6206041890895172990",
    "📧": "5456174900622412791",
    "✉️": "6206112371308500200",
    "📲": "5807750375033278838",
    "🔗": "6278397714577231283",
    "🎁": "6206027872121918710",
    "👥": "5249050854392091366",
    "💸": "5269531045165816230",
    "📋": "5460865451586250451",
    "▫️": "6158808005237281009",
    "😘": "6156703243628975744",
    "🌟": "6156604678424498789",
    "🔍": "5429571366384842791",
    "✈️": "6158743275785164209",
    "🌐": "6159148355625685932",
    "🕊": "6156440593493923335",
    "❓": "6159091967000057963",
    "🤑": "6258058785277873558",
    "♾": "6258154670422757760",
    "📹": "6258178962757784207",
    "⭐": "6258227852370514116",
    "🏁": "5222206157740847357",
    # ── Service icons (premium animated) ──────────────────────────
    "🟦": "5323261730283863478",   # Facebook
    "🟩": "5334998226636390258",   # WhatsApp
    "🔷": "5330237710655306682",   # Telegram
    "🟣": "5337155807752524558",   # Imo
    "🟠": "5319160079465857105",   # Instagram
    "🍏": "5334637951894722661",   # Apple
    "💌": "5348494358205207761",   # Email
    "🐦": "5330337435500951363",   # Twitter/X
    "📦": "5346056560537779652",   # Amazon
    # ── Country flags (premium animated) ─────────────────────────
    "🇦🇫": "5291937511591925566", "🇦🇽": "5294077418917616055",
    "🇦🇱": "5294202819077756005", "🇩🇿": "5294048127240655242",
    "🇦🇸": "5291994273879709721", "🇦🇩": "5294215205763434181",
    "🇦🇴": "5294516785482062829", "🇦🇮": "5292186323342350940",
    "🇦🇬": "5294005972136647964", "🇦🇷": "5292208210495689627",
    "🇦🇲": "5291978717508164018", "🇦🇼": "5294007002928798927",
    "🇦🇺": "5294444247779399477", "🇦🇹": "5291975174160145850",
    "🇦🇿": "5294323533428579078", "🇧🇸": "5294031587321600012",
    "🇧🇭": "5294108398516720753", "🇧🇩": "5291824687096027834",
    "🇧🇧": "5294526187165471742", "🇧🇾": "5294134426018536120",
    "🇧🇪": "5291774466043435275", "🇧🇿": "5294171848068584842",
    "🇧🇯": "5293984969746566866", "🇧🇹": "5294121983498277263",
    "🇧🇴": "5294201479047957700", "🇧🇼": "5294026179957772585",
    "🇧🇷": "5291892229751723900", "🇧🇳": "5292098293692650297",
    "🇧🇬": "5294308947719640437", "🇧🇫": "5294153164960848949",
    "🇧🇮": "5294051631933967760", "🇰🇭": "5294225191562400452",
    "🇨🇲": "5291997306126626950", "🇨🇦": "5292290347450259214",
    "🇨🇻": "5292203503211535593", "🇨🇫": "5294210571493724819",
    "🇹🇩": "5291780728105753403", "🇨🇱": "5294231037012888049",
    "🇨🇳": "5294068833277990704", "🇨🇴": "5294010206974397371",
    "🇰🇲": "5294351381996521508", "🇨🇬": "5294035229453865597",
    "🇨🇰": "5292098684534675100", "🇨🇷": "5292063805105263554",
    "🇨🇮": "5293991322003200135", "🇭🇷": "5291999676948569127",
    "🇨🇺": "5291963947115631526", "🇨🇾": "5294062721539526918",
    "🇨🇿": "5294242852467923382", "🇩🇰": "5294531860817268837",
    "🇩🇯": "5294127214768468283", "🇩🇲": "5294485513825178032",
    "🇩🇴": "5294522197140857947", "🇪🇨": "5292083733753517221",
    "🇪🇬": "5293992082212409502", "🇸🇻": "5294337307388695687",
    "🇬🇶": "5292170045416297012", "🇪🇷": "5291922054004625949",
    "🇪🇪": "5291951143818123103", "🇪🇹": "5292245976143124155",
    "🇪🇺": "5291992809295861098", "🇬🇮": "5292055799286224027",
    "🇬🇲": "5294399820637688352", "🇬🇱": "5292014752283774878",
    "🇫🇮": "5294049961191690629", "🇫🇷": "5291817660529533837",
    "🇬🇦": "5294321325815389139", "🇬🇪": "5294349389131697267",
    "🇩🇪": "5292013274815028523", "🇬🇭": "5294347396266873249",
    "🇬🇷": "5291948395039054764", "🇬🇼": "5294409819321550432",
    "🇬🇹": "5294336633078831209", "🇬🇳": "5291892096607739008",
    "🇬🇾": "5292062692708736193", "🇭🇹": "5292045130587462814",
    "🇭🇳": "5291901034434682297", "🇭🇰": "5292166459118606932",
    "🇭🇺": "5294229581018975260", "🇮🇸": "5294354358408859664",
    "🇮🇳": "5291933173674957761", "🇮🇷": "5294220170745630736",
    "🇮🇶": "5294325010897327367", "🇮🇪": "5294471971793293647",
    "🇮🇲": "5294318478252070646", "🇮🇱": "5294069056616289553",
    "🇮🇹": "5291826830284709120", "🇯🇲": "5294505107465982830",
    "🇯🇵": "5291799063321139445", "🇯🇪": "5291950280529697493",
    "🇯🇴": "5291988613112814801", "🇰🇿": "5294227175837290463",
    "🇰🇪": "5292111852904416801", "🇰🇮": "5294538934628405146",
    "🇰🇵": "5294193812531333564", "🇰🇷": "5294408281723262763",
    "🇰🇼": "5292066437920218075", "🇰🇬": "5292091954320922577",
    "🇲🇲": "5294254478944393569",
    "🇲🇬": "5291991568050312348",
    # ── Flags not in new list — kept from previous set ────────────
    "🇦🇪": "5224565851427976312", "🇧🇲": "5222482143749353810",
    "🇨🇭": "5224707263226194753", "🇪🇸": "5222024776976970940",
    "🇫🇲": "5222280486444873367", "🇬🇧": "5224518800061245598",
    "🇱🇨": "5222000927023577045", "🇱🇰": "5224277294050192388",
    "🇲🇦": "5224530035695693965", "🇲🇨": "5221937224068640464",
    "🇲🇩": "5224216473018314447", "🇲🇪": "5224463399278096980",
    "🇲🇭": "5224538449536624503", "🇲🇱": "5224322352552096671",
    "🇲🇳": "5224192257992701543", "🇲🇹": "5224731388057497620",
    "🇲🇺": "5224238347286752315", "🇲🇽": "5221971386238514431",
    "🇲🇿": "5222470388423864826", "🇳🇦": "5224690826386351746",
    "🇳🇪": "5222099049846420864", "🇳🇬": "5224723614166691638",
    "🇳🇱": "5224516489368841614", "🇳🇴": "5224465228934163949",
    "🇳🇵": "5222444378101925267", "🇳🇿": "5224573595254009705",
    "🇴🇲": "5222396686785066306", "🇵🇦": "5222111719999945107",
    "🇵🇪": "5224482026551258766", "🇵🇬": "5224500164198149905",
    "🇵🇭": "5222065042295376892", "🇵🇰": "5224637061985742245",
    "🇵🇱": "5224670399521892983", "🇵🇷": "5224220115150582423",
    "🇵🇸": "5222041677673282461", "🇵🇹": "5224404094369672274",
    "🇵🇾": "5222152565138929235", "🇶🇦": "5222225596762830469",
    "🇷🇴": "5222273794885826118", "🇷🇸": "5222145396838512729",
    "🇷🇺": "5280582975270963511", "🇷🇼": "5222449197055227754",
    "🇸🇦": "5224698145010624573", "🇸🇧": "5222290588207954120",
    "🇸🇨": "5224467496676896871", "🇸🇩": "5224372990216514135",
    "🇸🇪": "5222201098269373561", "🇸🇬": "5224194023224257181",
    "🇸🇮": "5224660718665607511", "🇸🇰": "5222401879400528047",
    "🇸🇱": "5224420995065983217", "🇸🇳": "5224358988623130949",
    "🇸🇴": "5222370504664428325", "🇸🇷": "5224567367551428669",
    "🇸🇸": "5224618146949773268", "🇸🇹": "5221953304426198315",
    "🇸🇿": "5224269666188274723", "🇹🇬": "5222408051268532030",
    "🇹🇭": "5224638530864556281", "🇹🇯": "5222217865821696536",
    "🇹🇲": "5224256935905208951", "🇹🇳": "5221991375016310330",
    "🇹🇷": "5224601903383457698", "🇹🇹": "5224391883777651050",
    "🇹🇿": "5224397364155923150", "🇺🇦": "5222250679371839695",
    "🇺🇬": "5222464040462200940", "🇺🇳": "5451772687993031127",
    "🇺🇸": "5224321781321442532", "🇺🇾": "5222466849370813232",
    "🇺🇿": "5222404546575219535", "🇻🇦": "5222420266155520507",
    "🇻🇨": "5224541228380467535", "🇻🇮": "5224395882392201810",
    "🇻🇳": "5222359651282071925", "🇻🇺": "5222126748090512778",
    "🇼🇸": "5224660353593387686", "🇾🇪": "5222300655611294950",
    "🇿🇦": "5224696216570309138", "🇿🇲": "5224646626877911277",
}

# Sorted by length (longest first) for greedy matching
_EMOJI_KEYS = sorted(PREMIUM_EMOJIS.keys(), key=len, reverse=True)

# ── Console forwarder: service emoji IDs (placeholder_char → custom_emoji_id) ──
CONSOLE_SERVICE_EMOJIS = {
    "facebook":  ("🟦", "5323261730283863478"),
    "fb":        ("🟦", "5323261730283863478"),
    "whatsapp":  ("🟩", "5334998226636390258"),
    "wa":        ("🟩", "5334998226636390258"),
    "telegram":  ("🔷", "5330237710655306682"),
    "tg":        ("🔷", "5330237710655306682"),
    "imo":       ("🟣", "5337155807752524558"),
    "instagram": ("🟠", "5319160079465857105"),
    "ig":        ("🟠", "5319160079465857105"),
    "apple":     ("🍏", "5334637951894722661"),
    "email":     ("💌", "5348494358205207761"),
    "gmail":     ("💌", "5348494358205207761"),
    "mail":      ("💌", "5348494358205207761"),
    "twitter":   ("🐦", "5330337435500951363"),
    "x":         ("🐦", "5330337435500951363"),
    "amazon":    ("📦", "5346056560537779652"),
}


# Number prefix → (flag_emoji, 2-letter-CC) — longest match wins
_CONSOLE_PREFIX_MAP = {
    # ── NANP 4-digit Caribbean / US territories (must be before "1") ──
    "1939":  ("🇵🇷", "PR"), "1787":  ("🇵🇷", "PR"),
    "1849":  ("🇩🇴", "DO"), "1829":  ("🇩🇴", "DO"), "1809":  ("🇩🇴", "DO"),
    "1876":  ("🇯🇲", "JM"), "1868":  ("🇹🇹", "TT"), "1869":  ("🇰🇳", "KN"),
    "1284":  ("🇻🇬", "VG"), "1340":  ("🇻🇮", "VI"), "1345":  ("🇰🇾", "KY"),
    "1441":  ("🇧🇲", "BM"), "1473":  ("🇬🇩", "GD"), "1649":  ("🇹🇨", "TC"),
    "1664":  ("🇲🇸", "MS"), "1670":  ("🇲🇵", "MP"), "1671":  ("🇬🇺", "GU"),
    "1684":  ("🇦🇸", "AS"), "1721":  ("🇸🇽", "SX"), "1758":  ("🇱🇨", "LC"),
    "1767":  ("🇩🇲", "DM"), "1784":  ("🇻🇨", "VC"), "1246":  ("🇧🇧", "BB"),
    "1242":  ("🇧🇸", "BS"), "1264":  ("🇦🇮", "AI"), "1268":  ("🇦🇬", "AG"),
    # ── Pakistan special prefixes (before "92") ──
    "22501": ("🇵🇰", "PK"), "22506": ("🇵🇰", "PK"),
    # ── Zone 1 ──
    "1":     ("🇺🇸", "US"),
    # ── Zone 2 — Africa ──
    "20":    ("🇪🇬", "EG"),
    "211":   ("🇸🇸", "SS"),
    "212":   ("🇲🇦", "MA"),
    "213":   ("🇩🇿", "DZ"),
    "216":   ("🇹🇳", "TN"),
    "218":   ("🇱🇾", "LY"),
    "220":   ("🇬🇲", "GM"),
    "221":   ("🇸🇳", "SN"),
    "222":   ("🇲🇷", "MR"),
    "223":   ("🇲🇱", "ML"),
    "224":   ("🇬🇳", "GN"),
    "225":   ("🇨🇮", "CI"),
    "226":   ("🇧🇫", "BF"),
    "227":   ("🇳🇪", "NE"),
    "228":   ("🇹🇬", "TG"),
    "229":   ("🇧🇯", "BJ"),
    "230":   ("🇲🇺", "MU"),
    "231":   ("🇱🇷", "LR"),
    "232":   ("🇸🇱", "SL"),
    "233":   ("🇬🇭", "GH"),
    "234":   ("🇳🇬", "NG"),
    "235":   ("🇹🇩", "TD"),
    "236":   ("🇨🇫", "CF"),
    "237":   ("🇨🇲", "CM"),
    "238":   ("🇨🇻", "CV"),
    "239":   ("🇸🇹", "ST"),
    "240":   ("🇬🇶", "GQ"),
    "241":   ("🇬🇦", "GA"),
    "242":   ("🇨🇬", "CG"),
    "243":   ("🇨🇩", "CD"),
    "244":   ("🇦🇴", "AO"),
    "245":   ("🇬🇼", "GW"),
    "246":   ("🇮🇴", "IO"),
    "247":   ("🇦🇨", "AC"),
    "248":   ("🇸🇨", "SC"),
    "249":   ("🇸🇩", "SD"),
    "250":   ("🇷🇼", "RW"),
    "251":   ("🇪🇹", "ET"),
    "252":   ("🇸🇴", "SO"),
    "253":   ("🇩🇯", "DJ"),
    "254":   ("🇰🇪", "KE"),
    "255":   ("🇹🇿", "TZ"),
    "256":   ("🇺🇬", "UG"),
    "257":   ("🇧🇮", "BI"),
    "258":   ("🇲🇿", "MZ"),
    "260":   ("🇿🇲", "ZM"),
    "261":   ("🇲🇬", "MG"),
    "262":   ("🇷🇪", "RE"),
    "263":   ("🇿🇼", "ZW"),
    "264":   ("🇳🇦", "NA"),
    "265":   ("🇲🇼", "MW"),
    "266":   ("🇱🇸", "LS"),
    "267":   ("🇧🇼", "BW"),
    "268":   ("🇸🇿", "SZ"),
    "269":   ("🇰🇲", "KM"),
    "27":    ("🇿🇦", "ZA"),
    "290":   ("🇸🇭", "SH"),
    "291":   ("🇪🇷", "ER"),
    "297":   ("🇦🇼", "AW"),
    "298":   ("🇫🇴", "FO"),
    "299":   ("🇬🇱", "GL"),
    # ── Zone 3 — Europe ──
    "30":    ("🇬🇷", "GR"),
    "31":    ("🇳🇱", "NL"),
    "32":    ("🇧🇪", "BE"),
    "33":    ("🇫🇷", "FR"),
    "34":    ("🇪🇸", "ES"),
    "350":   ("🇬🇮", "GI"),
    "351":   ("🇵🇹", "PT"),
    "352":   ("🇱🇺", "LU"),
    "353":   ("🇮🇪", "IE"),
    "354":   ("🇮🇸", "IS"),
    "355":   ("🇦🇱", "AL"),
    "356":   ("🇲🇹", "MT"),
    "357":   ("🇨🇾", "CY"),
    "358":   ("🇫🇮", "FI"),
    "359":   ("🇧🇬", "BG"),
    "36":    ("🇭🇺", "HU"),
    "370":   ("🇱🇹", "LT"),
    "371":   ("🇱🇻", "LV"),
    "372":   ("🇪🇪", "EE"),
    "373":   ("🇲🇩", "MD"),
    "374":   ("🇦🇲", "AM"),
    "375":   ("🇧🇾", "BY"),
    "376":   ("🇦🇩", "AD"),
    "377":   ("🇲🇨", "MC"),
    "378":   ("🇸🇲", "SM"),
    "380":   ("🇺🇦", "UA"),
    "381":   ("🇷🇸", "RS"),
    "382":   ("🇲🇪", "ME"),
    "383":   ("🇽🇰", "XK"),
    "385":   ("🇭🇷", "HR"),
    "386":   ("🇸🇮", "SI"),
    "387":   ("🇧🇦", "BA"),
    "389":   ("🇲🇰", "MK"),
    "39":    ("🇮🇹", "IT"),
    "40":    ("🇷🇴", "RO"),
    "41":    ("🇨🇭", "CH"),
    "420":   ("🇨🇿", "CZ"),
    "421":   ("🇸🇰", "SK"),
    "423":   ("🇱🇮", "LI"),
    "43":    ("🇦🇹", "AT"),
    "44":    ("🇬🇧", "GB"),
    "45":    ("🇩🇰", "DK"),
    "46":    ("🇸🇪", "SE"),
    "47":    ("🇳🇴", "NO"),
    "48":    ("🇵🇱", "PL"),
    "49":    ("🇩🇪", "DE"),
    # ── Zone 5 — Central/South America ──
    "500":   ("🇫🇰", "FK"),
    "501":   ("🇧🇿", "BZ"),
    "502":   ("🇬🇹", "GT"),
    "503":   ("🇸🇻", "SV"),
    "504":   ("🇭🇳", "HN"),
    "505":   ("🇳🇮", "NI"),
    "506":   ("🇨🇷", "CR"),
    "507":   ("🇵🇦", "PA"),
    "508":   ("🇵🇲", "PM"),
    "509":   ("🇭🇹", "HT"),
    "51":    ("🇵🇪", "PE"),
    "52":    ("🇲🇽", "MX"),
    "53":    ("🇨🇺", "CU"),
    "54":    ("🇦🇷", "AR"),
    "55":    ("🇧🇷", "BR"),
    "56":    ("🇨🇱", "CL"),
    "57":    ("🇨🇴", "CO"),
    "58":    ("🇻🇪", "VE"),
    "590":   ("🇬🇵", "GP"),
    "591":   ("🇧🇴", "BO"),
    "592":   ("🇬🇾", "GY"),
    "593":   ("🇪🇨", "EC"),
    "594":   ("🇬🇫", "GF"),
    "595":   ("🇵🇾", "PY"),
    "596":   ("🇲🇶", "MQ"),
    "597":   ("🇸🇷", "SR"),
    "598":   ("🇺🇾", "UY"),
    "599":   ("🇨🇼", "CW"),
    # ── Zone 6 — Southeast Asia / Pacific ──
    "60":    ("🇲🇾", "MY"),
    "61":    ("🇦🇺", "AU"),
    "62":    ("🇮🇩", "ID"),
    "63":    ("🇵🇭", "PH"),
    "64":    ("🇳🇿", "NZ"),
    "65":    ("🇸🇬", "SG"),
    "66":    ("🇹🇭", "TH"),
    "670":   ("🇹🇱", "TL"),
    "672":   ("🇳🇫", "NF"),
    "673":   ("🇧🇳", "BN"),
    "674":   ("🇳🇷", "NR"),
    "675":   ("🇵🇬", "PG"),
    "676":   ("🇹🇴", "TO"),
    "677":   ("🇸🇧", "SB"),
    "678":   ("🇻🇺", "VU"),
    "679":   ("🇫🇯", "FJ"),
    "680":   ("🇵🇼", "PW"),
    "681":   ("🇼🇫", "WF"),
    "682":   ("🇨🇰", "CK"),
    "683":   ("🇳🇺", "NU"),
    "685":   ("🇼🇸", "WS"),
    "686":   ("🇰🇮", "KI"),
    "687":   ("🇳🇨", "NC"),
    "688":   ("🇹🇻", "TV"),
    "689":   ("🇵🇫", "PF"),
    "690":   ("🇹🇰", "TK"),
    "691":   ("🇫🇲", "FM"),
    "692":   ("🇲🇭", "MH"),
    # ── Zone 7 — Russia / CIS ──
    "76":    ("🇰🇿", "KZ"),
    "77":    ("🇰🇿", "KZ"),
    "7":     ("🇷🇺", "RU"),
    # ── Zone 8 — East Asia ──
    "81":    ("🇯🇵", "JP"),
    "82":    ("🇰🇷", "KR"),
    "84":    ("🇻🇳", "VN"),
    "850":   ("🇰🇵", "KP"),
    "852":   ("🇭🇰", "HK"),
    "853":   ("🇲🇴", "MO"),
    "855":   ("🇰🇭", "KH"),
    "856":   ("🇱🇦", "LA"),
    "86":    ("🇨🇳", "CN"),
    "880":   ("🇧🇩", "BD"),
    "886":   ("🇹🇼", "TW"),
    # ── Zone 9 — South/Central Asia & Middle East ──
    "90":    ("🇹🇷", "TR"),
    "91":    ("🇮🇳", "IN"),
    "92":    ("🇵🇰", "PK"),
    "93":    ("🇦🇫", "AF"),
    "94":    ("🇱🇰", "LK"),
    "95":    ("🇲🇲", "MM"),
    "960":   ("🇲🇻", "MV"),
    "961":   ("🇱🇧", "LB"),
    "962":   ("🇯🇴", "JO"),
    "963":   ("🇸🇾", "SY"),
    "964":   ("🇮🇶", "IQ"),
    "965":   ("🇰🇼", "KW"),
    "966":   ("🇸🇦", "SA"),
    "967":   ("🇾🇪", "YE"),
    "968":   ("🇴🇲", "OM"),
    "970":   ("🇵🇸", "PS"),
    "971":   ("🇦🇪", "AE"),
    "972":   ("🇮🇱", "IL"),
    "973":   ("🇧🇭", "BH"),
    "974":   ("🇶🇦", "QA"),
    "975":   ("🇧🇹", "BT"),
    "976":   ("🇲🇳", "MN"),
    "977":   ("🇳🇵", "NP"),
    "98":    ("🇮🇷", "IR"),
    "992":   ("🇹🇯", "TJ"),
    "993":   ("🇹🇲", "TM"),
    "994":   ("🇦🇿", "AZ"),
    "995":   ("🇬🇪", "GE"),
    "996":   ("🇰🇬", "KG"),
    "998":   ("🇺🇿", "UZ"),
}
_CONSOLE_PREFIX_KEYS = sorted(_CONSOLE_PREFIX_MAP.keys(), key=len, reverse=True)


def detect_message_language(text):
    """Detect language of an SMS from script blocks and keyword patterns."""
    for ch in text:
        o = ord(ch)
        if 0x0980 <= o <= 0x09FF:
            return "Bengali"
        if 0x0900 <= o <= 0x097F:
            return "Hindi"
        if 0x0600 <= o <= 0x06FF:
            return "Arabic"
        if 0x4E00 <= o <= 0x9FFF:
            return "Chinese"
        if 0x0400 <= o <= 0x04FF:
            return "Russian"
        if 0x0E00 <= o <= 0x0E7F:
            return "Thai"
    # Keyword-based detection for Latin-script languages
    lower = text.lower()
    words = set(re.findall(r"[a-z'àâéèêëîïôùûüç]+", lower))
    _FRENCH_WORDS  = {"partagez","votre","vous","pas","est","les","des","avec","pour",
                      "ne","ce","cette","partager","valide","entrez","cliquez",
                      "bonjour","merci","notre","compte","sécurité","expirer","partage",
                      "confirmez","connectez","saisissez","vérification","réinitialisation"}
    _SPANISH_WORDS = {"tu","código","es","no","compartir","este","con","para","por","una",
                      "hola","gracias","cuenta","seguridad","comparte","verificación"}
    _PORTUGUESE_WORDS = {"seu","sua","não","para","código","com","uma","conta","segurança",
                         "compartilhe","obrigado","ola","brasil","verificação","validade"}
    _INDONESIAN_WORDS = {"kode","jangan","bagikan","anda","adalah","dengan","untuk","tidak",
                         "atau","ini","dari","akan","pada","keamanan","verifikasi"}
    scores = {
        "French":     len(words & _FRENCH_WORDS),
        "Spanish":    len(words & _SPANISH_WORDS),
        "Portuguese": len(words & _PORTUGUESE_WORDS),
        "Indonesian": len(words & _INDONESIAN_WORDS),
    }
    best_lang, best_score = max(scores.items(), key=lambda x: x[1])
    if best_score >= 1:
        return best_lang
    return "English"


def extract_otp_from_message(msg):
    """Return a nicely formatted OTP code from an SMS, or None.
    Handles: plain digits, space-separated (e.g. '536 718'), hyphen-separated (e.g. '742-998').
    """
    import re as _re
    # First: already hyphen-separated groups (e.g. "801-046", "1234-5678")
    hyphenated = _re.search(r'\b(\d{2,4})-(\d{2,4})\b', msg)
    if hyphenated:
        code = hyphenated.group(1) + hyphenated.group(2)
        if 4 <= len(code) <= 8:
            if len(code) == 6:
                return f"{code[:3]}-{code[3:]}"
            if len(code) == 8:
                return f"{code[:4]}-{code[4:]}"
            return code
    # Second: space-separated groups totalling 4-8 digits (e.g. "536 718", "237 086", "541 936")
    spaced = _re.search(r'\b(\d{2,4})\s(\d{2,4})\b', msg)
    if spaced:
        code = spaced.group(1) + spaced.group(2)
        if 4 <= len(code) <= 8:
            if len(code) == 6:
                return f"{code[:3]}-{code[3:]}"
            if len(code) == 8:
                return f"{code[:4]}-{code[4:]}"
            return code
    # Third: contiguous digits 4-8
    codes = _re.findall(r'\b(\d{4,8})\b', msg)
    if not codes:
        return None
    code = max(codes, key=len)
    if len(code) == 6:
        return f"{code[:3]}-{code[3:]}"
    if len(code) == 8:
        return f"{code[:4]}-{code[4:]}"
    return code


_SMS_SERVICE_PATTERNS = [
    ("instagram",  re.compile(r'instagram', re.I)),
    ("whatsapp",   re.compile(r'whatsapp|whats app', re.I)),
    ("telegram",   re.compile(r'telegram', re.I)),
    ("facebook",   re.compile(r'facebook|fb', re.I)),
    ("apple",      re.compile(r'apple|icloud|appleid', re.I)),
    ("imo",        re.compile(r'\bimo\b', re.I)),
    ("email",      re.compile(r'\bgmail\b|\byahoo\b|\boutlook\b|\bemail\b', re.I)),
]

def detect_service_from_sms(sid, msg):
    """Detect real service from SMS text; fall back to sid if no match."""
    for svc_key, pat in _SMS_SERVICE_PATTERNS:
        if pat.search(msg):
            return svc_key
    return (sid or "").lower().strip()


def get_country_from_prefix(num_str):
    """Return (flag_emoji, 2-letter-CC) from a number prefix string."""
    n = str(num_str).lstrip("+").replace(" ", "")
    for pref in _CONSOLE_PREFIX_KEYS:
        if n.startswith(pref):
            return _CONSOLE_PREFIX_MAP[pref]
    return ("🌍", "XX")


def _utf16_len(s):
    return sum(2 if ord(c) > 0xFFFF else 1 for c in s)


def _build_entities(text, emoji_overrides=None):
    """
    Parse Markdown-formatted text with premium emojis.
    Returns (plain_text, entities_list) for Telegram sendMessage entities param.
    Handles: *bold*, _italic_, `code`, and premium custom emojis.
    emoji_overrides: dict of {emoji_char: emoji_id} — overrides PREMIUM_EMOJIS for specific chars.
    """
    plain = []
    entities = []
    offset = [0]
    _overrides = emoji_overrides or {}
    _ov_keys   = sorted(_overrides.keys(), key=len, reverse=True)

    def scan(src):
        i = 0
        while i < len(src):
            # Bold *text*
            if src[i] == '*':
                j = src.find('*', i + 1)
                if j != -1:
                    start = offset[0]
                    scan(src[i + 1:j])
                    length = offset[0] - start
                    if length > 0:
                        entities.append({'type': 'bold', 'offset': start, 'length': length})
                    i = j + 1
                    continue
            # Italic _text_
            if src[i] == '_':
                j = src.find('_', i + 1)
                if j != -1:
                    start = offset[0]
                    scan(src[i + 1:j])
                    length = offset[0] - start
                    if length > 0:
                        entities.append({'type': 'italic', 'offset': start, 'length': length})
                    i = j + 1
                    continue
            # Code `text`
            if src[i] == '`':
                j = src.find('`', i + 1)
                if j != -1:
                    inner = src[i + 1:j]
                    start = offset[0]
                    ilen = _utf16_len(inner)
                    entities.append({'type': 'code', 'offset': start, 'length': ilen})
                    plain.append(inner)
                    offset[0] += ilen
                    i = j + 1
                    continue
            # Premium emoji — overrides first (can introduce new emoji IDs), then PREMIUM_EMOJIS
            matched     = None
            matched_eid = None
            for em in _ov_keys:
                if src[i:i + len(em)] == em:
                    matched     = em
                    matched_eid = _overrides[em]
                    break
            if not matched:
                for em in _EMOJI_KEYS:
                    if src[i:i + len(em)] == em:
                        matched     = em
                        matched_eid = _overrides.get(em) or PREMIUM_EMOJIS[em]
                        break
            if matched:
                elen = _utf16_len(matched)
                entities.append({
                    'type': 'custom_emoji',
                    'offset': offset[0],
                    'length': elen,
                    'custom_emoji_id': matched_eid
                })
                plain.append(matched)
                offset[0] += elen
                i += len(matched)
                continue
            # Regular character
            ch = src[i]
            plain.append(ch)
            offset[0] += _utf16_len(ch)
            i += 1

    scan(text)
    return ''.join(plain), entities


DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bot_data.json")

default_data = {
    "countries": [
        {
            "id": "36",
            "name": "🇨🇦 Canada",
            "services": []
        }
    ],
    "users":         {},
    "mail_services": []
}


# ============= DATA =============

import copy

_data_lock    = threading.Lock()   # protects _data_cache (in-memory only)
_disk_lock    = threading.Lock()   # serialises disk writes (held only during I/O)
_data_cache   = None               # live dict; callers get deepcopy — no JSON round-trip
_save_count   = 0                  # backup every N saves
_search_index = None               # pre-built service search index; rebuilt on save
_search_lock  = threading.Lock()   # protects _search_index

# ── Async write-behind queue ──────────────────────────────────────────────
# save_data() no longer blocks on disk I/O.  It puts the latest snapshot
# onto this queue and returns immediately.  A single background thread
# drains the queue with a short debounce (writes the LATEST pending
# snapshot every 0.5 s at most), so rapid bursts of saves collapse into
# one disk write instead of dozens.
_write_queue = queue.Queue()   # items: dict (the data snapshot to persist)

def _disk_writer_thread():
    """Background thread: drains _write_queue and persists to disk."""
    global _save_count
    while True:
        try:
            snapshot = _write_queue.get(block=True, timeout=1)
            # Drain any newer snapshots that arrived while we were waiting —
            # only the freshest one needs to be written.
            while True:
                try:
                    snapshot = _write_queue.get_nowait()
                except queue.Empty:
                    break
            # Now write the freshest snapshot to disk.
            with _disk_lock:
                try:
                    with _data_lock:
                        _save_count += 1
                        cur_count = _save_count
                    serialized = json.dumps(snapshot, ensure_ascii=False, separators=(',', ':'))
                    tmp = DATA_FILE + ".tmp"
                    with open(tmp, 'w', encoding='utf-8') as f:
                        f.write(serialized)
                    if cur_count % 20 == 1 and os.path.exists(DATA_FILE):
                        try:
                            shutil.copy2(DATA_FILE, DATA_FILE + ".bak")
                        except Exception:
                            pass
                    os.replace(tmp, DATA_FILE)
                except Exception as e:
                    print(f"[Writer] Save error: {e}")
        except queue.Empty:
            continue
        except Exception as e:
            print(f"[Writer] Thread error: {e}")

_writer_thread = threading.Thread(target=_disk_writer_thread, daemon=True, name="disk-writer")
_writer_thread.start()

# ── Force-join membership cache (TTL = 60 s) ─────────────
_join_cache      = {}   # {user_id: (checked_at_float, unjoined_list)}
_join_cache_lock = threading.Lock()
_JOIN_CACHE_TTL  = 60   # seconds


def _clear_join_cache(user_id):
    with _join_cache_lock:
        _join_cache.pop(user_id, None)


def get_unjoined_channels_cached(user_id, data):
    """Cached wrapper around get_unjoined_channels — re-checks at most every 60 s."""
    now = time.time()
    with _join_cache_lock:
        entry = _join_cache.get(user_id)
        if entry and (now - entry[0]) < _JOIN_CACHE_TTL:
            return entry[1]
    # Cache miss or expired — do the real (slow) HTTP check
    result = get_unjoined_channels(user_id, data)
    with _join_cache_lock:
        _join_cache[user_id] = (time.time(), result)
    return result


def _apply_data_defaults(data):
    """Apply all setdefault / migration logic to a freshly parsed data dict."""
    data.setdefault("countries", [])
    data.setdefault("s3_countries", [])
    data.setdefault("s4_countries", [])
    data.setdefault("mail_services", [])
    data.setdefault("users", {})
    data.setdefault("gift_codes", {})
    data.setdefault("referrals", {})
    data.setdefault("services_data", {})
    ds = data.setdefault("deposit_settings", {})
    ds.setdefault("upi_id",      UPI_ID)
    ds.setdefault("qr_url",      "https://t.me/KP_MODZ_2/55")
    ds.setdefault("min_deposit", MIN_DEPOSIT)
    ds.setdefault("enabled",     True)
    for c in data["countries"]:
        c.setdefault("services", [])
        for s in c["services"]:
            s.setdefault("price", 0)
            if s.get("api") in ("smsbower", None, ""):
                s["api"] = "s1"
            elif s.get("api") == "herosms":
                s["api"] = "s2"
    for c in data["s3_countries"]:
        c.setdefault("services", [])
        for s in c["services"]:
            s.setdefault("price", 0)
            s.setdefault("operator", "any")
    for c in data["s4_countries"]:
        c.setdefault("services", [])
        for s in c["services"]:
            s.setdefault("price", 0)
            s.setdefault("operator", "any")
    # Load saved API keys into APIS dict
    saved_keys = data.get("api_keys", {})
    for api_id, key in saved_keys.items():
        if api_id in APIS and key:
            APIS[api_id]["key"] = key
    # Load custom picker emojis saved by admin
    data.setdefault("custom_picker_emojis", [])
    for cid in data["custom_picker_emojis"]:
        if cid not in PICKER_ID_TO_CHAR:
            PICKER_EMOJIS_DATA.append(("●", cid))
            PICKER_ID_TO_CHAR[cid] = "●"
    return data


def load_data():
    global _data_cache
    with _data_lock:
        # ── Fast path: deepcopy the in-memory dict (no JSON round-trip) ──
        if _data_cache is not None:
            return copy.deepcopy(_data_cache)
    # ── Slow path: read from disk (no lock held during I/O) ───────────
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                raw = f.read().strip()
            if not raw:
                raise ValueError("Empty file")
            data = json.loads(raw)
            _apply_data_defaults(data)
        else:
            data = copy.deepcopy(default_data)
            with _disk_lock:
                with open(DATA_FILE, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
        with _data_lock:
            if _data_cache is None:   # only set if nobody else won the race
                _data_cache = copy.deepcopy(data)
        return data
    except Exception as e:
        print(f"Load error: {e}")
        bak = DATA_FILE + ".bak"
        if os.path.exists(bak):
            try:
                with open(bak, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                print("Restored from backup.")
                _apply_data_defaults(data)
                with _data_lock:
                    _data_cache = copy.deepcopy(data)
                return data
            except Exception:
                pass
        return copy.deepcopy(default_data)


def save_data(data):
    global _data_cache
    # 1. Update in-memory cache immediately — all subsequent load_data() calls
    #    in any thread will see the new data without touching disk at all.
    snapshot = copy.deepcopy(data)
    with _data_lock:
        _data_cache = snapshot
    _invalidate_search_index()   # force rebuild on next search
    # 2. Enqueue for async disk write — returns instantly, never blocks caller.
    #    The background _disk_writer_thread will persist the latest snapshot.
    try:
        _write_queue.put_nowait(copy.deepcopy(snapshot))
    except Exception as e:
        print(f"Save enqueue error: {e}")
    return True


def get_user(data, user_id):
    uid = str(user_id)
    if uid not in data["users"]:
        data["users"][uid] = {"balance": 0, "history": []}
    u = data["users"][uid]
    # Ensure balance is always a float — guards against corrupt JSON or None
    try:
        u["balance"] = float(u.get("balance") or 0)
    except (TypeError, ValueError):
        u["balance"] = 0.0
    return u


# ============= API CALLS =============

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/json,*/*",
}


def http_get(url, timeout=12):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode().strip()


def call_api(action, service="", country="", extra_params=None, api="s1"):
    cfg = APIS.get(api, APIS["s1"])
    params = f"api_key={cfg['key']}&action={action}"
    if service:
        params += f"&service={service}"
    if country:
        params += f"&country={country}"
    if extra_params:
        for k, v in extra_params.items():
            params += f"&{k}={v}"
    url = f"{cfg['url']}?{params}"
    try:
        return http_get(url, timeout=20)
    except Exception as e:
        return f"ERROR: {str(e)}"


MAIL_API_BASE = "https://smsbower.page/api/mail"

def call_mail_api(endpoint, **params):
    """Call SmsBower mail API. Returns parsed JSON dict or {"error": ...}."""
    params["api_key"] = APIS["s1"]["key"]
    qs = "&".join(f"{k}={v}" for k, v in params.items())
    url = f"{MAIL_API_BASE}/{endpoint}?{qs}"
    try:
        raw = http_get(url, timeout=20)
        try:
            return json.loads(raw)
        except Exception:
            return {"raw": raw}
    except Exception as e:
        return {"error": str(e)}


def get_live_prices(country_id, api="s1"):
    cfg = APIS.get(api, APIS["s1"])
    url = f"{cfg['url']}?api_key={cfg['key']}&action=getPricesV3&country={country_id}"
    try:
        raw  = http_get(url, timeout=15)
        if not raw:
            return {}
        full = json.loads(raw)
        country_data = full.get(str(country_id), {})
        result = {}
        for svc_id, providers in country_data.items():
            if not isinstance(providers, dict):
                continue
            cheapest    = None
            total_count = 0
            for prov in providers.values():
                price = prov.get("price", 0)
                count = prov.get("count", 0)
                total_count += count
                if cheapest is None or price < cheapest:
                    cheapest = price
            if cheapest is not None:
                result[svc_id] = {"cost": cheapest, "count": total_count}
        return result
    except Exception as e:
        try:
            url2 = f"{cfg['url']}?api_key={cfg['key']}&action=getPrices&country={country_id}"
            raw2  = http_get(url2, timeout=15)
            full2 = json.loads(raw2)
            country_data2 = full2.get(str(country_id), {})
            result2 = {}
            for svc_id, info in country_data2.items():
                if isinstance(info, dict):
                    result2[svc_id] = {"cost": info.get("cost", 0), "count": info.get("count", 0)}
                elif isinstance(info, (int, float)):
                    result2[svc_id] = {"cost": info, "count": 0}
            return result2
        except Exception as e2:
            print(f"Live price error ({api}): {e} / {e2}")
            return {}


# ============= TELEGRAM =============

def get_updates(offset=None, timeout=25):
    params = f"?timeout={timeout}"
    if offset:
        params += f"&offset={offset}"
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates{params}"
    try:
        with urllib.request.urlopen(url, timeout=timeout + 5) as resp:
            return json.loads(resp.read().decode()).get("result", [])
    except Exception:
        return []


def _parse_channel(ch_str):
    """Parse '@channel/329' or '-100.../329' into (chat_id, thread_id_or_None)."""
    s = str(ch_str).strip()
    if "/" in s:
        parts = s.rsplit("/", 1)
        try:
            return parts[0].strip(), int(parts[1].strip())
        except (ValueError, IndexError):
            pass
    return s, None


_tg_retry_until = 0   # global: epoch time until which we must not send


def send_message(chat_id, text, reply_markup=None, emoji_overrides=None):
    global _tg_retry_until
    # Telegram hard-limits message text to 4096 chars
    if len(text) > 4000:
        text = text[:4000] + "\n…"
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    plain, ents = _build_entities(text, emoji_overrides)
    parsed_id, thread_id = _parse_channel(chat_id)
    d = {"chat_id": parsed_id, "text": plain}
    if thread_id is not None:
        d["message_thread_id"] = thread_id
    if ents:
        d["entities"] = json.dumps(ents)
    if reply_markup:
        d["reply_markup"] = json.dumps(reply_markup)
    wait = _tg_retry_until - time.time()
    if wait > 0:
        time.sleep(min(wait, 30))
    try:
        req = urllib.request.Request(url, urllib.parse.urlencode(d).encode(), headers=HEADERS)
        resp = urllib.request.urlopen(req, timeout=10)
        result = json.loads(resp.read().decode())
        _r = result.get("result")
        return _r.get("message_id") if isinstance(_r, dict) else None
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        if e.code == 429:
            try:
                retry_after = json.loads(body).get("parameters", {}).get("retry_after", 30)
            except Exception:
                retry_after = 30
            _tg_retry_until = time.time() + retry_after
        else:
            print(f"Send error ({e.code}): {body[:300]}")
        return None
    except Exception as e:
        print(f"Send error: {e}")
        return None


def _forward_sms_to_channel(data, lines_text):
    """Forward a plain SMS block to the configured live SMS channel (legacy fallback)."""
    ch = data.get("tnevs_settings", {}).get("sms_channel", "").strip()
    if not ch:
        return
    try:
        send_message(ch, lines_text)
    except Exception as e:
        print(f"Channel forward error: {e}")


def forward_console_hit_to_channel(data, sid, rng, msg):
    """Forward a single console OTP hit to the channel with rich seven1tel-style design."""
    cfg = data.get("tnevs_settings", {})
    ch_all = cfg.get("sms_channel_all", "").strip()
    # Detect service for per-service channel routing
    svc_key_for_ch = detect_service_from_sms(sid, msg)
    ch_specific = ""
    if svc_key_for_ch in ("instagram", "ig"):
        ch_specific = cfg.get("sms_channel_ig", "").strip()
    elif svc_key_for_ch in ("facebook", "fb"):
        ch_specific = cfg.get("sms_channel_fb", "").strip()
    elif svc_key_for_ch in ("whatsapp", "wa"):
        ch_specific = cfg.get("sms_channel_wa", "").strip()
    channels = list(filter(None, dict.fromkeys([ch_all, ch_specific])))
    if not channels:
        return
    ch = channels[0]  # primary channel used below; we'll send to all after building markup
    # Service emoji — detect from SMS content first, fall back to sid
    svc_key   = detect_service_from_sms(sid, msg)
    svc_info  = CONSOLE_SERVICE_EMOJIS.get(svc_key)
    svc_ph    = svc_info[0] if svc_info else "📡"
    svc_eid   = svc_info[1] if svc_info else None
    svc_name  = (sid or "").upper()[:12]
    # Country
    flag, cc  = get_country_from_prefix(rng)
    # Language
    lang      = detect_message_language(msg)
    # OTP extraction
    otp       = extract_otp_from_message(msg)
    # Number display (prefix + XXX)
    num_disp  = f"{rng}XXX" if rng and not rng.endswith("XXX") else (rng or "???")
    # Header: flag CC  •  ServiceEmoji num  •  Language  (num in code entity = tap to copy)
    header = f"{flag} {cc}  •  {svc_ph} `{num_disp}`  •  {lang}"
    # Inline buttons
    markup = {"inline_keyboard": []}
    if otp:
        markup["inline_keyboard"].append(
            [{"text": f"{otp}", "copy_text": {"text": otp}, "style": "success", "icon_custom_emoji_id": "6206420230269310869"}]
        )
    markup["inline_keyboard"].append(
        [{"text": "𝐅𝐮𝐥𝐥 𝐌𝐚𝐬𝐬𝐚𝐠𝐞", "copy_text": {"text": msg}, "style": "primary", "icon_custom_emoji_id": "5337302974806922068"}]
    )
    sent_count = 0
    for target_ch in channels:
        try:
            send_message(target_ch, header, markup)
            sent_count += 1
        except Exception as e:
            print(f"Channel forward error ({target_ch}): {e}")
    if sent_count > 0:
        import datetime as _dt
        today = _dt.date.today().isoformat()
        _fresh = load_data()
        stats = _fresh.setdefault("tnevs_stats", {})
        stats["otp_received"]       = stats.get("otp_received", 0) + 1
        stats["otp_today_date"]     = today if stats.get("otp_today_date") != today else stats["otp_today_date"]
        if stats.get("otp_today_date") == today:
            stats["otp_today"]      = stats.get("otp_today", 0) + 1
        else:
            stats["otp_today_date"] = today
            stats["otp_today"]      = 1
        save_data(_fresh)


def edit_message(chat_id, message_id, text, reply_markup=None, emoji_overrides=None):
    # Telegram hard-limits message text to 4096 chars
    if len(text) > 4000:
        text = text[:4000] + "\n…"
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/editMessageText"
    plain, ents = _build_entities(text, emoji_overrides)
    d = {"chat_id": chat_id, "message_id": message_id, "text": plain}
    if ents:
        d["entities"] = json.dumps(ents)
    if reply_markup:
        d["reply_markup"] = json.dumps(reply_markup)
    try:
        urllib.request.urlopen(url, urllib.parse.urlencode(d).encode(), timeout=10)
        return True
    except Exception as e:
        print(f"Edit failed, sending new: {e}")
        return send_message(chat_id, text, reply_markup)


def delete_message(chat_id, message_id):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/deleteMessage"
    d = {"chat_id": chat_id, "message_id": message_id}
    try:
        urllib.request.urlopen(url, urllib.parse.urlencode(d).encode(), timeout=10)
        return True
    except Exception as e:
        print(f"Delete error: {e}")
        return False


def copy_message(to_chat_id, from_chat_id, message_id):
    """Copy a message preserving all formatting, premium emojis, photos, etc."""
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/copyMessage"
    d = {
        "chat_id":      to_chat_id,
        "from_chat_id": from_chat_id,
        "message_id":   message_id,
    }
    try:
        req  = urllib.request.Request(url, urllib.parse.urlencode(d).encode(), headers=HEADERS)
        resp = urllib.request.urlopen(req, timeout=10)
        return True
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        if e.code == 429:
            try:
                retry_after = json.loads(body).get("parameters", {}).get("retry_after", 5)
            except Exception:
                retry_after = 5
            time.sleep(retry_after)
        return False
    except Exception:
        return False


def send_photo(chat_id, photo_url, caption, reply_markup=None, emoji_overrides=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto"
    plain, ents = _build_entities(caption, emoji_overrides)
    d = {"chat_id": chat_id, "photo": photo_url, "caption": plain}
    if ents:
        d["caption_entities"] = json.dumps(ents)
    if reply_markup:
        d["reply_markup"] = json.dumps(reply_markup)
    try:
        urllib.request.urlopen(url, urllib.parse.urlencode(d).encode(), timeout=15)
        return True
    except Exception as e:
        print(f"Photo error: {e}")
        return False


def send_document(chat_id, file_content, filename, caption=None):
    import io
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
    boundary = "----WebKitFormBoundary7MA4YWx"
    body = b""
    body += f"--{boundary}\r\nContent-Disposition: form-data; name=\"chat_id\"\r\n\r\n{chat_id}\r\n".encode()
    if caption:
        body += f"--{boundary}\r\nContent-Disposition: form-data; name=\"caption\"\r\n\r\n{caption}\r\n".encode()
    body += (
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"document\"; "
        f"filename=\"{filename}\"\r\nContent-Type: text/plain; charset=utf-8\r\n\r\n"
    ).encode()
    body += file_content.encode("utf-8")
    body += f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(url, data=body, headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    try:
        urllib.request.urlopen(req, timeout=20)
        return True
    except Exception as e:
        print(f"Send document error: {e}")
        return False


def send_document_bytes(chat_id, file_bytes, filename, caption=None):
    """Send a binary file (e.g. xlsx) via Telegram sendDocument."""
    url      = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
    boundary = "----XlsxBoundaryVaultSms"
    body = b""
    body += f"--{boundary}\r\nContent-Disposition: form-data; name=\"chat_id\"\r\n\r\n{chat_id}\r\n".encode()
    if caption:
        body += f"--{boundary}\r\nContent-Disposition: form-data; name=\"caption\"\r\n\r\n{caption}\r\n".encode()
    body += (
        f"--{boundary}\r\n"
        f"Content-Disposition: form-data; name=\"document\"; filename=\"{filename}\"\r\n"
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode()
    body += file_bytes
    body += f"\r\n--{boundary}--\r\n".encode()
    req = urllib.request.Request(
        url, data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"}
    )
    try:
        urllib.request.urlopen(req, timeout=30)
        return True
    except Exception as e:
        print(f"Send document bytes error: {e}")
        return False


def build_user_export_xlsx(uid, user):
    """Generate xlsx bytes with purchase history + deposit history for a user."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None  # openpyxl not installed

    wb = openpyxl.Workbook()

    # ── HEADER STYLE ──────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style_row(row_cells, font=None, fill=None, align=None):
        for c in row_cells:
            if font:  c.font  = font
            if fill:  c.fill  = fill
            if align: c.alignment = align
            c.border = border

    # ══ Sheet 1: Purchase History ══════════════════════════════════
    ws1 = wb.active
    ws1.title = "Purchase History"

    headers1 = ["#", "Service Name", "Number", "Price (₹)", "Status", "OTP", "Date"]
    ws1.append(headers1)
    _style_row(ws1[1], font=hdr_font, fill=hdr_fill, align=hdr_align)
    ws1.row_dimensions[1].height = 22

    history = user.get("history", [])
    # Newest first
    for i, h in enumerate(reversed(history), 1):
        status_raw = h.get("status", "")
        if status_raw in ("active", "sms_received", "completed"):
            status = "✅ Success"
        elif status_raw == "cancelled":
            status = "❌ Cancelled"
        else:
            status = "⏳ Pending"
        otp  = h.get("sms_code") or ""
        row  = [i, h.get("service", ""), h.get("number", ""), h.get("price", 0), status, str(otp), h.get("timestamp", "")]
        ws1.append(row)
        row_cells = ws1[ws1.max_row]
        alt_fill  = PatternFill("solid", fgColor="EBF3FB") if i % 2 == 0 else None
        _style_row(row_cells, align=Alignment(horizontal="center", vertical="center"), fill=alt_fill)
        # Status colour
        status_cell = row_cells[4]
        if "Success" in status:
            status_cell.font = Font(color="217346", bold=True)
        elif "Cancelled" in status:
            status_cell.font = Font(color="C00000", bold=True)

    # Column widths
    col_widths1 = [5, 22, 18, 12, 14, 12, 18]
    for col, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ══ Sheet 2: Deposit History ════════════════════════════════════
    ws2 = wb.create_sheet("Deposit History")
    headers2 = ["#", "Amount (₹)", "UTR / Ref", "Status", "Date"]
    ws2.append(headers2)
    _style_row(ws2[1], font=hdr_font, fill=hdr_fill, align=hdr_align)
    ws2.row_dimensions[1].height = 22

    deposits = user.get("deposits", [])
    for i, d in enumerate(reversed(deposits), 1):
        status_d = d.get("status", "")
        status_label = "✅ Approved" if status_d == "approved" else ("❌ Declined" if status_d == "declined" else status_d.capitalize())
        row = [i, d.get("amount", 0), d.get("utr", ""), status_label, d.get("timestamp", "")]
        ws2.append(row)
        row_cells = ws2[ws2.max_row]
        alt_fill  = PatternFill("solid", fgColor="EBF3FB") if i % 2 == 0 else None
        _style_row(row_cells, align=Alignment(horizontal="center", vertical="center"), fill=alt_fill)
        status_cell = row_cells[3]
        if "Approved" in status_label:
            status_cell.font = Font(color="217346", bold=True)
        elif "Declined" in status_label:
            status_cell.font = Font(color="C00000", bold=True)

    col_widths2 = [5, 14, 22, 14, 18]
    for col, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_photo_file(chat_id, file_id, caption, reply_markup=None):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto"
    plain, ents = _build_entities(caption)
    d = {"chat_id": chat_id, "photo": file_id, "caption": plain}
    if ents:
        d["caption_entities"] = json.dumps(ents)
    if reply_markup:
        d["reply_markup"] = json.dumps(reply_markup)
    try:
        urllib.request.urlopen(url, urllib.parse.urlencode(d).encode(), timeout=15)
        return True
    except Exception as e:
        print(f"Photo file error: {e}")
        return False


def answer_callback(callback_id, text="✅", show_alert=False):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/answerCallbackQuery"
    try:
        urllib.request.urlopen(url, urllib.parse.urlencode(
            {"callback_query_id": callback_id, "text": text,
             "show_alert": "true" if show_alert else "false"}).encode())
    except Exception:
        pass


def remove_inline_keyboard(chat_id, message_id):
    """Remove buttons from an existing message (editMessageReplyMarkup)."""
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/editMessageReplyMarkup"
    try:
        urllib.request.urlopen(url, urllib.parse.urlencode({
            "chat_id":    chat_id,
            "message_id": message_id,
            "reply_markup": json.dumps({"inline_keyboard": []})
        }).encode(), timeout=10)
    except Exception:
        pass


# ============= FORCE JOIN HELPERS =============

def tg_get_chat_member(channel_id, user_id):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getChatMember"
    try:
        data = urllib.parse.urlencode({"chat_id": str(channel_id), "user_id": user_id}).encode()
        req  = urllib.request.Request(url, data, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=10) as r:
            res = json.loads(r.read())
        return res.get("result", {})
    except Exception as e:
        print(f"getChatMember error: {e}")
        return {}


def tg_create_invite_link(channel_id):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/createChatInviteLink"
    try:
        data = urllib.parse.urlencode({"chat_id": str(channel_id)}).encode()
        req  = urllib.request.Request(url, data, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=10) as r:
            res = json.loads(r.read())
        return res.get("result", {}).get("invite_link", "")
    except Exception as e:
        print(f"createChatInviteLink error: {e}")
        return ""


def tg_get_chat(channel_id):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getChat"
    try:
        data = urllib.parse.urlencode({"chat_id": str(channel_id)}).encode()
        req  = urllib.request.Request(url, data, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=10) as r:
            res = json.loads(r.read())
        return res.get("result", {})
    except Exception as e:
        print(f"getChat error: {e}")
        return {}


def get_unjoined_channels(user_id, data):
    """Return list of channel dicts that user has NOT joined."""
    channels = data.get("force_channels", [])
    unjoined = []
    for ch in channels:
        ch_id = ch.get("id")
        if not ch_id:
            continue
        try:
            member = tg_get_chat_member(ch_id, user_id)
            status = member.get("status", "left")
            if status in ("left", "kicked", ""):
                unjoined.append(ch)
        except Exception:
            pass
    return unjoined


def send_force_join_message(chat_id, unjoined):
    """Send force-join prompt with channel buttons and a verify button."""
    lines = [
        "🔒 *Bot Use locked* 🔒\n━━━━━━━━\n",
        "_Join All Chanel and unlock_\n",
    ]
    join_btns = [
        {"text": "Join", "url": ch.get("invite_link", ""), "style": "primary", "icon_custom_emoji_id": "5337255927735163754"}
        for ch in unjoined if ch.get("invite_link")
    ]
    # pair join buttons 2 per row
    rows = [join_btns[i:i+2] for i in range(0, len(join_btns), 2)]
    rows.append([{"text": "unlock", "callback_data": "verify_join", "style": "success", "icon_custom_emoji_id": "5956267526630412170"}])
    send_message(chat_id, "\n".join(lines), {"inline_keyboard": rows}, emoji_overrides={"🔒": "5296369303661067030"})


_CALLBACK_CLUES = {
    "admin_panel":   "⚙️ Opening Admin Panel...",
    "sm_panel":      "⚙️ Opening Service Manage...",
    "sm_more":       "⚙️ Opening more settings...",
    "admin_api_bal": "📡 Fetching API balances...",
    "buy_start":     "🛒 Loading available services...",
    "depo_submit":   "📤 Submit your payment proof",
    "noop":          "✅ SMS received successfully!",
    "nsrv_panel":    "🔵 Opening Number Server...",
    "nsrv_api":      "⚙️ Loading API settings...",
    "nsrv_services": "📋 Loading services...",
    "free_number_refresh": "📱 Fetching free numbers...",
    "set_cancel_time": "⏱ Opening cancel time settings...",
    "sct_s1":        "⏱ Setting cancel time for Server 1...",
    "sct_s2":        "⏱ Setting cancel time for Server 2...",
    "sct_s3":        "⏱ Setting cancel time for Server 3...",
    "sct_min":       "✅ Saving cancel time...",
    "sct_sec":       "✅ Saving cancel time...",
}

_CALLBACK_PREFIX_CLUES = [
    ("chk_",       "📩 Checking for SMS code..."),
    ("cnc_",       "❌ Cancelling number..."),
    ("bsn_",       "🔢 Fetching your number, please wait..."),
    ("bnn_",       "🔢 Fetching your number, please wait..."),
    ("bs_",        "🔢 Fetching your number, please wait..."),
    ("bsp_",       "🔢 Fetching your number, please wait..."),
    ("svc_",       "🌍 Loading countries for this service..."),
    ("admin_srv_", "📋 Loading country list..."),
    ("da_",        "✅ Approving deposit..."),
    ("dd_",        "❌ Rejecting deposit..."),
    ("aac_",       "➕ Add a new country"),
    ("adc_",       "🗑 Select a country to remove"),
    ("aas_",       "➕ Add a new service"),
    ("ads_",       "🗑 Select a service to remove"),
    ("asc_",       "📊 Loading services & live prices..."),
    ("dco_",       "🗑 Removing country..."),
    ("dse_",       "🗑 Removing service..."),
    ("epick_",     "✨ Emoji selected!"),
    ("mail_svc_",  "📧 Getting your mail address..."),
    ("mailcode_",  "📩 Checking for mail code..."),
    ("admin_mail", "📧 Loading mail services..."),
    ("mdel_",      "🗑 Removing mail service..."),
    ("nsvc|",      "🌍 Loading countries..."),
    ("nctr|",      "📱 Loading ranges..."),
    ("nrng|",      "🛒 Fetching your numbers..."),
    ("nsrv_s|",    "📁 Opening service..."),
    ("nsrv_c|",    "🌍 Opening country..."),
    ("nsrv_add_",  "➕ Adding..."),
    ("nsrv_del_",  "🗑 Deleting..."),
    ("nsrv_set_",  "✏️ Setting value..."),
    ("freenum_",          "📩 Checking messages..."),
    ("ctry_search_",      "🔍 Search countries..."),
    ("ctry_pg_picker_",   "📄 Select page..."),
    ("svc_pg_picker_",    "📄 Select page..."),
    ("ctry_multi_",       "🌍 Loading price options..."),
]


def _get_clue(callback_data):
    if callback_data in _CALLBACK_CLUES:
        return _CALLBACK_CLUES[callback_data]
    for prefix, clue in _CALLBACK_PREFIX_CLUES:
        if callback_data.startswith(prefix):
            return clue
    return "⚡ Loading..."


# ============= KEYBOARDS =============

UPI_ID = "krishpatel284@fam"
MIN_DEPOSIT = 20


ICON_EMOJI_ID = "6159119252927289026"
BACK_BTN_EMOJI_ID = "6206505206197261313"


def _btn(text, **kwargs):
    """Build an inline button (no custom emoji)."""
    return {"text": _strip_emoji(text), **kwargs}

def _back_btn(text="◀️ Back", **kwargs):
    """Build a Back button (no custom emoji)."""
    return {"text": _strip_emoji(text), **kwargs}


def _kb_btn(text, color=None):
    """Build a reply keyboard button with text_entities for premium emojis and optional color."""
    btn = {"text": text}
    if color:
        btn["style"] = color
    ents = []
    offset = 0
    i = 0
    while i < len(text):
        matched = next((em for em in _EMOJI_KEYS if text[i:i+len(em)] == em), None)
        if matched and matched in PREMIUM_EMOJIS:
            eid  = PREMIUM_EMOJIS[matched]
            elen = _utf16_len(matched)
            ents.append({"type": "custom_emoji", "offset": offset,
                         "length": elen, "custom_emoji_id": eid})
            offset += elen
            i += len(matched)
        else:
            offset += _utf16_len(text[i])
            i += 1
    if ents:
        btn["text_entities"] = ents
    return btn


def get_mail_services_inline(mail_services):
    """Inline keyboard listing all mail services."""
    buttons = []
    row = []
    for ms in mail_services:
        price = ms.get("price", 0)
        btn = {
            "text":          f"{ms['name']}  ₹{price}",
            "callback_data": f"mail_svc_{ms['id']}",
            "style":         ms.get("style", "primary"),
        }
        eid = ms.get("emoji_id", "")
        if eid:
            btn["icon_custom_emoji_id"] = eid
        row.append(btn)
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    if not buttons:
        buttons = [[{"text": "❌ No mail services", "callback_data": "noop"}]]
    return {"inline_keyboard": buttons}


def get_mail_action_inline(mail_id):
    """Inline keyboard shown after a mail address is issued."""
    return {
        "inline_keyboard": [
            [{"text": "📩 Get Mail Code", "callback_data": f"mailcode_{mail_id}", "style": "success"}],
            [{"text": "❌ Cancel & Refund", "callback_data": f"mailcancel_{mail_id}", "style": "danger"}],
        ]
    }


CONTROLLABLE_BUTTONS = [
    "Buy Number", "Get Mail", "Refer", "Deposit",
    "Account", "Active Numbers", "More",
    "Gift Code", "Temp Mail", "Free Number",
]

def get_button_states(data):
    states = data.get("button_states", {})
    return {btn: states.get(btn, True) for btn in CONTROLLABLE_BUTTONS}

def is_button_enabled(data, btn_name):
    return data.get("button_states", {}).get(btn_name, True)

def get_dm_inline(data):
    ds  = data.get("deposit_settings", {})
    on  = ds.get("enabled", True)
    lbl = "✅ Deposit ON" if on else "❌ Deposit OFF"
    sty = "success" if on else "danger"
    ch  = ds.get("notify_channel", "")
    ch_lbl = f"📢 Channel: {ch}" if ch else "📢 Set Notify Channel"
    return {"inline_keyboard": [
        [_btn("💳 Change UPI ID",       callback_data="dm_set_upi",     style="primary")],
        [_btn("🖼️ Change QR Code URL",  callback_data="dm_set_qr",      style="primary")],
        [_btn("💸 Change Min Deposit",  callback_data="dm_set_min",     style="primary")],
        [_btn(ch_lbl,                   callback_data="dm_set_channel", style="primary")],
        [_btn(lbl,                      callback_data="dm_toggle",      style=sty)],
        [_back_btn("◀️ Back",           callback_data="uh_main")],
    ]}


def get_btn_ctrl_inline(data):
    states = get_button_states(data)
    rows = []
    for btn in CONTROLLABLE_BUTTONS:
        on    = states[btn]
        style = "success" if on else "danger"
        label = f"{'✅' if on else '❌'} {btn}"
        rows.append([_btn(label, callback_data=f"btn_toggle_{btn}", style=style)])
    rows.append([_back_btn("◀️ Back", callback_data="uh_main")])
    return {"inline_keyboard": rows}


def get_main_keyboard(is_admin=False):
    rows = [
        [
            {"text": "Buy Number", "style": "success", "icon_custom_emoji_id": "5431499171045581032"},
            {"text": "Get Mail",   "style": "primary", "icon_custom_emoji_id": "5456174900622412791"},
        ],
        [
            {"text": "Refer",   "style": "success", "icon_custom_emoji_id": "6206027872121918710"},
            {"text": "Deposit", "style": "success", "icon_custom_emoji_id": "6199349120667226859"},
        ],
        [
            {"text": "Account",       "style": "primary", "icon_custom_emoji_id": "5258011929993026890"},
            {"text": "Active Numbers","style": "primary", "icon_custom_emoji_id": "5460865451586250451"},
        ],
        [
            {"text": "More", "style": "success", "icon_custom_emoji_id": "5337267511261960341"},
        ],
    ]
    if is_admin:
        rows.append([{"text": "Admin Panel", "style": "danger"}])
    return {"keyboard": rows, "resize_keyboard": True, "one_time_keyboard": False}


def get_more_keyboard():
    rows = [
        [
            {"text": "Gift Code",    "style": "success", "icon_custom_emoji_id": "5251227707026470504"},
            {"text": "Temp Mail",    "style": "danger",  "icon_custom_emoji_id": "5251635634430309777"},
        ],
        [
            {"text": "Free Number",  "style": "primary", "icon_custom_emoji_id": "5431499171045581032"},
        ],
        [
            {"text": "Back", "style": "primary", "icon_custom_emoji_id": "5816895683256390576"},
        ],
    ]
    return {"keyboard": rows, "resize_keyboard": True, "one_time_keyboard": False}


def get_admin_main_inline():
    return {
        "inline_keyboard": [
            [_btn("⚙️ Service Manage",       callback_data="sm_panel")],
            [_btn("🔵 2oo9 Number Server",    callback_data="nsrv_panel")],
            [_btn("📢 Force Join Channels",  callback_data="fj_admin")],
            [_btn("👥 User Handle",          callback_data="uh_main")],
            [_btn("📣 Broadcast",            callback_data="admin_broadcast", style="primary")],
        ]
    }


# ─── 2oo9 NUMBER SERVER ADMIN KEYBOARDS ──────────────────────────────────────

def get_nsrv_panel_inline():
    return {
        "inline_keyboard": [
            [_btn("⚙️ API Settings",    callback_data="nsrv_api")],
            [_btn("📋 Manage Services", callback_data="nsrv_services")],
            [{"text": "Manage Withdrawal", "callback_data": "manage_withdrawal", "style": "primary", "icon_custom_emoji_id": "6269567575349270928"}],
            [_btn("📡 Live SMS Channel", callback_data="sms_channel_panel")],
            [_back_btn("◀️ Back",       callback_data="admin_panel")],
        ]
    }


# ─── WITHDRAWAL KEYBOARD HELPERS ────────────────────────────────────────────

_WD_OVR = {          # placeholder chars → premium emoji IDs for user panel
    "◆": "5352877703043258544",
    "◇": "5276032951342088188",
    "◈": "6199349120667226859",
    "◉": "6199253428795872585",
}

def get_wd_settings(data):
    ws = data.get("withdrawal_settings", {})
    return {
        "enabled":         ws.get("enabled", True),
        "min_amount":      ws.get("min_amount", 50.0),
        "payment_methods": ws.get("payment_methods", []),
        "service_earn":    ws.get("service_earn", {}),
    }

def _wd_svc_earn(data, svc_id):
    return data.get("withdrawal_settings", {}).get("service_earn", {}).get(svc_id,
                                                                            {"enabled": False, "price": 1.0})

def get_manage_withdrawal_inline(data):
    ws  = get_wd_settings(data)
    on  = ws["enabled"]
    lbl = "✅ Withdrawal ON" if on else "❌ Withdrawal OFF"
    sty = "success" if on else "danger"
    return {"inline_keyboard": [
        [{"text": "💰 Update Price",       "callback_data": "mw_update_price", "style": "primary"}],
        [{"text": lbl,                     "callback_data": "mw_toggle",       "style": sty}],
        [{"text": "💵 Minimum Withdrawal", "callback_data": "mw_min",          "style": "primary"}],
        [{"text": "💳 Payment Method",     "callback_data": "mw_pm",           "style": "primary"}],
        [_back_btn("◀️ Back", callback_data="nsrv_panel")],
    ]}

def get_mw_update_price_inline(data):
    services = data.get("services_data", {})
    btns = []
    for sid, srv in services.items():
        ec    = _wd_svc_earn(data, sid)
        icon  = "✅" if ec.get("enabled") else "❌"
        price = ec.get("price", 1.0)
        btn   = {"text": f"{icon} {srv['name']}  •  ₹{price}", "callback_data": f"mw_svc|{sid}", "style": "primary"}
        eid   = _free_svc_icon_id(srv["name"])
        if eid:
            btn["icon_custom_emoji_id"] = eid
        btns.append([btn])
    if not btns:
        btns.append([{"text": "❌ No services configured", "callback_data": "noop"}])
    btns.append([_back_btn("◀️ Back", callback_data="manage_withdrawal")])
    return {"inline_keyboard": btns}

def get_mw_svc_inline(data, svc_id, svc_name):
    ec    = _wd_svc_earn(data, svc_id)
    on    = ec.get("enabled", False)
    lbl   = "✅ Earn ON" if on else "❌ Earn OFF"
    sty   = "success" if on else "danger"
    price = ec.get("price", 1.0)
    eid   = _free_svc_icon_id(svc_name)
    svc_btn = {"text": svc_name, "callback_data": "noop", "style": "primary"}
    if eid:
        svc_btn["icon_custom_emoji_id"] = eid
    return {"inline_keyboard": [
        [svc_btn],
        [{"text": lbl,                                       "callback_data": f"mw_earn_toggle|{svc_id}", "style": sty}],
        [{"text": f"💰 Change Price  (now ₹{price})",       "callback_data": f"mw_price|{svc_id}",       "style": "primary"}],
        [_back_btn("◀️ Back", callback_data="mw_update_price")],
    ]}

def get_mw_pm_inline(data):
    ws      = get_wd_settings(data)
    methods = ws["payment_methods"]
    rows = []
    for method, label in [("upi", "🏦 UPI"), ("bank", "🏛️ Bank"), ("binance", "₿ Binance")]:
        on  = method in methods
        lbl = f"{'✅' if on else '❌'} {label}"
        sty = "success" if on else "danger"
        rows.append([{"text": lbl, "callback_data": f"mw_pm_toggle|{method}", "style": sty}])
    rows.append([_back_btn("◀️ Back", callback_data="manage_withdrawal")])
    return {"inline_keyboard": rows}

def get_user_withdrawal_inline():
    return {"inline_keyboard": [
        [{"text": "Add Payment Method", "callback_data": "user_add_payment",  "style": "primary",  "icon_custom_emoji_id": "5352922460897452503"}],
        [{"text": "Withdrawal Now",     "callback_data": "user_withdraw_now", "style": "success",  "icon_custom_emoji_id": "5206610654747705715"}],
        [{"text": "◀️ Back",           "callback_data": "free_number_refresh","style": "danger",   "icon_custom_emoji_id": "5370726448959085259"}],
    ]}

def get_user_add_payment_inline(data):
    ws      = get_wd_settings(data)
    methods = ws["payment_methods"]
    rows    = []
    labels   = {"upi": "UPI",     "bank": "Bank",    "binance": "Binance"}
    pm_emojis = {"upi": "5807750375033278838", "bank": "6156548740770436100", "binance": "5454205576512818013"}
    for m in methods:
        btn = {"text": labels.get(m, m.title()), "callback_data": f"user_pm|{m}", "style": "primary"}
        if m in pm_emojis:
            btn["icon_custom_emoji_id"] = pm_emojis[m]
        rows.append([btn])
    if not rows:
        rows.append([{"text": "❌ No payment methods available", "callback_data": "noop"}])
    rows.append([_back_btn("◀️ Back", callback_data="user_withdrawal")])
    return {"inline_keyboard": rows}


def get_sms_channel_panel_inline(data):
    cfg = data.get("tnevs_settings", {})
    def _ch_disp(key):
        ch = cfg.get(key, "")
        return f"✅ {ch}" if ch else "❌ Not set"
    return {
        "inline_keyboard": [
            [_btn(f"📡 All OTP: {_ch_disp('sms_channel_all')}",       callback_data="sms_ch_all")],
            [_btn(f"🟠 Instagram OTP: {_ch_disp('sms_channel_ig')}",  callback_data="sms_ch_ig")],
            [_btn(f"🟦 Facebook OTP: {_ch_disp('sms_channel_fb')}",   callback_data="sms_ch_fb")],
            [_btn(f"🟩 WhatsApp OTP: {_ch_disp('sms_channel_wa')}",   callback_data="sms_ch_wa")],
            [_btn(f"👤 User SMS: {_ch_disp('sms_channel_user')}",      callback_data="sms_ch_user")],
            [_back_btn("◀️ Back",                                            callback_data="admin_panel")],
        ]
    }


def get_nsrv_api_inline(data):
    cfg      = data.get("tnevs_settings", {})
    key_disp = "✅ Configured" if cfg.get("api_key") else "❌ Not set"
    return {
        "inline_keyboard": [
            [_btn(f"🔑 API Key: {key_disp}", callback_data="nsrv_set_key")],
            [_back_btn("◀️ Back",                 callback_data="nsrv_panel")],
        ]
    }


def get_nsrv_services_inline(data):
    services = data.get("services_data", {})
    btns = []
    for sid, srv in services.items():
        cnt_count = len(srv.get("countries", {}))
        btns.append([_btn(f"📁 {srv['name']}  ({cnt_count} countries)", callback_data=f"nsrv_s|{sid}")])
    btns.append([_btn("➕ Add Service", callback_data="nsrv_add_srv", style="success")])
    btns.append([_back_btn("◀️ Back", callback_data="nsrv_panel")])
    return {"inline_keyboard": btns}


def get_nsrv_countries_inline(srv_id, srv_data):
    countries = srv_data.get("countries", {})
    btns = []
    for cnt_id, cnt in countries.items():
        flag     = get_country_flag(cnt["name"])
        rng_cnt  = len(cnt.get("ranges", {}))
        btns.append([_btn(f"{flag} {cnt['name']}  ({rng_cnt} ranges)", callback_data=f"nsrv_c|{srv_id}|{cnt_id}")])
    btns.append([
        _btn("➕ Add Country", callback_data=f"nsrv_add_cnt|{srv_id}", style="success"),
        _btn("🗑 Del Service",  callback_data=f"nsrv_del_srv|{srv_id}", style="danger"),
    ])
    btns.append([_back_btn("◀️ Back", callback_data="nsrv_services")])
    return {"inline_keyboard": btns}


def get_nsrv_ranges_inline(srv_id, cnt_id, cnt_data):
    ranges = cnt_data.get("ranges", {})
    btns   = []
    for rid, rng in ranges.items():
        btns.append([_btn(f"📱 {rng}  ✕ delete", callback_data=f"nsrv_del_rng|{srv_id}|{cnt_id}|{rid}", style="danger")])
    btns.append([
        _btn("➕ Add Range",    callback_data=f"nsrv_add_rng|{srv_id}|{cnt_id}", style="success"),
        _btn("🗑 Del Country", callback_data=f"nsrv_del_cnt|{srv_id}|{cnt_id}", style="danger"),
    ])
    btns.append([_back_btn("◀️ Back", callback_data=f"nsrv_s|{srv_id}")])
    return {"inline_keyboard": btns}


# ─── NEXA USER BUY FLOW KEYBOARDS ────────────────────────────────────────────

def get_nsrv_user_services_inline(data):
    services = data.get("services_data", {})
    btns = []
    row  = []
    for sid, srv in services.items():
        has_ranges = any(
            len(cnt.get("ranges", {})) > 0
            for cnt in srv.get("countries", {}).values()
        )
        if has_ranges:
            row.append({"text": srv["name"], "callback_data": f"nsvc|{sid}", "style": "primary"})
            if len(row) == 2:
                btns.append(row)
                row = []
    if row:
        btns.append(row)
    if not btns:
        btns = [[{"text": "❌ No services available", "callback_data": "noop"}]]
    return {"inline_keyboard": btns}


def get_nsrv_user_countries_inline(srv_id, srv_data):
    countries = srv_data.get("countries", {})
    btns      = []
    for cnt_id, cnt in countries.items():
        if cnt.get("ranges"):
            eid  = _country_emoji_id(cnt)
            name = _strip_emoji(cnt["name"]) if eid else cnt["name"]
            btn  = {"text": name, "callback_data": f"nctr|{srv_id}|{cnt_id}", "style": "primary"}
            if eid:
                btn["icon_custom_emoji_id"] = eid
            btns.append([btn])
    btns.append([{"text": "◀️ Back", "icon_custom_emoji_id": BACK_BTN_EMOJI_ID, "callback_data": "buy_start", "style": "danger"}])
    return {"inline_keyboard": btns}


def get_nsrv_user_ranges_inline(srv_id, cnt_id, cnt_data):
    ranges = cnt_data.get("ranges", {})
    btns   = []
    for rid, rng in ranges.items():
        btns.append([{"text": f"📱 {rng}", "callback_data": f"nrng|{srv_id}|{cnt_id}|{rid}", "style": "success"}])
    btns.append([{"text": "◀️ Back", "icon_custom_emoji_id": BACK_BTN_EMOJI_ID, "callback_data": f"nsvc|{srv_id}", "style": "danger"}])
    return {"inline_keyboard": btns}


def get_nsrv_active_inline(number, number_id, srv_id):
    return {
        "inline_keyboard": [
            [{"text": "🔴 Live Console", "callback_data": "tnevs_console", "style": "primary"}],
        ]
    }


# ─── FREE NUMBER SERVICE EMOJI MAP ───────────────────────────────────────────
FREE_SVC_EMOJI_IDS = {
    "facebook":  "5323261730283863478",
    "whatsapp":  "5334998226636390258",
    "telegram":  "5330237710655306682",
    "imo":       "5337155807752524558",
    "instagram": "5319160079465857105",
    "apple":     "5334637951894722661",
    "email":     "5348494358205207761",
}

def _free_svc_emoji(name):
    """Return custom_emoji_id for a free-number service name, or None."""
    key = str(name).lower().strip()
    for keyword, eid in FREE_SVC_EMOJI_IDS.items():
        if keyword in key:
            return eid
    return None

def _free_svc_icon_char(name):
    """Return the emoji character for a free-number service name (e.g. '🟩' for WhatsApp)."""
    key = str(name).lower().strip()
    for keyword, (ch, _) in CONSOLE_SERVICE_EMOJIS.items():
        if keyword in key:
            return ch
    return "📱"

def _free_svc_icon_id(name):
    """Return the premium emoji ID for a free-number service name, or None."""
    key = str(name).lower().strip()
    for keyword, (_, eid) in CONSOLE_SERVICE_EMOJIS.items():
        if keyword in key:
            return eid
    return None

# ─── FREE NUMBER KEYBOARDS ────────────────────────────────────────────────────

def get_free_number_inline(numbers):
    if not numbers:
        return {"inline_keyboard": [
            [{"text": "❌ No free numbers available now", "callback_data": "noop"}],
            [{"text": "🔄 Retry", "callback_data": "free_number_refresh", "style": "primary"}],
        ]}
    btns = []
    for n in numbers[:12]:
        if isinstance(n, dict):
            num     = n.get("number", str(n))
            country = n.get("country", n.get("location", ""))
            flag    = get_country_flag(country) if country else "📱"
        else:
            num  = str(n)
            flag = "📱"
        btns.append([{"text": f"{flag} {num}", "callback_data": f"freenum_{num}", "style": "primary"}])
    btns.append([{"text": "🔄 Refresh", "callback_data": "free_number_refresh", "style": "success"}])
    return {"inline_keyboard": btns}


def get_free_nexa_services_inline(data):
    services = data.get("services_data", {})
    btns = []
    row = []
    for sid, srv in services.items():
        has_ranges = any(
            len(cnt.get("ranges", {})) > 0
            for cnt in srv.get("countries", {}).values()
        )
        if has_ranges:
            btn = {"text": srv["name"], "callback_data": f"fn_svc|{sid}", "style": "primary"}
            eid = _free_svc_emoji(srv["name"])
            if eid:
                btn["icon_custom_emoji_id"] = eid
            row.append(btn)
            if len(row) == 2:
                btns.append(row)
                row = []
    if row:
        btns.append(row)
    btns.append([{"text": "Buy Custom Range", "callback_data": "fn_direct_custom_range", "style": "success", "icon_custom_emoji_id": "5337132498965010628"}])
    # Only show Withdrawal button when withdrawal is enabled
    if data.get("withdrawal_settings", {}).get("enabled", True):
        btns.append([{"text": "Withdrawal", "callback_data": "user_withdrawal", "style": "primary", "icon_custom_emoji_id": "6269567575349270928"}])
    return {"inline_keyboard": btns}


def _flag_emoji_to_cc(flag):
    """Convert a flag emoji like '🇲🇬' to its 2-letter ISO CC 'MG'."""
    try:
        chars = [c for c in flag if '\U0001F1E6' <= c <= '\U0001F1FF']
        if len(chars) >= 2:
            return chr(ord(chars[0]) - 0x1F1E6 + ord('A')) + chr(ord(chars[1]) - 0x1F1E6 + ord('A'))
    except Exception:
        pass
    return ""


def _cc_from_country_name(cnt_name):
    """Get 2-letter ISO CC for a country name like 'Ivory coast' or '🇨🇮 Ivory Coast'."""
    # First try: extract from flag emoji if present in the name
    for i, ch in enumerate(cnt_name):
        if '\U0001F1E6' <= ch <= '\U0001F1FF':
            if i + 1 < len(cnt_name):
                ch2 = cnt_name[i + 1]
                if '\U0001F1E6' <= ch2 <= '\U0001F1FF':
                    return chr(ord(ch) - 0x1F1E6 + ord('A')) + chr(ord(ch2) - 0x1F1E6 + ord('A'))
    # Second try: look up the name in COUNTRY_FLAGS → get flag → extract CC
    name_clean = _strip_emoji(cnt_name).lower().strip()
    flag = COUNTRY_FLAGS.get(name_clean)
    if not flag:
        for cname, f in COUNTRY_FLAGS.items():
            if len(cname) >= 3 and (cname in name_clean or name_clean in cname):
                flag = f
                break
    if flag:
        cc = _flag_emoji_to_cc(flag)
        if cc:
            return cc
    return ""


def _svc_name_matches_key(srv_name, svc_key):
    """Check if a services_data service name matches a console svc_key (e.g. 'ig'→'instagram')."""
    _aliases = {"ig": "instagram", "wa": "whatsapp", "fb": "facebook", "tg": "telegram", "tw": "twitter"}
    n = _strip_emoji(srv_name).lower().strip()
    k = svc_key.lower().strip()
    k_full = _aliases.get(k, k)
    return k_full in n or k in n


def _otp_age_str(age_secs):
    """Convert elapsed seconds into a compact label: '5s', '3m', '2h'."""
    age_secs = max(0, age_secs)
    if age_secs < 60:
        return f"{int(age_secs)}s"
    elif age_secs < 3600:
        return f"{int(age_secs / 60)}m"
    else:
        return f"{int(age_secs / 3600)}h"


def get_free_nexa_countries_inline(srv_id, srv_data, earn_price=0.0):
    countries = srv_data.get("countries", {})
    srv_name  = srv_data.get("name", "")
    entries = []  # (age_secs_or_inf, btn_row)
    now  = time.time()
    for cnt_id, cnt in countries.items():
        if cnt.get("ranges"):
            eid  = _country_emoji_id(cnt)
            name = _strip_emoji(cnt["name"]) if eid else cnt["name"]
            # OTP age badge: find most-recent console hit for this service+country
            cc = _cc_from_country_name(cnt["name"])
            age_label = ""
            sort_age  = float("inf")
            if cc:
                best_ts = None
                for (sk, scc), ts in list(_last_console_otp_time.items()):
                    if scc == cc and _svc_name_matches_key(srv_name, sk):
                        if best_ts is None or ts > best_ts:
                            best_ts = ts
                if best_ts is not None:
                    age = now - best_ts
                    if age < 7200:  # only show if within 2 hours
                        age_label = f" {_otp_age_str(age)}"
                        sort_age  = age
            # Build button label: "Country Name  20s  ₹0.30"
            price_label = f"  ₹{earn_price}" if earn_price > 0 else ""
            btn  = {"text": f"{name}{age_label}{price_label}", "callback_data": f"fn_ctr|{srv_id}|{cnt_id}", "style": "primary"}
            if eid:
                btn["icon_custom_emoji_id"] = eid
            entries.append((sort_age, [btn]))
    # Sort: freshest OTP (smallest age) first; no-OTP countries at bottom
    entries.sort(key=lambda x: x[0])
    btns = [row for _, row in entries]
    btns.append([{"text": "◀️ Back", "icon_custom_emoji_id": BACK_BTN_EMOJI_ID, "callback_data": "free_number_refresh", "style": "danger"}])
    return {"inline_keyboard": btns}


def get_free_nexa_ranges_inline(srv_id, cnt_id, cnt_data):
    ranges = cnt_data.get("ranges", {})
    eid = _country_emoji_id(cnt_data)
    btns = []
    for rid, rng in ranges.items():
        btn = {"text": rng, "callback_data": f"fn_rng|{srv_id}|{cnt_id}|{rid}", "style": "success"}
        if eid:
            btn["icon_custom_emoji_id"] = eid
        else:
            btn["text"] = f"📱 {rng}"
        btns.append([btn])
    btns.append([{"text": "Buy Custom Range", "callback_data": f"fn_custom_range|{srv_id}|{cnt_id}", "style": "success", "icon_custom_emoji_id": "5337132498965010628"}])
    btns.append([{"text": "◀️ Back", "icon_custom_emoji_id": BACK_BTN_EMOJI_ID, "callback_data": f"fn_svc|{srv_id}", "style": "danger"}])
    return {"inline_keyboard": btns}


def get_api_keys_panel_inline():
    """Show current API key status for all 3 servers."""
    def _key_disp(api_id):
        k = APIS[api_id]["key"]
        return f"✅ {k[:6]}...{k[-4:]}" if k else "❌ Not set"
    return {
        "inline_keyboard": [
            [_btn(f"🔵 S1 SmsBower  {_key_disp('s1')}", callback_data="change_api_s1", style="primary")],
            [_btn(f"🟢 S2 HeroSMS   {_key_disp('s2')}", callback_data="change_api_s2", style="primary")],
            [_btn(f"🟡 S3 SmsIndia  {_key_disp('s3')}", callback_data="change_api_s3", style="primary")],
            [_btn(f"🟤 S4 UotpStore {_key_disp('s4')}", callback_data="change_api_s4", style="primary")],
            [_btn(f"🟣 S5 SmsProvider {_key_disp('s5')}", callback_data="change_api_s5", style="primary")],
            [_back_btn("◀️ Back", callback_data="sm_more")],
        ]
    }


def get_service_manage_inline(data=None):
    return {
        "inline_keyboard": [
            [_btn("🔵 Server 1 (SmsBower)",    callback_data="admin_srv_s1")],
            [_btn("🟢 Server 2 (HeroSMS)",     callback_data="admin_srv_s2")],
            [_btn("🟡 Server 3 (SmsIndia)",    callback_data="admin_srv_s3")],
            [_btn("🟤 Server 4 (UotpStore)",   callback_data="admin_srv_s4")],
            [_btn("🟣 Server 5 (SmsProvider)", callback_data="admin_srv_s5")],
            [_btn("📧 Mail Services",           callback_data="admin_mail")],
            [_btn("⋯ More",                 callback_data="sm_more")],
            [_back_btn("◀️ Back",            callback_data="admin_panel")],
        ]
    }


def get_service_manage_more_inline(data=None):
    cfg = (data or {}).get("tnevs_settings", {})
    def _ch_disp(key):
        ch = cfg.get(key, "")
        return f"✅ {ch}" if ch else "❌ Not set"
    return {
        "inline_keyboard": [
            [_btn("🔑 API Keys",             callback_data="api_keys_panel")],
            [_btn("📡 API Balance",          callback_data="admin_api_bal")],
            [_btn(f"👤 User SMS Channel: {_ch_disp('sms_channel_user_buy')}", callback_data="sms_ch_user_buy")],
            [_btn("⏱ Change Cancel Time",   callback_data="set_cancel_time")],
            [_back_btn("◀️ Back",            callback_data="sm_panel")],
        ]
    }


def get_srv_countries_inline(countries, api, page=0):
    PAGE_SIZE = 20
    cfg = APIS[api]
    # Newest first
    items = list(reversed(countries))
    total = len(items)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    page_items = items[start:start + PAGE_SIZE]
    buttons = []
    row = []
    for c in page_items:
        row.append({"text": c["name"], "callback_data": f"asc_{api}_{c['id']}"})
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    # Navigation
    nav = []
    if page > 0:
        nav.append({"text": "Prev", "callback_data": f"admin_cntry_pg_{api}_{page-1}", "style": "primary", "icon_custom_emoji_id": "5370726448959085259"})
    if page < total_pages - 1:
        nav.append({"text": "Next", "callback_data": f"admin_cntry_pg_{api}_{page+1}", "style": "primary", "icon_custom_emoji_id": "6197296817789474264"})
    if nav:
        buttons.append(nav)
    buttons.append([
        {"text": f"{page+1}/{total_pages}", "callback_data": "noop", "style": "primary", "icon_custom_emoji_id": "5258477770735885832"},
        _btn("➕ Add Country", callback_data=f"aac_{api}"),
        _btn("🗑 Del Country",  callback_data=f"adc_{api}")
    ])
    buttons.append([_back_btn("◀️ Back", callback_data="sm_panel")])
    return {"inline_keyboard": buttons}


def get_srv_country_inline(api, country_id):
    return {
        "inline_keyboard": [
            [
                _btn("➕ Add Service", callback_data=f"aas_{api}_{country_id}"),
                _btn("🗑 Del Service",  callback_data=f"ads_{api}_{country_id}")
            ],
            [_back_btn("◀️ Back to Countries", callback_data=f"admin_srv_{api}")]
        ]
    }


def get_s3_country_services_inline(api, cid, svcs, page=0):
    """S3/S4 country view — unique services (dedup by name), one button each, paginated (20/page, 2/row)."""
    PAGE_SIZE = 20
    # Deduplicate: one entry per unique normalised service name (newest-first within dedup)
    seen_norms  = {}
    unique_items = []
    for orig_idx, s in enumerate(svcs):
        norm = s.get("name", "").strip().lower()
        if norm not in seen_norms:
            seen_norms[norm] = orig_idx
            unique_items.append((orig_idx, s))
    items       = list(reversed(unique_items))   # newest first
    total       = len(items)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page        = max(0, min(page, total_pages - 1))
    start       = page * PAGE_SIZE
    page_items  = items[start:start + PAGE_SIZE]
    all_btns = []
    for orig_idx, s in page_items:
        label = f"{s['name']}  ₹{s.get('price', 0)}"
        all_btns.append(_btn(label, callback_data=f"s3grp_{api}_{cid}_{s['id']}", style="primary"))
    # 2 per row
    buttons = []
    row = []
    for btn in all_btns:
        row.append(btn)
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    # Navigation
    nav = []
    if page > 0:
        nav.append({"text": "Prev", "callback_data": f"s3svc_pg_{api}_{cid}_{page-1}", "style": "primary", "icon_custom_emoji_id": "5370726448959085259"})
    if page < total_pages - 1:
        nav.append({"text": "Next", "callback_data": f"s3svc_pg_{api}_{cid}_{page+1}", "style": "primary", "icon_custom_emoji_id": "6197296817789474264"})
    if nav:
        buttons.append(nav)
    # Page + Search in one row
    buttons.append([
        {"text": f"{page+1}/{total_pages}", "callback_data": "noop",        "style": "primary", "icon_custom_emoji_id": "5258477770735885832"},
        {"text": "Search",                  "callback_data": "buy_search",   "style": "success", "icon_custom_emoji_id": "6206446249181189526"},
    ])
    buttons.append([_btn("➕ Add Service", callback_data=f"aas_{api}_{cid}", style="success")])
    buttons.append([_back_btn("◀️ Back", callback_data=f"admin_srv_{api}")])
    return {"inline_keyboard": buttons}


def get_s3_svc_group_inline(api, cid, svcs, target_norm, svc_id):
    """Shows all operator entries for a service group (same name) with Add Operator + Back."""
    group = [(idx, s) for idx, s in enumerate(svcs) if s.get("name", "").strip().lower() == target_norm]
    buttons = []
    for idx, s in group:
        op    = s.get("operator", "any")
        price = s.get("price", 0)
        mp    = s.get("max_price")
        mp_part = f"  🔸{mp}" if mp else ""
        label   = f"🔢 Op: {op}  ₹{price}{mp_part}"
        buttons.append([_btn(label, callback_data=f"s3opact_{api}_{cid}_{idx}", style="primary")])
    buttons.append([_btn("➕ Add Operator", callback_data=f"s3addop_{api}_{cid}_{svc_id}", style="success")])
    buttons.append([_back_btn("◀️ Back", callback_data=f"s3back_{api}_{cid}")])
    return {"inline_keyboard": buttons}


def get_s3_op_action_inline(api, cid, svc_idx, svc_id):
    """Update Price + Delete Operator panel for a single operator entry."""
    return {
        "inline_keyboard": [
            [_btn("💰 Update Price",    callback_data=f"s3updprice_{api}_{cid}_{svc_idx}", style="success")],
            [_btn("🗑 Delete Operator", callback_data=f"s3delop_{api}_{cid}_{svc_idx}",   style="danger")],
            [_back_btn("◀️ Back",       callback_data=f"s3grp_{api}_{cid}_{svc_id}")],
        ]
    }


def get_del_country_inline(countries, api):
    buttons = [[_btn(f"🗑 {c['name']}", callback_data=f"dco_{api}_{c['id']}")]
               for c in countries]
    buttons.append([_back_btn("◀️ Back", callback_data=f"admin_srv_{api}")])
    return {"inline_keyboard": buttons}


def get_del_service_inline(services, api, country_id):
    buttons = []
    for idx, s in enumerate(services):
        buttons.append([_btn(f"🗑 {s['name']} ₹{s.get('price',0)}",
                             callback_data=f"dse_{api}_{country_id}_{idx}")])
    buttons.append([_back_btn("◀️ Back", callback_data=f"asc_{api}_{country_id}")])
    return {"inline_keyboard": buttons}


def get_srv_services_inline_for_api(data, api, page=0):
    """Service-first view: list unique services for an API with country counts. Paginated (20/page, 2/row)."""
    PAGE_SIZE = 20
    seen = {}
    seen_order = []
    source = data.get("s3_countries", []) if api == "s3" else (data.get("s4_countries", []) if api == "s4" else data.get("countries", []))
    # Reversed so newest appears first
    for c in reversed(source):
        for s in reversed(c.get("services", [])):
            if s.get("api") == api:
                sid = s["id"]
                if sid not in seen:
                    seen[sid] = {"id": sid, "name": s["name"], "count": 0}
                    seen_order.append(sid)
                seen[sid]["count"] += 1
    all_btns = []
    for sid in seen_order:
        sinfo = seen[sid]
        cnt   = sinfo["count"]
        lbl   = f"{sinfo['name']}  ({cnt} {'country' if cnt == 1 else 'countries'})"
        all_btns.append(_btn(lbl, callback_data=f"asrvsvc_{api}_{sinfo['id']}"))
    total       = len(all_btns)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page        = max(0, min(page, total_pages - 1))
    start       = page * PAGE_SIZE
    page_btns   = all_btns[start:start + PAGE_SIZE]
    buttons = []
    row = []
    for btn in page_btns:
        row.append(btn)
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    # Navigation
    nav = []
    if page > 0:
        nav.append({"text": "Prev", "callback_data": f"admin_srvsvc_pg_{api}_{page-1}", "style": "primary", "icon_custom_emoji_id": "5370726448959085259"})
    if page < total_pages - 1:
        nav.append({"text": "Next", "callback_data": f"admin_srvsvc_pg_{api}_{page+1}", "style": "primary", "icon_custom_emoji_id": "6197296817789474264"})
    if nav:
        buttons.append(nav)
    buttons.append([
        {"text": f"{page+1}/{total_pages}", "callback_data": "noop", "style": "primary", "icon_custom_emoji_id": "5258477770735885832"},
        _btn("➕ Add Service", callback_data=f"aas_{api}"),
        _btn("🗑 Del Service",  callback_data=f"adss_{api}")
    ])
    buttons.append([_back_btn("◀️ Back", callback_data="sm_panel")])
    return {"inline_keyboard": buttons}


def get_srv_service_countries_inline(data, api, svc_id, page=0):
    """Countries list under a specific service (paginated, 40 per page to stay within Telegram limits)."""
    PAGE_SIZE = 40
    svc_name = ""
    entries = []
    for c in data.get("countries", []):
        for s in c.get("services", []):
            if s.get("api") == api and s["id"] == svc_id:
                if not svc_name:
                    svc_name = s["name"]
                flag = get_country_flag(c["name"])
                entries.append({"cid": c["id"], "cname": c["name"], "price": s.get("price", 0), "flag": flag})
                break
    total       = len(entries)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page        = max(0, min(page, total_pages - 1))
    start       = page * PAGE_SIZE
    page_entries = entries[start:start + PAGE_SIZE]
    buttons = []
    for e in page_entries:
        # Display only — not a delete action
        buttons.append([_btn(f"{e['flag']} {e['cname']}  ₹{e['price']}", callback_data=f"vcsc_{api}_{svc_id}_{e['cid']}")])
    # Pagination nav
    nav = []
    if page > 0:
        nav.append(_btn(f"◀ Prev", callback_data=f"asrvsvc_pg_{api}_{svc_id}_{page-1}"))
    if page < total_pages - 1:
        nav.append(_btn(f"Next ▶", callback_data=f"asrvsvc_pg_{api}_{svc_id}_{page+1}"))
    if nav:
        buttons.append(nav)
    if total_pages > 1:
        buttons.append([_btn(f"Page {page+1}/{total_pages}  ({total} countries)", callback_data="noop")])
    buttons.append([
        _btn("➕ Add Country", callback_data=f"aasc_{api}_{svc_id}"),
        _btn("🗑 Del Country",  callback_data=f"dcsc_{api}_{svc_id}")
    ])
    buttons.append([
        _btn("🗑 Del Service", callback_data=f"adss_{api}_{svc_id}"),
        _back_btn("◀️ Back", callback_data=f"admin_srv_{api}")
    ])
    return {"inline_keyboard": buttons}


def get_del_country_from_service_inline(data, api, svc_id, page=0):
    """Delete-selection view: list countries under a service with 🗑 delete buttons (paginated)."""
    PAGE_SIZE = 40
    entries = []
    for c in data.get("countries", []):
        for s in c.get("services", []):
            if s.get("api") == api and s["id"] == svc_id:
                flag = get_country_flag(c["name"])
                entries.append({"cid": c["id"], "cname": c["name"], "price": s.get("price", 0), "flag": flag})
                break
    total       = len(entries)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page        = max(0, min(page, total_pages - 1))
    start       = page * PAGE_SIZE
    page_entries = entries[start:start + PAGE_SIZE]
    buttons = []
    for e in page_entries:
        buttons.append([_btn(
            f"🗑 {e['flag']} {e['cname']}  ₹{e['price']}",
            callback_data=f"adsc_{api}_{svc_id}_{e['cid']}"
        )])
    # Pagination nav
    nav = []
    if page > 0:
        nav.append(_btn(f"◀ Prev", callback_data=f"delcpg_{api}_{svc_id}_{page-1}"))
    if page < total_pages - 1:
        nav.append(_btn(f"Next ▶", callback_data=f"delcpg_{api}_{svc_id}_{page+1}"))
    if nav:
        buttons.append(nav)
    if total_pages > 1:
        buttons.append([_btn(f"Page {page+1}/{total_pages}  ({total} total)", callback_data="noop")])
    buttons.append([_back_btn("◀️ Back", callback_data=f"asrvsvc_{api}_{svc_id}")])
    return {"inline_keyboard": buttons}


def get_number_action_inline(request_id, cid=None, svc_id=None, api=None):
    buy_cb = f"bnn_{cid}_{svc_id}_{api}" if (cid and svc_id and api) else "buy_start"
    return {
        "inline_keyboard": [
            [{"text": "Get SMS Code",    "callback_data": f"chk_{request_id}", "style": "success", "icon_custom_emoji_id": "6235270479079281272"}],
            [{"text": "Cancel & Refund", "callback_data": f"cnc_{request_id}", "style": "danger",  "icon_custom_emoji_id": "5420130255174145507"}],
            [{"text": "Buy New Number",  "callback_data": buy_cb,              "style": "primary", "icon_custom_emoji_id": "5440841102871517055"}],
        ]
    }


def get_sms_received_inline(cid=None, svc_id=None, api=None):
    buy_cb = f"bnn_{cid}_{svc_id}_{api}" if (cid and svc_id and api) else "buy_start"
    return {
        "inline_keyboard": [
            [_btn("✅ SMS Received!", callback_data="noop")],
            [{"text": "Buy New Number", "callback_data": buy_cb, "style": "primary", "icon_custom_emoji_id": "5440841102871517055"}],
        ]
    }


_LEADING_EMOJI_RE = re.compile(
    r'^(?:'
    r'[\U0001F1E0-\U0001F1FF]{2}'          # regional indicator pairs (flags)
    r'|[\U0001F300-\U0001FAFF]'             # misc symbols, pictographs, transport
    r'|[\U00010000-\U0010FFFF]'             # any remaining high-plane code point
    r'|[\u2600-\u26FF]'                     # misc symbols
    r'|[\u2700-\u27BF]'                     # dingbats
    r'|[\u2300-\u23FF]'                     # misc technical
    r'|[\u25A0-\u25FF]'                     # geometric shapes
    r'|[\u2B00-\u2BFF]'                     # misc symbols/arrows
    r'|[\uFE00-\uFE0F]'                     # variation selectors
    r'|\u200D'                              # zero-width joiner
    r')+\s*'
)

def _strip_emoji(name):
    """Remove ALL leading emoji characters and spaces from a string.
    Handles flags, standard emoji, variation selectors, ZWJ sequences,
    skin-tone modifiers, and premium emoji keys.
    """
    # Loop to handle multiple consecutive emoji (e.g. "🟩📱 text")
    while True:
        stripped = _LEADING_EMOJI_RE.sub('', name)
        if stripped == name:
            break
        name = stripped
    return name.strip()


SERVICE_DEFAULT_EMOJI_IDS = {
    "facebook":  "5323261730283863478",
    "fb":        "5323261730283863478",
    "whatsapp":  "5334998226636390258",
    "wa":        "5334998226636390258",
    "telegram":  "5330237710655306682",
    "tg":        "5330237710655306682",
    "imo":       "5337155807752524558",
    "instagram": "5319160079465857105",
    "ig":        "5319160079465857105",
    "apple":     "5334637951894722661",
    "email":     "5348494358205207761",
    "mail":      "5348494358205207761",
    "twitter":   "5330337435500951363",
    "x":         "5330337435500951363",
    "amazon":    "5346056560537779652",
}

def _svc_emoji_id(svc):
    """Return the correct premium emoji_id for a service dict.
    SERVICE_DEFAULT_EMOJI_IDS always wins for known services (by id or name),
    then falls back to the stored emoji_id, then PREMIUM_EMOJIS.
    """
    name   = svc.get("name", "").lower().strip()
    svc_id = svc.get("id", "").lower().strip()
    # strip all leading emoji/symbol characters to get the plain name
    clean_name = re.sub(r'^[\U0001F300-\U0001FAFF\u2600-\u26FF\u2700-\u27BF\s]+', '', name).strip()
    # 1. known-service override (highest priority)
    for key, eid in SERVICE_DEFAULT_EMOJI_IDS.items():
        if svc_id == key or clean_name == key or clean_name.startswith(key):
            return eid
    # 2. stored emoji_id from bot_data
    if svc.get("emoji_id"):
        return svc["emoji_id"]
    # 3. derive from PREMIUM_EMOJIS map
    for em in _EMOJI_KEYS:
        if name.startswith(em) and em in PREMIUM_EMOJIS:
            return PREMIUM_EMOJIS[em]
    return ""


def _build_search_index(data):
    """Build a deduplicated list of service entries for fast text search.
    Called once at startup and whenever data is saved (invalidated)."""
    seen    = {}   # svc_id -> entry dict (first occurrence wins for dedup)
    sources = (
        data.get("countries",    []) +
        data.get("s3_countries", []) +
        data.get("s4_countries", [])
    )
    for c in sources:
        for s in c.get("services", []):
            svc_id = s.get("id", "")
            if not svc_id or svc_id in seen:
                continue
            raw_name   = s.get("name", "")
            name_lower = _strip_emoji(raw_name).lower().strip()
            seen[svc_id] = {
                "svc_id":     svc_id,
                "name":       raw_name,
                "name_lower": name_lower,
                "emoji_id":   _svc_emoji_id(s),
                "style":      s.get("style", "primary"),
            }
    return list(seen.values())


def _get_search_index(data):
    """Return the cached search index, building it lazily if needed."""
    global _search_index
    with _search_lock:
        if _search_index is None:
            _search_index = _build_search_index(data)
        return _search_index


def _invalidate_search_index():
    """Call after any data change so the next search rebuilds the index."""
    global _search_index
    with _search_lock:
        _search_index = None


def _country_emoji_id(country):
    """Return the stored premium emoji_id for a country dict, or derive it from name."""
    if country.get("emoji_id"):
        return country["emoji_id"]
    name = country.get("name", "")
    flag = get_country_flag(name)
    if flag:
        fid = _flag_premium_id(flag)
        if fid:
            return fid
    for em in _EMOJI_KEYS:
        if name.startswith(em) and em in PREMIUM_EMOJIS:
            return PREMIUM_EMOJIS[em]
    if flag and flag in PREMIUM_EMOJIS:
        return PREMIUM_EMOJIS[flag]
    return ""


def get_style_picker_inline():
    """3-button style picker for admin use."""
    return {
        "inline_keyboard": [[
            {"text": "🟢 Success", "callback_data": "stylepick_success"},
            {"text": "🔵 Primary", "callback_data": "stylepick_primary"},
            {"text": "🔴 Danger",  "callback_data": "stylepick_danger"},
        ]]
    }


def get_buy_server_picker_inline(data):
    """Show available servers (S1/S2/S3/S4) as first step in Buy Number flow."""
    countries    = data.get("countries", [])
    s3_countries = data.get("s3_countries", [])
    s4_countries = data.get("s4_countries", [])
    btns = []
    for api_key, cfg in APIS.items():
        if api_key == "s3":
            svcs = [s for c in s3_countries for s in c.get("services", [])]
        elif api_key == "s4":
            svcs = [s for c in s4_countries for s in c.get("services", [])]
        else:
            svcs = [s for c in countries for s in c.get("services", []) if s.get("api") == api_key]
        if svcs:
            btns.append([{
                "text":               cfg["short"],
                "callback_data":      f"buy_srv|{api_key}",
                "style":              "primary",
                "icon_custom_emoji_id": cfg["emoji_id"],
            }])
    if not btns:
        btns = [[{"text": "❌ No servers available", "callback_data": "noop"}]]
    return {"inline_keyboard": btns}


def get_all_services_inline(countries, page=0):
    """Step 1: Show all unique services with pagination (max 20/page, 2 per row). Newest first.
    Services with the same name (different servers/IDs) are merged into one button."""
    PAGE_SIZE = 20
    # Deduplicate by normalised name so same-named services across servers appear once.
    # seen_name maps normalised_name → first svc_id (used as callback representative)
    seen_name  = {}   # normalised_name → first svc_id
    seen_meta  = {}   # first svc_id    → {name, emoji_id, style}
    name_order = []   # insertion order of normalised names
    # Traverse reversed so newest-added appears first
    for c in reversed(countries):
        for s in reversed(c.get("services", [])):
            norm = _strip_emoji(s["name"]).strip().lower()
            if norm not in seen_name:
                seen_name[norm]         = s["id"]
                seen_meta[s["id"]]      = {
                    "name":     s["name"],
                    "emoji_id": _svc_emoji_id(s),
                    "style":    s.get("style", "primary"),
                }
                name_order.append(norm)
    # Build flat list of button entries
    all_btns = []
    for norm in name_order:
        svc_id = seen_name[norm]
        info   = seen_meta[svc_id]
        eid    = info["emoji_id"]
        style  = info.get("style", "primary")
        text   = _strip_emoji(info["name"]) if eid else info["name"]
        btn    = {"text": text, "callback_data": f"svc_{svc_id}", "style": style}
        if eid:
            btn["icon_custom_emoji_id"] = eid
        all_btns.append(btn)
    total       = len(all_btns)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page        = max(0, min(page, total_pages - 1))
    start       = page * PAGE_SIZE
    page_btns   = all_btns[start:start + PAGE_SIZE]
    # Arrange 2 per row
    buttons = []
    row = []
    for btn in page_btns:
        row.append(btn)
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    # Navigation row
    nav = []
    if page > 0:
        nav.append({"text": f"Prev", "callback_data": f"svc_pg_{page-1}", "style": "primary", "icon_custom_emoji_id": "5370726448959085259"})
    if page < total_pages - 1:
        nav.append({"text": f"Next", "callback_data": f"svc_pg_{page+1}", "style": "primary", "icon_custom_emoji_id": "6197296817789474264"})
    if nav:
        buttons.append(nav)
    # Page indicator + Search
    buttons.append([
        {"text": f"{page+1}/{total_pages}", "callback_data": f"svc_pg_picker_{page}", "style": "primary", "icon_custom_emoji_id": "5258477770735885832"},
        {"text": "Search",                     "callback_data": "buy_search",             "style": "success", "icon_custom_emoji_id": "6206446249181189526"},
    ])
    return {"inline_keyboard": buttons}


def get_countries_for_service_inline(service_id, countries, page=0):
    """Step 2: Show all country+server options for a service (matched by name, not just ID).
    S1 multi-price: countries with multiple S1 price entries are grouped into ONE button
    showing a price range; clicking it opens a sub-list of individual prices.
    S4: deduplicated by (country_id, api, operator). Pagination: 10/page, newest first."""
    PAGE_SIZE = 10
    # Resolve the canonical name from service_id so we can match across servers
    target_name = None
    for c in countries:
        for s in c.get("services", []):
            if s["id"] == service_id:
                target_name = _strip_emoji(s["name"]).strip().lower()
                break
        if target_name is not None:
            break
    # Auto Buy button at the very top
    auto_buy_btn = [{"text": "Auto Buy", "callback_data": f"ab_{service_id}", "style": "success",
                     "icon_custom_emoji_id": "5224607267797606837",
                     "right_icon_custom_emoji_id": "5224607267797606837"}]

    # ── Collect entries ──────────────────────────────────────────────────────
    # s1_groups: country_id → {"country": c, "svc_cfg": ..., "prices": [{"price", "svc_id"}]}
    s1_groups        = {}
    s1_group_order   = []        # preserve insertion order for newest-first reversal later
    non_s1_entries   = []
    seen_s4_cid_api_op = set()

    for c in countries:
        for s in c.get("services", []):
            if not (s["id"] == service_id or
                    (target_name and _strip_emoji(s["name"]).strip().lower() == target_name)):
                continue
            api      = s.get("api", "s1")
            operator = s.get("operator", "")
            cfg      = APIS.get(api, APIS["s1"])
            price    = s.get("price", 0)
            cname    = _strip_emoji(c["name"])
            srv_icon  = cfg.get("icon", "")
            srv_short = cfg.get("short", api.upper())
            ceid      = _country_emoji_id(c)
            seid      = cfg.get("emoji_id")
            has_custom_icon = bool(ceid or seid)
            srv_prefix = "" if has_custom_icon else srv_icon

            if api == "s1":
                # Group all S1 price options for this country together
                cid = c["id"]
                if cid not in s1_groups:
                    s1_groups[cid] = {
                        "country":    c,
                        "cname":      cname,
                        "srv_prefix": srv_prefix,
                        "srv_short":  srv_short,
                        "ceid":       ceid,
                        "seid":       seid,
                        "prices":     [],
                    }
                    s1_group_order.append(cid)
                s1_groups[cid]["prices"].append({"price": price, "svc_id": s["id"]})
            elif api == "s4":
                key = (c["id"], api, operator)
                if key in seen_s4_cid_api_op:
                    continue
                seen_s4_cid_api_op.add(key)
                label   = f"{cname}  ₹{price}  {srv_prefix}{srv_short}"
                op_part = f"_{operator}" if operator else ""
                btn     = {"text": label, "callback_data": f"bs_{c['id']}_{s['id']}_{api}{op_part}",
                           "style": c.get("style", "primary")}
                if ceid:   btn["icon_custom_emoji_id"] = ceid
                elif seid: btn["icon_custom_emoji_id"] = seid
                non_s1_entries.append([btn])
            else:
                label = f"{cname}  ₹{price}  {srv_prefix}{srv_short}"
                btn   = {"text": label, "callback_data": f"bs_{c['id']}_{s['id']}_{api}",
                         "style": c.get("style", "primary")}
                if ceid:   btn["icon_custom_emoji_id"] = ceid
                elif seid: btn["icon_custom_emoji_id"] = seid
                non_s1_entries.append([btn])

    # Build S1 entries (grouped)
    s1_entries = []
    for cid in s1_group_order:
        grp    = s1_groups[cid]
        c      = grp["country"]
        prices = grp["prices"]  # list of {"price", "svc_id"}
        ceid   = grp["ceid"]
        seid   = grp["seid"]
        srv_prefix = grp["srv_prefix"]
        srv_short  = grp["srv_short"]
        cname      = grp["cname"]

        if len(prices) == 1:
            # Single price → direct buy button
            p   = prices[0]
            btn = {"text": f"{cname}  ₹{p['price']}  {srv_prefix}{srv_short}",
                   "callback_data": f"bs_{cid}_{p['svc_id']}_s1",
                   "style": c.get("style", "primary")}
        else:
            # Multiple prices → grouped button showing lowest price only
            min_price = min(p["price"] for p in prices)
            btn = {"text": f"{cname}  ₹{min_price}  {srv_prefix}{srv_short}",
                   "callback_data": f"ctry_multi_{service_id}_{cid}",
                   "style": c.get("style", "primary")}
        if ceid:   btn["icon_custom_emoji_id"] = ceid
        elif seid: btn["icon_custom_emoji_id"] = seid
        s1_entries.append([btn])

    # Merge: non-S1 first (preserve original order), then S1 grouped; then reverse all
    entries = non_s1_entries + s1_entries
    entries = list(reversed(entries))

    total       = len(entries)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page        = max(0, min(page, total_pages - 1))
    start       = page * PAGE_SIZE
    buttons     = ([auto_buy_btn] if entries else []) + entries[start:start + PAGE_SIZE]

    # Navigation row
    nav = []
    if page > 0:
        nav.append({"text": "Prev", "callback_data": f"ctry_pg_{service_id}_{page-1}",
                    "style": "primary", "icon_custom_emoji_id": "5370726448959085259"})
    if page < total_pages - 1:
        nav.append({"text": "Next", "callback_data": f"ctry_pg_{service_id}_{page+1}",
                    "style": "primary", "icon_custom_emoji_id": "6197296817789474264"})
    if nav:
        buttons.append(nav)
    # Page indicator + Search
    buttons.append([
        {"text": f"{page+1}/{total_pages}", "callback_data": f"ctry_pg_picker_{service_id}_{page}",
         "style": "primary", "icon_custom_emoji_id": "5258477770735885832"},
        {"text": "Search", "callback_data": f"ctry_search_{service_id}",
         "style": "success", "icon_custom_emoji_id": "6206446249181189526"},
    ])
    # Back button
    buttons.append([
        {"text": "Back", "callback_data": "buy_all",
         "style": "danger", "icon_custom_emoji_id": "5253456541650025535"},
    ])
    return {"inline_keyboard": buttons}


def _ctry_total_pages(service_id, all_countries, page_size=10):
    """Count total pages for a service's country list, matching the S1-grouped logic in
    get_countries_for_service_inline so the page picker always shows the correct count."""
    target_name = None
    for c in all_countries:
        for s in c.get("services", []):
            if s["id"] == service_id:
                target_name = _strip_emoji(s["name"]).strip().lower()
                break
        if target_name:
            break
    seen_s4   = set()
    seen_s1_c = set()   # S1 country IDs already counted (one entry per country after grouping)
    count     = 0
    for c in all_countries:
        for s in c.get("services", []):
            sname = _strip_emoji(s.get("name", "")).strip().lower()
            if not (s["id"] == service_id or (target_name and sname == target_name)):
                continue
            api_c = s.get("api", "s1")
            op    = s.get("operator", "")
            if api_c == "s1":
                if c["id"] not in seen_s1_c:
                    seen_s1_c.add(c["id"])
                    count += 1
            elif api_c == "s4":
                key = (c["id"], api_c, op)
                if key not in seen_s4:
                    seen_s4.add(key)
                    count += 1
            else:
                count += 1
    return max(1, (count + page_size - 1) // page_size)


def get_page_picker_inline(total_pages, current_page, pg_cb_prefix, picker_nav_prefix, service_id_part="", picker_page=0):
    """Show a grid of clickable page numbers.
    If total_pages <= 30: show all at once, no Prev/Next.
    If total_pages > 30: chunk into groups of 30 with custom-emoji Prev/Next."""
    PICKER_SIZE = 30
    start = picker_page * PICKER_SIZE
    end   = min(start + PICKER_SIZE, total_pages)
    buttons = []
    row = []
    for pg in range(start, end):
        btn = {
            "text":          str(pg + 1),
            "callback_data": f"{pg_cb_prefix}{pg}",
            "style":         "success" if pg == current_page else "primary",
        }
        row.append(btn)
        if len(row) == 5:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    # Prev / Next only when total_pages > 30
    if total_pages > 30:
        nav = []
        if picker_page > 0:
            nav.append({"text": "Prev", "callback_data": f"{picker_nav_prefix}{service_id_part}{picker_page - 1}",
                        "style": "danger",  "icon_custom_emoji_id": "5253997076169115797"})
        if end < total_pages:
            nav.append({"text": "Next", "callback_data": f"{picker_nav_prefix}{service_id_part}{picker_page + 1}",
                        "style": "success", "icon_custom_emoji_id": "6239913824057563150"})
        if nav:
            buttons.append(nav)
    return {"inline_keyboard": buttons}


def get_submit_payment_inline():
    return {"inline_keyboard": [[{
        "text": "Submit Payment Proof",
        "callback_data": "depo_submit",
        "style": "primary",
        "icon_custom_emoji_id": "6199249357166874962"
    }]]}


def get_admin_depo_inline(user_id, amount, utr):
    return {"inline_keyboard": [[
        _btn(f"✅ Approve ₹{amount}", callback_data=f"da_{user_id}_{amount}_{utr}"),
        _btn("❌ Reject",              callback_data=f"dd_{user_id}")
    ]]}


# All real premium emoji IDs provided by the bot owner.
# Format: (display_char, premium_id)
# Each entry becomes one button in the picker.
# icon_custom_emoji_id shows the actual premium animated version on the button.
PICKER_EMOJIS_DATA = [
    ("●", "5334807341109908955"),
    ("●", "5334759662677957452"),
    ("●", "5337010556253543833"),
    ("●", "5336814486701514414"),
    ("●", "5337155807752524558"),
    ("●", "5334868205091459431"),
    ("●", "5334637951894722661"),
    ("●", "5334530732331143967"),
    ("●", "5334944492300573096"),
    ("●", "5334584041465222978"),
    ("●", "5334769042886533147"),
    ("●", "5335010201005231986"),
    ("●", "5334763399299506604"),
    ("●", "5334880948259427772"),
    ("●", "5334590977837403844"),
    ("●", "5334780609233460393"),
    ("●", "5337102391244263212"),
    ("●", "5336983442125001376"),
    ("●", "5337172996211648018"),
    ("●", "5337302974806922068"),
    ("●", "5337255927735163754"),
    ("●", "5337132498965010628"),
    ("●", "5336850036145823599"),
    ("●", "5336997731481193790"),
    ("●", "5336944168944047463"),
    ("●", "5336972142066047577"),
    ("●", "5337267511261960341"),
    ("●", "5336879280578138635"),
    ("●", "5339213256001102461"),
    ("●", "5341526903343847656"),
    ("●", "5343835156207669196"),
    ("●", "5343900431120633475"),
    ("●", "5343958606452658750"),
    ("●", "5343646336560434469"),
    ("●", "5345970940364770589"),
    ("●", "5348390922507817684"),
    ("●", "5346335574498251610"),
    ("●", "5346250203433309111"),
    ("●", "5348486915026884464"),
    ("●", "5346311574221000149"),
    ("●", "5348075478634766440"),
    ("●", "5348360226376556295"),
    ("●", "5350655873511362387"),
    ("●", "5467538555158943525"),
    ("●", "5244837092042750681"),
    ("●", "5246762912428603768"),
    ("●", "5397782960512444700"),
    ("●", "5264919878082509254"),
    ("●", "5416081784641168838"),
    ("●", "5411225014148014586"),
    ("●", "5282843764451195532"),
    ("●", "5337080053119336309"),
    ("●", "5370803230089436998"),
    ("●", "5370847309338798719"),
    ("●", "5372849034681672140"),
    ("●", "5371078915450233360"),
    ("●", "5370691118558110746"),
    ("●", "5384466886857614542"),
    ("●", "5370700219593812158"),
    ("●", "5370957204667003973"),
    ("●", "5370563845792227799"),
    ("●", "5373303446516560657"),
    ("●", "5375131156964468313"),
    ("●", "5373098855749414389"),
    ("●", "5375116747349188794"),
    ("●", "5384300435400058254"),
    ("●", "5388614120163533453"),
    ("●", "6206150489143252478"),
    ("●", "6190336264940559752"),
    ("●", "6206479140040743133"),
    ("●", "6206508629286196237"),
    ("●", "6206122090819490939"),
    ("●", "6203738495639360972"),
    ("●", "6206112371308500200"),
    ("●", "6156611653451387536"),
    ("●", "5300949181907616610"),
    ("●", "5341511312612540947"),
    ("●", "5386351922234020397"),
    ("●", "5456135459937733007"),
    ("●", "5456470737969758983"),
    ("●", "5323589698281570818"),
    ("●", "5323510761077636002"),
    ("●", "5323262125420871113"),
    ("●", "5300762754557167595"),
    ("●", "5330032308139343696"),
    ("●", "5330411970363410701"),
    ("●", "5190610896547760543"),
    ("●", "5203934795633020471"),
    ("●", "5330470948854314862"),
    ("●", "5330348752739779452"),
    ("●", "5330071645744807685"),
    ("●", "5454218568788886144"),
    ("●", "5332556245670837192"),
    ("●", "5332645924587978526"),
    ("●", "5330029486345830013"),
    ("●", "5332536368562192411"),
    ("●", "5204129207327671929"),
    ("●", "5454365533979825405"),
    ("●", "5332656211034652890"),
    ("●", "5332414412965824001"),
    ("●", "5454152160004555979"),
    ("●", "5190786822703173195"),
    ("●", "5453963997487317976"),
    ("●", "5190949975625845488"),
    ("●", "5190537663060394979"),
    ("●", "5192859659819567730"),
    ("●", "5303368360891794776"),
    ("●", "5206491830182494639"),
    ("●", "5453980756449705907"),
    ("●", "5190489756995175984"),
    ("●", "5332726004253212527"),
    ("●", "5330491732201062615"),
    ("●", "5192792422106547477"),
    ("●", "5193062876197177282"),
    ("●", "5454355243238179258"),
    ("●", "5454370215494175251"),
    ("●", "5206364123624914906"),
    ("●", "5451814813032268395"),
    ("●", "5204403355090177407"),
    ("●", "5204233252910418010"),
    ("●", "5204072788637272243"),
    ("●", "5190511227536692244"),
    ("●", "5193031329662389224"),
    ("●", "5451925459979750066"),
    ("●", "5204162201266439163"),
    ("●", "5204053607313330638"),
    ("●", "5206296293206411360"),
    ("●", "5204122051912159295"),
    ("●", "5454008471873669821"),
    ("●", "5454100049166357274"),
    ("●", "5454081378943518859"),
    ("●", "5456157115162839306"),
    ("●", "5456174900622412791"),
    ("●", "5453895913665743004"),
    ("●", "5454275610249546209"),
    ("●", "5453902695419104908"),
    ("●", "5454051000639834518"),
    ("●", "5341513541700569186"),
    ("●", "5454150033995744549"),
    ("●", "5300880007164350042"),
    ("●", "5330511196992846570"),
    ("●", "5332543287754506739"),
    ("●", "5330285088439550100"),
    ("●", "5332571548639314320"),
    ("●", "5332307700208388836"),
    ("●", "5330319224839618247"),
    ("●", "5454205520678242808"),
    ("●", "5453962468478962129"),
    ("●", "5192687732278708280"),
    ("●", "5190842343745406919"),
    ("●", "5190608405466726992"),
    ("●", "5192986434369242588"),
    ("●", "5453941478973785670"),
    ("●", "5314490264604142366"),
    ("●", "5190668814181746392"),
    ("●", "5300973955278982398"),
    ("●", "5310168299143723285"),
    ("●", "5332581392704360215"),
    ("●", "5332441144842277218"),
    ("●", "5463219974132746636"),
    ("●", "5462911088674753179"),
    ("●", "5330329455451721164"),
    ("●", "5332275187305958945"),
    ("●", "5377562086979161643"),
    ("●", "5462948059753237765"),
    ("●", "5463072798488411534"),
    ("●", "5332639202964161592"),
    ("●", "5332687864943626828"),
    ("●", "5355127939488815433"),
    ("●", "5314378500965145730"),
    ("●", "5936299373267783023"),
    ("●", "5938138877695888929"),
    ("●", "5938213545702329918"),
    ("●", "5938201339405274907"),
    ("●", "5938173838729678430"),
    ("●", "5935868283105316300"),
    ("●", "5936011043523269585"),
    ("●", "5936275226961645297"),
    ("●", "5938035098401116139"),
    ("●", "5938063389350695244"),
    ("●", "5938362168750640752"),
    ("●", "5938436162447217183"),
    ("●", "5938078052369043009"),
    ("●", "5936135812323216910"),
    ("●", "5936085221903437430"),
    ("●", "5938048670497770888"),
    ("●", "5938101648919368057"),
    ("●", "5938472854352825982"),
    ("●", "5935856879967145175"),
    ("●", "5935787872727601571"),
    ("●", "5938250499600946517"),
    ("●", "5935846116779101893"),
    ("●", "5938216676733488820"),
    ("●", "5938302971216400756"),
    ("●", "5940754379635166866"),
    ("●", "5936278860503977292"),
    ("●", "5938111239581339701"),
    ("●", "5938164479995941036"),
    ("●", "5935982473400815138"),
    ("●", "5940292305578627456"),
    ("●", "5323379047315555501"),
    ("●", "5334954057192719331"),
    ("●", "5335012820935261564"),
    ("●", "5323764984486837459"),
    ("●", "5325612636467903082"),
    ("●", "5323261730283863478"),
    ("●", "5330081305126255700"),
    ("●", "5334750591707005005"),
    ("●", "5319160079465857105"),
    ("●", "5334933574493683027"),
    ("●", "5325726234057915652"),
    ("●", "5323608076446613036"),
    ("●", "5325647210954637197"),
    ("●", "5334998226636390258"),
    ("●", "5330337435500951363"),
    ("●", "5334955749409834455"),
    ("●", "5332394707655869572"),
    ("●", "5346242859039209592"),
    ("●", "5346309375197725525"),
    ("●", "5357184657592962696"),
    ("●", "5318911503938634641"),
    ("●", "5346134750417403743"),
    ("●", "5346242644290846992"),
    ("●", "5346251367369425932"),
    ("●", "5345799562579688546"),
    ("●", "5345937224871461019"),
    ("●", "5346259862814734771"),
    ("●", "5345844509412444249"),
    ("●", "5346319945112240722"),
]

# Reverse lookup: premium_id → display_char
PICKER_ID_TO_CHAR = {pid: char for char, pid in PICKER_EMOJIS_DATA}

# ── Country flag picker (A-Z flags only) ──────────────────
COUNTRY_FLAG_PICKER_EMOJIS_DATA = [
    # A–K (from user-provided premium IDs)
    ("🇦🇫", "5291937511591925566"), ("🇦🇱", "5294202819077756005"), ("🇩🇿", "5294048127240655242"),
    ("🇦🇸", "5291994273879709721"), ("🇦🇩", "5294215205763434181"), ("🇦🇴", "5294516785482062829"),
    ("🇦🇮", "5292186323342350940"), ("🇦🇬", "5294005972136647964"), ("🇦🇷", "5292208210495689627"),
    ("🇦🇲", "5291978717508164018"), ("🇦🇼", "5294007002928798927"), ("🇦🇺", "5294444247779399477"),
    ("🇦🇹", "5291975174160145850"), ("🇦🇿", "5294323533428579078"), ("🇧🇸", "5294031587321600012"),
    ("🇧🇭", "5294108398516720753"), ("🇧🇩", "5291824687096027834"), ("🇧🇧", "5294526187165471742"),
    ("🇧🇾", "5294134426018536120"), ("🇧🇪", "5291774466043435275"), ("🇧🇿", "5294171848068584842"),
    ("🇧🇯", "5293984969746566866"), ("🇧🇹", "5294121983498277263"), ("🇧🇴", "5294201479047957700"),
    ("🇧🇼", "5294026179957772585"), ("🇧🇷", "5291892229751723900"), ("🇧🇳", "5292098293692650297"),
    ("🇧🇬", "5294308947719640437"), ("🇧🇫", "5294153164960848949"), ("🇧🇮", "5294051631933967760"),
    ("🇰🇭", "5294225191562400452"), ("🇨🇲", "5291997306126626950"), ("🇨🇦", "5292290347450259214"),
    ("🇨🇻", "5292203503211535593"), ("🇨🇫", "5294210571493724819"), ("🇹🇩", "5291780728105753403"),
    ("🇨🇱", "5294231037012888049"), ("🇨🇳", "5294068833277990704"), ("🇨🇴", "5294010206974397371"),
    ("🇰🇲", "5294351381996521508"), ("🇨🇬", "5294035229453865597"), ("🇨🇰", "5292098684534675100"),
    ("🇨🇷", "5292063805105263554"), ("🇨🇮", "5293991322003200135"), ("🇭🇷", "5291999676948569127"),
    ("🇨🇺", "5291963947115631526"), ("🇨🇾", "5294062721539526918"), ("🇨🇿", "5294242852467923382"),
    ("🇩🇰", "5294531860817268837"), ("🇩🇯", "5294127214768468283"), ("🇩🇲", "5294485513825178032"),
    ("🇩🇴", "5294522197140857947"), ("🇪🇨", "5292083733753517221"), ("🇪🇬", "5293992082212409502"),
    ("🇸🇻", "5294337307388695687"), ("🇬🇶", "5292170045416297012"), ("🇪🇷", "5291922054004625949"),
    ("🇪🇪", "5291951143818123103"), ("🇪🇹", "5292245976143124155"), ("🇪🇺", "5291992809295861098"),
    ("🇫🇮", "5294049961191690629"), ("🇫🇷", "5291817660529533837"), ("🇬🇦", "5294321325815389139"),
    ("🇬🇪", "5294349389131697267"), ("🇩🇪", "5292013274815028523"), ("🇬🇭", "5294347396266873249"),
    ("🇬🇮", "5292055799286224027"), ("🇬🇱", "5292014752283774878"), ("🇬🇲", "5294399820637688352"),
    ("🇬🇷", "5291948395039054764"), ("🇬🇼", "5294409819321550432"), ("🇬🇹", "5294336633078831209"),
    ("🇬🇳", "5291892096607739008"), ("🇬🇾", "5292062692708736193"), ("🇭🇹", "5292045130587462814"),
    ("🇭🇳", "5291901034434682297"), ("🇭🇰", "5292166459118606932"), ("🇭🇺", "5294229581018975260"),
    ("🇮🇸", "5294354358408859664"), ("🇮🇳", "5291933173674957761"), ("🇮🇷", "5294220170745630736"),
    ("🇮🇶", "5294325010897327367"), ("🇮🇪", "5294471971793293647"), ("🇮🇱", "5294069056616289553"),
    ("🇮🇹", "5291826830284709120"), ("🇯🇲", "5294505107465982830"), ("🇯🇵", "5291799063321139445"),
    ("🇯🇴", "5291988613112814801"), ("🇰🇿", "5294227175837290463"), ("🇰🇪", "5292111852904416801"),
    ("🇰🇮", "5294538934628405146"), ("🇰🇵", "5294193812531333564"), ("🇰🇷", "5294408281723262763"),
    ("🇰🇼", "5292066437920218075"), ("🇰🇬", "5292091954320922577"),
    # L–Z (existing entries)
    ("🇱🇨", "5222000927023577045"), ("🇱🇰", "5224277294050192388"),
    ("🇲🇦", "5224530035695693965"), ("🇲🇩", "5224216473018314447"), ("🇲🇪", "5224463399278096980"),
    ("🇲🇭", "5224538449536624503"), ("🇲🇱", "5224322352552096671"), ("🇫🇲", "5222280486444873367"),
    ("🇲🇨", "5221937224068640464"), ("🇲🇳", "5224192257992701543"), ("🇧🇲", "5222482143749353810"),
    ("🇲🇹", "5224731388057497620"), ("🇲🇺", "5224238347286752315"), ("🇲🇽", "5221971386238514431"),
    ("🇲🇿", "5222470388423864826"), ("🇳🇦", "5224690826386351746"), ("🇳🇪", "5222099049846420864"),
    ("🇳🇬", "5224723614166691638"), ("🇳🇱", "5224516489368841614"), ("🇳🇴", "5224465228934163949"),
    ("🇳🇵", "5222444378101925267"), ("🇳🇿", "5224573595254009705"), ("🇴🇲", "5222396686785066306"),
    ("🇵🇦", "5222111719999945107"), ("🇵🇬", "5224500164198149905"), ("🇵🇰", "5224637061985742245"),
    ("🇵🇪", "5224482026551258766"), ("🇵🇭", "5222065042295376892"), ("🇵🇱", "5224670399521892983"),
    ("🇵🇷", "5224220115150582423"), ("🇵🇸", "5222041677673282461"), ("🇵🇹", "5224404094369672274"),
    ("🇵🇾", "5222152565138929235"), ("🇶🇦", "5222225596762830469"), ("🇷🇴", "5222273794885826118"),
    ("🇷🇸", "5222145396838512729"), ("🇷🇺", "5280582975270963511"), ("🇷🇼", "5222449197055227754"),
    ("🇸🇦", "5224698145010624573"), ("🇸🇧", "5222290588207954120"), ("🇸🇨", "5224467496676896871"),
    ("🇸🇩", "5224372990216514135"), ("🇸🇪", "5222201098269373561"), ("🇸🇬", "5224194023224257181"),
    ("🇸🇮", "5224660718665607511"), ("🇸🇰", "5222401879400528047"), ("🇸🇱", "5224420995065983217"),
    ("🇸🇳", "5224358988623130949"), ("🇸🇴", "5222370504664428325"), ("🇸🇷", "5224567367551428669"),
    ("🇸🇸", "5224618146949773268"), ("🇸🇹", "5221953304426198315"), ("🇸🇻", "5294337307388695687"),
    ("🇨🇭", "5224707263226194753"), ("🇹🇭", "5224638530864556281"), ("🇹🇯", "5222217865821696536"),
    ("🇹🇲", "5224256935905208951"), ("🇹🇳", "5221991375016310330"), ("🇹🇬", "5222408051268532030"),
    ("🇹🇹", "5224391883777651050"), ("🇹🇿", "5224397364155923150"), ("🇺🇦", "5222250679371839695"),
    ("🇺🇬", "5222464040462200940"), ("🇬🇧", "5224518800061245598"), ("🇺🇳", "5451772687993031127"),
    ("🇺🇸", "5224321781321442532"), ("🇺🇾", "5222466849370813232"), ("🇺🇿", "5222404546575219535"),
    ("🇻🇦", "5222420266155520507"), ("🇻🇨", "5224541228380467535"), ("🇻🇮", "5224395882392201810"),
    ("🇻🇳", "5222359651282071925"), ("🇻🇺", "5222126748090512778"), ("🇼🇸", "5224660353593387686"),
    ("🇾🇪", "5222300655611294950"), ("🇿🇦", "5224696216570309138"), ("🇿🇲", "5224646626877911277"),
    ("🇦🇪", "5224565851427976312"), ("🇹🇷", "5224601903383457698"),
]

COUNTRY_FLAG_PICKER_ID_TO_CHAR = {pid: char for char, pid in COUNTRY_FLAG_PICKER_EMOJIS_DATA}
# Forward lookup: flag char → premium picker ID
COUNTRY_FLAG_CHAR_TO_PICKER_ID = {char: pid for char, pid in COUNTRY_FLAG_PICKER_EMOJIS_DATA}
# Merge into main reverse-lookup so epick_ handler resolves these IDs too
PICKER_ID_TO_CHAR.update(COUNTRY_FLAG_PICKER_ID_TO_CHAR)

# ── Mail-specific emoji picker data ───────────────────────
MAIL_PICKER_EMOJIS_DATA = [
    ("📧", "5323261730283863478"),
    ("📧", "5337155807752524558"),
    ("📧", "5319160079465857105"),
    ("📧", "5334637951894722661"),
    ("🖥",  "5334880948259427772"),
    ("📧", "5335010201005231986"),
    ("📧", "5334944492300573096"),
    ("🍏", "5337132498965010628"),
    ("💸", "5348212415077064131"),
    ("🥏", "5345970940364770589"),
    ("🥚", "5348390922507817684"),
    ("⚒",  "5346250203433309111"),
    ("⚗",  "5346311574221000149"),
    ("🛴", "5348075478634766440"),
    ("😕", "5348360226376556295"),
    ("✨", "5352552689983067014"),
    ("👤", "5269334245469362034"),
    ("🎥", "5269514462297104098"),
    ("📞", "5269657987219232606"),
    ("✈️", "5269215244810491516"),
    ("📸", "5269226403135528711"),
    ("🍎", "5269759232483303288"),
    ("💬", "5443038326535759644"),
    ("🛡",  "5251203410396458957"),
    ("📱", "6156948967297913424"),
    ("📱", "6156588752685765320"),
    ("📱", "6156673548225090260"),
    ("📱", "6159184811308093485"),
    ("📱", "6156561806060950710"),
    ("📱", "6159103318598619839"),
    ("📱", "6159119252927289026"),
    ("📱", "6156667045644604564"),
    ("📱", "6156667685594731199"),
    ("📱", "6156546644826397723"),
    ("📱", "6156447963657803606"),
    ("📱", "6156940824039920226"),
    ("🔥", "5251560184739818113"),
    ("👑", "5253809746875542344"),
    ("⚡", "5251494085193132432"),
    ("🤔", "5251386049585768540"),
    ("🔔", "5251684060186569219"),
    ("🎩", "5251660768578927322"),
    ("🖤", "5251411875224119501"),
    ("🧡", "5253618530636559531"),
    ("😌", "5253468455889301185"),
    ("✅", "5251685816828194329"),
    ("⭐", "5251363999223671175"),
    ("💀", "5251514997388897766"),
    ("🍭", "5253601909113122325"),
    ("💎", "5253770675558049501"),
    ("💎", "5253811739740366123"),
    ("🔤", "5269734347442779086"),
    ("🤖", "5314391089514291948"),
    ("🎁", "5316520749048012362"),
    ("☎️", "5318765591014678496"),
    ("📩", "5319185587276619437"),
]

MAIL_PICKER_ID_TO_CHAR = {pid: char for char, pid in MAIL_PICKER_EMOJIS_DATA}


def get_mail_emoji_picker_inline():
    """Inline keyboard for mail service emoji selection (4 per row) + Skip."""
    buttons = []
    # Custom ID button is FIRST (top of keyboard)
    buttons.append([
        {"text": "✏️ Custom ID", "callback_data": "mail_epick_custom", "style": "primary"},
        {"text": "⏭ Skip (no emoji)", "callback_data": "mail_epick_skip"},
    ])
    row = []
    for char, pid in MAIL_PICKER_EMOJIS_DATA:
        row.append({
            "text": char,
            "callback_data": f"mail_epick_{pid}",
            "icon_custom_emoji_id": pid,
        })
        if len(row) == 4:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    return {"inline_keyboard": buttons}


def get_emoji_picker_inline():
    """
    Inline keyboard with all real premium emoji IDs (4 per row) + Skip.
    Each button uses icon_custom_emoji_id so Telegram shows the actual
    premium animated version on the left side of the button.
    callback_data = epick_<numeric_id>  (not the emoji char)
    """
    buttons = []
    # Custom ID button is FIRST (top of keyboard)
    buttons.append([
        {"text": "✏️ Custom ID", "callback_data": "epick_custom", "style": "primary"},
        {"text": "⏭ Skip (no emoji)", "callback_data": "epick_skip"},
    ])
    row = []
    for char, pid in PICKER_EMOJIS_DATA:
        row.append({
            "text": char,
            "callback_data": f"epick_{pid}",
            "icon_custom_emoji_id": pid,
        })
        if len(row) == 4:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    return {"inline_keyboard": buttons}


def get_country_emoji_picker_inline(page=0):
    """
    Flag-only emoji picker for the 'add country' step.
    Shows country flag premium emojis in 4-per-row layout.
    Paginated (24 flags per page = 96 buttons max) to stay within Telegram's ~100 button limit.
    """
    PAGE_SIZE = 96  # 24 rows × 4 cols
    total     = len(COUNTRY_FLAG_PICKER_EMOJIS_DATA)
    pages     = max(1, -(-total // PAGE_SIZE))  # ceiling division
    start     = page * PAGE_SIZE
    end       = min(start + PAGE_SIZE, total)
    buttons   = []
    # Custom ID button is FIRST (top of keyboard)
    buttons.append([
        {"text": "✏️ Custom ID", "callback_data": "epick_custom", "style": "primary"},
        {"text": "⏭ Skip (no emoji)", "callback_data": "epick_skip"},
    ])
    row       = []
    for char, pid in COUNTRY_FLAG_PICKER_EMOJIS_DATA[start:end]:
        row.append({
            "text": char,
            "callback_data": f"epick_{pid}",
        })
        if len(row) == 4:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    nav = []
    if page > 0:
        nav.append({"text": "◀️ Prev", "callback_data": f"epick_pg_{page - 1}", "style": "danger"})
    if page < pages - 1:
        nav.append({"text": "Next ▶️", "callback_data": f"epick_pg_{page + 1}", "style": "primary"})
    if nav:
        buttons.append(nav)
    return {"inline_keyboard": buttons}


def get_auto_flag_keyboard(cname):
    """
    Smart flag auto-suggest keyboard.
    Detects the country flag from cname, returns a single-flag keyboard.
    Returns None if no flag detected (caller should fall back to full picker).
    """
    flag_char = get_country_flag(cname)
    if not flag_char or flag_char == "🌍":
        return None
    pid = COUNTRY_FLAG_CHAR_TO_PICKER_ID.get(flag_char)
    if not pid:
        return None
    return {"inline_keyboard": [
        [{"text": flag_char, "callback_data": f"epick_{pid}"}],
        [{"text": "🔍 Pick Different Flag", "callback_data": "epick_pg_0"}],
        [{"text": "⏭ Skip (no flag)", "callback_data": "epick_skip"}],
    ]}


# ============= STATE & BACKGROUND =============

user_states  = {}
_user_locks  = {}          # chat_id → threading.Lock()
_ulocks_meta = threading.Lock()   # protects _user_locks dict itself

def _get_user_lock(chat_id):
    with _ulocks_meta:
        if chat_id not in _user_locks:
            _user_locks[chat_id] = threading.Lock()
        return _user_locks[chat_id]

pending_sms  = {}
_pending_sms_lock = threading.Lock()   # protects pending_sms dict
pending_mail = {}   # mail_id → {chat_id, service, mail_address}
temp_mails    = {}   # chat_id → {address, password, token}
pending_tmail = {}   # chat_id → {address, token, msg_id, started_at, known_ids}
_pending_tmail_lock = threading.Lock()
pending_free_nums = {}   # number → {chat_id, msg_id, srv_name, started_at, known_msgs}
_pending_free_lock = threading.Lock()
pending_nsrv_nums = {}   # number → {chat_id, msg_id, data_snap, started_at}
_pending_nsrv_lock = threading.Lock()
_seen_console_hits = set()   # fingerprints of already-forwarded console OTPs
_last_console_otp_time = {}  # (svc_key, country_cc) → unix timestamp of last OTP
TMAIL_EXPIRE_SECS = 3600   # 1 hour


# ============= TEMP MAIL (mail.tm) =============

MAILTM_BASE = "https://api.mail.tm"

def _mailtm_post(path, payload):
    try:
        body = json.dumps(payload).encode()
        req  = urllib.request.Request(
            f"{MAILTM_BASE}{path}", data=body,
            headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except Exception as e:
        return {"error": str(e)}

def _mailtm_get(path, token=None):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        req = urllib.request.Request(f"{MAILTM_BASE}{path}", headers=headers)
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except Exception as e:
        return {"error": str(e)}

def create_temp_mail():
    domains = _mailtm_get("/domains")
    if "error" in domains or not domains.get("hydra:member"):
        return None
    domain  = domains["hydra:member"][0]["domain"]
    import random, string
    user    = ''.join(random.choices(string.ascii_lowercase + string.digits, k=10))
    address = f"{user}@{domain}"
    passwd  = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
    acc = _mailtm_post("/accounts", {"address": address, "password": passwd})
    if "error" in acc and "hydra:description" not in str(acc):
        return None
    tok = _mailtm_post("/token", {"address": address, "password": passwd})
    token = tok.get("token")
    if not token:
        return None
    return {"address": address, "password": passwd, "token": token}

def fetch_temp_inbox(token):
    msgs = _mailtm_get("/messages", token=token)
    return msgs.get("hydra:member", [])

def fetch_temp_message(token, msg_id):
    return _mailtm_get(f"/messages/{msg_id}", token=token)

def get_temp_mail_inline(chat_id):
    return {"inline_keyboard": [
        [{"text": "📩 Check Inbox", "callback_data": f"tmail_check_{chat_id}", "style": "primary"}],
        [{"text": "🔄 New Email",   "callback_data": f"tmail_new_{chat_id}",   "style": "danger"}],
    ]}

def get_tmail_read_inline(chat_id, msg_id):
    return {"inline_keyboard": [
        [{"text": "📖 Read Full Message", "callback_data": f"tmail_read_{msg_id}", "style": "primary"}],
        [{"text": "📩 Back to Inbox",     "callback_data": f"tmail_check_{chat_id}", "style": "danger"}],
    ]}


def get_api_for_request(data, chat_id, req_id):
    uid = str(chat_id)
    for h in data.get("users", {}).get(uid, {}).get("history", []):
        if h.get("req_id") == req_id:
            return h.get("api", "s1")
    return "s1"


AUTO_CANCEL_SECS  = 600   # 10 minutes — auto-cancel all numbers
MANUAL_CANCEL_WAIT = 180  # 3 minutes — default min wait before manual cancel (overridden by admin)
TIMER_EDIT_EVERY  = 15    # edit message every 15 seconds

def get_cancel_wait(data, api_id=None):
    """Return cancel wait time in seconds. Per-service if api_id given, else global fallback."""
    if api_id:
        per = data.get("cancel_wait", {})
        secs = per.get(api_id)
        if secs is not None:
            try:
                return max(0, int(secs))
            except (TypeError, ValueError):
                pass
    mins = data.get("cancel_wait_mins")
    try:
        return max(0, int(mins)) * 60
    except (TypeError, ValueError):
        return MANUAL_CANCEL_WAIT


def _cw_disp_str(secs):
    """Format seconds into human-readable cancel wait string."""
    if secs == 0:
        return "instant"
    m, s = divmod(int(secs), 60)
    return f"{m}m {s}s" if s else f"{m}m"


def get_cancel_time_srv_inline(data):
    """Service selection keyboard for changing per-service cancel time."""
    cw = data.get("cancel_wait", {})
    global_secs = get_cancel_wait(data)
    def _disp(api_id):
        secs = cw.get(api_id)
        if secs is None:
            return f"default ({_cw_disp_str(global_secs)})"
        return _cw_disp_str(secs)
    return {
        "inline_keyboard": [
            [_btn(f"🔵 Server 1 (SmsBower)   ⏱ {_disp('s1')}", callback_data="sct_s1")],
            [_btn(f"🟢 Server 2 (HeroSMS)    ⏱ {_disp('s2')}", callback_data="sct_s2")],
            [_btn(f"🟡 Server 3 (SmsIndia)   ⏱ {_disp('s3')}", callback_data="sct_s3")],
            [_btn(f"🟤 Server 4 (UotpStore)  ⏱ {_disp('s4')}", callback_data="sct_s4")],
            [_btn(f"🟣 Server 5 (SmsProvider)⏱ {_disp('s5')}", callback_data="sct_s5")],
            [_back_btn("◀️ Back", callback_data="sm_more")],
        ]
    }

_last_edit      = {}              # req_id → last edit timestamp
_last_edit_lock = threading.Lock()  # protects _last_edit
_seen_updates   = set()           # dedup: update_ids already processed


def _number_msg(info, remaining):
    mins     = int(remaining // 60)
    secs     = int(remaining % 60)
    svc      = info.get("service", "")
    ctr      = info.get("country", "")
    svc_icon = _free_svc_icon_char(svc)
    ctr_flag = get_country_flag(ctr)
    return (
        f"✅ *Number Ready!*\n━━━━━━━━\n\n"
        f"{ctr_flag} {svc_icon} `{info.get('number','')}`\n"
        f"━━━━━━━━\n"
        f"⏳ Waiting for SMS...\n"
        f"⏱ Auto-cancel in *{mins}:{secs:02d}*"
    )


def _do_auto_cancel(req_id, info):
    chat_id = info["chat_id"]
    api     = info.get("api", "s1")
    msg_id  = info.get("msg_id")
    price   = info.get("price", 0)
    try:
        call_api("setStatus", extra_params={"id": req_id, "status": 8}, api=api)
    except Exception:
        pass
    # Acquire user lock to prevent race condition with user-triggered actions
    with _get_user_lock(chat_id):
        d   = load_data()
        uid = str(chat_id)
        already_refunded = False
        for h in d.get("users", {}).get(uid, {}).get("history", []):
            if h.get("req_id") == req_id:
                if h.get("status") in ("cancelled", "refunded"):
                    already_refunded = True
                else:
                    h["status"] = "cancelled"
                break
        user = get_user(d, chat_id)
        if price > 0 and not already_refunded:
            user["balance"] = round(user["balance"] + price, 2)
        save_data(d)
    cfg = APIS.get(api, APIS["s1"])
    txt = (
        f"⏰ *Auto-Cancelled*\n━━━━━━━━\n\n"
        f"📞 `{info.get('number','')}`\n"
        f"{cfg['icon']} {cfg['short']}\n\n"
        f"💰 ₹{price} refunded  |  Balance: ₹{user['balance']}\n\n"
        f"_No SMS received within 10 minutes._"
    )
    if msg_id:
        edit_message(chat_id, msg_id, txt)
    else:
        send_message(chat_id, txt)


# ============= 2oo9.cloud NUMBER API =============

TNEVS_BASE = "https://api.2oo9.cloud/MXS47FLFX0U/tnevs/@public/api"

def get_tnevs_settings(data):
    return data.get("tnevs_settings", {"api_key": ""})


def _tnevs_headers(api_key):
    return {**HEADERS, "mauthapi": api_key, "Content-Type": "application/json"}


def tnevs_get_number(data, rid):
    """POST /getnum — allocate one number from a range. Returns {"number": {...}} or {"error": ...}."""
    api_key = get_tnevs_settings(data).get("api_key", "")
    if not api_key:
        return {"error": "2oo9 API key not configured. Contact admin."}
    try:
        url  = f"{TNEVS_BASE}/getnum"
        body = json.dumps({"rid": str(rid)}).encode()
        req  = urllib.request.Request(url, body, headers=_tnevs_headers(api_key))
        with urllib.request.urlopen(req, timeout=25) as resp:
            r = json.loads(resp.read().decode())
        if r.get("meta", {}).get("code") == 200:
            return {"number": r.get("data", {})}
        return {"error": r.get("message", str(r))}
    except urllib.error.HTTPError as e:
        try:
            body_err = json.loads(e.read().decode())
            return {"error": body_err.get("message", f"HTTP {e.code}")}
        except Exception:
            return {"error": f"HTTP {e.code}: {e.reason}"}
    except Exception as e:
        return {"error": str(e)}


def tnevs_get_my_otps(data):
    """GET /success-otp — your last 50 OTPs. Returns {"otps": [...]} or {"error": ...}."""
    api_key = get_tnevs_settings(data).get("api_key", "")
    if not api_key:
        return {"error": "2oo9 API key not configured."}
    try:
        req = urllib.request.Request(f"{TNEVS_BASE}/success-otp", headers=_tnevs_headers(api_key))
        with urllib.request.urlopen(req, timeout=15) as resp:
            r = json.loads(resp.read().decode())
        if r.get("meta", {}).get("code") == 200:
            return {"otps": r.get("data", {}).get("otps", [])}
        return {"error": r.get("message", str(r))}
    except urllib.error.HTTPError as e:
        return {"error": f"HTTP {e.code}: {e.reason}"}
    except Exception as e:
        return {"error": str(e)}


def tnevs_live_access(data):
    """GET /liveaccess — recently active services & ranges."""
    api_key = get_tnevs_settings(data).get("api_key", "")
    if not api_key:
        return {"error": "2oo9 API key not configured."}
    try:
        req = urllib.request.Request(f"{TNEVS_BASE}/liveaccess", headers=_tnevs_headers(api_key))
        with urllib.request.urlopen(req, timeout=15) as resp:
            r = json.loads(resp.read().decode())
        if r.get("meta", {}).get("code") == 200:
            return {"services": r.get("data", {}).get("services", [])}
        return {"error": r.get("message", str(r))}
    except urllib.error.HTTPError as e:
        return {"error": f"HTTP {e.code}: {e.reason}"}
    except Exception as e:
        return {"error": str(e)}


def tnevs_console(data):
    """GET /console — global live feed last 15 min."""
    api_key = get_tnevs_settings(data).get("api_key", "")
    if not api_key:
        return {"error": "2oo9 API key not configured."}
    try:
        req = urllib.request.Request(f"{TNEVS_BASE}/console", headers=_tnevs_headers(api_key))
        with urllib.request.urlopen(req, timeout=15) as resp:
            r = json.loads(resp.read().decode())
        if r.get("meta", {}).get("code") == 200:
            return {"hits": r.get("data", {}).get("hits", [])}
        return {"error": r.get("message", str(r))}
    except urllib.error.HTTPError as e:
        return {"error": f"HTTP {e.code}: {e.reason}"}
    except Exception as e:
        return {"error": str(e)}


# ============= FREE NUMBER API (HTML scraping) =============

_FREE_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)


def _guess_country_from_number(number):
    n = str(number).lstrip("+")
    if n.startswith("1"):   return "USA"
    if n.startswith("44"):  return "UK"
    if n.startswith("91"):  return "India"
    if n.startswith("49"):  return "Germany"
    if n.startswith("33"):  return "France"
    if n.startswith("7"):   return "Russia"
    if n.startswith("55"):  return "Brazil"
    if n.startswith("86"):  return "China"
    if n.startswith("39"):  return "Italy"
    if n.startswith("34"):  return "Spain"
    if n.startswith("61"):  return "Australia"
    if n.startswith("81"):  return "Japan"
    if n.startswith("82"):  return "South Korea"
    if n.startswith("62"):  return "Indonesia"
    if n.startswith("63"):  return "Philippines"
    if n.startswith("66"):  return "Thailand"
    if n.startswith("84"):  return "Vietnam"
    return "Unknown"


def _fetch_html_free(url, timeout=15):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": _FREE_UA,
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
            "Accept-Language": "en-US,en;q=0.9",
            "Cache-Control": "no-cache",
        }
    )
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read().decode("utf-8", errors="ignore")


def _strip_tags(html_fragment):
    return re.sub(r"<[^>]+>", "", html_fragment).strip()


def get_free_numbers():
    """
    Scrape free public SMS numbers from multiple sites.
    Returns list of dicts: [{number, country}, ...]
    Confirmed URL patterns from live testing:
      receive-sms.cc  → hrefs like /UK-Phone-Number/447915895765
      hs3x.com        → hrefs like read-sms-16477976798.html
    """
    results = []

    # ── Source 1: receive-sms.cc ─────────────────────────────────
    try:
        html = _fetch_html_free("https://receive-sms.cc/", timeout=15)
        # hrefs: /UK-Phone-Number/447915895765
        for m in re.finditer(
            r'href=["\']([^"\']*?-Phone-Number/(\d{9,15}))["\']',
            html, re.IGNORECASE
        ):
            path, digits = m.group(1), m.group(2)
            country_raw  = re.search(r'/([A-Za-z][A-Za-z\-]+)-Phone-Number/', path)
            country      = country_raw.group(1).replace("-", " ") if country_raw else _guess_country_from_number("+" + digits)
            results.append({"number": "+" + digits, "country": country, "_src": "rcc", "_path": path})
        if results:
            seen, unique = set(), []
            for r in results:
                if r["number"] not in seen:
                    seen.add(r["number"]); unique.append(r)
            return unique[:15]
    except Exception:
        pass

    # ── Source 2: hs3x.com ────────────────────────────────────────
    try:
        html = _fetch_html_free("https://hs3x.com/", timeout=15)
        # hrefs: read-sms-16477976798.html
        for m in re.finditer(r'read-sms-(\d{9,15})\.html', html, re.IGNORECASE):
            digits = m.group(1)
            results.append({
                "number":  "+" + digits,
                "country": _guess_country_from_number("+" + digits),
                "_src":    "hs3x",
                "_path":   f"read-sms-{digits}.html",
            })
        if results:
            seen, unique = set(), []
            for r in results:
                if r["number"] not in seen:
                    seen.add(r["number"]); unique.append(r)
            return unique[:15]
    except Exception:
        pass

    return []


def _rcc_url_for_number(digits):
    """Find the receive-sms.cc page URL for a given digit string."""
    try:
        html = _fetch_html_free("https://receive-sms.cc/", timeout=12)
        m = re.search(
            rf'href=["\']([^"\']*?-Phone-Number/{re.escape(digits)})["\']',
            html, re.IGNORECASE
        )
        if m:
            path = m.group(1)
            return "https://receive-sms.cc" + (path if path.startswith("/") else "/" + path)
    except Exception:
        pass
    return None


def get_free_number_messages(number):
    """
    Scrape SMS messages for a free public number.
    Returns list of dicts: [{from, body, time}, ...]
    """
    digits = re.sub(r"\D", "", str(number))

    URLS_TO_TRY = []

    # receive-sms.cc: look up exact page URL from homepage
    rcc = _rcc_url_for_number(digits)
    if rcc:
        URLS_TO_TRY.append(rcc)

    # hs3x.com
    URLS_TO_TRY.append(f"https://hs3x.com/read-sms-{digits}.html")

    for url in URLS_TO_TRY:
        try:
            html     = _fetch_html_free(url, timeout=15)
            messages = []

            # Strategy 1: <tr> table rows  (sender | message | time)
            _HEADER_WORDS = {"from", "sender", "message", "text", "sms", "body",
                             "time", "date", "received", "number", "to", "content"}
            rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.DOTALL | re.IGNORECASE)
            for row in rows:
                # skip <th> header rows entirely
                if re.search(r"<th[\s>]", row, re.IGNORECASE):
                    continue
                cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL | re.IGNORECASE)
                cells = [re.sub(r"\s+", " ", _strip_tags(c)).strip() for c in cells]
                cells = [c for c in cells if c and len(c) > 1]
                if len(cells) < 2:
                    continue
                body = cells[1]
                # skip rows that are just column headers (all cells are single header words)
                if all(c.lower().strip() in _HEADER_WORDS for c in cells):
                    continue
                # body must have real content — more than a single header word, at least 4 chars
                if len(body) < 4 or body.lower().strip() in _HEADER_WORDS:
                    continue
                messages.append({
                    "from": cells[0],
                    "body": body,
                    "time": cells[-1] if len(cells) > 2 else "",
                })

            # Strategy 2: generic blocks with sms/msg class
            if not messages:
                blocks = re.findall(
                    r'class=["\'][^"\']*(?:sms|msg|message|inbox|panel-body)[^"\']*["\'][^>]*>'
                    r'(.*?)</(?:div|article|li|section)>',
                    html, re.DOTALL | re.IGNORECASE
                )
                for block in blocks:
                    text = re.sub(r"\s+", " ", _strip_tags(block)).strip()
                    if len(text) > 8:
                        messages.append({"from": "?", "body": text[:250], "time": ""})

            if messages:
                return messages[:8]
        except Exception:
            continue

    return []


def auto_sms_checker():
    while True:
        try:
            time.sleep(5)
            with _pending_sms_lock:
                if not pending_sms:
                    continue
                snapshot = list(pending_sms.items())
            now     = time.time()
            expired = []
            for req_id, info in snapshot:
                chat_id   = info.get("chat_id")
                started   = info.get("started_at", time.time())
                api       = info.get("api", "s1")
                msg_id    = info.get("msg_id")
                elapsed   = now - started
                remaining = AUTO_CANCEL_SECS - elapsed

                # ── Auto-cancel at 10 min ─────────────────────────
                if elapsed >= AUTO_CANCEL_SECS:
                    try:
                        _do_auto_cancel(req_id, info)
                    except Exception as e:
                        print(f"AutoCancel error ({req_id}): {e}")
                    expired.append(req_id)
                    continue

                # ── Edit message with live timer ──────────────────
                with _last_edit_lock:
                    last_t = _last_edit.get(req_id, 0)
                if chat_id and msg_id and (now - last_t) >= TIMER_EDIT_EVERY:
                    try:
                        edit_message(chat_id, msg_id,
                            _number_msg(info, remaining),
                            get_number_action_inline(req_id, info.get("cid"), info.get("svc_id"), info.get("api", "s1")))
                        with _last_edit_lock:
                            _last_edit[req_id] = now
                    except Exception:
                        pass

                # ── Poll for SMS ──────────────────────────────────
                try:
                    resp = call_api("getStatus", extra_params={"id": req_id}, api=api)
                    if "STATUS_OK" in resp:
                        code = resp.split(":")[1] if ":" in resp else "?"
                        # Acquire user lock to prevent race condition with user-triggered actions
                        with _get_user_lock(chat_id):
                            d   = load_data()
                            uid = str(chat_id)
                            for h in d.get("users", {}).get(uid, {}).get("history", []):
                                if h.get("req_id") == req_id:
                                    h["sms_code"] = code
                                    h["status"]   = "sms_received"
                            save_data(d)
                        sms_txt = (
                            f"📨 *SMS Code Received!*\n━━━━━━━━\n\n"
                            f"🔑 Code: `{code}`\n\n━━━━━━━━"
                        )
                        if chat_id:
                            if msg_id:
                                edit_message(chat_id, msg_id, sms_txt, get_sms_received_inline(info.get("cid"), info.get("svc_id"), info.get("api", "s1")))
                            else:
                                send_message(chat_id, sms_txt, get_sms_received_inline(info.get("cid"), info.get("svc_id"), info.get("api", "s1")))
                        try:
                            _buy_ch = d.get("tnevs_settings", {}).get("sms_channel_user_buy", "").strip()
                            if _buy_ch:
                                _usr = d.get("users", {}).get(str(chat_id), {})
                                _uname = _usr.get("username", "")
                                _fname = _usr.get("first_name", "")
                                _user_disp = f"@{_uname}" if _uname else (_fname or str(chat_id))
                                _fwd_txt = (
                                    f"📨 *SMS Received*\n━━━━━━━━\n\n"
                                    f"👤 User: {_user_disp}\n"
                                    f"🆔 ID: `{chat_id}`\n"
                                    f"📞 Number: `{info.get('number', '?')}`\n"
                                    f"📱 Service: {info.get('service', '?')}\n"
                                    f"🌍 Country: {info.get('country', '?')}\n"
                                    f"🔑 Code: `{code}`"
                                )
                                send_message(_buy_ch, _fwd_txt)
                        except Exception:
                            pass
                        expired.append(req_id)
                    elif "STATUS_CANCEL" in resp:
                        expired.append(req_id)
                except Exception as e:
                    print(f"AutoSMS error ({req_id}): {e}")

            with _pending_sms_lock:
                for r in expired:
                    pending_sms.pop(r, None)
                    with _last_edit_lock:
                        _last_edit.pop(r, None)
        except Exception as e:
            import traceback; print(f"[AutoSMS thread error] {e} — {traceback.format_exc()[-200:]}")


def auto_tmail_checker():
    while True:
        try:
            time.sleep(5)
            with _pending_tmail_lock:
                if not pending_tmail:
                    continue
                snapshot = list(pending_tmail.items())
            now     = time.time()
            expired = []
            for cid, info in snapshot:
                try:
                    elapsed  = now - info.get("started_at", now)
                    msg_id   = info.get("msg_id")
                    token    = info.get("token", "")
                    address  = info.get("address", "")
                    known    = info.get("known_ids", set())

                    if elapsed >= TMAIL_EXPIRE_SECS:
                        if msg_id:
                            try:
                                edit_message(cid, msg_id,
                                    f"📧 *Temp Mail*\n━━━━━━━━\n\n"
                                    f"✉️ `{address}`\n\n"
                                    f"⏰ _Session expired (1 hour)_",
                                    {"inline_keyboard": [[{"text": "🔄 New Email", "callback_data": f"tmail_new_{cid}", "style": "danger", "icon_custom_emoji_id": "5384466886857614542"}]]}
                                )
                            except Exception:
                                pass
                        expired.append(cid)
                        continue

                    msgs = fetch_temp_inbox(token)
                    new  = [m for m in msgs if m.get("id") not in known]
                    if not new:
                        continue

                    m       = new[0]
                    msg_id2 = m.get("id", "")
                    known.add(msg_id2)
                    info["known_ids"] = known

                    full  = fetch_temp_message(token, msg_id2)
                    frm   = full.get("from", {}).get("address", "?")
                    subj  = full.get("subject", "(no subject)")
                    body  = ""
                    if full.get("text"):
                        body = full["text"].strip()
                    elif full.get("html"):
                        raw  = full["html"]
                        if isinstance(raw, list):
                            raw = "\n".join(raw)
                        body = re.sub(r"<[^>]+>", "", raw).strip()
                    body = body[:1500] if body else "_No content_"

                    txt = (
                        f"📧 *Temp Mail Ready!*\n━━━━━━━━\n\n"
                        f"✉️ `{address}`\n\n"
                        f"━━━━━━━━\n"
                        f"📨 *SMS Received!*\n\n"
                        f"📤 From: `{frm}`\n"
                        f"📌 *{subj}*\n\n"
                        f"{body}"
                    )
                    markup = {"inline_keyboard": [[{"text": "🔄 New Email", "callback_data": f"tmail_new_{cid}", "style": "danger", "icon_custom_emoji_id": "5384466886857614542"}]]}
                    if msg_id:
                        try:
                            edit_message(cid, msg_id, txt, markup)
                        except Exception:
                            send_message(cid, txt, markup)
                    else:
                        send_message(cid, txt, markup)
                except Exception as e:
                    print(f"AutoTMail error ({cid}): {e}")

            with _pending_tmail_lock:
                for r in expired:
                    pending_tmail.pop(r, None)
                    temp_mails.pop(r, None)
        except Exception as e:
            import traceback; print(f"[AutoTMail thread error] {e} — {traceback.format_exc()[-200:]}")


AUTO_NUM_EXPIRE_SECS = 600   # 10 minutes


def auto_free_num_checker():
    """Background thread: polls 2oo9 success-otp every 5s for free numbers and auto-delivers OTP."""
    while True:
        try:
            time.sleep(5)
            with _pending_free_lock:
                if not pending_free_nums:
                    continue
                snapshot_free = list(pending_free_nums.items())
            now     = time.time()
            expired = []

            try:
                d      = load_data()
                result = tnevs_get_my_otps(d)
            except Exception as e:
                print(f"AutoFreeNum fetch error: {e}")
                continue

            if "error" in result:
                continue

            otps = result.get("otps", [])

            # Pre-build earn price map from already-loaded d (no extra load_data calls)
            _svc_earn_map  = d.get("withdrawal_settings", {}).get("service_earn", {})
            _services_data = d.get("services_data", {})
            _usr_ch_fwd    = d.get("tnevs_settings", {}).get("sms_channel_user", "").strip()

            # Accumulate all earn credits here; apply in ONE save at the end
            _pending_credits = {}  # uid_str → {amount, count, today}

            for number, info in snapshot_free:
                chat_id   = info.get("chat_id")
                msg_id    = info.get("msg_id")
                known     = info.get("known_msgs", set())
                elapsed   = now - info.get("started_at", now)

                if elapsed >= AUTO_NUM_EXPIRE_SECS:
                    expired.append(number)
                    continue

                try:
                    clean_num = number.lstrip("+")
                    matched   = [o for o in otps if clean_num in str(o.get("number", "")).lstrip("+")]

                    new = []
                    for o in matched:
                        key = str(o.get("number", "")) + str(o.get("message", ""))[:60]
                        if key not in known:
                            known.add(key)
                            new.append(o)
                    info["known_msgs"] = known

                    if not new:
                        continue

                    # Resolve service earn price using already-loaded data
                    _srv_name_check = (info.get("srv_name", "") or "").lower()
                    _matched_svc_id = None
                    for _sid, _sd in _services_data.items():
                        if _sd.get("name", "").lower() == _srv_name_check:
                            _matched_svc_id = _sid
                            break
                    _earn_price = 0.0
                    if _matched_svc_id:
                        _ec = _svc_earn_map.get(_matched_svc_id, {})
                        if _ec.get("enabled", False):
                            _ep = float(_ec.get("price", 0))
                            if _ep > 0:
                                _earn_price = _ep

                    for o in new[:5]:
                        msg_text = (o.get("message", "") or "")[:300]
                        num_str  = o.get("number", number)
                        otp_code = extract_otp_from_message(msg_text)
                        srv_name = info.get("srv_name", "")
                        detected = detect_service_from_sms(srv_name, msg_text)
                        svc_eid  = _free_svc_icon_id(detected or srv_name)

                        # Accumulate earn credit (applied in one batch save below)
                        if _earn_price > 0:
                            uid_str = str(chat_id)
                            cr = _pending_credits.setdefault(uid_str, {"amount": 0.0, "count": 0, "today": 0})
                            cr["amount"] = round(cr["amount"] + _earn_price, 2)
                            cr["count"]  += 1
                            cr["today"]  += 1

                        # Message text: flag + service + number + earned amount
                        flag, _   = get_country_from_prefix(str(num_str).lstrip("+"))
                        svc_label = (detected or srv_name or "").capitalize()
                        txt       = f"{flag} *{svc_label}*  `{num_str}`"
                        if _earn_price > 0:
                            txt += f"\n💰 Earned : ₹{_earn_price}"

                        inline = []
                        if otp_code:
                            clean_otp = otp_code.replace("-", "")
                            otp_btn = {
                                "text":      clean_otp,
                                "copy_text": {"text": clean_otp},
                                "style":     "success",
                                "icon_custom_emoji_id": svc_eid if svc_eid else "6206420230269310869"
                            }
                            inline.append([otp_btn])
                        send_message(chat_id, txt, {"inline_keyboard": inline} if inline else None)

                        # Forward to user OTP channel
                        if _usr_ch_fwd:
                            try:
                                _usr = d.get("users", {}).get(str(chat_id), {})
                                _uname = _usr.get("username", "")
                                _fname = _usr.get("first_name", "")
                                _user_disp = f"@{_uname}" if _uname else (_fname or str(chat_id))
                                send_message(_usr_ch_fwd,
                                    f"👤 {_user_disp} (ID: `{chat_id}`)\n" + txt,
                                    {"inline_keyboard": inline} if inline else None)
                            except Exception:
                                pass
                    expired.append(number)
                except Exception as e:
                    print(f"AutoFreeNum error ({number}): {e}")

            # Apply all accumulated earn credits in ONE load+save
            if _pending_credits:
                try:
                    _earn_d    = load_data()
                    _today_str = datetime.date.today().isoformat()
                    for uid_str, cr in _pending_credits.items():
                        _u = _earn_d.setdefault("users", {}).setdefault(uid_str, {"balance": 0, "history": []})
                        _u["earn_balance"]    = round(float(_u.get("earn_balance", 0)) + cr["amount"], 2)
                        _u["otp_earned_total"] = _u.get("otp_earned_total", 0) + cr["count"]
                        if _u.get("otp_earned_date") != _today_str:
                            _u["otp_earned_date"]  = _today_str
                            _u["otp_earned_today"] = 0
                        _u["otp_earned_today"] = _u.get("otp_earned_today", 0) + cr["today"]
                    save_data(_earn_d)
                except Exception as _earn_err:
                    print(f"OTP earn credit error: {_earn_err}")

            with _pending_free_lock:
                for n in expired:
                    pending_free_nums.pop(n, None)
        except Exception as e:
            import traceback; print(f"[AutoFreeNum thread error] {e} — {traceback.format_exc()[-200:]}")


def auto_nsrv_num_checker():
    """Background thread: polls 2oo9 success-otp every 5s and auto-delivers OTP to the user."""
    while True:
        try:
            time.sleep(5)
            with _pending_nsrv_lock:
                if not pending_nsrv_nums:
                    continue
                snapshot = list(pending_nsrv_nums.items())
            now     = time.time()
            expired = []
            for number, info in snapshot:
                chat_id   = info.get("chat_id")
                msg_id    = info.get("msg_id")
                data_snap = info.get("data_snap", {})
                elapsed   = now - info.get("started_at", now)

                if elapsed >= AUTO_NUM_EXPIRE_SECS:
                    expired.append(number)
                    continue

                try:
                    result = tnevs_get_my_otps(data_snap)
                    if "error" in result:
                        continue
                    otps      = result.get("otps", [])
                    clean_num = number.lstrip("+")
                    matched   = [o for o in otps if clean_num in str(o.get("number", "")).lstrip("+")]
                    if not matched:
                        continue
                    lines = []
                    for o in matched[:5]:
                        msg_text = o.get("message", "")[:200]
                        num_str  = o.get("number", "")
                        lines.append(f"📞 `{num_str}`\n💬 {msg_text}")
                    txt = (
                        f"📨 *OTP Received!*\n━━━━━━━━\n\n"
                        + "\n\n".join(lines)
                        + "\n\n━━━━━━━━"
                    )
                    markup = {"inline_keyboard": [[{"text": "🔴 Live Console", "callback_data": "tnevs_console", "style": "primary"}]]}
                    if msg_id:
                        try:
                            edit_message(chat_id, msg_id, txt, markup)
                        except Exception:
                            send_message(chat_id, txt, markup)
                    else:
                        send_message(chat_id, txt, markup)
                    expired.append(number)
                except Exception as e:
                    print(f"AutoNsrvNum error ({number}): {e}")

            with _pending_nsrv_lock:
                for n in expired:
                    pending_nsrv_nums.pop(n, None)
        except Exception as e:
            import traceback
            print(f"[AutoNsrvNum thread error] {e} — {traceback.format_exc()[-200:]}")


def auto_console_forwarder():
    """Background thread: polls 2oo9 live console every 10s and forwards NEW hits to the channel."""
    global _seen_console_hits
    while True:
        try:
            time.sleep(10)
            d   = load_data()
            cfg = d.get("tnevs_settings", {})
            api_key = cfg.get("api_key", "").strip()
            # Check if ANY channel is configured (legacy or new)
            has_channel = any([
                cfg.get("sms_channel", "").strip(),
                cfg.get("sms_channel_all", "").strip(),
                cfg.get("sms_channel_ig", "").strip(),
                cfg.get("sms_channel_fb", "").strip(),
                cfg.get("sms_channel_wa", "").strip(),
            ])
            if not api_key or not has_channel:
                continue
            result = tnevs_console(d)
            if "error" in result:
                continue
            hits = result.get("hits", [])
            for h in hits:
                sid = h.get("sid", "")
                rng = h.get("range", "")
                msg = h.get("message", "")
                # Always update OTP time tracking regardless of seen status
                _sk = detect_service_from_sms(sid, msg)
                _, _cc = get_country_from_prefix(rng)
                if _cc and _cc != "XX":
                    _last_console_otp_time[(_sk, _cc)] = time.time()
                fp  = f"{sid}|{rng}|{msg[:60]}"
                if fp in _seen_console_hits:
                    continue
                _seen_console_hits.add(fp)
                # Keep memory bounded
                if len(_seen_console_hits) > 1000:
                    _seen_console_hits = set(list(_seen_console_hits)[-500:])
                forward_console_hit_to_channel(d, sid, rng, msg)
        except Exception as e:
            import traceback
            print(f"AutoConsole error: {e}\n{traceback.format_exc()[:300]}")


def show_account(chat_id, data):
    user    = get_user(data, chat_id)
    balance = user.get("balance", 0)
    history = user.get("history", [])
    total   = len(history)
    text    = (
        "👤 *My Account*\n━━━━━━━━\n\n"
        f"💳 *Balance:*  ₹{balance}\n\n━━━━━━━━"
    )
    btn_label = f"📋 Show History ({total})" if total else "📋 History (empty)"
    markup = {"inline_keyboard": [[{"text": btn_label, "callback_data": "acc_history"}]]}
    send_message(chat_id, text, markup)


def _history_text_and_overrides(history):
    """Build (text, emoji_overrides) for last 10 history entries."""
    if not history:
        return "📭 _No purchase history yet._", {}
    lines        = [f"📋 *Purchase History (last 10):*\n"]
    emoji_ovr    = {}
    for h in reversed(history[-10:]):
        status   = h.get("status", "active")
        icon     = "🟢" if status == "active" else "🔴" if status == "cancelled" else "📨"
        sms      = f"\n   🔑 Code: `{h['sms_code']}`" if h.get("sms_code") else ""
        svc_name = h.get("service", "")
        eid      = h.get("emoji_id", "")
        # collect override: find emoji char at start of service name
        if eid:
            for em in _EMOJI_KEYS:
                if svc_name.startswith(em):
                    emoji_ovr[em] = eid
                    break
        lines.append(
            f"{icon} `{h.get('number','N/A')}`\n"
            f"   {svc_name}  •  ₹{h.get('price',0)}{sms}\n"
            f"   🕐 {h.get('timestamp','')}\n"
        )
    return "\n".join(lines), emoji_ovr


# ============= HELPERS =============

def get_api_countries(data, api):
    """Return the correct countries list for the given API key."""
    if api == "s3":
        return data.get("s3_countries", [])
    if api == "s4":
        return data.get("s4_countries", [])
    return data.get("countries", [])

def set_api_countries(data, api, new_list):
    """Save back the correct countries list for the given API key."""
    if api == "s3":
        data["s3_countries"] = new_list
    elif api == "s4":
        data["s4_countries"] = new_list
    else:
        data["countries"] = new_list


# ============= USER HANDLE HELPERS =============

def find_user_by_identifier(data, identifier):
    """Find user by numeric ID or @username. Returns (uid_str, user_obj) or (None, None)."""
    identifier = identifier.strip().lstrip("@")
    users = data.get("users", {})
    if identifier.isdigit():
        if identifier in users:
            return identifier, users[identifier]
        return None, None
    identifier_lower = identifier.lower()
    for uid, u in users.items():
        if u.get("username", "").lower() == identifier_lower:
            return uid, u
    return None, None


def build_all_users_export_xlsx(users_dict):
    """Generate xlsx with all users summary: ID, Username, Name, Balance, Total Deposit, Total OTP (success+cancelled), Success, Cancelled, Pending, Ban."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None  # openpyxl not installed

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "All Users"

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ctr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style_row(row_cells, font=None, fill=None, align=None):
        for c in row_cells:
            if font:  c.font  = font
            if fill:  c.fill  = fill
            if align: c.alignment = align
            c.border = border

    headers = [
        "User ID", "Username", "Name",
        "Balance (₹)", "Total Deposit (₹)",
        "Total OTP", "✅ Success", "❌ Cancelled", "⏳ Pending",
        "Banned"
    ]
    ws.append(headers)
    _style_row(ws[1], font=hdr_font, fill=hdr_fill, align=hdr_align)
    ws.row_dimensions[1].height = 24

    alt_fill = PatternFill("solid", fgColor="EBF3FB")

    for i, (uid, u) in enumerate(users_dict.items(), 1):
        history   = u.get("history", [])
        deposits  = u.get("deposits", [])
        success   = sum(1 for h in history if h.get("status") in ("active", "sms_received", "completed"))
        cancelled = sum(1 for h in history if h.get("status") == "cancelled")
        pending   = len(history) - success - cancelled
        total_otp = success + cancelled
        total_dep = sum(d.get("amount", 0) for d in deposits if d.get("status") == "approved")
        username  = f"@{u.get('username')}" if u.get("username") else "—"
        name      = u.get("first_name", "—") or "—"
        banned    = "🚫 Yes" if u.get("banned") else "—"

        row = [uid, username, name, u.get("balance", 0), round(total_dep, 2),
               total_otp, success, cancelled, pending, banned]
        ws.append(row)
        row_cells = ws[ws.max_row]
        _style_row(row_cells, align=ctr_align, fill=(alt_fill if i % 2 == 0 else None))
        # Highlight banned
        if u.get("banned"):
            row_cells[9].font = Font(color="C00000", bold=True)

    col_widths = [16, 20, 18, 14, 18, 12, 12, 14, 12, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_user_profile_text(uid, user):
    history      = user.get("history", [])
    mail_history = user.get("mail_history", [])
    deposits     = user.get("deposits", [])
    total        = len(history)
    success      = sum(1 for h in history if h.get("status") in ("active", "sms_received", "completed"))
    cancelled    = sum(1 for h in history if h.get("status") == "cancelled")
    pending      = total - success - cancelled
    total_dep    = sum(d.get("amount", 0) for d in deposits if d.get("status") == "approved")
    banned       = user.get("banned", False)
    username     = user.get("username", "")
    fname        = user.get("first_name", "")
    ban_line     = "\n🚫 *BANNED* 🚫\n" if banned else ""
    return (
        f"👤 *User Profile*\n━━━━━━━━{ban_line}\n"
        f"🆔 ID: `{uid}`\n"
        f"👤 Name: {fname or '_Unknown_'}\n"
        f"🔗 Username: {'@' + username if username else '_No username_'}\n"
        f"💳 Balance: ₹{user.get('balance', 0)}\n"
        f"━━━━━━━━\n"
        f"📊 *Purchase Stats*\n"
        f"📦 Total: {total}  |  ✅ Success: {success}\n"
        f"❌ Cancelled: {cancelled}  |  ⏳ Pending: {pending}\n"
        f"━━━━━━━━\n"
        f"💰 Total Deposited: ₹{round(total_dep, 2)}\n"
        f"📧 Mail Orders: {len(mail_history)}"
    )


def build_user_action_inline(uid, user):
    banned  = user.get("banned", False)
    ban_btn = _btn("✅ Unban", callback_data=f"uh_unban_{uid}") if banned else _btn("🚫 Ban", callback_data=f"uh_ban_{uid}")
    return {"inline_keyboard": [
        [_btn("➕ Bal Add", callback_data=f"uh_badd_{uid}"), _btn("➖ Bal Minus", callback_data=f"uh_bmin_{uid}")],
        [ban_btn, _btn("🔄 Refresh", callback_data=f"uh_view_{uid}")],
        [_btn("📤 Export All Data", callback_data=f"uh_export_{uid}", style="success")],
        [_back_btn("◀️ Back", callback_data="uh_main")],
    ]}


# ============= CALLBACK HANDLER =============

def process_callback(chat_id, callback_data, callback_id, message_id, first_name="", username=""):
    try:
        _process_callback_inner(chat_id, callback_data, callback_id, message_id, first_name, username)
    except Exception as _cb_err:
        import traceback
        print(f"⚠️ Callback error [{callback_data[:30]}]: {_cb_err}\n{traceback.format_exc()}")
        try:
            answer_callback(callback_id, "⚠️ An error occurred. Please try again.")
        except Exception:
            pass

def _process_callback_inner(chat_id, callback_data, callback_id, message_id, first_name="", username=""):
    if not callback_data:
        return
    is_admin = (chat_id == ADMIN_ID)

    # ── Early check: cnc_ wait → popup alert (must fire BEFORE answer_callback) ──
    if callback_data.startswith("cnc_"):
        req_id  = callback_data[4:]
        pending = pending_sms.get(req_id)
        if pending:
            _cw_data = load_data()
            _cw_secs = get_cancel_wait(_cw_data, pending.get("api", "s1"))
            if _cw_secs > 0:
                elapsed = time.time() - pending.get("started_at", 0)
                if elapsed < _cw_secs:
                    wait = int(_cw_secs - elapsed)
                    m, s = divmod(wait, 60)
                    _cw_label = f"{_cw_secs // 60} minutes" if _cw_secs % 60 == 0 else f"{_cw_secs // 60}m {_cw_secs % 60}s"
                    answer_callback(
                        callback_id,
                        f"⏳ Wait {m}:{s:02d} more. (Limit: {_cw_label})",
                        show_alert=True
                    )
                    return

    answer_callback(callback_id, _get_clue(callback_data))
    data      = load_data()
    countries = data.get("countries", [])   # s1/s2 only

    def rep(text, markup=None, emoji_overrides=None):
        edit_message(chat_id, message_id, text, markup, emoji_overrides)

    # bsn_ = "Buy Again" — edits the existing failed message in place
    _bsn_mode = callback_data.startswith("bsn_")
    if _bsn_mode:
        callback_data = "bs_" + callback_data[4:]

    # bnn_ = "Buy New Number" — sends ONE new message, then edits it for updates
    _bnn_mode = callback_data.startswith("bnn_")
    if _bnn_mode:
        callback_data = "bs_" + callback_data[4:]
        _bnn_sent_mid = [None]
        def rep(text, markup=None, emoji_overrides=None):  # noqa: F811
            if _bnn_sent_mid[0] is None:
                mid = send_message(chat_id, text, markup, emoji_overrides)
                _bnn_sent_mid[0] = mid
                return mid
            else:
                return edit_message(chat_id, _bnn_sent_mid[0], text, markup, emoji_overrides)

    # ── SAVE USER PROFILE (username / name) ──────────────────
    if not is_admin:
        _uo = get_user(data, chat_id)
        _ch = False
        if username and _uo.get("username") != username.lower():
            _uo["username"] = username.lower(); _ch = True
        if first_name and _uo.get("first_name") != first_name:
            _uo["first_name"] = first_name; _ch = True
        if _ch:
            save_data(data)

    # ── BAN CHECK ────────────────────────────────────────────
    _uh_exempt = callback_data.startswith("uh_") or callback_data == "verify_join" or callback_data.startswith("fj_")
    if not is_admin and not _uh_exempt:
        if data.get("users", {}).get(str(chat_id), {}).get("banned"):
            send_message(chat_id, "🚫 *You have been banned from this bot.*\n\n_Contact admin if you think this is a mistake._")
            return

    # ── FORCE JOIN GATE (non-admin, skip verify_join & fj_* callbacks) ──
    _fj_exempt = callback_data == "verify_join" or callback_data.startswith("fj_")
    if not is_admin and not _fj_exempt:
        unjoined = get_unjoined_channels_cached(chat_id, data)
        if unjoined:
            send_force_join_message(chat_id, unjoined)
            return

    # ── VERIFY JOIN ───────────────────────────────────────────
    if callback_data == "verify_join":
        _clear_join_cache(chat_id)
        unjoined = get_unjoined_channels(chat_id, data)
        if unjoined:
            names = ", ".join(ch.get("name", "?") for ch in unjoined)
            rep(
                f"❌ *Not joined yet!*\n━━━━━━━━\n\n"
                "_Join the channel(s) then verify again._",
                {"inline_keyboard": [
                    *[[{"text": "Join", "url": ch.get("invite_link", ""), "style": "primary", "icon_custom_emoji_id": "5337255927735163754"}
                       for ch in unjoined[i:i+2] if ch.get("invite_link")]
                      for i in range(0, len(unjoined), 2)],
                    [{"text": "unlock", "callback_data": "verify_join", "style": "success", "icon_custom_emoji_id": "5956267526630412170"}],
                ]}
            )
        else:
            rep(
                "✅ *All channels joined!*\n━━━━━━━━\n\n"
                "_All bot features are now unlocked. Use the menu below._"
            )
            send_message(chat_id,
                "🤖 *Virtual Numbers Bot*\n━━━━━━━━\n\nUse the menu below 👇",
                get_main_keyboard(is_admin))
        return

    # ── ADMIN: FORCE JOIN MANAGEMENT ─────────────────────────
    if callback_data == "fj_admin":
        if not is_admin:
            return
        channels = data.get("force_channels", [])
        lines    = ["📢 *Force Join Channels*\n━━━━━━━━\n"]
        buttons  = []
        if channels:
            for ch in channels:
                lines.append(f"• *{ch.get('name','?')}*  `{ch.get('id','')}`")
                buttons.append([
                    _btn(f"🗑 {ch.get('name','?')}", callback_data=f"fj_del_{ch['uid']}"),
                    _btn("🔗 New Link", callback_data=f"fj_link_{ch['uid']}"),
                ])
        else:
            lines.append("_No channels added yet._")
        buttons.append([_btn("➕ Add Channel", callback_data="fj_add")])
        buttons.append([_back_btn("◀️ Back", callback_data="admin_panel")])
        rep("\n".join(lines), {"inline_keyboard": buttons})
        return

    if callback_data == "fj_add":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "fj_add_ch"}
        rep(
            "➕ *Add Channel*\n━━━━━━━━\n\n"
            "Send the channel *ID* or *@username*:\n\n"
            "📌 Examples:\n"
            "`-1001234567890`\n"
            "`@mychannel`\n\n"
            "⚠️ Make sure the bot is *Admin* of the channel!\n\n"
            "_/cancel to abort_"
        )
        return

    if callback_data.startswith("fj_del_"):
        if not is_admin:
            return
        uid = callback_data[7:]
        channels = data.get("force_channels", [])
        data["force_channels"] = [c for c in channels if c.get("uid") != uid]
        save_data(data)
        rep("🗑 *Channel removed.*",
            {"inline_keyboard": [[_back_btn("◀️ Back", callback_data="fj_admin")]]})
        return

    if callback_data.startswith("fj_link_"):
        if not is_admin:
            return
        uid      = callback_data[8:]
        channels = data.get("force_channels", [])
        ch_obj   = next((c for c in channels if c.get("uid") == uid), None)
        if not ch_obj:
            rep("❌ Channel not found.")
            return
        rep("⏳ *Generating new invite link...*")
        new_link = tg_create_invite_link(ch_obj["id"])
        if new_link:
            ch_obj["invite_link"] = new_link
            save_data(data)
            rep(
                f"✅ *New Invite Link created!*\n━━━━━━━━\n\n"
                f"📢 *{ch_obj['name']}*\n🔗 `{new_link}`",
                {"inline_keyboard": [[_back_btn("◀️ Back", callback_data="fj_admin")]]}
            )
        else:
            rep("❌ Could not create link. Is the bot Admin of the channel?",
                {"inline_keyboard": [[_back_btn("◀️ Back", callback_data="fj_admin")]]})
        return

    # ── ADMIN: USER HANDLE ───────────────────────────────────
    if callback_data == "uh_main":
        if not is_admin:
            return
        rep(
            "👥 *User Handle*\n━━━━━━━━\n\n"
            "_Search users, add/minus balance, ban/unban all from here:_",
            {"inline_keyboard": [
                [_btn("🔍 User Search / Profile", callback_data="uh_search")],
                [_btn("➕ Balance Add",  callback_data="uh_badd"), _btn("➖ Balance Minus", callback_data="uh_bmin")],
                [_btn("🚫 Ban User",    callback_data="uh_ban"),  _btn("✅ Unban User",   callback_data="uh_unban")],
                [_btn("🔘 Button On/Off",    callback_data="btn_ctrl",  style="primary")],
                [_btn("💰 Deposit Manage",   callback_data="dm_main",   style="primary")],
                [_btn("📊 Export All Users", callback_data="uh_export_all", style="success")],
                [_back_btn("◀️ Back",        callback_data="admin_panel")],
            ]}
        )
        return

    if callback_data == "admin_broadcast":
        if not is_admin:
            return
        d = load_data()
        user_count = len(d.get("users", {}))
        user_states[chat_id] = {"action": "broadcast"}
        rep(
            f"📣 *Broadcast*\n━━━━━━━━\n\n"
            f"👥 Will send to *{user_count}* users\n\n"
            f"Send your message now:\n"
            f"_(Supports text, photos, premium emojis, @mentions, bold, italic, etc.)_\n\n"
            f"_/cancel to abort_",
            {"inline_keyboard": [[_back_btn("◀️ Back", callback_data="admin_panel")]]}
        )
        return

    if callback_data == "uh_search":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "uh_search"}
        rep("🔍 *User Search*\n━━━━━━━━\n\nSend the user's @username or User ID:\n\n_/cancel to abort_")
        return

    if callback_data == "uh_badd":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "uh_bal_add", "step": "user"}
        rep("➕ *Balance Add*\n━━━━━━━━\n\nSend the user's @username or User ID:\n\n_/cancel to abort_")
        return

    if callback_data == "uh_bmin":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "uh_bal_min", "step": "user"}
        rep("➖ *Balance Minus*\n━━━━━━━━\n\nSend the user's @username or User ID:\n\n_/cancel to abort_")
        return

    if callback_data == "uh_ban":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "uh_ban_user"}
        rep("🚫 *Ban User*\n━━━━━━━━\n\nSend the user's @username or User ID:\n\n_/cancel to abort_")
        return

    if callback_data == "uh_unban":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "uh_unban_user"}
        rep("✅ *Unban User*\n━━━━━━━━\n\nSend the user's @username or User ID:\n\n_/cancel to abort_")
        return

    if callback_data == "dm_main":
        if not is_admin:
            return
        d = load_data()
        ds = d.get("deposit_settings", {})
        ch = ds.get("notify_channel", "")
        rep(
            "💰 *Deposit Manage*\n━━━━━━━━\n\n"
            f"💳 UPI ID : `{ds.get('upi_id', UPI_ID)}`\n"
            f"🖼️ QR URL : `{ds.get('qr_url', 'N/A')}`\n"
            f"💸 Min Deposit : ₹{ds.get('min_deposit', MIN_DEPOSIT)}\n"
            f"📢 Notify Channel : `{ch if ch else 'Not set'}`\n"
            f"🔘 Status : {'✅ ON' if ds.get('enabled', True) else '❌ OFF'}",
            get_dm_inline(d)
        )
        return

    if callback_data == "dm_set_upi":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "dm_set_upi"}
        rep("💳 *Change UPI ID*\n━━━━━━━━\n\nSend the new UPI ID:\n\n_/cancel to abort_")
        return

    if callback_data == "dm_set_qr":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "dm_set_qr"}
        rep("🖼️ *Change QR Code URL*\n━━━━━━━━\n\nSend the new QR image URL or Telegram media link:\n\n_/cancel to abort_")
        return

    if callback_data == "dm_set_min":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "dm_set_min"}
        rep("💸 *Change Minimum Deposit*\n━━━━━━━━\n\nSend the new minimum amount (number only):\n\n_/cancel to abort_")
        return

    if callback_data == "dm_set_channel":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "dm_set_channel"}
        rep(
            "📢 *Set Notify Channel*\n━━━━━━━━\n\n"
            "Send the channel username or ID.\n\n"
            "Example: `@mychannel` or `-1001234567890`\n\n"
            "_Make sure the bot is added as admin in the channel._\n\n"
            "_/cancel to abort_"
        )
        return

    if callback_data == "dm_toggle":
        if not is_admin:
            return
        d  = load_data()
        ds = d.setdefault("deposit_settings", {})
        ds["enabled"] = not ds.get("enabled", True)
        save_data(d)
        status = "✅ ON" if ds["enabled"] else "❌ OFF"
        ch = ds.get("notify_channel", "")
        rep(
            f"💰 *Deposit Manage*\n━━━━━━━━\n\n"
            f"💳 UPI ID : `{ds.get('upi_id', UPI_ID)}`\n"
            f"🖼️ QR URL : `{ds.get('qr_url', 'N/A')}`\n"
            f"💸 Min Deposit : ₹{ds.get('min_deposit', MIN_DEPOSIT)}\n"
            f"📢 Notify Channel : `{ch if ch else 'Not set'}`\n"
            f"🔘 Status : {status}",
            get_dm_inline(d)
        )
        return

    if callback_data == "btn_ctrl":
        if not is_admin:
            return
        d = load_data()
        rep(
            "🔘 *Button On/Off*\n━━━━━━━━\n\n"
            "_Tap a button to toggle On/Off._\n"
            "✅ = On   ❌ = Off",
            get_btn_ctrl_inline(d)
        )
        return

    if callback_data.startswith("btn_toggle_"):
        if not is_admin:
            return
        btn_name = callback_data[11:]
        if btn_name not in CONTROLLABLE_BUTTONS:
            return
        d = load_data()
        states = d.setdefault("button_states", {})
        states[btn_name] = not states.get(btn_name, True)
        save_data(d)
        rep(
            "🔘 *Button On/Off*\n━━━━━━━━\n\n"
            "_Tap a button to toggle On/Off._\n"
            "✅ = On   ❌ = Off",
            get_btn_ctrl_inline(d)
        )
        return

    if callback_data.startswith("uh_view_"):
        if not is_admin:
            return
        uid = callback_data[8:]
        d   = load_data()
        u   = d.get("users", {}).get(uid)
        if not u:
            rep("❌ User not found.")
            return
        rep(build_user_profile_text(uid, u), build_user_action_inline(uid, u))
        return

    if callback_data.startswith("uh_ban_"):
        if not is_admin:
            return
        uid = callback_data[7:]
        d   = load_data()
        u   = d.get("users", {}).get(uid)
        if not u:
            rep("❌ User not found.")
            return
        u["banned"] = True
        save_data(d)
        uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
        rep(
            f"🚫 *Banned!*\n━━━━━━━━\n\n{uname} has been banned.",
            {"inline_keyboard": [
                [_btn("🔄 Refresh Profile", callback_data=f"uh_view_{uid}")],
                [_btn("◀️ User Handle",     callback_data="uh_main")],
            ]}
        )
        try:
            send_message(int(uid), "🚫 *You have been banned from this bot.*\n\n_Contact admin if you think this is a mistake._")
        except Exception:
            pass
        return

    if callback_data.startswith("uh_unban_"):
        if not is_admin:
            return
        uid = callback_data[9:]
        d   = load_data()
        u   = d.get("users", {}).get(uid)
        if not u:
            rep("❌ User not found.")
            return
        u["banned"] = False
        save_data(d)
        uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
        rep(
            f"✅ *Unbanned!*\n━━━━━━━━\n\n{uname} has been unbanned.",
            {"inline_keyboard": [
                [_btn("🔄 Refresh Profile", callback_data=f"uh_view_{uid}")],
                [_btn("◀️ User Handle",     callback_data="uh_main")],
            ]}
        )
        try:
            send_message(int(uid), "✅ *Your ban has been lifted. You can use the bot again.*")
        except Exception:
            pass
        return

    if callback_data.startswith("uh_badd_"):
        if not is_admin:
            return
        uid = callback_data[8:]
        d   = load_data()
        u   = d.get("users", {}).get(uid, {})
        user_states[chat_id] = {"action": "uh_bal_add", "step": "amount", "target_uid": uid}
        uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
        rep(
            f"➕ *Balance Add*\n━━━━━━━━\n\n"
            f"👤 User: {uname}\n"
            f"💳 Current Balance: ₹{u.get('balance', 0)}\n\n"
            f"How much to add?\n\n_/cancel to abort_"
        )
        return

    if callback_data.startswith("uh_bmin_"):
        if not is_admin:
            return
        uid = callback_data[8:]
        d   = load_data()
        u   = d.get("users", {}).get(uid, {})
        user_states[chat_id] = {"action": "uh_bal_min", "step": "amount", "target_uid": uid}
        uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
        rep(
            f"➖ *Balance Minus*\n━━━━━━━━\n\n"
            f"👤 User: {uname}\n"
            f"💳 Current Balance: ₹{u.get('balance', 0)}\n\n"
            f"How much to deduct?\n\n_/cancel to abort_"
        )
        return

    # ── EXPORT USER DATA ──────────────────────────────────────
    if callback_data == "uh_export_all":
        if not is_admin:
            return
        d     = load_data()
        users = d.get("users", {})
        if not users:
            rep("❌ No users found.")
            return
        rep(f"⏳ Generating export for *{len(users)}* users…")
        try:
            xlsx_bytes = build_all_users_export_xlsx(users)
            if not xlsx_bytes:
                rep("❌ Export failed: `openpyxl` library not installed on server.")
                return
            filename   = "all_users_export.xlsx"
            caption    = f"📊 All Users Export\n👥 Total users: {len(users)}"
            send_document_bytes(chat_id, xlsx_bytes, filename, caption)
        except Exception as ex:
            rep(f"❌ Export failed: `{ex}`")
        return

    if callback_data.startswith("uh_export_"):
        if not is_admin:
            return
        uid = callback_data[10:]
        d   = load_data()
        u   = d.get("users", {}).get(uid)
        if not u:
            rep("❌ User not found.")
            return
        uname  = f"@{u.get('username','')}" if u.get("username") else uid
        fname  = u.get("first_name", "Unknown")
        h_cnt  = len(u.get("history", []))
        dep_cnt = len(u.get("deposits", []))
        rep(f"⏳ Generating export for {uname}…")
        try:
            xlsx_bytes = build_user_export_xlsx(uid, u)
            if not xlsx_bytes:
                rep("❌ Export failed: `openpyxl` library not installed on server.")
                return
            filename   = f"user_{uid}_data.xlsx"
            caption    = (
                f"📤 Export — {fname} ({uname})\n"
                f"🆔 ID: {uid}\n"
                f"📦 Purchases: {h_cnt}  |  💰 Deposits: {dep_cnt}"
            )
            send_document_bytes(chat_id, xlsx_bytes, filename, caption)
        except Exception as ex:
            rep(f"❌ Export failed: `{ex}`")
        return

    # ── ADMIN PANEL ──────────────────────────────────────────
    if callback_data == "admin_panel":
        if not is_admin:
            return
        rep("⚙️ *Admin Panel*\n━━━━━━━━", get_admin_main_inline())
        return

    if callback_data == "sm_panel":
        if not is_admin:
            return
        rep("⚙️ *Service Manage*\n━━━━━━━━\n\n_Manage servers and services:_", get_service_manage_inline(data))
        return

    if callback_data == "sm_more":
        if not is_admin:
            return
        rep("⚙️ *More Settings*\n━━━━━━━━\n\n_API keys, balances, channels & cancel time:_", get_service_manage_more_inline(data))
        return

    if callback_data == "set_cancel_time":
        if not is_admin:
            return
        rep(
            "⏱ *Change Cancel Time*\n━━━━━━━━\n\n"
            "Select which server to change cancel wait time for:",
            get_cancel_time_srv_inline(data)
        )
        return

    if callback_data in ("sct_s1", "sct_s2", "sct_s3", "sct_s4", "sct_s5"):
        if not is_admin:
            return
        api_id = callback_data[4:]   # "s1", "s2", "s3", "s4", "s5"
        cfg_api = APIS[api_id]
        _cw_secs = get_cancel_wait(data, api_id)
        user_states[chat_id] = {"action": "sct_input", "api_id": api_id}
        rep(
            f"⏱ *Cancel Time — {cfg_api['icon']} {cfg_api['name']}*\n━━━━━━━━\n\n"
            f"Current: *{_cw_disp_str(_cw_secs)}*\n\n"
            f"Send a number (e.g. `3`) — then choose *Minutes* or *Seconds*:\n\n"
            f"_/cancel to abort_"
        )
        return

    if callback_data in ("sct_min", "sct_sec"):
        if not is_admin:
            return
        st = user_states.get(chat_id, {})
        if st.get("action") != "sct_confirm":
            rep("⚠️ Session expired. Please start again.", get_cancel_time_srv_inline(data))
            return
        api_id = st.get("api_id", "s1")
        num    = st.get("num", 0)
        secs   = num * 60 if callback_data == "sct_min" else num
        cfg_api = APIS.get(api_id, APIS["s1"])
        user_states.pop(chat_id, None)
        d = load_data()
        d.setdefault("cancel_wait", {})[api_id] = secs
        save_data(d)
        unit_label = "minutes" if callback_data == "sct_min" else "seconds"
        rep(
            f"✅ *Cancel Time Updated!*\n━━━━━━━━\n\n"
            f"{cfg_api['icon']} *{cfg_api['name']}*\n"
            f"New cancel wait: *{num} {unit_label}* ({_cw_disp_str(secs)})",
            get_cancel_time_srv_inline(d)
        )
        return

    if callback_data == "sms_ch_user_buy":
        if not is_admin:
            return
        cfg = data.get("tnevs_settings", {})
        ch = cfg.get("sms_channel_user_buy", "")
        disp = f"`{ch}`" if ch else "_Not set_"
        user_states[chat_id] = {"action": "nsrv_set", "field": "sms_channel_user_buy"}
        rep(
            f"👤 *User SMS Channel (Bought Numbers)*\n━━━━━━━━\n\n"
            f"Current: {disp}\n\n"
            f"When a user buys a number and receives SMS, it will be forwarded here with user info.\n\n"
            f"Send your channel username or ID:\n"
            f"Example: `@mychannel`  or  `-1001234567890`\n\n"
            f"Send `-` to clear the channel.\n\n_/cancel to abort_"
        )
        return

    if callback_data == "admin_api_bal":
        if not is_admin:
            return
        rep("⏳ *Checking balances...*")
        lines = ["📡 *API Balances*\n━━━━━━━━\n"]
        for api_id, cfg in APIS.items():
            r   = call_api("getBalance", api=api_id)
            bal = r.replace("ACCESS_BALANCE:", "").strip() if "ACCESS_BALANCE" in r else r
            lines.append(f"{cfg['icon']} *{cfg['name']}:*  `{bal}`")
        rep("\n".join(lines), {"inline_keyboard": [[_back_btn("◀️ Back", callback_data="sm_more")]]})
        return

    # ── API KEYS PANEL ────────────────────────────────────────
    if callback_data == "api_keys_panel":
        if not is_admin:
            return
        rep(
            "🔑 *API Keys*\n━━━━━━━━\n\n"
            "Server এর API key পরিবর্তন করতে সেই server এ tap করুন:",
            get_api_keys_panel_inline()
        )
        return

    if callback_data in ("change_api_s1", "change_api_s2", "change_api_s3", "change_api_s4", "change_api_s5"):
        if not is_admin:
            return
        api_id = callback_data.replace("change_api_", "")
        cfg    = APIS[api_id]
        cur    = cfg["key"]
        disp   = f"`{cur[:6]}...{cur[-4:]}`" if cur else "_Not set_"
        user_states[chat_id] = {"action": "change_api_key", "api_id": api_id}
        rep(
            f"{cfg['icon']} *{cfg['name']} — API Key পরিবর্তন*\n━━━━━━━━\n\n"
            f"Current Key: {disp}\n\n"
            f"নতুন API key পাঠান:\n\n"
            f"_/cancel to abort_"
        )
        return

    # ── ADMIN: MAIL SERVICES ──────────────────────────────────
    if callback_data == "admin_mail":
        if not is_admin:
            return
        mail_services = data.get("mail_services", [])
        lines = ["📧 *Mail Services*\n━━━━━━━━\n"]
        for i, ms in enumerate(mail_services):
            domain = ms.get("domain", "-")
            price  = ms.get("price", 0)
            lines.append(f"{i+1}. `{ms['id']}` — {ms['name']} | Domain: `{domain}` | ₹{price}")
        if not mail_services:
            lines.append("_No mail services yet._")
        btns = [
            [_btn("➕ Add Mail Service", callback_data="mail_add")],
        ]
        for i, ms in enumerate(mail_services):
            btns.append([_btn(f"🗑 Del: {ms['name']}", callback_data=f"mdel_{i}")])
        btns.append([_back_btn("◀️ Back", callback_data="sm_panel")])
        rep("\n".join(lines), {"inline_keyboard": btns})
        return

    if callback_data == "mail_add":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "add_mail", "step": "id"}
        rep(
            "📧 *Add Mail Service*\n━━━━━━━━\n\n"
            "Step 1/5 — Send the *Service ID*\n"
            "_(e.g. `gm` for Gmail, `ya` for Yahoo)_\n\n"
            "_/cancel to abort_"
        )
        return

    if callback_data.startswith("mdel_"):
        if not is_admin:
            return
        try:
            idx = int(callback_data[5:])
        except Exception:
            return
        mail_services = data.get("mail_services", [])
        if 0 <= idx < len(mail_services):
            removed = mail_services.pop(idx)
            data["mail_services"] = mail_services
            save_data(data)
            rep(f"🗑 *Removed:* {removed['name']}", get_admin_main_inline())
        else:
            rep("❌ Not found.")
        return

    # ── SERVER COUNTRY LIST ───────────────────────────────────
    if callback_data.startswith("admin_srv_"):
        if not is_admin:
            return
        api = callback_data.replace("admin_srv_", "")
        if api not in APIS:
            return
        cfg = APIS[api]
        # S3/S4: country-first flow
        if api in ("s3", "s4"):
            api_countries = get_api_countries(data, api)
            rep(
                f"{cfg['icon']} *{cfg['name']}*\n━━━━━━━━\n\nSelect a country:",
                get_srv_countries_inline(api_countries, api)
            )
        else:
            rep(
                f"{cfg['icon']} *{cfg['name']}*\n━━━━━━━━\n\nSelect a service:",
                get_srv_services_inline_for_api(data, api)
            )
        return

    # ── ADMIN COUNTRY LIST PAGE NAV ───────────────────────────────────────
    if callback_data.startswith("admin_cntry_pg_"):
        if not is_admin:
            return
        rest  = callback_data[15:]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        api, page_str = parts
        try:
            page = int(page_str)
        except ValueError:
            return
        if api not in APIS:
            return
        cfg = APIS[api]
        api_countries = get_api_countries(data, api)
        rep(
            f"{cfg['icon']} *{cfg['name']}*\n━━━━━━━━\n\nSelect a country:",
            get_srv_countries_inline(api_countries, api, page=page)
        )
        return

    # ── ADMIN SERVICE LIST PAGE NAV (S1/S2/S5) ────────────────────────────
    if callback_data.startswith("admin_srvsvc_pg_"):
        if not is_admin:
            return
        rest  = callback_data[16:]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        api, page_str = parts
        try:
            page = int(page_str)
        except ValueError:
            return
        if api not in APIS:
            return
        cfg = APIS[api]
        rep(
            f"{cfg['icon']} *{cfg['name']}*\n━━━━━━━━\n\nSelect a service:",
            get_srv_services_inline_for_api(data, api, page=page)
        )
        return

    # ── ADD COUNTRY (per server — country is shared, just sets context) ──
    if callback_data.startswith("aac_"):
        if not is_admin:
            return
        api = callback_data.replace("aac_", "")
        user_states[chat_id] = {"action": "add_country", "step": "id", "api": api}
        rep(
            "➕ *Add Country*\n━━━━━━━━\n\n"
            "Send the country ID:\nExample: `12`\n\n_/cancel to abort_"
        )
        return

    # ── DELETE COUNTRY ────────────────────────────────────────
    if callback_data.startswith("adc_"):
        if not is_admin:
            return
        api          = callback_data.replace("adc_", "")
        api_countries = get_api_countries(data, api)
        if not api_countries:
            rep("❌ No countries to remove.")
            return
        rep(
            "🗑 *Remove Country*\n━━━━━━━━\n\nSelect:",
            get_del_country_inline(api_countries, api)
        )
        return

    if callback_data.startswith("dco_"):
        if not is_admin:
            return
        rest          = callback_data[4:]
        api, cid      = rest.split("_", 1)
        api_countries = get_api_countries(data, api)
        set_api_countries(data, api, [c for c in api_countries if c["id"] != cid])
        save_data(data)
        rep("✅ *Country removed!*", get_srv_countries_inline(get_api_countries(data, api), api))
        return

    # ── VIEW COUNTRY (server-specific services) ───────────────
    if callback_data.startswith("asc_"):
        if not is_admin:
            return
        rest          = callback_data[4:]
        api, cid      = rest.split("_", 1)
        api_countries = get_api_countries(data, api)
        country       = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        cfg      = APIS.get(api, APIS["s1"])
        svcs     = [s for s in country.get("services", []) if s.get("api") == api]
        if svcs:
            if api in ("s3", "s4"):
                def _s34_svc_line(s):
                    line = f"  ▸ `{s['id']}` — {s['name']}  ₹{s.get('price',0)}"
                    if s.get("max_price"):
                        line += f"  🔸max:₹{s['max_price']}"
                    op = s.get("operator")
                    if op:
                        line += f"  🔢op:`{op}`"
                    return line
                svc_list = "\n".join([_s34_svc_line(s) for s in svcs])
            else:
                svc_list = "\n".join([f"  ▸ `{s['id']}` — {s['name']}  ₹{s.get('price',0)}" for s in svcs])
        else:
            svc_list = "  _No services added yet_"

        if api in ("s3", "s4"):
            live_text = f"ℹ️ Live price lookup not available for {APIS[api]['name']}."
        else:
            rep(f"⏳ _Fetching live prices..._")
            live = get_live_prices(cid, api=api)
            if live:
                top = sorted(live.items(), key=lambda x: x[1].get("cost", 0))[:12]
                price_lines = [f"  `{sid}` — ${info['cost']:.3f}  ({info['count']:,} avail)" for sid, info in top]
                live_text = f"📡 *Live Prices ({cfg['name']}):*\n" + "\n".join(price_lines)
            else:
                live_text = "⚠️ Could not fetch live prices."

        rep(
            f"{cfg['icon']} *{country['name']}*  •  {cfg['short']}\n"
            f"━━━━━━━━\n\n"
            f"📱 *Services:*\n{svc_list}\n\n"
            f"━━━━━━━━\n"
            f"{live_text}",
            get_s3_country_services_inline(api, cid, svcs) if api in ("s3", "s4") else get_srv_country_inline(api, cid)
        )
        return

    # ── S3/S4 SERVICE GROUP VIEW — shows all operators for a service ──────────
    if callback_data.startswith("s3grp_"):
        if not is_admin:
            return
        rest = callback_data[6:]
        api, cid, svc_id = rest.split("_", 2)
        api_countries = get_api_countries(data, api)
        country = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        svcs = [s for s in country.get("services", []) if s.get("api") == api]
        # Find the service name from svc_id
        target = next((s for s in svcs if s["id"] == svc_id), None)
        if not target:
            rep("❌ Service not found.")
            return
        target_norm = target["name"].strip().lower()
        cfg = APIS.get(api, APIS["s1"])
        op_count = sum(1 for s in svcs if s.get("name", "").strip().lower() == target_norm)
        rep(
            f"{cfg['icon']} *{target['name']}*  •  {cfg['short']}\n"
            f"━━━━━━━━\n\n"
            f"🌍 Country: {country['name']}\n"
            f"🔢 Operators: {op_count}\n\n"
            f"Select an operator to manage, or add a new one:",
            get_s3_svc_group_inline(api, cid, svcs, target_norm, svc_id)
        )
        return

    # ── S3/S4 OPERATOR ACTION — Update Price + Delete ──────────────────────
    if callback_data.startswith("s3opact_"):
        if not is_admin:
            return
        rest  = callback_data[8:]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        svc_idx = int(parts[1])
        api_cid = parts[0]
        api, cid = api_cid.split("_", 1)
        api_countries = get_api_countries(data, api)
        country = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        svcs = [s for s in country.get("services", []) if s.get("api") == api]
        if svc_idx >= len(svcs):
            rep("❌ Operator not found.")
            return
        s   = svcs[svc_idx]
        cfg = APIS.get(api, APIS["s1"])
        op_line = f"\n🔢 Operator: `{s['operator']}`" if s.get("operator") else ""
        mp_line = f"\n🔸 Max Price: `{s.get('max_price')}`" if s.get("max_price") else ""
        rep(
            f"{cfg['icon']} *{s['name']}*  •  {cfg['short']}\n"
            f"━━━━━━━━\n\n"
            f"💰 Price: ₹{s.get('price', 0)}"
            f"{op_line}{mp_line}\n\n"
            f"🌍 Country: {country['name']}\n\n"
            f"Select an action:",
            get_s3_op_action_inline(api, cid, svc_idx, s["id"])
        )
        return

    # ── S3/S4 BACK TO COUNTRY SERVICE LIST ────────────────────
    # ── S3/S4 SERVICE LIST PAGE NAV ──────────────────────────────────────
    if callback_data.startswith("s3svc_pg_"):
        if not is_admin:
            return
        rest  = callback_data[9:]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        api_cid, page_str = parts
        try:
            page = int(page_str)
        except ValueError:
            return
        api, cid = api_cid.split("_", 1)
        api_countries = get_api_countries(data, api)
        country = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        cfg  = APIS.get(api, APIS["s1"])
        svcs = [s for s in country.get("services", []) if s.get("api") == api]
        rep(
            f"{cfg['icon']} *{country['name']}*  •  {cfg['short']}\n"
            f"━━━━━━━━\n\nSelect a service:",
            get_s3_country_services_inline(api, cid, svcs, page=page)
        )
        return

    if callback_data.startswith("s3back_"):
        if not is_admin:
            return
        rest = callback_data[7:]
        api, cid = rest.split("_", 1)
        api_countries = get_api_countries(data, api)
        country = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        cfg  = APIS.get(api, APIS["s1"])
        svcs = [s for s in country.get("services", []) if s.get("api") == api]
        rep(
            f"{cfg['icon']} *{country['name']}*  •  {cfg['short']}\n"
            f"━━━━━━━━\n\nSelect a service:",
            get_s3_country_services_inline(api, cid, svcs)
        )
        return

    # ── S3/S4 ADD OPERATOR ────────────────────────────────────
    if callback_data.startswith("s3addop_"):
        if not is_admin:
            return
        rest  = callback_data[8:]
        parts = rest.split("_", 2)   # [api, cid] or [api, cid, svc_id]
        api   = parts[0]
        cid   = parts[1] if len(parts) > 1 else ""
        cfg   = APIS.get(api, APIS["s1"])
        state = {"action": "s3_add_operator", "api": api, "country_id": cid}
        if len(parts) == 3:
            # Called from group view — pre-fill name from existing svc_id
            pre_svc_id = parts[2]
            api_countries = get_api_countries(data, api)
            country = next((c for c in api_countries if c["id"] == cid), None)
            pre_name = ""
            pre_sid  = pre_svc_id
            if country:
                svcs = [s for s in country.get("services", []) if s.get("api") == api]
                target = next((s for s in svcs if s["id"] == pre_svc_id), None)
                if target:
                    pre_name = target["name"]
                    pre_sid  = target["id"]
            state["svc_name"] = pre_name
            if api == "s3":
                # S3: every operator has a UNIQUE service ID — always ask for it
                state["step"] = "svc_id"
                user_states[chat_id] = state
                rep(
                    f"➕ *Add Operator*  [{cfg['icon']} {cfg['short']}]\n"
                    f"━━━━━━━━\n\n"
                    f"Service: *{pre_name}*\n\n"
                    f"1️⃣ *Send the Service ID for this operator:*\n"
                    f"Example: `amazon`, `amz2`\n\n_/cancel to abort_"
                )
            else:
                # S4: all operators share the SAME service ID — pre-fill it
                state["svc_id"] = pre_sid
                state["step"]   = "price"
                user_states[chat_id] = state
                rep(
                    f"➕ *Add Operator*  [{cfg['icon']} {cfg['short']}]\n"
                    f"━━━━━━━━\n\n"
                    f"Service: *{pre_name}*\n\n"
                    f"1️⃣ *Send the price (₹):*\nExample: `20`\n\n_/cancel to abort_"
                )
        else:
            # Called standalone — ask for name first
            state["step"] = "name"
            user_states[chat_id] = state
            rep(
                f"➕ *Add Operator*  [{cfg['icon']} {cfg['short']}]\n"
                f"━━━━━━━━\n\n"
                f"1️⃣ *Send the service name:*\nExample: `📱 JioMart`\n\n_/cancel to abort_"
            )
        return

    # ── S3/S4 UPDATE PRICE (from operator action panel) ───────
    if callback_data.startswith("s3updprice_"):
        if not is_admin:
            return
        rest  = callback_data[11:]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        try:
            svc_idx = int(parts[1])
        except (ValueError, IndexError):
            rep("❌ Invalid data.")
            return
        api_cid = parts[0]
        api, cid = api_cid.split("_", 1)
        api_countries = get_api_countries(data, api)
        country = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        svcs = [s for s in country.get("services", []) if s.get("api") == api]
        if svc_idx >= len(svcs):
            rep("❌ Service not found.")
            return
        s   = svcs[svc_idx]
        cfg = APIS.get(api, APIS["s1"])
        user_states[chat_id] = {
            "action": "s3_update_price", "step": "price",
            "api": api, "country_id": cid, "svc_idx": svc_idx, "svc_id": s["id"]
        }
        rep(
            f"💰 *Update Price*  [{cfg['icon']} {cfg['short']}]\n"
            f"━━━━━━━━\n\n"
            f"Service: *{s['name']}*\n"
            f"Operator: `{s.get('operator', 'any')}`\n"
            f"Current price: ₹{s.get('price', 0)}\n\n"
            f"Send new price:\nExample: `25`\n\n_/cancel to abort_"
        )
        return

    # ── S3/S4 DELETE OPERATOR ─────────────────────────────────
    if callback_data.startswith("s3delop_"):
        if not is_admin:
            return
        rest  = callback_data[8:]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        try:
            svc_idx = int(parts[1])
        except (ValueError, IndexError):
            rep("❌ Invalid data.")
            return
        api_cid = parts[0]
        api, cid = api_cid.split("_", 1)
        fresh_data    = load_data()
        api_countries = get_api_countries(fresh_data, api)
        country       = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        svcs = [s for s in country.get("services", []) if s.get("api") == api]
        if svc_idx >= len(svcs):
            rep("❌ Service not found.")
            return
        deleted = svcs[svc_idx]
        del_norm = deleted.get("name", "").strip().lower()
        del_svc_id = deleted.get("id", "")
        # Remove the svc_idx-th api-matching service from country's full list
        api_svc_count = 0
        new_svcs      = []
        for s in country.get("services", []):
            if s.get("api") == api:
                if api_svc_count == svc_idx:
                    api_svc_count += 1
                    continue  # skip = delete
                api_svc_count += 1
            new_svcs.append(s)
        country["services"] = new_svcs
        set_api_countries(fresh_data, api, api_countries)
        save_data(fresh_data)
        cfg      = APIS.get(api, APIS["s1"])
        remaining = [s for s in country.get("services", []) if s.get("api") == api]
        # Check if siblings still exist for this service name → go back to group, else service list
        siblings = [s for s in remaining if s.get("name", "").strip().lower() == del_norm]
        if siblings:
            target_norm = del_norm
            rep(
                f"✅ *Operator Deleted!*\n\n"
                f"Removed: *{deleted['name']}*  op:`{deleted.get('operator', '')}`  ₹{deleted.get('price', 0)}\n"
                f"🌍 {country['name']}\n{cfg['icon']} {cfg['name']}\n\n"
                f"Remaining operators:",
                get_s3_svc_group_inline(api, cid, remaining, target_norm, del_svc_id)
            )
        else:
            rep(
                f"✅ *Operator Deleted!*\n\n"
                f"Removed: *{deleted['name']}*  ₹{deleted.get('price', 0)}\n"
                f"🌍 {country['name']}\n{cfg['icon']} {cfg['name']}",
                get_s3_country_services_inline(api, cid, remaining)
            )
        return

    # ── ADD SERVICE (new: service first, then country) ────────
    if callback_data.startswith("aas_"):
        if not is_admin:
            return
        rest = callback_data[4:]
        # New flow: aas_{api} (no country_id) — add service at top level
        if rest in APIS:
            api = rest
            user_states[chat_id] = {"action": "add_service_flow", "phase": 1, "step": "svc_id", "api": api}
            cfg = APIS.get(api, APIS["s1"])
            rep(
                f"➕ *Add Service*  [{cfg['icon']} {cfg['short']}]\n"
                f"━━━━━━━━\n\n"
                f"1️⃣ Send the service ID:\nExample: `wa`, `tg`, `ig`\n\n_/cancel to abort_"
            )
            return
        # Old compat: aas_{api}_{cid}
        if "_" in rest:
            api, cid = rest.split("_", 1)
            user_states[chat_id] = {"action": "add_service", "step": "id", "api": api, "country_id": cid}
            cfg = APIS.get(api, APIS["s1"])
            extra = "\n🔢 _(Operator ID will be asked after price)_" if cfg.get("operator") else ""
            rep(
                f"➕ *Add Service*  [{cfg['icon']} {cfg['short']}]\n"
                f"━━━━━━━━\n\n"
                f"Send the service ID:\nExample: `wa`{extra}\n\n_/cancel to abort_"
            )
            return
        return

    # ── VIEW SERVICE PAGINATION (countries list next/prev pages) ──
    if callback_data.startswith("asrvsvc_pg_"):
        if not is_admin:
            return
        # format: asrvsvc_pg_{api}_{svc_id}_{page}
        rest_pg = callback_data[len("asrvsvc_pg_"):]
        parts_pg = rest_pg.rsplit("_", 1)
        if len(parts_pg) < 2:
            return
        api_svc_pg, pg_str = parts_pg
        try:
            pg = int(pg_str)
        except ValueError:
            pg = 0
        # api_svc_pg is like "s1_wa"
        api_svc_parts = api_svc_pg.split("_", 1)
        if len(api_svc_parts) < 2 or api_svc_parts[0] not in APIS:
            return
        api, svc_id = api_svc_parts
        cfg = APIS[api]
        svc_name = ""
        total_count = 0
        for c in data.get("countries", []):
            for s in c.get("services", []):
                if s.get("api") == api and s["id"] == svc_id:
                    if not svc_name:
                        svc_name = s["name"]
                    total_count += 1
                    break
        rep(
            f"{cfg['icon']} *{svc_name or svc_id}*  •  {cfg['short']}\n"
            f"━━━━━━━━\n\n"
            f"🌍 *Countries:* {total_count} total\n\n"
            f"_Tap ➕ to add a country_",
            get_srv_service_countries_inline(data, api, svc_id, page=pg)
        )
        return

    # ── VIEW SERVICE (countries under it) ─────────────────────
    if callback_data.startswith("asrvsvc_"):
        if not is_admin:
            return
        rest = callback_data[8:]
        api, svc_id = rest.split("_", 1)
        if api not in APIS:
            return
        cfg = APIS[api]
        svc_name = ""
        total_count = 0
        for c in data.get("countries", []):
            for s in c.get("services", []):
                if s.get("api") == api and s["id"] == svc_id:
                    if not svc_name:
                        svc_name = s["name"]
                    total_count += 1
                    break
        rep(
            f"{cfg['icon']} *{svc_name or svc_id}*  •  {cfg['short']}\n"
            f"━━━━━━━━\n\n"
            f"🌍 *Countries:* {total_count} total\n\n"
            f"_Tap ➕ to add a country_",
            get_srv_service_countries_inline(data, api, svc_id, page=0)
        )
        return

    # ── ADD COUNTRY TO EXISTING SERVICE ───────────────────────
    if callback_data.startswith("aasc_"):
        if not is_admin:
            return
        rest = callback_data[5:]
        api, svc_id = rest.split("_", 1)
        if api not in APIS:
            return
        svc_name = ""
        svc_emoji_id = ""
        svc_style = "primary"
        for c in data.get("countries", []):
            for s in c.get("services", []):
                if s.get("api") == api and s["id"] == svc_id:
                    svc_name     = s["name"]
                    svc_emoji_id = s.get("emoji_id", "")
                    svc_style    = s.get("style", "primary")
                    break
            if svc_name:
                break
        cfg = APIS.get(api, APIS["s1"])
        user_states[chat_id] = {
            "action": "add_country_to_service",
            "step": "cid",
            "api": api,
            "svc_id": svc_id,
            "svc_name": svc_name,
            "svc_emoji_id": svc_emoji_id,
            "svc_style": svc_style
        }
        rep(
            f"➕ *Add Country*\n"
            f"Service: *{svc_name or svc_id}*  [{cfg['icon']} {cfg['short']}]\n"
            f"━━━━━━━━\n\n"
            f"Send the country ID:\nExample: `12`, `6`\n\n_/cancel to abort_"
        )
        return

    # ── DELETE SERVICE (from all countries) ───────────────────
    if callback_data.startswith("adss_"):
        if not is_admin:
            return
        rest = callback_data[5:]
        if "_" in rest:
            # adss_{api}_{svc_id} — delete specific service from all countries
            api, svc_id = rest.split("_", 1)
            if api not in APIS:
                return
            fresh_data = load_data()
            removed = 0
            for c in fresh_data.get("countries", []):
                before = len(c["services"])
                c["services"] = [s for s in c["services"] if not (s.get("api") == api and s["id"] == svc_id)]
                removed += before - len(c["services"])
            save_data(fresh_data)
            cfg = APIS[api]
            rep(
                f"✅ *Service Deleted!*\n━━━━━━━━\n\n"
                f"Removed `{svc_id}` from {removed} countries.\n{cfg['icon']} {cfg['name']}",
                get_srv_services_inline_for_api(fresh_data, api)
            )
        else:
            # adss_{api} — show list of services to pick from
            api = rest
            if api not in APIS:
                return
            seen = {}
            for c in data.get("countries", []):
                for s in c.get("services", []):
                    if s.get("api") == api:
                        sid = s["id"]
                        if sid not in seen:
                            seen[sid] = s["name"]
            if not seen:
                rep("❌ No services to delete.", get_srv_services_inline_for_api(data, api))
                return
            buttons = [[_btn(f"🗑 {name}", callback_data=f"adss_{api}_{sid}")] for sid, name in seen.items()]
            buttons.append([_back_btn("◀️ Back", callback_data=f"admin_srv_{api}")])
            rep(
                "🗑 *Delete Service*\n━━━━━━━━\n\nSelect service to remove from ALL countries:",
                {"inline_keyboard": buttons}
            )
        return

    # ── DELETE COUNTRY FROM SERVICE ────────────────────────────
    # ── VIEW COUNTRY IN SERVICE (display only, no delete) ─────
    if callback_data.startswith("vcsc_"):
        if not is_admin:
            return
        # Just a display button — no action needed, ignore silently
        return

    # ── DEL COUNTRY FROM SERVICE — selection screen ────────────
    # ── DEL COUNTRY PAGINATION ──────────────────────────────────
    if callback_data.startswith("delcpg_"):
        if not is_admin:
            return
        # format: delcpg_{api}_{svc_id}_{page}
        rest_delcpg = callback_data[len("delcpg_"):]
        parts_delcpg = rest_delcpg.rsplit("_", 1)
        if len(parts_delcpg) < 2:
            return
        api_svc_delcpg, pg_str_delcpg = parts_delcpg
        try:
            pg_delcpg = int(pg_str_delcpg)
        except ValueError:
            pg_delcpg = 0
        api_svc_parts_delcpg = api_svc_delcpg.split("_", 1)
        if len(api_svc_parts_delcpg) < 2 or api_svc_parts_delcpg[0] not in APIS:
            return
        api_delcpg, svc_id_delcpg = api_svc_parts_delcpg
        cfg_delcpg = APIS[api_delcpg]
        svc_name_delcpg = ""
        for c in data.get("countries", []):
            for s in c.get("services", []):
                if s.get("api") == api_delcpg and s["id"] == svc_id_delcpg:
                    svc_name_delcpg = s["name"]
                    break
            if svc_name_delcpg:
                break
        rep(
            f"🗑 *Remove Country*\n"
            f"Service: *{svc_name_delcpg or svc_id_delcpg}*  [{cfg_delcpg['icon']} {cfg_delcpg['short']}]\n"
            f"━━━━━━━━\n\n"
            f"Select a country to remove:",
            get_del_country_from_service_inline(data, api_delcpg, svc_id_delcpg, page=pg_delcpg)
        )
        return

    if callback_data.startswith("dcsc_"):
        if not is_admin:
            return
        rest = callback_data[5:]
        api, svc_id = rest.split("_", 1)
        if api not in APIS:
            return
        # Check if there are countries to delete
        has_countries = any(
            s.get("api") == api and s["id"] == svc_id
            for c in data.get("countries", [])
            for s in c.get("services", [])
        )
        if not has_countries:
            rep("❌ No countries to remove.", get_srv_service_countries_inline(data, api, svc_id))
            return
        cfg = APIS[api]
        # Find the service name for display
        svc_name = ""
        for c in data.get("countries", []):
            for s in c.get("services", []):
                if s.get("api") == api and s["id"] == svc_id:
                    svc_name = s["name"]
                    break
            if svc_name:
                break
        rep(
            f"🗑 *Remove Country*\n"
            f"Service: *{svc_name or svc_id}*  [{cfg['icon']} {cfg['short']}]\n"
            f"━━━━━━━━\n\n"
            f"Select a country to remove:",
            get_del_country_from_service_inline(data, api, svc_id)
        )
        return

    # ── DELETE COUNTRY FROM SERVICE ────────────────────────────
    if callback_data.startswith("adsc_"):
        if not is_admin:
            return
        rest  = callback_data[5:]
        parts = rest.split("_", 2)
        if len(parts) < 3:
            return
        api, svc_id, cid = parts
        fresh_data = load_data()
        cname = ""
        for c in fresh_data.get("countries", []):
            if c["id"] == cid:
                cname = c.get("name", cid)
                c["services"] = [s for s in c["services"] if not (s.get("api") == api and s["id"] == svc_id)]
                break
        save_data(fresh_data)
        rep(
            f"✅ *Country removed!*\n{get_country_flag(cname)} {cname}",
            get_del_country_from_service_inline(fresh_data, api, svc_id)
        )
        return

    # ── DELETE SERVICE (old compat: per-country) ──────────────
    if callback_data.startswith("ads_"):
        if not is_admin:
            return
        rest          = callback_data[4:]
        api, cid      = rest.split("_", 1)
        api_countries = get_api_countries(data, api)
        country       = next((c for c in api_countries if c["id"] == cid), None)
        if not country:
            rep("❌ Country not found.")
            return
        svcs = [s for s in country.get("services", []) if s.get("api") == api]
        if not svcs:
            rep("❌ No services to remove.", get_srv_country_inline(api, cid))
            return
        rep(
            "🗑 *Remove Service*\n━━━━━━━━\n\nSelect:",
            get_del_service_inline(svcs, api, cid)
        )
        return

    if callback_data.startswith("dse_"):
        if not is_admin:
            return
        parts    = callback_data[4:].split("_", 2)
        if len(parts) < 3:
            return
        api, cid, idx_str = parts
        try:
            idx = int(idx_str)
        except Exception:
            return
        api_countries = get_api_countries(data, api)
        for c in api_countries:
            if c["id"] == cid:
                svcs_of_api = [s for s in c["services"] if s.get("api") == api]
                if 0 <= idx < len(svcs_of_api):
                    c["services"].remove(svcs_of_api[idx])
                break
        save_data(data)
        rep("✅ *Service removed!*", get_srv_country_inline(api, cid))
        return

    # ── EMOJI PICKER PAGINATION ───────────────────────────────
    if callback_data.startswith("epick_pg_"):
        if not is_admin:
            return
        state = user_states.get(chat_id)
        if not state or state.get("step") != "emoji":
            return
        try:
            pg = int(callback_data[9:])
        except ValueError:
            pg = 0
        action = state.get("action")
        phase  = state.get("phase", 1)
        use_flag_picker = (
            action == "add_country" or
            action == "add_country_to_service" or
            (action == "add_service_flow" and phase == 2)
        )
        if use_flag_picker:
            rep(
                f"✨ *Pick a flag emoji* (page {pg + 1})\n_(tap one, or skip)_",
                get_country_emoji_picker_inline(page=pg)
            )
        else:
            rep(
                f"✨ *Pick an emoji* (page {pg + 1})\n_(tap one, or skip)_",
                get_emoji_picker_inline()
            )
        return

    # ── EMOJI PICKER ──────────────────────────────────────────
    if callback_data.startswith("epick_"):
        if not is_admin:
            return
        state = user_states.get(chat_id)
        if not state or state.get("step") != "emoji":
            return
        raw      = callback_data[6:]  # numeric ID, "skip", or "custom"
        if raw == "custom":
            state["step"] = "epick_custom_input"
            rep(
                "✏️ *Custom Emoji ID*\n\n"
                "Emoji ID টাইপ করুন:\n"
                "Example: `5267598011867491533`\n\n"
                "_/cancel to abort_"
            )
            return
        if raw == "skip":
            chosen   = ""  # no emoji char
            emoji_id = ""  # no premium ID
        else:
            chosen   = PICKER_ID_TO_CHAR.get(raw, "")  # display char
            emoji_id = raw                              # numeric premium ID
        action = state.get("action")

        if action == "add_country":
            state["emoji"]    = chosen
            state["emoji_id"] = emoji_id
            state["step"]     = "style"
            cname = state.get("country_name", "")
            rep(
                f"✅ Emoji: {chosen if chosen else '_(none)_'}\nName: *{cname}*\n\n"
                f"🎨 *Pick a button style:*",
                get_style_picker_inline()
            )
            return

        if action == "add_service":
            state["emoji"]    = chosen
            state["emoji_id"] = emoji_id
            state["step"]     = "style"
            svc_name = state.get("service_name", "")
            rep(
                f"✅ Emoji: {chosen if chosen else '_(none)_'}\nName: *{svc_name}*\n\n"
                f"🎨 *Pick a button style:*",
                get_style_picker_inline()
            )
            return

        # add_service_flow phase 1 → service emoji
        if action == "add_service_flow" and state.get("phase", 1) == 1:
            state["svc_emoji"]    = chosen
            state["svc_emoji_id"] = emoji_id
            state["step"]         = "style"
            svc_name = state.get("svc_name", "")
            rep(
                f"✅ Emoji: {chosen if chosen else '_(none)_'}\nName: *{svc_name}*\n\n"
                f"🎨 *Pick a service button style:*",
                get_style_picker_inline()
            )
            return

        # add_service_flow phase 2 → country flag emoji
        if action == "add_service_flow" and state.get("phase", 1) == 2:
            state["cemoji"]    = chosen
            state["cemoji_id"] = emoji_id
            state["step"]      = "style"
            cname = state.get("cname", "")
            rep(
                f"✅ Flag: {chosen if chosen else '_(none)_'}\nCountry: *{cname}*\n\n"
                f"🎨 *Pick a country button style:*",
                get_style_picker_inline()
            )
            return

        # add_country_to_service → country flag emoji
        if action == "add_country_to_service":
            state["cemoji"]    = chosen
            state["cemoji_id"] = emoji_id
            state["step"]      = "style"
            cname = state.get("cname", "")
            rep(
                f"✅ Flag: {chosen if chosen else '_(none)_'}\nCountry: *{cname}*\n\n"
                f"🎨 *Pick a country button style:*",
                get_style_picker_inline()
            )
            return

    # ── MAIL EMOJI PICKER ─────────────────────────────────────
    if callback_data.startswith("mail_epick_"):
        if not is_admin:
            return
        state = user_states.get(chat_id)
        if not state or state.get("step") != "emoji" or state.get("action") != "add_mail":
            return
        raw = callback_data[11:]
        if raw == "custom":
            state["step"] = "epick_custom_input"
            rep(
                "✏️ *Custom Emoji ID*\n\n"
                "Emoji ID টাইপ করুন:\n"
                "Example: `5267598011867491533`\n\n"
                "_/cancel to abort_"
            )
            return
        if raw == "skip":
            state["emoji_id"] = ""
        else:
            state["emoji_id"] = raw
        state["step"] = "domain"
        rep(
            f"✅ Emoji {'set ✨' if raw != 'skip' else 'skipped'}\n\n"
            f"Now send the *Domain* (or send `-` for none):\n"
            f"Example: `gmail.com`"
        )
        return

    # ── STYLE PICKER ──────────────────────────────────────────
    if callback_data.startswith("stylepick_"):
        if not is_admin:
            return
        state = user_states.get(chat_id)
        if not state or state.get("step") != "style":
            return
        style        = callback_data[10:]  # "primary", "success", or "danger"
        action       = state.get("action")
        style_labels = {"success": "🟢 Success", "primary": "🔵 Primary", "danger": "🔴 Danger"}
        state["style"] = style

        if action == "add_country":
            api      = state.get("api", "s1")
            cid      = state.get("country_id")
            cname    = state.get("country_name", "")
            chosen   = state.get("emoji", "")
            emoji_id = state.get("emoji_id", "")
            full_name = f"{chosen} {cname}".strip() if chosen else cname
            # Re-load data fresh right before saving to avoid race-condition overwrites
            fresh_data    = load_data()
            api_countries = get_api_countries(fresh_data, api)
            if any(c["id"] == cid for c in api_countries):
                # ID already exists — tell admin clearly instead of silently skipping
                rep(
                    f"⚠️ *Country ID `{cid}` already exists!*\n\n"
                    f"Each country needs a unique ID.\n"
                    f"Use ➕ Add Country again with a different ID.",
                    get_srv_countries_inline(api_countries, api)
                )
                user_states.pop(chat_id, None)
                return
            api_countries.append({
                "id": cid, "name": full_name,
                "services": [], "emoji_id": emoji_id, "style": style
            })
            set_api_countries(fresh_data, api, api_countries)
            save_data(fresh_data)
            user_states.pop(chat_id, None)
            rep(
                f"✅ *Country Added!*\n`{cid}` — {full_name}\n"
                f"Style: {style_labels.get(style, style)}",
                get_srv_countries_inline(get_api_countries(fresh_data, api), api)
            )
            return

        # add_service_flow phase 1 — service style picked, transition to phase 2 (country)
        if action == "add_service_flow" and state.get("phase", 1) == 1:
            state["svc_style"] = style
            state.pop("style", None)
            state["phase"] = 2
            state["step"]  = "cid"
            cfg = APIS.get(state.get("api", "s1"), APIS["s1"])
            svc_name = state.get("svc_name", "")
            rep(
                f"✅ Service: *{svc_name}*  Style: {style_labels.get(style, style)}\n\n"
                f"━━━━━━━━\n"
                f"🌍 Now add the first country for this service.\n\n"
                f"2️⃣ Send the country ID:\nExample: `12`, `6`\n\n_/cancel to abort_"
            )
            return

        # add_service_flow phase 2 — country style picked, ask for price
        if action == "add_service_flow" and state.get("phase", 1) == 2:
            state["cstyle"] = style
            state.pop("style", None)
            state["step"]   = "price"
            cname = state.get("cname", "")
            rep(
                f"✅ Country: *{cname}*  Style: {style_labels.get(style, style)}\n\n"
                f"💰 Now send the price in ₹:\nExample: `20`"
            )
            return

        # add_country_to_service — country style picked, ask for price
        if action == "add_country_to_service":
            state["cstyle"] = style
            state.pop("style", None)
            state["step"]   = "price"
            cname = state.get("cname", "")
            rep(
                f"✅ Country: *{cname}*  Style: {style_labels.get(style, style)}\n\n"
                f"💰 Now send the price in ₹:\nExample: `20`"
            )
            return

        # add_service (old flow) → next step is price
        state["step"] = "price"
        svc_name = state.get("service_name", "")
        rep(
            f"✅ Style: *{style_labels.get(style, style)}*\n"
            f"Name: *{svc_name}*\n\n"
            f"💰 Now send the price in ₹:\nExample: `20`"
        )
        return

    # ── USER: BUY START ───────────────────────────────────────
    if callback_data == "acc_history":
        user         = get_user(data, chat_id)
        history      = user.get("history", [])
        mail_history = user.get("mail_history", [])
        if not history and not mail_history:
            rep("📭 _No purchase history yet._")
            return
        lines = [f"Purchase History — User {chat_id}\n{'='*40}\n"]
        for h in reversed(history):
            status = h.get("status", "active")
            sms    = f"\n  OTP: {h['sms_code']}" if h.get("sms_code") else ""
            lines.append(
                f"\n[{h.get('timestamp','')}]\n"
                f"Number : {h.get('number','N/A')}\n"
                f"Service: {h.get('service','')}\n"
                f"Country: {h.get('country','')}\n"
                f"Price  : Rs.{h.get('price',0)}\n"
                f"Status : {status}{sms}\n"
                f"{'-'*30}"
            )
        if mail_history:
            lines.append(f"\n\nMail History\n{'='*40}")
            for h in reversed(mail_history):
                lines.append(
                    f"\n[{h.get('timestamp','')}]\n"
                    f"Mail   : {h.get('mail_address','')}\n"
                    f"Service: {h.get('service','')}\n"
                    f"Status : {h.get('status','')}\n"
                    f"{'-'*30}"
                )
        content = "\n".join(lines)
        send_document(chat_id, content, f"history_{chat_id}.txt")
        return

    if callback_data == "acc_back":
        user    = get_user(data, chat_id)
        balance = user.get("balance", 0)
        history = user.get("history", [])
        total   = len(history)
        text    = (
            "👤 *My Account*\n━━━━━━━━\n\n"
            f"💳 *Balance:*  ₹{balance}\n\n━━━━━━━━"
        )
        btn_label = f"📋 Show History ({total})" if total else "📋 History (empty)"
        markup = {"inline_keyboard": [[{"text": btn_label, "callback_data": "acc_history"}]]}
        rep(text, markup)
        return

    if callback_data == "more_gift":
        rep("🎁 *Gift Code*\n━━━━━━━━\n\n_Coming soon..._")
        return

    if callback_data == "gift_claim":
        user_states[chat_id] = {"action": "claim_gift", "step": "code"}
        rep(
            "🎁 *Claim Gift Code*\n━━━━━━━━\n\n"
            "Enter your gift code below:\n\n_/cancel to abort_"
        )
        return

    if callback_data == "gift_create":
        if chat_id != ADMIN_ID:
            rep("❌ Admin only.")
            return
        user_states[chat_id] = {"action": "create_gift", "step": "amount"}
        rep(
            "➕ *Create Gift Code*\n━━━━━━━━\n\nInter amount ₹\n\n_/cancel to abort_",
            emoji_overrides={"➕": "6206185428702206246"}
        )
        return

    if callback_data == "more_transfer":
        rep("💸 *Transfer Balance*\n━━━━━━━━\n\n_Coming soon..._")
        return

    if callback_data.startswith("tmail_new_") or callback_data == "tmail_create":
        pending_tmail.pop(chat_id, None)
        creating_id = send_message(chat_id, "⏳ *Creating temp email...*", emoji_overrides={"⏳": "5350719881408966477"})
        mail = create_temp_mail()
        if creating_id:
            delete_message(chat_id, creating_id)
        if not mail:
            send_message(chat_id, "❌ *Failed to create temp email.* Try again later.")
            return
        temp_mails[chat_id] = mail
        msg_id = send_message(chat_id,
            f"📧 *Temp Mail Ready!*\n━━━━━━━━\n\n"
            f"✉️ `{mail['address']}`\n\n"
            f"⏳ *Waiting for SMS...*",
            {"inline_keyboard": [[{"text": "🔄 New Email", "callback_data": f"tmail_new_{chat_id}", "style": "danger", "icon_custom_emoji_id": "5384466886857614542"}]]},
            emoji_overrides={"⏳": "5208619406657082341"}
        )
        with _pending_tmail_lock:
            pending_tmail[chat_id] = {
                "address":    mail["address"],
                "token":      mail["token"],
                "msg_id":     msg_id,
                "started_at": time.time(),
                "known_ids":  set(),
            }
        return

    if callback_data.startswith("tmail_check_"):
        mail = temp_mails.get(chat_id)
        if not mail:
            rep("❌ No active temp email. Create one first.")
            return
        rep("📡 *Checking inbox...*")
        messages = fetch_temp_inbox(mail["token"])
        if not messages:
            rep(
                f"📭 *Inbox Empty*\n━━━━━━━━\n\n"
                f"✉️ `{mail['address']}`\n\n_No messages yet._",
                get_temp_mail_inline(chat_id)
            )
            return
        rep(
            f"📬 *{len(messages)} message(s)*\n━━━━━━━━\n\n"
            f"✉️ `{mail['address']}`"
        )
        for m in messages[:5]:
            frm   = m.get("from", {}).get("address", "?")
            subj  = m.get("subject", "(no subject)")
            intro = m.get("intro", "")
            seen  = "✅" if m.get("seen") else "🔵"
            msg_id = m.get("id", "")
            rep(
                f"{seen} *{subj}*\n"
                f"📤 `{frm}`\n"
                f"_{intro[:120]}_",
                get_tmail_read_inline(chat_id, msg_id)
            )
        return

    if callback_data.startswith("tmail_read_"):
        msg_id = callback_data[11:]
        mail   = temp_mails.get(chat_id)
        if not mail:
            rep("❌ Session expired. Create a new temp email.")
            return
        rep("📖 *Loading message...*")
        full = fetch_temp_message(mail["token"], msg_id)
        if "error" in full or not full.get("id"):
            rep("❌ Could not load message. Try again.")
            return
        frm   = full.get("from", {}).get("address", "?")
        subj  = full.get("subject", "(no subject)")
        # Try text body first, fallback to HTML stripped
        body  = ""
        if full.get("text"):
            body = full["text"].strip()
        elif full.get("html"):
            raw = full["html"]
            if isinstance(raw, list):
                raw = "\n".join(raw)
            import re as _re
            body = _re.sub(r"<[^>]+>", "", raw).strip()
        body = body[:2000] if body else "_No content_"
        rep(
            f"📧 *{subj}*\n━━━━━━━━\n\n"
            f"📤 From: `{frm}`\n\n"
            f"{body}",
            {"inline_keyboard": [[{"text": "📩 Back to Inbox", "callback_data": f"tmail_check_{chat_id}", "style": "primary"}]]}
        )
        return

    # ── PAGE PICKER: country list ─────────────────────────────────────────────
    # ctry_pg_picker_{service_id}_{current_page}
    if callback_data.startswith("ctry_pg_picker_") and not callback_data.startswith("ctry_pg_picker_nav_"):
        rest  = callback_data[len("ctry_pg_picker_"):]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        service_id, pg_str = parts
        try:
            current_page = int(pg_str)
        except ValueError:
            return
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        total_pages = _ctry_total_pages(service_id, all_countries)
        markup = get_page_picker_inline(
            total_pages, current_page,
            pg_cb_prefix       = f"ctry_pg_{service_id}_",
            picker_nav_prefix  = "ctry_pg_picker_nav_",
            service_id_part    = f"{service_id}_{current_page}_",
            picker_page        = 0,
        )
        rep("📄 *Select Page*\n━━━━━━━━", markup, {"📄": "5258477770735885832"})
        return

    # ctry_pg_picker_nav_{service_id}_{current_page}_{picker_page}
    if callback_data.startswith("ctry_pg_picker_nav_"):
        rest  = callback_data[len("ctry_pg_picker_nav_"):]
        parts = rest.rsplit("_", 2)
        if len(parts) != 3:
            return
        service_id, cur_pg_str, picker_pg_str = parts
        try:
            current_page = int(cur_pg_str)
            picker_page  = int(picker_pg_str)
        except ValueError:
            return
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        total_pages = _ctry_total_pages(service_id, all_countries)
        markup = get_page_picker_inline(
            total_pages, current_page,
            pg_cb_prefix       = f"ctry_pg_{service_id}_",
            picker_nav_prefix  = "ctry_pg_picker_nav_",
            service_id_part    = f"{service_id}_{current_page}_",
            picker_page        = picker_page,
        )
        rep("📄 *Select Page*\n━━━━━━━━", markup, {"📄": "5258477770735885832"})
        return

    # ── PAGE PICKER: service list ─────────────────────────────────────────────
    # svc_pg_picker_{current_page}
    if callback_data.startswith("svc_pg_picker_") and not callback_data.startswith("svc_pg_picker_nav_"):
        try:
            current_page = int(callback_data[len("svc_pg_picker_"):])
        except ValueError:
            return
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        seen_norm = set()
        svc_count = 0
        for c in reversed(all_countries):
            for s in reversed(c.get("services", [])):
                norm = _strip_emoji(s["name"]).strip().lower()
                if norm not in seen_norm:
                    seen_norm.add(norm)
                    svc_count += 1
        PAGE_SIZE   = 20
        total_pages = max(1, (svc_count + PAGE_SIZE - 1) // PAGE_SIZE)
        markup = get_page_picker_inline(
            total_pages, current_page,
            pg_cb_prefix       = "svc_pg_",
            picker_nav_prefix  = "svc_pg_picker_nav_",
            service_id_part    = f"{current_page}_",
            picker_page        = 0,
        )
        rep("📄 *Select Page*\n━━━━━━━━", markup, {"📄": "5258477770735885832"})
        return

    # svc_pg_picker_nav_{current_page}_{picker_page}
    if callback_data.startswith("svc_pg_picker_nav_"):
        rest  = callback_data[len("svc_pg_picker_nav_"):]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        cur_pg_str, picker_pg_str = parts
        try:
            current_page = int(cur_pg_str)
            picker_page  = int(picker_pg_str)
        except ValueError:
            return
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        seen_norm = set()
        svc_count = 0
        for c in reversed(all_countries):
            for s in reversed(c.get("services", [])):
                norm = _strip_emoji(s["name"]).strip().lower()
                if norm not in seen_norm:
                    seen_norm.add(norm)
                    svc_count += 1
        PAGE_SIZE   = 20
        total_pages = max(1, (svc_count + PAGE_SIZE - 1) // PAGE_SIZE)
        markup = get_page_picker_inline(
            total_pages, current_page,
            pg_cb_prefix       = "svc_pg_",
            picker_nav_prefix  = "svc_pg_picker_nav_",
            service_id_part    = f"{current_page}_",
            picker_page        = picker_page,
        )
        rep("📄 *Select Page*\n━━━━━━━━", markup, {"📄": "5258477770735885832"})
        return

    # ── S1 MULTI-PRICE: country group sub-list ───────────────────────────────
    # ctry_multi_{service_id}_{country_id}
    if callback_data.startswith("ctry_multi_"):
        rest  = callback_data[len("ctry_multi_"):]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        service_id, country_id = parts
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        # Resolve canonical name
        target_name = None
        for c in all_countries:
            for s in c.get("services", []):
                if s["id"] == service_id:
                    target_name = _strip_emoji(s["name"]).strip().lower()
                    break
            if target_name:
                break
        # Find country
        country = next((c for c in all_countries if c["id"] == country_id), None)
        if not country:
            rep("❌ Country not found.")
            return
        # Collect all S1 price options for this service in this country
        price_options = []
        for s in country.get("services", []):
            if s.get("api", "s1") != "s1":
                continue
            sname = _strip_emoji(s.get("name", "")).strip().lower()
            if s["id"] == service_id or (target_name and sname == target_name):
                price_options.append(s)
        if not price_options:
            rep("❌ No price options found.")
            return
        # Sort by price ascending
        price_options = sorted(price_options, key=lambda x: x.get("price", 0))
        cname    = _strip_emoji(country["name"])
        cfg_s1   = APIS.get("s1", {})
        srv_short = cfg_s1.get("short", "S1")
        ceid     = _country_emoji_id(country)
        seid     = cfg_s1.get("emoji_id")
        buttons  = []
        for s in price_options:
            price = s.get("price", 0)
            max_p = s.get("max_price")
            label = f"₹{price}  •  {srv_short}"
            # Use bsp_ (buy specific price) so each tier gets a unique callback
            cb = (f"bsp_{country_id}_{s['id']}_s1_{max_p}"
                  if max_p is not None else f"bs_{country_id}_{s['id']}_s1")
            btn   = {"text": label, "callback_data": cb,
                     "style": country.get("style", "primary")}
            if ceid:   btn["icon_custom_emoji_id"] = ceid
            elif seid: btn["icon_custom_emoji_id"] = seid
            buttons.append([btn])
        # Back button returns to service's country list
        buttons.append([{"text": "◀ Back", "callback_data": f"svc_{service_id}", "style": "danger",
                         "icon_custom_emoji_id": "5370726448959085259"}])
        rep(f"🌍 *{cname}* — Choose Price\n━━━━━━━━", {"inline_keyboard": buttons})
        return

    if callback_data == "buy_start":
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        all_svcs = [s for c in all_countries for s in c.get("services", [])]
        if not all_svcs:
            rep("❌ No services available yet. Contact admin.")
            return
        rep(
            "📱 *Select a Service*\n━━━━━━━━\n\nChoose the service you need:",
            get_all_services_inline(all_countries)
        )
        return

    # svc_pg_{page} — paginate all-services list
    if callback_data.startswith("svc_pg_"):
        try:
            page = int(callback_data[7:])
        except ValueError:
            return
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        all_svcs = [s for c in all_countries for s in c.get("services", [])]
        if not all_svcs:
            rep("❌ No services available yet.")
            return
        rep(
            "📱 *Select a Service*\n━━━━━━━━\n\nChoose the service you need:",
            get_all_services_inline(all_countries, page=page)
        )
        return

    # svc_{service_id} → show countries for that service
    if callback_data.startswith("svc_"):
        svc_id        = callback_data[4:]
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        markup        = get_countries_for_service_inline(svc_id, all_countries)
        if not markup["inline_keyboard"]:
            rep("❌ No countries available for this service.")
            return
        # get service name + emoji info from first match
        svc_name     = svc_id
        svc_obj      = None
        for c in all_countries:
            for s in c.get("services", []):
                if s["id"] == svc_id:
                    svc_name = s["name"]
                    svc_obj  = s
                    break
            if svc_obj:
                break
        # build emoji override so message uses the service's specific premium emoji_id
        emoji_ovr = {}
        if svc_obj:
            eid = _svc_emoji_id(svc_obj)
            if eid:
                # find the emoji char at the start of svc_name
                for em in _EMOJI_KEYS:
                    if svc_name.startswith(em):
                        emoji_ovr[em] = eid
                        break
        rep(
            f"🌍 *Select Country*\n━━━━━━━━\n\n"
            f"*{svc_name}*",
            markup,
            emoji_ovr or None
        )
        return

    # ctry_pg_{service_id}_{page} — paginate country list
    if callback_data.startswith("ctry_pg_"):
        rest  = callback_data[8:]
        parts = rest.rsplit("_", 1)
        if len(parts) != 2:
            return
        svc_id, page_str = parts
        try:
            page = int(page_str)
        except ValueError:
            return
        all_countries = countries + data.get("s3_countries", []) + data.get("s4_countries", [])
        markup = get_countries_for_service_inline(svc_id, all_countries, page=page)
        svc_name = svc_id
        svc_obj  = None
        for c in all_countries:
            for s in c.get("services", []):
                if s["id"] == svc_id:
                    svc_name = s["name"]
                    svc_obj  = s
                    break
            if svc_obj:
                break
        emoji_ovr = {}
        if svc_obj:
            eid = _svc_emoji_id(svc_obj)
            if eid:
                for em in _EMOJI_KEYS:
                    if svc_name.startswith(em):
                        emoji_ovr[em] = eid
                        break
        rep(
            f"🌍 *Select Country*\n━━━━━━━━\n\n*{svc_name}*",
            markup,
            emoji_ovr or None
        )
        return

    # ab_{service_id} — Auto Buy: try each server in price order until one succeeds
    if callback_data.startswith("ab_"):
        service_id    = callback_data[3:]
        all_countries = data.get("countries", []) + data.get("s3_countries", []) + data.get("s4_countries", [])

        # Collect all matching entries (same logic as get_countries_for_service_inline)
        target_name = None
        for c in all_countries:
            for s in c.get("services", []):
                if s["id"] == service_id:
                    target_name = _strip_emoji(s["name"]).strip().lower()
                    break
            if target_name is not None:
                break

        candidates = []
        seen_s4_op = set()
        for c in all_countries:
            for s in c.get("services", []):
                sname = _strip_emoji(s["name"]).strip().lower()
                if s["id"] == service_id or (target_name and sname == target_name):
                    api_c    = s.get("api", "s1")
                    operator = s.get("operator", "")
                    if api_c == "s4":
                        key = (c["id"], api_c, operator)
                        if key in seen_s4_op:
                            continue
                        seen_s4_op.add(key)
                    candidates.append({
                        "cid":      c["id"],
                        "country":  c,
                        "service":  s,
                        "api":      api_c,
                        "operator": operator,
                        "price":    s.get("price", 0),
                    })

        if not candidates:
            rep("❌ No servers available for this service.")
            return

        # Sort cheapest first
        candidates.sort(key=lambda x: x["price"])

        user = get_user(data, chat_id)
        rep(f"\u26a1 *Auto Buy*\n\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\n\nTrying {len(candidates)} server(s)...")

        success = False
        tried   = []
        for cand in candidates:
            cid      = cand["cid"]
            api      = cand["api"]
            svc_obj  = cand["service"]
            svc_id_c = svc_obj["id"]
            operator = cand["operator"]
            price    = cand["price"]
            country  = cand["country"]
            cfg      = APIS.get(api, APIS["s1"])

            if user["balance"] < price:
                tried.append(f"❌ {_strip_emoji(country['name'])} {cfg['short']} — Insufficient balance (₹{price})")
                continue

            op_extra = {}
            if operator:
                op_extra["operator"] = operator
            if api == "s1":
                mp = svc_obj.get("max_price")
                mn = svc_obj.get("min_price")
                if mn is not None: op_extra["minPrice"] = mn
                if mp is not None: op_extra["maxPrice"] = mp
            use_v2 = (api == "s2" and svc_obj.get("max_price") is not None)
            if use_v2:
                op_extra["maxPrice"] = svc_obj["max_price"]
            if api in ("s3", "s4"):
                mp = svc_obj.get("max_price")
                if mp is not None: op_extra["maxPrice"] = mp
            op_extra = op_extra if op_extra else None

            action_name = "getNumberV2" if use_v2 else "getNumber"
            response    = call_api(action_name, svc_id_c, cid, api=api, extra_params=op_extra)

            # Parse response
            number, req_id = "", ""
            if use_v2:
                try:
                    rj = json.loads(response)
                    number = str(rj.get("phoneNumber") or rj.get("phone") or "")
                    req_id = str(rj.get("activationId") or rj.get("id") or "")
                except Exception:
                    pass
            elif "ACCESS_NUMBER" in (response or ""):
                rp = response.split(":")
                if len(rp) >= 3:
                    req_id = rp[1]
                    number = rp[2]

            if number and req_id:
                user["balance"] -= price
                ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
                user["history"].append({
                    "number":    number,
                    "req_id":    req_id,
                    "service":   svc_obj["name"],
                    "emoji_id":  _svc_emoji_id(svc_obj),
                    "country":   country["name"],
                    "price":     price,
                    "status":    "active",
                    "sms_code":  None,
                    "timestamp": ts,
                    "api":       api,
                    "cid":       cid,
                    "svc_id":    svc_id_c,
                })
                save_data(data)
                info_entry = {
                    "chat_id":       chat_id,
                    "price":         price,
                    "started_at":    time.time(),
                    "api":           api,
                    "cid":           cid,
                    "svc_id":        svc_id_c,
                    "msg_id":        message_id,
                    "number":        number,
                    "service":       svc_obj["name"],
                    "country":       country["name"],
                    "balance_after": user["balance"],
                }
                with _pending_sms_lock:
                    pending_sms[req_id] = info_entry
                new_mid = rep(
                    _number_msg(info_entry, AUTO_CANCEL_SECS),
                    get_number_action_inline(req_id, cid, svc_id_c, api)
                )
                if isinstance(new_mid, int):
                    info_entry["msg_id"] = new_mid
                success = True
                break
            else:
                err_short = (response or "No number available")[:80]
                tried.append(f"❌ {_strip_emoji(country['name'])} {cfg['short']} — {err_short}")

        if not success:
            tried_txt = "\n".join(tried) if tried else "No servers available."
            rep(
                f"\u274c *Auto Buy Failed*\n\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\n\nAll servers tried:\n{tried_txt}",
                {"inline_keyboard": [[{
                    "text": "Try Again",
                    "callback_data": f"ab_{service_id}",
                    "style": "success",
                    "icon_custom_emoji_id": "5224607267797606837",
                    "right_icon_custom_emoji_id": "5224607267797606837"
                }]]}
            )
        return

    # bs_{country_id}_{service_id}_{api}  (S4: bs_{country_id}_{service_id}_{api}_{operator})
    # bsp_ = "buy specific price" — same as bs_ but carries max_price for exact tier matching
    # Normalise it to bs_ after extracting the price filter so all buy logic stays in one place.
    _s1_max_price_filter = None
    if callback_data.startswith("bsp_"):
        _bsp_parts = callback_data[4:].split("_", 3)
        if len(_bsp_parts) == 4:
            _s1_max_price_filter = _bsp_parts[3]
            callback_data = "bs_" + "_".join(_bsp_parts[:3])

    if callback_data.startswith("bs_"):
        rest   = callback_data[3:]
        parts  = rest.split("_", 3)
        if len(parts) < 3:
            return
        cid     = parts[0]
        svc_id  = parts[1]
        api     = parts[2]
        op_filter = parts[3] if len(parts) > 3 else None

        api_countries = get_api_countries(data, api)
        country = next((c for c in api_countries if c["id"] == cid), None)
        service = None
        if country:
            if op_filter is not None:
                # S4: match by service_id, api AND operator
                service = next(
                    (s for s in country.get("services", [])
                     if s["id"] == svc_id and s.get("api") == api and str(s.get("operator", "")) == op_filter),
                    None
                )
            elif _s1_max_price_filter is not None:
                # S1 multi-price: match exact tier by max_price to avoid always picking cheapest
                try:
                    _target_mp = float(_s1_max_price_filter)
                    service = next(
                        (s for s in country.get("services", [])
                         if s["id"] == svc_id and s.get("api") == api
                         and abs((s.get("max_price") if s.get("max_price") is not None else -1) - _target_mp) < 0.00001),
                        None
                    )
                except Exception:
                    pass
            if not service:
                service = next(
                    (s for s in country.get("services", []) if s["id"] == svc_id and s.get("api") == api),
                    None
                )
        if not service:
            rep("❌ Service not found.")
            return

        price    = service.get("price", 0)
        cfg      = APIS.get(api, APIS["s1"])
        user     = get_user(data, chat_id)

        if user["balance"] < price:
            rep(
                f"❌ *Insufficient Balance*\n━━━━━━━━\n\n"
                f"💰 Required:  ₹{price}\n"
                f"💳 Your balance:  ₹{user['balance']}\n\n"
                f"_Please deposit funds first._"
            )
            return

        operator_id = service.get("operator")
        op_extra = {}
        if operator_id:
            op_extra["operator"] = operator_id

        # S1 (SmsBower): minPrice + maxPrice via getNumber
        if api == "s1":
            min_price = service.get("min_price")
            max_price = service.get("max_price")
            if min_price is not None:
                op_extra["minPrice"] = min_price
            if max_price is not None:
                op_extra["maxPrice"] = max_price

        # S2 (HeroSMS): maxPrice via getNumberV2 (returns JSON)
        use_v2 = (api == "s2" and service.get("max_price") is not None)
        if use_v2:
            op_extra["maxPrice"] = service["max_price"]

        # S3/S4: maxPrice only via getNumber (API does not support minPrice)
        if api in ("s3", "s4"):
            max_price = service.get("max_price")
            if max_price is not None:
                op_extra["maxPrice"] = max_price

        op_extra = op_extra if op_extra else None

        rep(
            f"⏳ *Getting your number...*\n\n"
            f"{service['name']}  •  ₹{price}\n"
            f"{country['name']}"
        )

        action_name = "getNumberV2" if use_v2 else "getNumber"
        response = call_api(action_name, svc_id, cid, api=api, extra_params=op_extra)

        # Parse getNumberV2 JSON response (HeroSMS S2)
        if use_v2:
            try:
                rjson = json.loads(response)
                number = str(rjson.get("phoneNumber") or rjson.get("phone") or "")
                req_id = str(rjson.get("activationId") or rjson.get("id") or "")
            except Exception:
                number, req_id = "", ""
            if number and req_id:
                user["balance"] -= price
                ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
                user["history"].append({
                    "number":    number,
                    "req_id":    req_id,
                    "service":   service["name"],
                    "emoji_id":  _svc_emoji_id(service),
                    "country":   country["name"],
                    "price":     price,
                    "status":    "active",
                    "sms_code":  None,
                    "timestamp": ts,
                    "api":       api,
                    "cid":       cid,
                    "svc_id":    svc_id,
                })
                save_data(data)
                info_entry = {
                    "chat_id":       chat_id,
                    "price":         price,
                    "started_at":    time.time(),
                    "api":           api,
                    "cid":           cid,
                    "svc_id":        svc_id,
                    "msg_id":        message_id,
                    "number":        number,
                    "service":       service["name"],
                    "country":       country["name"],
                    "balance_after": user["balance"],
                }
                with _pending_sms_lock:
                    pending_sms[req_id] = info_entry
                new_mid = rep(
                    _number_msg(info_entry, AUTO_CANCEL_SECS),
                    get_number_action_inline(req_id, cid, svc_id, api)
                )
                if isinstance(new_mid, int):
                    info_entry["msg_id"] = new_mid
            else:
                err = response[:200] if response else "Unknown error"
                rep(
                    f"❌ *Failed to get number*\n\n`{err}`",
                    {"inline_keyboard": [
                        [{
                            "text": "🔄 Buy Again",
                            "callback_data": f"bsn_{cid}_{svc_id}_{api}",
                            "style": "primary",
                            "icon_custom_emoji_id": "6197478323107402845"
                        }],
                        [{
                            "text": "◀️ Back",
                            "callback_data": f"svc_{svc_id}",
                            "style": "danger"
                        }]
                    ]}
                )

        elif response and "ACCESS_NUMBER" in response:
            rp = response.split(":")
            if len(rp) >= 3:
                number = rp[2]
                req_id = rp[1]
                user["balance"] -= price
                ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
                user["history"].append({
                    "number":    number,
                    "req_id":    req_id,
                    "service":   service["name"],
                    "emoji_id":  _svc_emoji_id(service),
                    "country":   country["name"],
                    "price":     price,
                    "status":    "active",
                    "sms_code":  None,
                    "timestamp": ts,
                    "api":       api,
                    "cid":       cid,
                    "svc_id":    svc_id,
                })
                save_data(data)
                info_entry = {
                    "chat_id":       chat_id,
                    "price":         price,
                    "started_at":    time.time(),
                    "api":           api,
                    "cid":           cid,
                    "svc_id":        svc_id,
                    "msg_id":        message_id,
                    "number":        number,
                    "service":       service["name"],
                    "country":       country["name"],
                    "balance_after": user["balance"],
                }
                with _pending_sms_lock:
                    pending_sms[req_id] = info_entry
                new_mid = rep(
                    _number_msg(info_entry, AUTO_CANCEL_SECS),
                    get_number_action_inline(req_id, cid, svc_id, api)
                )
                if isinstance(new_mid, int):
                    info_entry["msg_id"] = new_mid
            else:
                rep(f"✅ {response}")
        else:
            rep(
                f"❌ *Failed to get number*\n\n`{(response or '')[:150]}`",
                {"inline_keyboard": [
                    [{
                        "text": "🔄 Buy Again",
                        "callback_data": f"bsn_{cid}_{svc_id}_{api}",
                        "style": "primary",
                        "icon_custom_emoji_id": "6197478323107402845"
                    }],
                    [{
                        "text": "◀️ Back",
                        "callback_data": f"svc_{svc_id}",
                        "style": "danger"
                    }]
                ]}
            )
        return

    # ── SMS CHECK ─────────────────────────────────────────────
    if callback_data.startswith("chk_"):
        req_id  = callback_data[4:]
        pending = pending_sms.get(req_id, {})
        _pcid   = pending.get("cid")
        _psvc   = pending.get("svc_id")
        _papi   = pending.get("api", "s1")
        rep("📡 *Checking for SMS...*")
        api  = get_api_for_request(data, chat_id, req_id)
        resp = call_api("getStatus", extra_params={"id": req_id}, api=api)
        if "STATUS_OK" in resp:
            code = resp.split(":")[1] if ":" in resp else "?"
            uid  = str(chat_id)
            for h in data.get("users", {}).get(uid, {}).get("history", []):
                if h.get("req_id") == req_id:
                    h["sms_code"] = code
                    h["status"]   = "sms_received"
            save_data(data)
            rep(
                f"📨 *SMS Code Received!*\n━━━━━━━━\n\n"
                f"🔑 Code: `{code}`\n\n━━━━━━━━",
                get_sms_received_inline(_pcid, _psvc, _papi)
            )
            try:
                _buy_ch = data.get("tnevs_settings", {}).get("sms_channel_user_buy", "").strip()
                if _buy_ch:
                    _usr_info = data.get("users", {}).get(str(chat_id), {})
                    _uname = _usr_info.get("username", "")
                    _fname = _usr_info.get("first_name", "")
                    _user_disp = f"@{_uname}" if _uname else (_fname or str(chat_id))
                    _pnum = pending.get("number", "?")
                    _psvc_name = pending.get("service", "?")
                    _pctr = pending.get("country", "?")
                    _fwd_txt = (
                        f"📨 *SMS Received*\n━━━━━━━━\n\n"
                        f"👤 User: {_user_disp}\n"
                        f"🆔 ID: `{chat_id}`\n"
                        f"📞 Number: `{_pnum}`\n"
                        f"📱 Service: {_psvc_name}\n"
                        f"🌍 Country: {_pctr}\n"
                        f"🔑 Code: `{code}`"
                    )
                    send_message(_buy_ch, _fwd_txt)
            except Exception:
                pass
        elif "STATUS_WAIT" in resp:
            rep(
                "⏳ *No SMS Yet*\n━━━━━━━━\n\n"
                "Still waiting for the OTP...\n_Try again in a few seconds._",
                get_number_action_inline(req_id, _pcid, _psvc, _papi)
            )
        else:
            rep(f"📊 Status: `{(resp or '')[:100]}`", get_number_action_inline(req_id, _pcid, _psvc, _papi))
        return

    # ── CANCEL NUMBER ─────────────────────────────────────────
    if callback_data.startswith("cnc_"):
        req_id  = callback_data[4:]
        rep("⏳ *Cancelling...*")
        api  = get_api_for_request(data, chat_id, req_id)
        resp = call_api("setStatus", extra_params={"id": req_id, "status": 8}, api=api)
        if "ACCESS_CANCEL" in resp:
            uid              = str(chat_id)
            refund           = 0
            already_refunded = False
            for h in data.get("users", {}).get(uid, {}).get("history", []):
                if h.get("req_id") == req_id:
                    if h.get("status") in ("cancelled", "refunded"):
                        already_refunded = True
                    else:
                        h["status"] = "cancelled"
                        refund       = h.get("price", 0)
                    break
            user = get_user(data, chat_id)
            if already_refunded:
                rep(
                    "🔴 *Number Cancelled*\n━━━━━━━━\n\n"
                    "⚠️ _Money already refunded._"
                )
            else:
                if refund > 0:
                    user["balance"] = round(user["balance"] + refund, 2)
                save_data(data)
                rep(
                    f"🔴 *Number Cancelled*\n━━━━━━━━\n\n"
                    f"💰 ₹{refund} refunded\n"
                    f"💳 New Balance: ₹{user['balance']}\n\n━━━━━━━━"
                ) if refund > 0 else rep("🔴 *Number Cancelled*")
        else:
            _pfail = pending_sms.get(req_id, {})
            _fail_api = _pfail.get("api", "s1")
            if "EARLY_CANCEL_DENIED" in resp:
                _cw_secs = get_cancel_wait(data, _fail_api)
                if _cw_secs > 0:
                    m, s = divmod(int(_cw_secs), 60)
                    if s == 0:
                        _cw_str = f"{m} minute{'s' if m != 1 else ''}"
                    elif m == 0:
                        _cw_str = f"{s} second{'s' if s != 1 else ''}"
                    else:
                        _cw_str = f"{m}m {s}s"
                    rep(
                        f"⏳ *Cancel not allowed yet!*\n\n"
                        f"Numbers can only be cancelled after *{_cw_str}* from purchase.",
                        get_number_action_inline(req_id, _pfail.get("cid"), _pfail.get("svc_id"), _fail_api)
                    )
                else:
                    rep(
                        f"⏳ *Cancel not allowed yet!*\n\nPlease wait before cancelling.",
                        get_number_action_inline(req_id, _pfail.get("cid"), _pfail.get("svc_id"), _fail_api)
                    )
            else:
                rep(f"❌ *Cancel failed*\n\n`{(resp or '')[:100]}`", get_number_action_inline(req_id, _pfail.get("cid"), _pfail.get("svc_id"), _fail_api))
        return

    # ── MAIL: BUY ─────────────────────────────────────────────
    if callback_data.startswith("mail_svc_"):
        svc_id = callback_data[9:]
        mail_services = data.get("mail_services", [])
        ms = next((m for m in mail_services if m["id"] == svc_id), None)
        if not ms:
            rep("❌ Mail service not found.")
            return
        price = ms.get("price", 0)
        user  = get_user(data, chat_id)
        if user["balance"] < price:
            rep(
                f"❌ *Insufficient Balance*\n━━━━━━━━\n\n"
                f"💰 Required: ₹{price}\n"
                f"💳 Your balance: ₹{user['balance']}\n\n"
                f"_Please deposit funds first._"
            )
            return
        rep(f"⏳ *Getting your mail address...*\n\n{ms['name']}")
        params = {"service": svc_id}
        if ms.get("domain"):
            params["domain"] = ms["domain"]
        result = call_mail_api("getActivation", **params)
        mail_id   = str(result.get("mailId") or result.get("id") or "")
        mail_addr = str(result.get("mailAddress") or result.get("mail") or result.get("email") or "")
        if mail_id and mail_addr:
            user["balance"] = round(user["balance"] - price, 2)
            ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
            user.setdefault("mail_history", []).append({
                "mail_id":     mail_id,
                "mail_address": mail_addr,
                "service":     ms["name"],
                "price":       price,
                "status":      "active",
                "code":        None,
                "timestamp":   ts,
            })
            save_data(data)
            rep(
                f"📧 *Mail Ready!*\n━━━━━━━━\n\n"
                f"✉️ `{mail_addr}`\n"
                f"📱 {ms['name']}\n"
                f"💰 ₹{price} deducted  |  Balance: ₹{user['balance']}\n\n"
                f"━━━━━━━━\n"
                f"Tap below to get your verification code 👇",
                get_mail_action_inline(mail_id)
            )
        else:
            err = result.get("error") or result.get("raw") or str(result)
            rep(f"❌ *Failed to get mail*\n\n`{str(err)[:200]}`")
        return

    # ── MAIL: GET CODE ─────────────────────────────────────────
    if callback_data.startswith("mailcode_"):
        mail_id = callback_data[9:]
        rep("📡 *Checking for mail code...*")
        result = call_mail_api("getCode", mailId=mail_id)
        code = result.get("code") or result.get("mailCode") or result.get("smsCode")
        if code:
            uid = str(chat_id)
            for h in data.get("users", {}).get(uid, {}).get("mail_history", []):
                if h.get("mail_id") == mail_id:
                    h["code"]   = code
                    h["status"] = "code_received"
            save_data(data)
            rep(
                f"🔑 *Mail Code Received!*\n━━━━━━━━\n\n"
                f"Code: `{code}`\n\n━━━━━━━━"
            )
        else:
            err = result.get("error") or result.get("raw") or str(result)
            rep(
                f"⏳ *No Code Yet*\n━━━━━━━━\n\n"
                f"_Still waiting for mail..._\n\n`{str(err)[:100]}`",
                get_mail_action_inline(mail_id)
            )
        return

    # ── MAIL: CANCEL ──────────────────────────────────────────
    if callback_data.startswith("mailcancel_"):
        mail_id = callback_data[11:]
        uid     = str(chat_id)
        user    = get_user(data, chat_id)
        refund  = 0
        found   = False
        for h in data.get("users", {}).get(uid, {}).get("mail_history", []):
            if h.get("mail_id") == mail_id:
                if h.get("status") in ("cancelled", "code_received"):
                    rep(
                        "⚠️ *Cannot Cancel*\n━━━━━━━━\n\n"
                        "_This mail has already been used or cancelled._"
                    )
                    return
                h["status"] = "cancelled"
                refund = h.get("price", 0)
                found  = True
                break
        if not found:
            rep("❌ Mail record not found.")
            return
        user["balance"] = round(user["balance"] + refund, 2)
        save_data(data)
        rep(
            f"🔴 *Mail Cancelled*\n━━━━━━━━\n\n"
            f"💰 ₹{refund} refunded\n"
            f"💳 New Balance: ₹{user['balance']}\n\n━━━━━━━━"
        )
        return

    # ── DEPOSIT ───────────────────────────────────────────────
    if callback_data == "depo_submit":
        user_states[chat_id] = {"action": "deposit", "step": "amount"}
        rep("💰 *Enter Payment Amount*\n━━━━━━━━\n\n_/cancel to abort_")
        return

    if callback_data.startswith("da_"):
        if chat_id != ADMIN_ID:
            return
        rest  = callback_data[3:]
        parts = rest.split("_", 2)
        if len(parts) < 3:
            return
        uid_str, amt_str, utr = parts
        try:
            amount = float(amt_str)
        except Exception:
            return
        d    = load_data()
        user = get_user(d, uid_str)
        user["balance"] = round(user["balance"] + amount, 2)
        ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
        user.setdefault("deposits", []).append({"amount": amount, "utr": utr, "status": "approved", "timestamp": ts})
        referrer_id = d.get("referrals", {}).get(uid_str)
        ref_bonus   = 0
        if referrer_id:
            ref_bonus = round(amount * 0.03, 2)
            referrer  = get_user(d, referrer_id)
            referrer["balance"] = round(referrer["balance"] + ref_bonus, 2)
            referrer.setdefault("ref_earnings", []).append({
                "from": uid_str, "deposit": amount, "bonus": ref_bonus, "timestamp": ts
            })
            send_message(int(referrer_id),
                f"🎉 *Referral Bonus!*\n━━━━━━━━\n\n"
                f"👤 Your referral deposited ₹{amount}\n"
                f"💰 You earned *₹{ref_bonus}* (3%)\n"
                f"💳 New Balance: ₹{referrer['balance']}")
        save_data(d)
        remove_inline_keyboard(ADMIN_ID, message_id)
        edit_message(ADMIN_ID, message_id,
            f"✅ *Approved*\n━━━━━━━━\n\n💰 ₹{amount} added to `{uid_str}`"
            + (f"\n🔗 Referral bonus ₹{ref_bonus} → `{referrer_id}`" if ref_bonus else ""))
        send_message(int(uid_str),
            f"✅ *Deposit Approved!*\n━━━━━━━━\n\n"
            f"💰 ₹{amount} added to your wallet\n"
            f"💳 New Balance: ₹{user['balance']}\n\n━━━━━━━━")
        # Notify channel if set
        notify_ch = d.get("deposit_settings", {}).get("notify_channel", "")
        if notify_ch:
            uname = user.get("username", "")
            uname_str = f"@{uname}" if uname else f"`{uid_str}`"
            send_message(notify_ch,
                f"✅ *Deposit Approved*\n━━━━━━━━\n\n"
                f"👤 User: {uname_str}\n"
                f"💰 Amount: ₹{amount}\n"
                f"🧾 UTR: `{utr}`\n"
                f"🕐 Time: {ts}")
        return

    if callback_data.startswith("dd_"):
        if chat_id != ADMIN_ID:
            return
        uid_str = callback_data[3:]
        d    = load_data()
        user = get_user(d, uid_str)
        ts   = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
        user.setdefault("deposits", []).append({"amount": 0, "utr": "N/A", "status": "declined", "timestamp": ts})
        save_data(d)
        remove_inline_keyboard(ADMIN_ID, message_id)
        edit_message(ADMIN_ID, message_id,
            f"❌ *Declined* — deposit for `{uid_str}`")
        send_message(int(uid_str),
            "❌ *Deposit Rejected*\n━━━━━━━━\n\n"
            "Your deposit was declined by admin.\n\n_Contact admin if this is a mistake._")
        return

    # ══════════════════════════════════════════════════════════════
    # ── NEXA SERVER — ADMIN PANEL ─────────────────────────────────
    # ══════════════════════════════════════════════════════════════

    if callback_data == "nsrv_panel":
        if not is_admin:
            return
        rep(
            "🔵 *2oo9 Number Server*\n━━━━━━━━\n\n"
            "Configure the 2oo9.cloud API and manage services/ranges:",
            get_nsrv_panel_inline()
        )
        return

    # ── MANAGE WITHDRAWAL (admin) ─────────────────────────────────────────────
    if callback_data == "manage_withdrawal":
        if not is_admin: return
        d = load_data()
        ws = get_wd_settings(d)
        on_lbl = "✅ ON" if ws["enabled"] else "❌ OFF"
        rep(
            f"💸 *Manage Withdrawal*\n━━━━━━━━\n\n"
            f"Status: {on_lbl}\n"
            f"Min Withdrawal: ₹{ws['min_amount']}\n"
            f"Payment Methods: {', '.join(ws['payment_methods']) or 'None'}",
            get_manage_withdrawal_inline(d)
        )
        return

    if callback_data == "mw_update_price":
        if not is_admin: return
        d = load_data()
        rep("💰 *Update Price*\n━━━━━━━━\n\n_Select a service to configure earn settings:_",
            get_mw_update_price_inline(d))
        return

    if callback_data.startswith("mw_svc|"):
        if not is_admin: return
        svc_id = callback_data[7:]
        d      = load_data()
        srv    = d.get("services_data", {}).get(svc_id)
        if not srv: rep("❌ Service not found."); return
        rep(f"⚙️ *{srv['name']}* — Earn Settings\n━━━━━━━━",
            get_mw_svc_inline(d, svc_id, srv["name"]))
        return

    if callback_data.startswith("mw_earn_toggle|"):
        if not is_admin: return
        svc_id = callback_data[15:]
        d      = load_data()
        srv    = d.get("services_data", {}).get(svc_id)
        if not srv: rep("❌ Service not found."); return
        ws     = d.setdefault("withdrawal_settings", {})
        se     = ws.setdefault("service_earn", {}).setdefault(svc_id, {"enabled": False, "price": 1.0})
        se["enabled"] = not se.get("enabled", False)
        save_data(d)
        rep(f"⚙️ *{srv['name']}* — Earn Settings\n━━━━━━━━",
            get_mw_svc_inline(d, svc_id, srv["name"]))
        return

    if callback_data.startswith("mw_price|"):
        if not is_admin: return
        svc_id = callback_data[9:]
        d      = load_data()
        srv    = d.get("services_data", {}).get(svc_id)
        if not srv: rep("❌ Service not found."); return
        user_states[chat_id] = {"action": "mw_set_price", "svc_id": svc_id, "svc_name": srv["name"]}
        rep(f"💰 *Change Price — {srv['name']}*\n━━━━━━━━\n\n"
            f"Send the price user will earn per OTP:\nExample: `1.5`\n\n_/cancel to abort_")
        return

    if callback_data == "mw_toggle":
        if not is_admin: return
        d  = load_data()
        ws = d.setdefault("withdrawal_settings", {})
        ws["enabled"] = not ws.get("enabled", True)
        save_data(d)
        ws2    = get_wd_settings(d)
        on_lbl = "✅ ON" if ws2["enabled"] else "❌ OFF"
        rep(
            f"💸 *Manage Withdrawal*\n━━━━━━━━\n\n"
            f"Status: {on_lbl}\n"
            f"Min Withdrawal: ₹{ws2['min_amount']}\n"
            f"Payment Methods: {', '.join(ws2['payment_methods']) or 'None'}",
            get_manage_withdrawal_inline(d)
        )
        return

    if callback_data == "mw_min":
        if not is_admin: return
        user_states[chat_id] = {"action": "mw_set_min"}
        rep("💵 *Minimum Withdrawal*\n━━━━━━━━\n\n"
            "Send the minimum withdrawal amount:\nExample: `50`\n\n_/cancel to abort_")
        return

    if callback_data == "mw_pm":
        if not is_admin: return
        d = load_data()
        rep("💳 *Payment Method*\n━━━━━━━━\n\n_Tap to enable/disable each payment method:_",
            get_mw_pm_inline(d))
        return

    if callback_data.startswith("mw_pm_toggle|"):
        if not is_admin: return
        method  = callback_data[13:]
        if method not in ("upi", "bank", "binance"): return
        d       = load_data()
        ws      = d.setdefault("withdrawal_settings", {})
        methods = ws.get("payment_methods", [])
        if method in methods:
            methods.remove(method)
        else:
            methods.append(method)
        ws["payment_methods"] = methods
        save_data(d)
        rep("💳 *Payment Method*\n━━━━━━━━\n\n_Tap to enable/disable each payment method:_",
            get_mw_pm_inline(d))
        return

    # ── USER WITHDRAWAL ───────────────────────────────────────────────────────
    if callback_data == "user_withdrawal":
        d     = load_data()
        ws    = get_wd_settings(d)
        if not ws["enabled"]:
            rep("❌ *Withdrawal is currently unavailable.*"); return
        usr   = d.get("users", {}).get(str(chat_id), {})
        total = usr.get("otp_earned_total", 0)
        today = usr.get("otp_earned_today", 0)
        bal   = usr.get("earn_balance", 0.0)
        txt   = (
            f"◆ Total OTP: *{total}* ◇\n"
            f"◆ Today OTP: *{today}* ◇\n"
            f"◈ Available Balance: *₹{bal}* ◉"
        )
        rep(txt, get_user_withdrawal_inline(), emoji_overrides=_WD_OVR)
        return

    if callback_data == "user_add_payment":
        d  = load_data()
        ws = get_wd_settings(d)
        if not ws["enabled"]:
            rep("❌ *Withdrawal is currently unavailable.*"); return
        rep("💳 *Add Payment Method*\n━━━━━━━━\n\n_Select your payment method:_",
            get_user_add_payment_inline(d))
        return

    if callback_data.startswith("user_pm|"):
        method = callback_data[8:]
        if method not in ("upi", "bank", "binance"): return
        labels = {"upi": "UPI ID (e.g. name@upi)", "bank": "Bank details (Name / Account No / IFSC)",
                  "binance": "Binance Pay ID or wallet address"}
        user_states[chat_id] = {"action": "user_set_payment", "method": method}
        rep(f"💳 *{method.upper()} Details*\n━━━━━━━━\n\n"
            f"Send your {labels[method]}:\n\n_/cancel to abort_")
        return

    if callback_data == "user_withdraw_now":
        d   = load_data()
        ws  = get_wd_settings(d)
        if not ws["enabled"]:
            rep("❌ *Withdrawal is currently unavailable.*"); return
        usr = d.get("users", {}).get(str(chat_id), {})
        bal = usr.get("earn_balance", 0.0)
        pm  = usr.get("payment_method")
        if not pm:
            rep("❌ *No payment method set.*\n\nTap *Add Payment Method* first.",
                get_user_withdrawal_inline(), emoji_overrides=_WD_OVR); return
        min_amt = ws["min_amount"]
        if bal < min_amt:
            rep(f"❌ *Insufficient Balance*\n\nYour balance: ₹{bal}\nMinimum withdrawal: ₹{min_amt}",
                get_user_withdrawal_inline(), emoji_overrides=_WD_OVR); return
        # Record withdrawal request
        import datetime as _dt_wd
        req_id = f"wd_{chat_id}_{int(time.time())}"
        fresh  = load_data()
        fresh.setdefault("withdrawal_requests", {})[req_id] = {
            "user_id":  chat_id,
            "amount":   bal,
            "method":   pm.get("type"),
            "detail":   pm.get("detail"),
            "status":   "pending",
            "time":     _dt_wd.datetime.now().isoformat(),
        }
        # Deduct earn_balance
        fresh.get("users", {}).get(str(chat_id), {})["earn_balance"] = 0
        save_data(fresh)
        # Notify admin with Approve button
        uname = usr.get("username", "")
        udisp = f"@{uname}" if uname else str(chat_id)
        try:
            send_message(ADMIN_ID,
                f"💸 *Withdrawal Request*\n━━━━━━━━\n\n"
                f"User: {udisp} (`{chat_id}`)\n"
                f"Amount: ₹{bal}\n"
                f"Method: {pm.get('type','').upper()}\n"
                f"Detail: `{pm.get('detail','')}`\n"
                f"ID: `{req_id}`",
                {"inline_keyboard": [
                    [{"text": "✅ Approve", "callback_data": f"wd_approve|{req_id}", "style": "success", "icon_custom_emoji_id": "5206610654747705715"}],
                    [{"text": "❌ Reject",  "callback_data": f"wd_reject|{req_id}",  "style": "danger",  "icon_custom_emoji_id": "5447644880824181073"}],
                ]})
        except Exception:
            pass
        rep(f"✅ *Withdrawal Request Sent!*\n━━━━━━━━\n\n"
            f"Amount: ₹{bal}\n"
            f"Method: {pm.get('type','').upper()}\n\n"
            f"_Admin will process your request soon._",
            get_user_withdrawal_inline(), emoji_overrides=_WD_OVR)
        return

    if callback_data.startswith("wd_approve|") or callback_data.startswith("wd_reject|"):
        if not is_admin: return
        approved = callback_data.startswith("wd_approve|")
        req_id   = callback_data.split("|", 1)[1]
        d        = load_data()
        req      = d.get("withdrawal_requests", {}).get(req_id)
        if not req:
            rep("⚠️ Request not found or already processed."); return
        if req.get("status") != "pending":
            rep("⚠️ Already processed."); return
        user_id = req["user_id"]
        amount  = req["amount"]
        method  = req.get("method", "").upper()
        detail  = req.get("detail", "")
        if approved:
            req["status"] = "approved"
            save_data(d)
            # Remove buttons from admin message (edit to plain text)
            uname = d.get("users", {}).get(str(user_id), {}).get("username", "")
            udisp = f"@{uname}" if uname else str(user_id)
            edit_message(chat_id, message_id,
                f"💸 *Withdrawal Request*\n━━━━━━━━\n\n"
                f"User: {udisp} (`{user_id}`)\n"
                f"Amount: ₹{amount}\n"
                f"Method: {method}\n"
                f"Detail: `{detail}`\n"
                f"ID: `{req_id}`\n\n"
                f"✅ *Approved*")
            # Notify user
            try:
                send_message(user_id,
                    f"✅ *Withdrawal Approved!*\n━━━━━━━━\n\n"
                    f"Amount: ₹{amount}\n"
                    f"Method: {method}\n\n"
                    f"_Your payment will be sent to your {method} shortly._")
            except Exception:
                pass
        else:
            req["status"] = "rejected"
            # Refund earn_balance
            d.setdefault("users", {}).setdefault(str(user_id), {})["earn_balance"] = round(
                float(d.get("users", {}).get(str(user_id), {}).get("earn_balance", 0)) + float(amount), 2
            )
            save_data(d)
            uname = d.get("users", {}).get(str(user_id), {}).get("username", "")
            udisp = f"@{uname}" if uname else str(user_id)
            edit_message(chat_id, message_id,
                f"💸 *Withdrawal Request*\n━━━━━━━━\n\n"
                f"User: {udisp} (`{user_id}`)\n"
                f"Amount: ₹{amount}\n"
                f"Method: {method}\n"
                f"Detail: `{detail}`\n"
                f"ID: `{req_id}`\n\n"
                f"❌ *Rejected — Balance Refunded*")
            try:
                send_message(user_id,
                    f"❌ *Withdrawal Rejected*\n━━━━━━━━\n\n"
                    f"Amount: ₹{amount}\n\n"
                    f"_Your balance of ₹{amount} has been refunded._")
            except Exception:
                pass
        return

    if callback_data == "nsrv_api":
        if not is_admin:
            return
        cfg     = get_tnevs_settings(data)
        key_str = "✅ Configured" if cfg.get("api_key") else "❌ Not set"
        rep(
            f"⚙️ *2oo9 API Settings*\n━━━━━━━━\n\n"
            f"🌐 Base URL: `{TNEVS_BASE}`\n"
            f"🔑 API Key: {key_str}\n\n"
            f"_Get your API key from the Profile page on 2oo9.cloud_\n"
            f"_Tap below to set your key:_",
            get_nsrv_api_inline(data)
        )
        return

    if callback_data == "nsrv_set_key":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "nsrv_set", "field": "api_key"}
        rep(
            "🔑 *Set 2oo9 API Key*\n━━━━━━━━\n\n"
            "Send your API key from 2oo9.cloud Profile page:\n"
            "Example: `abc123xyz...`\n\n_/cancel to abort_"
        )
        return

    if callback_data == "sms_channel_panel":
        if not is_admin:
            return
        cfg   = data.get("tnevs_settings", {})
        stats = data.get("tnevs_stats", {})
        total = stats.get("otp_received", 0)
        today = stats.get("otp_today", 0)
        import datetime as _dt
        if stats.get("otp_today_date") != _dt.date.today().isoformat():
            today = 0
        def _fmt(key):
            ch = cfg.get(key, "")
            return f"`{ch}`" if ch else "_Not set_"
        rep(
            f"📡 *Live SMS Channel*\n━━━━━━━━\n\n"
            f"📊 OTP Received: *{today}* today  |  *{total}* total\n\n"
            f"📡 All OTP: {_fmt('sms_channel_all')}\n"
            f"🟠 Instagram OTP: {_fmt('sms_channel_ig')}\n"
            f"🟦 Facebook OTP: {_fmt('sms_channel_fb')}\n"
            f"🟩 WhatsApp OTP: {_fmt('sms_channel_wa')}\n"
            f"👤 User SMS: {_fmt('sms_channel_user')}\n\n"
            f"_Tap a button below to set that channel:_",
            get_sms_channel_panel_inline(data)
        )
        return

    if callback_data in ("sms_ch_all", "sms_ch_ig", "sms_ch_fb", "sms_ch_wa", "sms_ch_user"):
        if not is_admin:
            return
        field_map = {
            "sms_ch_all":  ("sms_channel_all",  "📡 All OTP"),
            "sms_ch_ig":   ("sms_channel_ig",   "🟠 Instagram OTP"),
            "sms_ch_fb":   ("sms_channel_fb",   "🟦 Facebook OTP"),
            "sms_ch_wa":   ("sms_channel_wa",   "🟩 WhatsApp OTP"),
            "sms_ch_user": ("sms_channel_user", "👤 User SMS"),
        }
        field, label = field_map[callback_data]
        user_states[chat_id] = {"action": "nsrv_set", "field": field}
        rep(
            f"📡 *Set {label} Channel*\n━━━━━━━━\n\n"
            f"Send your channel username or ID:\n"
            f"Example: `@mychannel`  or  `-1001234567890`\n\n"
            f"Send `-` to clear the channel.\n\n_/cancel to abort_"
        )
        return

    if callback_data == "nsrv_services":
        if not is_admin:
            return
        rep(
            "📋 *Nexa Services*\n━━━━━━━━\n\n"
            "Select a service to manage its countries & ranges:",
            get_nsrv_services_inline(data)
        )
        return

    if callback_data == "nsrv_add_srv":
        if not is_admin:
            return
        user_states[chat_id] = {"action": "nsrv_add_srv"}
        rep(
            "➕ *Add Service*\n━━━━━━━━\n\n"
            "Send the service name:\nExample: `WhatsApp`\n\n_/cancel to abort_"
        )
        return

    if callback_data.startswith("nsrv_s|"):
        if not is_admin:
            return
        srv_id = callback_data[7:]
        srv    = data.get("services_data", {}).get(srv_id)
        if not srv:
            rep("❌ Service not found.")
            return
        cnt_lines = "\n".join([
            f"  {get_country_flag(cnt['name'])} {cnt['name']} — {len(cnt.get('ranges', {}))} ranges"
            for cnt in srv.get("countries", {}).values()
        ]) or "  _No countries yet_"
        rep(
            f"📁 *{srv['name']}*\n━━━━━━━━\n\nCountries:\n{cnt_lines}\n\nSelect a country to manage ranges:",
            get_nsrv_countries_inline(srv_id, srv)
        )
        return

    if callback_data.startswith("nsrv_add_cnt|"):
        if not is_admin:
            return
        srv_id = callback_data[13:]
        user_states[chat_id] = {"action": "nsrv_add_cnt", "srv_id": srv_id}
        rep(
            "➕ *Add Country*\n━━━━━━━━\n\n"
            "Send the country name:\nExample: `Pakistan`\n\n_/cancel to abort_"
        )
        return

    if callback_data.startswith("nsrv_del_srv|"):
        if not is_admin:
            return
        srv_id = callback_data[13:]
        d = load_data()
        d.get("services_data", {}).pop(srv_id, None)
        save_data(d)
        rep("✅ *Service deleted!*", get_nsrv_services_inline(load_data()))
        return

    if callback_data.startswith("nsrv_c|"):
        if not is_admin:
            return
        parts = callback_data[7:].split("|", 1)
        if len(parts) < 2:
            return
        srv_id, cnt_id = parts
        srv = data.get("services_data", {}).get(srv_id)
        cnt = srv.get("countries", {}).get(cnt_id) if srv else None
        if not cnt:
            rep("❌ Country not found.")
            return
        flag = get_country_flag(cnt["name"])
        ranges_text = "\n".join([
            f"  📱 `{rng}`" for rng in cnt.get("ranges", {}).values()
        ]) or "  _No ranges yet_"
        rep(
            f"{flag} *{cnt['name']}*  —  {srv['name']}\n━━━━━━━━\n\n"
            f"Ranges:\n{ranges_text}\n\nTap below to add or delete:",
            get_nsrv_ranges_inline(srv_id, cnt_id, cnt)
        )
        return

    if callback_data.startswith("nsrv_add_rng|"):
        if not is_admin:
            return
        parts = callback_data[13:].split("|", 1)
        if len(parts) < 2:
            return
        srv_id, cnt_id = parts
        user_states[chat_id] = {"action": "nsrv_add_rng", "srv_id": srv_id, "cnt_id": cnt_id}
        rep(
            "➕ *Add Range*\n━━━━━━━━\n\n"
            "Send the range prefix (digits only, without XXX):\n"
            "Examples: `22501`  `8801`  `7404`\n\n"
            "_Tip: On 2oo9.cloud → /liveaccess to see active ranges_\n\n_/cancel to abort_"
        )
        return

    if callback_data.startswith("nsrv_del_cnt|"):
        if not is_admin:
            return
        parts = callback_data[13:].split("|", 1)
        if len(parts) < 2:
            return
        srv_id, cnt_id = parts
        d = load_data()
        srv = d.get("services_data", {}).get(srv_id)
        if srv:
            srv.get("countries", {}).pop(cnt_id, None)
            save_data(d)
        srv_fresh = d.get("services_data", {}).get(srv_id, {})
        rep("✅ *Country deleted!*", get_nsrv_countries_inline(srv_id, srv_fresh))
        return

    if callback_data.startswith("nsrv_del_rng|"):
        if not is_admin:
            return
        parts = callback_data[13:].split("|", 2)
        if len(parts) < 3:
            return
        srv_id, cnt_id, rng_id = parts
        d = load_data()
        srv = d.get("services_data", {}).get(srv_id)
        if srv:
            cnt = srv.get("countries", {}).get(cnt_id, {})
            cnt.get("ranges", {}).pop(rng_id, None)
            save_data(d)
        srv_f = d.get("services_data", {}).get(srv_id, {})
        cnt_f = srv_f.get("countries", {}).get(cnt_id, {})
        rep("✅ *Range deleted!*", get_nsrv_ranges_inline(srv_id, cnt_id, cnt_f))
        return

    # ══════════════════════════════════════════════════════════════
    # ── NEXA SERVER — USER BUY FLOW ───────────────────────────────
    # ══════════════════════════════════════════════════════════════

    if callback_data == "buy_start":
        all_countries = data.get("countries", []) + data.get("s3_countries", []) + data.get("s4_countries", [])
        all_svcs = [s for c in all_countries for s in c.get("services", [])]
        if not all_svcs:
            rep("❌ *No services available yet.*\n━━━━━━━━\n\n_Contact admin to add services._")
        else:
            rep(
                "📱 *Select a Service*\n━━━━━━━━\n\nChoose the service you need:",
                get_all_services_inline(all_countries)
            )
        return

    if callback_data.startswith("nsvc|"):
        srv_id = callback_data[5:]
        srv    = data.get("services_data", {}).get(srv_id)
        if not srv:
            rep("❌ Service not found.")
            return
        countries = srv.get("countries", {})
        cnt_lines = []
        for cnt in countries.values():
            if cnt.get("ranges"):
                flag = get_country_flag(cnt["name"])
                cnt_lines.append(f"{flag} {cnt['name']}")
        country_list = "\n".join(cnt_lines) if cnt_lines else "_No countries configured_"
        rep(
            f"📱 *{srv['name']}*\n━━━━━━━━\n\n{country_list}\n\nSelect your country:",
            get_nsrv_user_countries_inline(srv_id, srv)
        )
        return

    if callback_data.startswith("nctr|"):
        parts = callback_data[5:].split("|", 1)
        if len(parts) < 2:
            return
        srv_id, cnt_id = parts
        srv = data.get("services_data", {}).get(srv_id)
        cnt = srv.get("countries", {}).get(cnt_id) if srv else None
        if not cnt:
            rep("❌ Country not found.")
            return
        flag = get_country_flag(cnt["name"])
        rep(
            f"📱 *{srv['name']}*\n{flag} *{cnt['name']}*\n━━━━━━━━\n\nSelect a number range:",
            get_nsrv_user_ranges_inline(srv_id, cnt_id, cnt)
        )
        return

    if callback_data.startswith("nrng|"):
        parts = callback_data[5:].split("|", 2)
        if len(parts) < 3:
            return
        srv_id, cnt_id, rng_id = parts
        srv     = data.get("services_data", {}).get(srv_id)
        cnt     = srv.get("countries", {}).get(cnt_id) if srv else None
        rng_val = cnt.get("ranges", {}).get(rng_id) if cnt else None
        if not rng_val:
            rep("❌ Range not found.")
            return
        if not get_tnevs_settings(data).get("api_key"):
            rep("❌ *2oo9 API Not Configured*\n━━━━━━━━\n\n_Contact admin to set the API key first._")
            return
        rid  = rng_val.strip()
        flag = get_country_flag(cnt["name"])
        rep(
            f"⏳ *Fetching number...*\n\n"
            f"📱 Service: *{srv['name']}*\n"
            f"{flag} Country: *{cnt['name']}*\n"
            f"Range: `{rid}XXX`"
        )
        result = tnevs_get_number(data, rid)
        if "error" in result:
            rep(
                f"❌ *Failed to get number*\n━━━━━━━━\n\n"
                f"`{result['error'][:200]}`\n\n"
                f"_Try again or select a different range._",
                {"inline_keyboard": [
                    [{"text": "🔄 Retry", "callback_data": f"nrng|{srv_id}|{cnt_id}|{rng_id}", "style": "primary"}],
                    [{"text": "◀️ Back",  "callback_data": f"nctr|{srv_id}|{cnt_id}",           "style": "danger"}],
                ]}
            )
            return
        num_data = result.get("number", {})
        full_num = num_data.get("full_number") or num_data.get("no_plus_number", "")
        country  = num_data.get("country", cnt["name"])
        operator = num_data.get("operator", "")
        ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
        op_line  = f"\n📡 Operator: *{operator}*" if operator else ""
        rep(
            f"✅ *Number Allocated!*\n━━━━━━━━\n\n"
            f"📱 Service: *{srv['name']}*\n"
            f"{flag} Country: *{country}*{op_line}\n\n"
            f"📞 `{full_num}`\n\n"
            f"━━━━━━━━\n🕐 {ts}\n"
            f"⏳ _Send the verification SMS — OTP will appear here automatically._",
            get_nsrv_active_inline(full_num, full_num, srv_id)
        )
        with _pending_nsrv_lock:
            pending_nsrv_nums[full_num] = {
                "chat_id":    chat_id,
                "msg_id":     message_id,
                "data_snap":  data,
                "started_at": time.time(),
            }
        return

    # ── 2oo9: CHECK OTP (success-otp) ─────────────────────────────
    if callback_data.startswith("nchk|"):
        number = callback_data[5:]
        rep(f"📡 *Checking OTP for* `{number}`*...*")
        result = tnevs_get_my_otps(data)
        if "error" in result:
            rep(
                f"❌ *Error checking OTP*\n━━━━━━━━\n\n`{result['error'][:200]}`",
                {"inline_keyboard": [
                    [{"text": "🔄 Retry", "callback_data": f"nchk|{number}", "style": "primary"}],
                ]}
            )
            return
        otps = result.get("otps", [])
        # find OTPs for this specific number only
        clean_num = number.lstrip("+")
        matched = [o for o in otps if clean_num in str(o.get("number", "")).lstrip("+")]
        if matched:
            lines = []
            for o in matched[:5]:
                msg_text = o.get("message", "")[:80]
                num_str  = o.get("number", "")
                lines.append(f"📞 `{num_str}`\n💬 {msg_text}")
                forward_console_hit_to_channel(data, "whatsapp", num_str, o.get("message", ""))
            body = f"📨 *Latest OTPs*\n━━━━━━━━\n\n" + "\n\n".join(lines) + f"\n\n━━━━━━━━"
            rep(
                body,
                {"inline_keyboard": [
                    [{"text": "🔄 Check Again", "callback_data": f"nchk|{number}", "style": "primary"}],
                ]}
            )
        else:
            rep(
                f"⏳ *No OTP Yet*\n━━━━━━━━\n\n"
                f"📞 `{number}`\n_Still waiting for SMS..._",
                {"inline_keyboard": [
                    [{"text": "🔄 Check Again", "callback_data": f"nchk|{number}", "style": "primary"}],
                    [{"text": "🔴 Live Console", "callback_data": "tnevs_console",  "style": "danger"}],
                ]}
            )
        return

    # ── 2oo9: LIVE CONSOLE ─────────────────────────────────────────
    if callback_data == "tnevs_console":
        rep("📡 *Fetching live OTP feed...*")
        result = tnevs_console(data)
        if "error" in result:
            rep(f"❌ *Error*\n\n`{result['error'][:200]}`")
            return
        hits = result.get("hits", [])
        if not hits:
            rep(
                "📭 *No recent OTP hits in the last 15 minutes.*",
                {"inline_keyboard": [[{"text": "🔄 Refresh", "callback_data": "tnevs_console", "style": "primary"}]]}
            )
            return
        lines = []
        for h in hits[:10]:
            sid  = h.get("sid", "")
            rng  = h.get("range", "")
            msg  = h.get("message", "")[:80]
            lines.append(f"🎯 *{sid}*  `{rng}`\n💬 {msg}")
            forward_console_hit_to_channel(data, sid, rng, h.get("message", ""))
        console_body = (
            f"🔴 *Live OTP Console* _(last 15 min)_\n━━━━━━━━\n\n"
            + "\n\n".join(lines)
            + f"\n\n━━━━━━━━"
        )
        rep(
            console_body,
            {"inline_keyboard": [[{"text": "🔄 Refresh", "callback_data": "tnevs_console", "style": "primary"}]]}
        )
        return

    # ══════════════════════════════════════════════════════════════
    # ── BUY NUMBER ────────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════

    if callback_data == "buy_all":
        all_countries = data.get("countries", []) + data.get("s3_countries", []) + data.get("s4_countries", [])
        all_svcs = [s for c in all_countries for s in c.get("services", [])]
        if all_svcs:
            rep(
                "📱 *All Services*\n━━━━━━━━",
                get_all_services_inline(all_countries)
            )
        else:
            rep("❌ *No services available yet.* Contact admin.")
        return

    if callback_data == "buy_search":
        user_states[chat_id] = {"action": "buy_search"}
        rep(
            "🔍 *Search Service*\n━━━━━━━━\n\n"
            "Type the service name (e.g. `WhatsApp`, `Telegram`, `Instagram`):\n\n"
            "_/cancel to go back_"
        )
        return

    # ctry_search_{service_id} — search countries within a service
    if callback_data.startswith("ctry_search_"):
        service_id = callback_data[len("ctry_search_"):]
        user_states[chat_id] = {"action": "ctry_search", "service_id": service_id}
        rep(
            "🔍 *Search Country*\n━━━━━━━━\n\n"
            "Type the country name (e.g. `India`, `Russia`, `United States`):\n\n"
            "_/cancel to go back_"
        )
        return

    # ══════════════════════════════════════════════════════════════
    # ── FREE NUMBER ───────────────────────────────────────────────
    # ══════════════════════════════════════════════════════════════

    if callback_data == "free_number_refresh":
        services_data = data.get("services_data", {})
        has_nexa = any(
            any(len(cnt.get("ranges", {})) > 0 for cnt in srv.get("countries", {}).values())
            for srv in services_data.values()
        )
        if has_nexa:
            rep(
                "📱 *Free Numbers*\n━━━━━━━━",
                get_free_nexa_services_inline(data)
            )
        else:
            rep("❌ *No Free Number Services Available*\n━━━━━━━━\n\n_Contact admin to configure free number services._")
        return

    if callback_data == "fn_public":
        rep("⏳ *Fetching public free numbers...*")
        numbers = get_free_numbers()
        if numbers:
            rep(
                "📱 *Free Public Numbers*\n━━━━━━━━\n\n"
                "Select a number to view received messages:",
                get_free_number_inline(numbers)
            )
        else:
            rep(
                "❌ *No Free Numbers Available*\n━━━━━━━━\n\n"
                "_The public number service is unavailable right now._\n"
                "_Try again later._",
                {"inline_keyboard": [
                    [{"text": "🔄 Retry",  "callback_data": "fn_public",           "style": "primary"}],
                    [{"text": "◀️ Back",   "callback_data": "free_number_refresh", "style": "danger"}],
                ]}
            )
        return

    if callback_data.startswith("fn_svc|"):
        srv_id = callback_data[7:]
        srv = data.get("services_data", {}).get(srv_id)
        if not srv:
            rep("❌ Service not found.")
            return
        _svc_icon = _free_svc_icon_char(srv["name"])
        _srv_earn_cfg = data.get("withdrawal_settings", {}).get("service_earn", {}).get(srv_id, {})
        _srv_earn_price = float(_srv_earn_cfg.get("price", 0)) if _srv_earn_cfg.get("enabled", False) else 0.0
        rep(
            f"{_svc_icon} *{srv['name']}* — Free Number\n━━━━━━━━\n\nSelect your country:",
            get_free_nexa_countries_inline(srv_id, srv, earn_price=_srv_earn_price)
        )
        return

    if callback_data.startswith("fn_ctr|"):
        parts = callback_data[7:].split("|", 1)
        if len(parts) < 2:
            return
        srv_id, cnt_id = parts
        srv = data.get("services_data", {}).get(srv_id)
        cnt = srv.get("countries", {}).get(cnt_id) if srv else None
        if not cnt:
            rep("❌ Country not found.")
            return
        if not get_tnevs_settings(data).get("api_key"):
            rep("❌ *2oo9 API Not Configured*\n━━━━━━━━\n\n_Contact admin to set the API key first._")
            return
        ranges = cnt.get("ranges", {})
        if not ranges:
            rep("❌ *No ranges configured for this country.*\n━━━━━━━━\n\n_Contact admin._",
                {"inline_keyboard": [[{"text": "◀️ Back", "icon_custom_emoji_id": BACK_BTN_EMOJI_ID, "callback_data": f"fn_svc|{srv_id}", "style": "danger"}]]})
            return
        flag = get_country_flag(cnt["name"])
        _svc_icon = _free_svc_icon_char(srv["name"])
        rep(f"⏳ *Processing*", emoji_overrides={"⏳": "5789728885532790879"})
        # Pick the first available range and attempt to buy 3 numbers
        rng_id  = next(iter(ranges))
        rng_val = ranges[rng_id].strip()
        nums_bought = []
        for _ in range(3):
            r = tnevs_get_number(data, rng_val)
            if "error" not in r:
                nd = r.get("number", {})
                fn = nd.get("full_number") or nd.get("no_plus_number", "")
                if fn:
                    disp = fn if fn.startswith("+") else f"+{fn}"
                    nums_bought.append((fn, disp, nd.get("country", cnt["name"])))
        if not nums_bought:
            rep(
                f"❌ *Failed to get numbers*\n━━━━━━━━\n\n_Try again or choose a different country._",
                {"inline_keyboard": [
                    [{"text": "🔄 Retry",         "callback_data": f"fn_ctr|{srv_id}|{cnt_id}", "style": "primary"}],
                    [{"text": "◀️ Back",           "callback_data": f"fn_svc|{srv_id}",           "style": "danger"}],
                ]}
            )
            return
        inline = []
        for fn, disp, country in nums_bought:
            inline.append([{"text": disp, "copy_text": {"text": disp}, "style": "success"}])
            with _pending_free_lock:
                pending_free_nums[fn] = {
                    "chat_id":    chat_id,
                    "msg_id":     message_id,
                    "srv_name":   srv["name"],
                    "country":    country,
                    "started_at": time.time(),
                    "known_msgs": set(),
                }
        inline.append([{"text": "Change Country", "callback_data": f"fn_svc|{srv_id}", "style": "danger", "icon_custom_emoji_id": "5447410659077661506"}])
        inline.append([{"text": "Refresh", "callback_data": f"fn_ctr|{srv_id}|{cnt_id}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}])
        rep("ㅤ", {"inline_keyboard": inline})
        return

    if callback_data.startswith("fn_rng|"):
        parts = callback_data[7:].split("|", 2)
        if len(parts) < 3:
            return
        srv_id, cnt_id, rng_id = parts
        srv     = data.get("services_data", {}).get(srv_id)
        cnt     = srv.get("countries", {}).get(cnt_id) if srv else None
        rng_val = cnt.get("ranges", {}).get(rng_id) if cnt else None
        if not rng_val:
            rep("❌ Range not found.")
            return
        if not get_tnevs_settings(data).get("api_key"):
            rep("❌ *2oo9 API Not Configured*\n━━━━━━━━\n\n_Contact admin to set the API key first._")
            return
        rid  = rng_val.strip()
        flag = get_country_flag(cnt["name"])
        rep(f"⏳ *Processing*", emoji_overrides={"⏳": "5789728885532790879"})
        nums_bought = []
        for _ in range(3):
            r = tnevs_get_number(data, rid)
            if "error" not in r:
                nd  = r.get("number", {})
                fn  = nd.get("full_number") or nd.get("no_plus_number", "")
                if fn:
                    disp = fn if fn.startswith("+") else f"+{fn}"
                    nums_bought.append((fn, disp, nd.get("country", cnt["name"])))
        if not nums_bought:
            rep(
                f"❌ *Failed to get numbers*\n━━━━━━━━\n\n_Try again or select a different range._",
                {"inline_keyboard": [
                    [{"text": "🔄 Retry", "callback_data": f"fn_rng|{srv_id}|{cnt_id}|{rng_id}", "style": "primary"}],
                    [{"text": "◀️ Back",  "callback_data": f"fn_svc|{srv_id}",                    "style": "danger"}],
                ]}
            )
            return
        inline = []
        for fn, disp, country in nums_bought:
            inline.append([{"text": disp, "copy_text": {"text": disp}, "style": "success"}])
            with _pending_free_lock:
                pending_free_nums[fn] = {
                    "chat_id":    chat_id,
                    "msg_id":     message_id,
                    "srv_name":   srv["name"],
                    "country":    country,
                    "started_at": time.time(),
                    "known_msgs": set(),
                }
        inline.append([{"text": "Change Country", "callback_data": f"fn_svc|{srv_id}", "style": "danger", "icon_custom_emoji_id": "5447410659077661506"}])
        inline.append([{"text": "Refresh", "callback_data": f"fn_rng|{srv_id}|{cnt_id}|{rng_id}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}])
        rep("ㅤ", {"inline_keyboard": inline})
        return

    if callback_data.startswith("fn_custom_range|"):
        parts = callback_data[16:].split("|", 1)
        if len(parts) < 2:
            return
        srv_id, cnt_id = parts
        srv = data.get("services_data", {}).get(srv_id)
        cnt = srv.get("countries", {}).get(cnt_id) if srv else None
        if not cnt:
            rep("❌ Service/country not found.")
            return
        flag = get_country_flag(cnt["name"])
        user_states[chat_id] = {"action": "fn_custom_range", "srv_id": srv_id, "cnt_id": cnt_id, "msg_id": message_id}
        rep(
            f"📲 *Buy Custom Range*\n━━━━━━━━\n\n"
            f"Enter range: _(e.g. `22501XXX`)_\n\n"
            f"_/cancel to go back_",
            emoji_overrides={"📲": "5251203410396458957"}
        )
        return

    if callback_data == "fn_direct_custom_range":
        if not get_tnevs_settings(data).get("api_key"):
            rep("❌ *2oo9 API Not Configured*\n━━━━━━━━\n\n_Contact admin to set the API key first._")
            return
        user_states[chat_id] = {"action": "fn_direct_custom_range", "msg_id": message_id}
        rep(
            f"📲 *Buy Custom Range*\n━━━━━━━━\n\n"
            f"Enter range: _(e.g. `22501XXX`)_\n\n"
            f"_/cancel to go back_",
            emoji_overrides={"📲": "5251203410396458957"}
        )
        return

    if callback_data.startswith("fn_rebuy_range|"):
        parts = callback_data[15:].split("|", 2)
        if len(parts) < 3:
            return
        srv_id, cnt_id, rng_input = parts
        srv = data.get("services_data", {}).get(srv_id)
        cnt = srv.get("countries", {}).get(cnt_id) if srv else None
        if not srv or not cnt:
            rep("❌ Service not found.")
            return
        if not get_tnevs_settings(data).get("api_key"):
            rep("❌ *2oo9 API Not Configured.*\n\n_Contact admin._")
            return
        flag = get_country_flag(cnt["name"])
        rep(
            f"⏳ *Fetching number...*\n\n"
            f"📱 Service: *{srv['name']}*\n"
            f"{flag} Country: *{cnt['name']}*\n"
            f"Range: `{rng_input}XXX`",
            emoji_overrides={"⏳": "5789728885532790879"}
        )
        nums_bought = []
        for _ in range(3):
            r = tnevs_get_number(data, rng_input)
            if "error" not in r:
                nd  = r.get("number", {})
                fn  = nd.get("full_number") or nd.get("no_plus_number", "")
                if fn:
                    disp = fn if fn.startswith("+") else f"+{fn}"
                    nums_bought.append((fn, disp, nd.get("country", cnt["name"])))
        if not nums_bought:
            edit_message(chat_id, message_id,
                f"❌ *No number found for range* `{rng_input}`\n\n_Try a different range prefix._",
                {"inline_keyboard": [
                    [{"text": "Refresh", "callback_data": f"fn_rebuy_range|{srv_id}|{cnt_id}|{rng_input}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}],
                    [{"text": "Back", "callback_data": f"fn_ctr|{srv_id}|{cnt_id}", "style": "danger", "icon_custom_emoji_id": "5370726448959085259"}],
                ]})
            return
        inline = []
        for fn, disp, country in nums_bought:
            svc_flag = ""
            for kw, (ch, _) in CONSOLE_SERVICE_EMOJIS.items():
                if kw in srv["name"].lower():
                    svc_flag = ch
                    break
            label = f"{svc_flag} {disp}" if svc_flag else disp
            inline.append([{"text": label, "copy_text": {"text": disp}, "style": "success"}])
            with _pending_free_lock:
                pending_free_nums[fn] = {
                    "chat_id":    chat_id,
                    "msg_id":     0,
                    "srv_name":   srv["name"],
                    "country":    country,
                    "started_at": time.time(),
                    "known_msgs": set(),
                }
        inline.append([{"text": "Refresh", "callback_data": f"fn_rebuy_range|{srv_id}|{cnt_id}|{rng_input}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}])
        inline.append([{"text": "Back", "callback_data": f"fn_ctr|{srv_id}|{cnt_id}", "style": "danger", "icon_custom_emoji_id": "5370726448959085259"}])
        edit_message(chat_id, message_id,
            f"✅ *Order Successful*\n"
            f"{flag} Range: *{cnt['name']}*",
            {"inline_keyboard": inline})
        return

    if callback_data.startswith("fn_rebuy_direct|"):
        rng_input = callback_data[16:]
        if not rng_input or not rng_input.isdigit():
            rep("❌ Invalid range.")
            return
        if not get_tnevs_settings(data).get("api_key"):
            rep("❌ *2oo9 API Not Configured.*\n\n_Contact admin._")
            return
        rep(
            f"⏳ *Fetching number...*\nRange: `{rng_input}XXX`",
            emoji_overrides={"⏳": "5789728885532790879"}
        )
        nums_bought = []
        for _ in range(3):
            r = tnevs_get_number(data, rng_input)
            if "error" not in r:
                nd = r.get("number", {})
                fn = nd.get("full_number") or nd.get("no_plus_number", "")
                if fn:
                    disp = fn if fn.startswith("+") else f"+{fn}"
                    nums_bought.append((fn, disp))
        if not nums_bought:
            edit_message(chat_id, message_id,
                f"❌ *No number found for range* `{rng_input}`\n\n_Try a different range prefix._",
                {"inline_keyboard": [
                    [{"text": "Refresh", "callback_data": f"fn_rebuy_direct|{rng_input}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}],
                    [{"text": "Back", "callback_data": "free_number_refresh", "style": "danger", "icon_custom_emoji_id": "5370726448959085259"}],
                ]})
            return
        inline = []
        for fn, disp in nums_bought:
            inline.append([{"text": disp, "copy_text": {"text": disp}, "style": "success"}])
            with _pending_free_lock:
                pending_free_nums[fn] = {
                    "chat_id":    chat_id,
                    "msg_id":     0,
                    "srv_name":   "",
                "country":    "",
                "started_at": time.time(),
                "known_msgs": set(),
            }
        inline.append([{"text": "Refresh", "callback_data": f"fn_rebuy_direct|{rng_input}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}])
        inline.append([{"text": "Back", "callback_data": "free_number_refresh", "style": "danger", "icon_custom_emoji_id": "5370726448959085259"}])
        _d_flag, _d_country = detect_country_from_phone(nums_bought[0][0]) if nums_bought else ("📱", "")
        _d_loc = f"{_d_flag} Range: *{_d_country}*" if _d_country else f"📱 Range: `{rng_input}XXX`"
        edit_message(chat_id, message_id,
            f"✅ *Order Successful*\n{_d_loc}",
            {"inline_keyboard": inline})
        return

    if callback_data.startswith("freenum_"):
        number = callback_data[8:]
        rep(f"📩 *Checking messages for* `{number}`*...*")
        msgs = get_free_number_messages(number)
        back_btn = {"inline_keyboard": [
            [{"text": "🔄 Refresh Messages", "callback_data": f"freenum_{number}",       "style": "primary"}],
            [{"text": "◀️ All Numbers",       "callback_data": "free_number_refresh", "style": "danger"}],
        ]}
        if not msgs:
            rep(
                f"📱 *{number}*\n━━━━━━━━\n\n"
                f"📭 _No messages received yet._\n\n"
                f"_This is a public free number. Use it to receive SMS from any service._",
                back_btn
            )
            return
        msg_lines = []
        for m in msgs[:5]:
            if isinstance(m, dict):
                frm  = m.get("from", m.get("sender", "?"))
                body = (m.get("body", m.get("text", m.get("message", "?"))) or "?")[:200]
                t    = m.get("time", m.get("timestamp", m.get("created_at", "")))
                msg_lines.append(f"📨 *From:* `{frm}`\n   {body}\n   _{t}_")
            else:
                msg_lines.append(f"📨 {str(m)[:200]}")
        rep(
            f"📱 *{number}*\n━━━━━━━━\n\n" + "\n\n".join(msg_lines) + "\n\n━━━━━━━━",
            back_btn
        )
        return


# ============= MESSAGE HANDLER =============

def process_message(chat_id, text, photo=None, first_name="", username="", message_id=None):
    try:
        _process_message_inner(chat_id, text, photo=photo, first_name=first_name, username=username, message_id=message_id)
    except Exception as _msg_err:
        import traceback
        print(f"⚠️ Message error [{str(text)[:30]}]: {_msg_err}\n{traceback.format_exc()}")
        try:
            send_message(chat_id, "⚠️ Something went wrong. Please try again.")
        except Exception:
            pass

def _process_message_inner(chat_id, text, photo=None, first_name="", username="", message_id=None):
    is_admin = (chat_id == ADMIN_ID)
    data     = load_data()

    # ── SAVE USER PROFILE (username / name) ──────────────────
    if not is_admin:
        _uo = get_user(data, chat_id)
        _ch = False
        if username and _uo.get("username") != username.lower():
            _uo["username"] = username.lower(); _ch = True
        if first_name and _uo.get("first_name") != first_name:
            _uo["first_name"] = first_name; _ch = True
        if _ch:
            save_data(data)

    # ── BAN CHECK ────────────────────────────────────────────
    if not is_admin:
        if data.get("users", {}).get(str(chat_id), {}).get("banned"):
            send_message(chat_id, "🚫 *You have been banned from this bot.*\n\n_Contact admin if you think this is a mistake._")
            return

    # ── FORCE JOIN GATE ───────────────────────────────────────
    if not is_admin:
        unjoined = get_unjoined_channels_cached(chat_id, data)
        if unjoined:
            send_force_join_message(chat_id, unjoined)
            return

    # ── DISABLED BUTTON CHECK (runs before state handler) ────
    if not is_admin and text in set(CONTROLLABLE_BUTTONS) and not is_button_enabled(data, text):
        send_message(chat_id,
            "🚫 *This option is currently disabled.*\n━━━━━━━━\n\n"
            "_Please contact admin for assistance._",
            {"inline_keyboard": [[{
                "text":                "@ITZ_EKBALL",
                "url":                 "https://t.me/ITZ_EKBALL",
                "style":               "danger",
                "icon_custom_emoji_id": "5314504236132747481",
            }]]})
        return

    if chat_id in user_states:
        state  = user_states[chat_id]
        action = state.get("action")

        if text == "/cancel":
            user_states.pop(chat_id, None)
            send_message(chat_id, "❌ *Cancelled.*", get_main_keyboard(is_admin))
            return

        if action == "broadcast":
            if not is_admin:
                user_states.pop(chat_id, None)
                return
            if not message_id:
                send_message(chat_id, "⚠️ Could not read message. Please send again:")
                return
            d = load_data()
            users = d.get("users", {})
            uid_str = str(chat_id)
            total = len(users)
            status_id = send_message(chat_id, f"⏳ *Broadcasting to {total} users...*")
            sent = 0
            failed = 0
            for uid in list(users.keys()):
                if uid == uid_str:
                    continue
                ok = copy_message(int(uid), chat_id, message_id)
                if ok:
                    sent += 1
                else:
                    failed += 1
                time.sleep(0.05)
            user_states.pop(chat_id, None)
            result_text = (
                f"✅ *Broadcast Complete!*\n━━━━━━━━\n\n"
                f"📤 Sent: *{sent}*\n"
                f"❌ Failed: *{failed}*\n"
                f"👥 Total users: *{total}*"
            )
            if status_id:
                edit_message(chat_id, status_id, result_text)
            else:
                send_message(chat_id, result_text)
            return

        if action == "dm_set_upi":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None); return
            new_upi = text.strip()
            if not new_upi:
                send_message(chat_id, "⚠️ Empty — send a valid UPI ID:")
                return
            d = load_data()
            d.setdefault("deposit_settings", {})["upi_id"] = new_upi
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id, f"✅ *UPI ID updated!*\n\nNew: `{new_upi}`", get_dm_inline(d))
            return

        if action == "dm_set_qr":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None); return
            new_url = text.strip()
            if not new_url:
                send_message(chat_id, "⚠️ Empty — send a valid URL:")
                return
            d = load_data()
            d.setdefault("deposit_settings", {})["qr_url"] = new_url
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id, f"✅ *QR URL updated!*\n\nNew: `{new_url}`", get_dm_inline(d))
            return

        if action == "dm_set_min":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None); return
            try:
                new_min = float(text.strip())
                if new_min <= 0:
                    raise ValueError
            except ValueError:
                send_message(chat_id, "⚠️ Invalid — send a positive number like `50`:")
                return
            d = load_data()
            d.setdefault("deposit_settings", {})["min_deposit"] = new_min
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id, f"✅ *Minimum deposit updated!*\n\nNew: ₹{new_min}", get_dm_inline(d))
            return

        if action == "dm_set_channel":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None); return
            new_ch = text.strip()
            if not new_ch:
                send_message(chat_id, "⚠️ Empty — send a channel username like `@mychannel`:")
                return
            d = load_data()
            d.setdefault("deposit_settings", {})["notify_channel"] = new_ch
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id,
                f"✅ *Notify Channel set!*\n\nChannel: `{new_ch}`\n\n"
                "_Make sure the bot is admin in that channel._",
                get_dm_inline(d))
            return

        # ── WITHDRAWAL ADMIN TEXT HANDLERS ────────────────────────────────────
        if action == "mw_set_price":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None); return
            try:
                new_price = float(text.strip())
                if new_price < 0: raise ValueError
            except ValueError:
                send_message(chat_id, "⚠️ Invalid — send a number like `1.5`:"); return
            svc_id   = state.get("svc_id")
            svc_name = state.get("svc_name", "")
            d = load_data()
            ws = d.setdefault("withdrawal_settings", {})
            se = ws.setdefault("service_earn", {}).setdefault(svc_id, {"enabled": False, "price": 1.0})
            se["price"] = new_price
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id, f"✅ *Price Updated!*\n\n{svc_name}: ₹{new_price} per OTP",
                         get_mw_svc_inline(d, svc_id, svc_name))
            return

        if action == "mw_set_min":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None); return
            try:
                new_min = float(text.strip())
                if new_min <= 0: raise ValueError
            except ValueError:
                send_message(chat_id, "⚠️ Invalid — send a positive number like `50`:"); return
            d = load_data()
            d.setdefault("withdrawal_settings", {})["min_amount"] = new_min
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id, f"✅ *Minimum Withdrawal updated!*\n\nNew: ₹{new_min}",
                         get_manage_withdrawal_inline(d))
            return

        # ── WITHDRAWAL USER TEXT HANDLERS ──────────────────────────────────
        if action == "user_set_payment":
            method  = state.get("method")
            detail  = text.strip()
            if not detail:
                send_message(chat_id, "⚠️ Empty — send your payment detail:"); return
            d = load_data()
            usr = d.setdefault("users", {}).setdefault(str(chat_id), {"balance": 0, "history": []})
            usr["payment_method"] = {"type": method, "detail": detail}
            save_data(d)
            user_states.pop(chat_id, None)
            labels = {"upi": "UPI", "bank": "Bank", "binance": "Binance"}
            usr2 = d.get("users", {}).get(str(chat_id), {})
            bal  = usr2.get("earn_balance", 0.0)
            tot  = usr2.get("otp_earned_total", 0)
            tod  = usr2.get("otp_earned_today", 0)
            # Send simple confirmation (no mixed entities) then show withdrawal panel
            send_message(chat_id,
                f"✅ *{labels.get(method,'Payment')} Saved!*\n\nDetail: `{detail}`")
            txt_wd = (
                f"◆ Total OTP: *{tot}* ◇\n"
                f"◆ Today OTP: *{tod}* ◇\n"
                f"◈ Available Balance: *₹{bal}* ◉"
            )
            send_message(chat_id, txt_wd,
                get_user_withdrawal_inline(), emoji_overrides=_WD_OVR)
            return

        if action == "deposit":
            step = state.get("step")
            _dep_min = load_data().get("deposit_settings", {}).get("min_deposit", MIN_DEPOSIT)
            if step == "amount":
                try:
                    amount = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid amount.* Send a number like `50`:")
                    return
                if amount < _dep_min:
                    send_message(chat_id, f"⚠️ *Too low!* Minimum is ₹{_dep_min}:")
                    return
                state["amount"] = amount
                state["step"]   = "utr"
                send_message(chat_id,
                    f"✅ Amount: *₹{amount}*\n\n"
                    "🔢 *Send your UTR / Transaction ID*\n_(12-digit number)_",
                    emoji_overrides={"✅": "6307741669877880674"})
                return
            if step == "utr":
                utr = text.strip()
                if len(utr) < 12:
                    send_message(chat_id, "⚠️ *Invalid UTR* — must be at least 12 digits:")
                    return
                state["utr"]  = utr
                state["step"] = "screenshot"
                send_message(chat_id,
                    f"✅ UTR: `{utr}`\n\n📸 *Send Payment Screenshot*\n\n_(⚠️ Fake = ban 🚫)_",
                    emoji_overrides={"✅": "6307741669877880674", "🚫": "5240241223632954241"})
                return
            if step == "screenshot":
                if not photo:
                    send_message(chat_id, "⚠️ *Please send an image* — not text.")
                    return
                amount  = state.get("amount")
                utr     = state.get("utr")
                file_id = photo[-1]["file_id"]
                caption = (
                    f"🔔 *New Deposit Request*\n━━━━━━━━\n\n"
                    f"👤 User: `{chat_id}`\n"
                    f"💰 Amount: ₹{amount}\n"
                    f"🧾 UTR: `{utr}`"
                )
                send_photo_file(ADMIN_ID, file_id, caption, get_admin_depo_inline(chat_id, amount, utr))
                user_states.pop(chat_id, None)
                send_message(chat_id,
                    "✅ *Request Submitted!*\n━━━━━━━━\n\n"
                    "⏳ Waiting for admin approval...\n_You'll be notified once approved._",
                    emoji_overrides={"✅": "6307741669877880674", "⏳": "5350427505805238170"})
                return

        if action == "ctry_search":
            user_states.pop(chat_id, None)
            service_id    = state.get("service_id", "")
            query         = text.strip().lower()
            safe_query    = re.sub(r'[*_`]', '', text.strip())
            countries_all = data.get("countries", []) + data.get("s3_countries", []) + data.get("s4_countries", [])
            # Resolve canonical service name
            target_name = None
            for c in countries_all:
                for s in c.get("services", []):
                    if s["id"] == service_id:
                        target_name = _strip_emoji(s["name"]).strip().lower()
                        break
                if target_name:
                    break
            # Collect matching country entries for this service
            matched_entries = []
            seen_s1_countries = set()   # one S1 group-button per country
            seen_s4 = set()
            for c in countries_all:
                cname_lower = _strip_emoji(c.get("name", "")).lower()
                if not (cname_lower.startswith(query) or query in cname_lower):
                    continue
                for s in c.get("services", []):
                    if not (s["id"] == service_id or (target_name and _strip_emoji(s.get("name","")).strip().lower() == target_name)):
                        continue
                    api      = s.get("api", "s1")
                    operator = s.get("operator", "")
                    if api == "s4":
                        key = (c["id"], api, operator)
                        if key in seen_s4:
                            continue
                        seen_s4.add(key)
                    cfg       = APIS.get(api, APIS["s1"])
                    price     = s.get("price", 0)
                    cname     = _strip_emoji(c["name"])
                    srv_short = cfg.get("short", api.upper())
                    srv_icon  = cfg.get("icon", "")
                    ceid      = _country_emoji_id(c)
                    seid      = cfg.get("emoji_id")
                    has_icon  = bool(ceid or seid)
                    srv_prefix = "" if has_icon else srv_icon
                    if api == "s1":
                        if c["id"] in seen_s1_countries:
                            continue   # already added one S1 group button for this country
                        # Check if this country has multiple S1 prices
                        all_s1 = [sv for sv in c.get("services", [])
                                  if sv.get("api","s1") == "s1" and
                                  (sv["id"] == service_id or (target_name and _strip_emoji(sv.get("name","")).strip().lower() == target_name))]
                        if len(all_s1) > 1:
                            min_p = min(sv.get("price",0) for sv in all_s1)
                            label = f"{cname}  ₹{min_p}  {srv_prefix}{srv_short}"
                            cb    = f"ctry_multi_{service_id}_{c['id']}"
                        else:
                            label = f"{cname}  ₹{price}  {srv_prefix}{srv_short}"
                            cb    = f"bs_{c['id']}_{s['id']}_s1"
                        btn = {"text": label, "callback_data": cb, "style": c.get("style","primary")}
                        if ceid:   btn["icon_custom_emoji_id"] = ceid
                        elif seid: btn["icon_custom_emoji_id"] = seid
                        matched_entries.append([btn])
                        seen_s1_countries.add(c["id"])
                    else:
                        op_part = f"_{operator}" if api == "s4" and operator else ""
                        label   = f"{cname}  ₹{price}  {srv_prefix}{srv_short}"
                        btn     = {"text": label, "callback_data": f"bs_{c['id']}_{s['id']}_{api}{op_part}", "style": c.get("style","primary")}
                        if ceid:   btn["icon_custom_emoji_id"] = ceid
                        elif seid: btn["icon_custom_emoji_id"] = seid
                        matched_entries.append([btn])
            if not matched_entries:
                send_message(chat_id,
                    f"🔍 *Country Search*\n━━━━━━━━\n\n❌ No country found matching: {safe_query}",
                    {"inline_keyboard": [[
                        {"text": "🔍 Search Again", "callback_data": f"ctry_search_{service_id}", "style": "primary"},
                        {"text": "◀ Back",          "callback_data": f"svc_{service_id}",         "style": "danger", "icon_custom_emoji_id": "5253456541650025535"},
                    ]]})
                return
            # Telegram allows max ~100 inline buttons — cap at 50 results to be safe
            _total_ctry_results = len(matched_entries)
            if _total_ctry_results > 50:
                matched_entries = matched_entries[:50]
                matched_entries.append([{"text": f"⚠️ Showing 50 of {_total_ctry_results} — search more specifically", "callback_data": "noop"}])
            matched_entries.append([{"text": "◀ Back", "callback_data": f"svc_{service_id}", "style": "danger", "icon_custom_emoji_id": "5253456541650025535"}])
            send_message(chat_id,
                f"🔍 *Country Search*\n━━━━━━━━\n\nFound *{min(_total_ctry_results, 50)}* result(s):",
                {"inline_keyboard": matched_entries})
            return

        if action == "buy_search":
            user_states.pop(chat_id, None)
            query = text.strip().lower()
            # Sanitize query for safe markdown embedding (remove * _ ` chars)
            safe_query = re.sub(r'[*_`]', '', text.strip())
            countries_all = data.get("countries", []) + data.get("s3_countries", []) + data.get("s4_countries", [])
            matched = {}
            try:
                index = _get_search_index(data)
                for entry in index:
                    nl  = entry["name_lower"]
                    sid = entry["svc_id"].lower()
                    if nl.startswith(query) or query in nl or sid.startswith(query):
                        matched[entry["svc_id"]] = {
                            "name":     entry["name"],
                            "emoji_id": entry["emoji_id"],
                            "style":    entry["style"],
                        }
            except Exception as _se:
                print(f"Search error: {_se}")
            if not matched:
                send_message(chat_id,
                    f"🔍 *Search*\n━━━━━━━━\n\n"
                    f"❌ No service found matching: {safe_query}",
                    {"inline_keyboard": [[{"text": "🔍 Search Again", "callback_data": "buy_search", "style": "primary", "icon_custom_emoji_id": "5429571366384842791"}, {"text": "All Services", "callback_data": "buy_all", "style": "success", "icon_custom_emoji_id": "5431499171045581032"}]]})
                return
            if len(matched) == 1:
                svc_id   = list(matched.keys())[0]
                svc_name = matched[svc_id]["name"]
                try:
                    markup = get_countries_for_service_inline(svc_id, countries_all, page=0)
                except Exception as _me:
                    print(f"Service inline error: {_me}")
                    send_message(chat_id, f"❌ Error loading service. Try again.")
                    return
                if not markup["inline_keyboard"]:
                    send_message(chat_id, f"❌ No countries available for this service.")
                    return
                send_message(chat_id, f"🌍 *Select Country*\n━━━━━━━━\n\n*{_strip_emoji(svc_name).strip()}*", markup)
                return
            buttons = []
            row = []
            for svc_id, info in matched.items():
                eid   = info.get("emoji_id", "")
                style = info.get("style", "primary")
                label = _strip_emoji(info["name"]) if eid else info["name"]
                btn   = {"text": label, "callback_data": f"svc_{svc_id}", "style": style}
                if eid:
                    btn["icon_custom_emoji_id"] = eid
                row.append(btn)
                if len(row) == 2:
                    buttons.append(row)
                    row = []
            if row:
                buttons.append(row)
            send_message(chat_id,
                f"🔍 *Search: _{text}_*\n━━━━━━━━\n\nFound *{len(matched)}* services:",
                {"inline_keyboard": buttons})
            return

        if action == "claim_gift":
            code = text.strip().upper()
            gift_codes = data.get("gift_codes", {})
            if code not in gift_codes:
                send_message(chat_id,
                    "❌ *Invalid Gift Code*\n━━━━━━━━\n\n"
                    "_This code does not exist. Check and try again._\n\n_/cancel to abort_")
                return
            gc       = gift_codes[code]
            used_by  = gc.get("used_by") if isinstance(gc.get("used_by"), list) else ([] if gc.get("used_by") is None else [gc["used_by"]])
            max_uses = gc.get("max_uses", 1)
            uid_str  = str(chat_id)
            if uid_str in used_by:
                send_message(chat_id,
                    "❌ *Already Claimed*\n━━━━━━━━\n\n"
                    "_You have already claimed this gift code._")
                user_states.pop(chat_id, None)
                return
            if len(used_by) >= max_uses:
                send_message(chat_id,
                    "❌ *Code Exhausted*\n━━━━━━━━\n\n"
                    "_This gift code has already been fully claimed._")
                user_states.pop(chat_id, None)
                return
            amount = gc.get("amount", 0)
            user   = get_user(data, chat_id)
            user["balance"] = round(user["balance"] + amount, 2)
            ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
            used_by.append(uid_str)
            gc["used_by"] = used_by
            gc["last_used_at"] = ts
            save_data(data)
            user_states.pop(chat_id, None)
            remaining = max_uses - len(used_by)
            send_message(chat_id,
                f"✅ *Gift Code Claimed!*\n━━━━━━━━\n\n"
                f"🎁 Code: `{code}`\n"
                f"💰 ₹{amount} added to your wallet!\n"
                f"💳 New Balance: ₹{user['balance']}\n\n━━━━━━━━")
            return

        if action == "create_gift":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            step = state.get("step")
            if step == "amount":
                try:
                    amount = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid amount.* Send a number like `50`:")
                    return
                state["amount"] = amount
                state["step"]   = "max_uses"
                send_message(chat_id,
                    f"✅ Amount: *₹{amount}*\n\n"
                    "Step 2: Enter the *number of users* who can claim this code:\n\n"
                    "Example: `1` (single use) or `100` (100 users)\n\n_/cancel to abort_")
                return
            if step == "max_uses":
                try:
                    max_uses = int(text.strip())
                    if max_uses < 1:
                        raise ValueError
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid number.* Send a positive integer like `1` or `50`:")
                    return
                import random, string
                amount       = state.get("amount", 0)
                total_cost   = round(amount * max_uses, 2)
                admin_user   = get_user(data, chat_id)
                if admin_user["balance"] < total_cost:
                    user_states.pop(chat_id, None)
                    send_message(chat_id,
                        f"❌ *Insufficient Balance*\n━━━━━━━━\n\n"
                        f"💰 Required:  ₹{total_cost}  ({max_uses} × ₹{amount})\n"
                        f"💳 Your balance:  ₹{admin_user['balance']}\n\n"
                        f"_Add funds first, then try again._")
                    return
                while True:
                    part1 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
                    part2 = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
                    code  = f"GIFT-{part1}-{part2}"
                    if code not in data.get("gift_codes", {}):
                        break
                admin_user["balance"] = round(admin_user["balance"] - total_cost, 2)
                ts         = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
                gift_codes = data.setdefault("gift_codes", {})
                gift_codes[code] = {"amount": amount, "max_uses": max_uses, "used_by": [], "created_at": ts}
                save_data(data)
                user_states.pop(chat_id, None)
                send_message(chat_id,
                    f"✅ *Gift Code Created!*\n━━━━━━━━\n\n"
                    f"🎁 Code: `{code}`\n"
                    f"💰 Value per user: ₹{amount}\n"
                    f"👥 Max Users: {max_uses}\n"
                    f"💸 Total deducted: ₹{total_cost}\n"
                    f"💳 Remaining balance: ₹{admin_user['balance']}\n\n"
                    f"_Share this code with users to give them balance!_",
                    emoji_overrides={"✅": "6298670698948724690", "🎁": "6307351429149365689", "💰": "6156906412761946453", "👥": "5249050854392091366", "💸": "5269531045165816230", "💳": "6305298855688672996"})
                return

        if action == "fn_custom_range":
            srv_id   = state.get("srv_id")
            cnt_id   = state.get("cnt_id")
            prompt_id = state.get("msg_id")
            user_states.pop(chat_id, None)
            if message_id:
                delete_message(chat_id, message_id)
            rng_input = text.strip().lstrip("+").upper()
            if rng_input.endswith("XXX"):
                rng_input = rng_input[:-3]
            rng_input = rng_input.lower()
            if not rng_input or not rng_input.isdigit():
                if prompt_id:
                    edit_message(chat_id, prompt_id,
                        "⚠️ *Invalid range.* Digits only (e.g. `22501`).\n\n_/cancel to abort_")
                else:
                    send_message(chat_id,
                        "⚠️ *Invalid range.* Please enter digits only (e.g. `9930`).\n\n_/cancel to abort_")
                user_states[chat_id] = {"action": "fn_custom_range", "srv_id": srv_id, "cnt_id": cnt_id, "msg_id": prompt_id}
                return
            srv  = data.get("services_data", {}).get(srv_id)
            cnt  = srv.get("countries", {}).get(cnt_id) if srv else None
            if not srv or not cnt:
                send_message(chat_id, "❌ Service not found. Please start over.")
                return
            if not get_tnevs_settings(data).get("api_key"):
                send_message(chat_id, "❌ *2oo9 API Not Configured.*\n\n_Contact admin._")
                return
            flag = get_country_flag(cnt["name"])
            if prompt_id:
                edit_message(chat_id, prompt_id,
                    f"⏳ *Fetching number...*\nRange: `{rng_input}XXX`",
                    emoji_overrides={"⏳": "5789728885532790879"})
            nums_bought = []
            for _ in range(3):
                r = tnevs_get_number(data, rng_input)
                if "error" not in r:
                    nd  = r.get("number", {})
                    fn  = nd.get("full_number") or nd.get("no_plus_number", "")
                    if fn:
                        disp = fn if fn.startswith("+") else f"+{fn}"
                        nums_bought.append((fn, disp, nd.get("country", cnt["name"])))
            if not nums_bought:
                err_markup = {"inline_keyboard": [
                    [{"text": "Try Again", "callback_data": f"fn_custom_range|{srv_id}|{cnt_id}", "style": "success", "icon_custom_emoji_id": "5337132498965010628"}],
                    [{"text": "◀️ Back",   "callback_data": f"fn_ctr|{srv_id}|{cnt_id}",         "style": "danger"}],
                ]}
                if prompt_id:
                    edit_message(chat_id, prompt_id,
                        f"❌ *No number found for range* `{rng_input}`\n\n_Try a different range prefix._",
                        err_markup)
                else:
                    send_message(chat_id,
                        f"❌ *No number found for range* `{rng_input}`\n\n_Try a different range prefix._",
                        err_markup)
                return
            inline = []
            for fn, disp, country in nums_bought:
                svc_flag = ""
                for kw, (ch, _) in CONSOLE_SERVICE_EMOJIS.items():
                    if kw in (srv.get("name") or "").lower():
                        svc_flag = ch
                        break
                label = f"{svc_flag} {disp}" if svc_flag else disp
                inline.append([{"text": label, "copy_text": {"text": disp}, "style": "success"}])
                pending_free_nums[fn] = {
                    "chat_id":    chat_id,
                    "msg_id":     0,
                    "srv_name":   srv.get("name", ""),
                    "country":    country,
                    "started_at": time.time(),
                    "known_msgs": set(),
                }
            inline.append([{"text": "Refresh", "callback_data": f"fn_rebuy_range|{srv_id}|{cnt_id}|{rng_input}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}])
            inline.append([{"text": "Back", "callback_data": f"fn_ctr|{srv_id}|{cnt_id}", "style": "danger", "icon_custom_emoji_id": "5370726448959085259"}])
            result_text = f"✅ *Order Successful*\n{flag} Range: *{cnt['name']}*"
            if prompt_id:
                edit_message(chat_id, prompt_id, result_text, {"inline_keyboard": inline})
            else:
                send_message(chat_id, result_text, {"inline_keyboard": inline})
            return

        if action == "fn_direct_custom_range":
            prompt_id = state.get("msg_id")
            user_states.pop(chat_id, None)
            if message_id:
                delete_message(chat_id, message_id)
            rng_input = text.strip().lstrip("+").upper()
            if rng_input.endswith("XXX"):
                rng_input = rng_input[:-3]
            rng_input = rng_input.lower()
            if not rng_input or not rng_input.isdigit():
                if prompt_id:
                    edit_message(chat_id, prompt_id,
                        "⚠️ *Invalid range.* Digits only (e.g. `22501`).\n\n_/cancel to abort_")
                else:
                    send_message(chat_id,
                        "⚠️ *Invalid range.* Please enter digits only (e.g. `9930`).\n\n_/cancel to abort_")
                user_states[chat_id] = {"action": "fn_direct_custom_range", "msg_id": prompt_id}
                return
            if not get_tnevs_settings(data).get("api_key"):
                send_message(chat_id, "❌ *2oo9 API Not Configured.*\n\n_Contact admin._")
                return
            if prompt_id:
                edit_message(chat_id, prompt_id,
                    f"⏳ *Fetching number...*\nRange: `{rng_input}XXX`",
                    emoji_overrides={"⏳": "5789728885532790879"})
            nums_bought = []
            for _ in range(3):
                r = tnevs_get_number(data, rng_input)
                if "error" not in r:
                    nd = r.get("number", {})
                    fn = nd.get("full_number") or nd.get("no_plus_number", "")
                    if fn:
                        disp = fn if fn.startswith("+") else f"+{fn}"
                        nums_bought.append((fn, disp))
            err_markup = {"inline_keyboard": [
                [{"text": "Try Again", "callback_data": "fn_direct_custom_range", "style": "success", "icon_custom_emoji_id": "5337132498965010628"}],
                [{"text": "◀️ Back",   "callback_data": "free_number_refresh",    "style": "danger"}],
            ]}
            if not nums_bought:
                if prompt_id:
                    edit_message(chat_id, prompt_id,
                        f"❌ *No number found for range* `{rng_input}`\n\n_Try a different range prefix._",
                        err_markup)
                else:
                    send_message(chat_id,
                        f"❌ *No number found for range* `{rng_input}`\n\n_Try a different range prefix._",
                        err_markup)
                return
            inline = []
            for fn, disp in nums_bought:
                inline.append([{"text": disp, "copy_text": {"text": disp}, "style": "success"}])
                pending_free_nums[fn] = {
                    "chat_id":    chat_id,
                    "msg_id":     0,
                    "srv_name":   "",
                    "country":    "",
                    "started_at": time.time(),
                    "known_msgs": set(),
                }
            inline.append([{"text": "Refresh", "callback_data": f"fn_rebuy_direct|{rng_input}", "style": "primary", "icon_custom_emoji_id": "5260491539167073671"}])
            inline.append([{"text": "Back", "callback_data": "free_number_refresh", "style": "danger", "icon_custom_emoji_id": "5370726448959085259"}])
            _d_flag, _d_country = detect_country_from_phone(nums_bought[0][0]) if nums_bought else ("📱", "")
            _d_loc = f"{_d_flag} Range: *{_d_country}*" if _d_country else f"📱 Range: `{rng_input}XXX`"
            result_text = f"✅ *Order Successful*\n{_d_loc}"
            if prompt_id:
                edit_message(chat_id, prompt_id, result_text, {"inline_keyboard": inline})
            else:
                send_message(chat_id, result_text, {"inline_keyboard": inline})
            return

        if action == "fj_add_ch":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            ch_input = text.strip()
            if not ch_input:
                send_message(chat_id, "⚠️ *Cannot send empty.* Send the Channel ID or @username:\n\n_/cancel to abort_")
                return
            send_message(chat_id, "⏳ *Verifying channel...*")
            ch_info = tg_get_chat(ch_input)
            if not ch_info or ch_info.get("type") not in ("channel", "supergroup", "group"):
                send_message(chat_id,
                    "❌ *Channel not found!*\n\n"
                    "• Is the bot an *Admin* of the channel?\n"
                    "• Is the ID/username correct?\n\n_Try again or /cancel_")
                return
            ch_title    = ch_info.get("title", ch_input)
            ch_id_final = str(ch_info.get("id", ch_input))
            invite_link = tg_create_invite_link(ch_id_final)
            uid = str(uuid.uuid4())[:8]
            d   = load_data()
            d.setdefault("force_channels", []).append({
                "uid":         uid,
                "id":          ch_id_final,
                "name":        ch_title,
                "invite_link": invite_link,
            })
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id,
                f"✅ *Channel added!*\n━━━━━━━━\n\n"
                f"📢 *{ch_title}*\n"
                f"🆔 ID: `{ch_id_final}`\n"
                f"🔗 Link: {invite_link or '_not generated_'}\n\n"
                f"_Users must now join this channel to use the bot._",
                {"inline_keyboard": [[_btn("📢 Force Join Manage", callback_data="fj_admin")]]}
            )
            return

        # ── ADMIN: USER HANDLE STATES ────────────────────────────
        if action == "uh_search":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            user_states.pop(chat_id, None)
            uid, u = find_user_by_identifier(data, text)
            if not u:
                send_message(chat_id, f"❌ *User not found!*\n\n`{text.strip()}` — no user with this ID/username.")
                return
            send_message(chat_id, build_user_profile_text(uid, u), build_user_action_inline(uid, u))
            return

        if action == "uh_ban_user":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            user_states.pop(chat_id, None)
            uid, u = find_user_by_identifier(data, text)
            if not u:
                send_message(chat_id, "❌ *User not found!*")
                return
            u["banned"] = True
            save_data(data)
            uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
            send_message(chat_id,
                f"🚫 *Banned!*\n━━━━━━━━\n\n{uname} has been banned.",
                {"inline_keyboard": [
                    [_btn("🔄 View Profile",  callback_data=f"uh_view_{uid}")],
                    [_btn("◀️ User Handle",   callback_data="uh_main")],
                ]})
            try:
                send_message(int(uid), "🚫 *You have been banned from this bot.*\n\n_Contact admin if you think this is a mistake._")
            except Exception:
                pass
            return

        if action == "uh_unban_user":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            user_states.pop(chat_id, None)
            uid, u = find_user_by_identifier(data, text)
            if not u:
                send_message(chat_id, "❌ *User not found!*")
                return
            u["banned"] = False
            save_data(data)
            uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
            send_message(chat_id,
                f"✅ *Unbanned!*\n━━━━━━━━\n\n{uname} has been unbanned.",
                {"inline_keyboard": [
                    [_btn("🔄 View Profile",  callback_data=f"uh_view_{uid}")],
                    [_btn("◀️ User Handle",   callback_data="uh_main")],
                ]})
            try:
                send_message(int(uid), "✅ *Your ban has been lifted. You can use the bot again.*")
            except Exception:
                pass
            return

        if action == "uh_bal_add":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            step = state.get("step")
            if step == "user":
                uid, u = find_user_by_identifier(data, text)
                if not u:
                    send_message(chat_id, "❌ *User not found!* Try again:\n\n_/cancel to abort_")
                    return
                state["target_uid"] = uid
                state["step"]       = "amount"
                uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
                send_message(chat_id,
                    f"➕ *Balance Add*\n━━━━━━━━\n\n"
                    f"👤 User: {uname}\n"
                    f"💳 Current: ₹{u.get('balance', 0)}\n\n"
                    f"How much to add?\n\n_/cancel to abort_")
                return
            if step == "amount":
                try:
                    amount = float(text.strip())
                    if amount <= 0:
                        raise ValueError
                except Exception:
                    send_message(chat_id, "⚠️ Enter a valid amount (e.g. `50`):")
                    return
                uid = state["target_uid"]
                user_states.pop(chat_id, None)
                d = load_data()
                u = get_user(d, uid)
                u["balance"] = round(u["balance"] + amount, 2)
                ts = datetime.datetime.now(IST).strftime("%Y-%m-%d %H:%M")
                u.setdefault("deposits", []).append({"amount": amount, "utr": "admin_add", "status": "approved", "timestamp": ts})
                save_data(d)
                uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
                send_message(chat_id,
                    f"✅ *Balance Added!*\n━━━━━━━━\n\n"
                    f"👤 User: {uname}\n"
                    f"➕ Added: ₹{amount}\n"
                    f"💳 New Balance: ₹{u['balance']}",
                    {"inline_keyboard": [
                        [_btn("🔄 View Profile", callback_data=f"uh_view_{uid}")],
                        [_btn("👥 User Handle",  callback_data="uh_main")],
                    ]})
                try:
                    send_message(int(uid),
                        f"💰 *Balance Added!*\n━━━━━━━━\n\n"
                        f"➕ ₹{amount} has been added to your wallet\n"
                        f"💳 New Balance: ₹{u['balance']}")
                except Exception:
                    pass
                return

        if action == "uh_bal_min":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            step = state.get("step")
            if step == "user":
                uid, u = find_user_by_identifier(data, text)
                if not u:
                    send_message(chat_id, "❌ *User not found!* Try again:\n\n_/cancel to abort_")
                    return
                state["target_uid"] = uid
                state["step"]       = "amount"
                uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
                send_message(chat_id,
                    f"➖ *Balance Minus*\n━━━━━━━━\n\n"
                    f"👤 User: {uname}\n"
                    f"💳 Current: ₹{u.get('balance', 0)}\n\n"
                    f"How much to deduct?\n\n_/cancel to abort_")
                return
            if step == "amount":
                try:
                    amount = float(text.strip())
                    if amount <= 0:
                        raise ValueError
                except Exception:
                    send_message(chat_id, "⚠️ Enter a valid amount (e.g. `50`):")
                    return
                uid = state["target_uid"]
                user_states.pop(chat_id, None)
                d = load_data()
                u = get_user(d, uid)
                u["balance"] = round(max(0, u["balance"] - amount), 2)
                save_data(d)
                uname = f"@{u.get('username','')}" if u.get("username") else f"`{uid}`"
                send_message(chat_id,
                    f"✅ *Balance Deducted!*\n━━━━━━━━\n\n"
                    f"👤 User: {uname}\n"
                    f"➖ Deducted: ₹{amount}\n"
                    f"💳 New Balance: ₹{u['balance']}",
                    {"inline_keyboard": [
                        [_btn("🔄 View Profile", callback_data=f"uh_view_{uid}")],
                        [_btn("👥 User Handle",  callback_data="uh_main")],
                    ]})
                try:
                    send_message(int(uid),
                        f"💳 *Balance Updated*\n━━━━━━━━\n\n"
                        f"➖ ₹{amount} has been deducted from your wallet\n"
                        f"💳 New Balance: ₹{u['balance']}")
                except Exception:
                    pass
                return

        if action == "add_country":
            api = state.get("api", "s1")
            if state.get("step") == "id":
                state["country_id"] = text.strip()
                state["step"]       = "name"
                send_message(chat_id, f"✅ ID: `{text.strip()}`\n\nNow send the country name:\nExample: `🇨🇦 Canada`")
                return
            elif state.get("step") == "name":
                cname_resolved = _expand_country_name(text.strip())
                state["country_name"] = cname_resolved
                state["step"]         = "emoji"
                auto_kb = get_auto_flag_keyboard(cname_resolved)
                if auto_kb:
                    send_message(chat_id,
                        f"✅ Name: *{cname_resolved}*\n\n"
                        f"🏳️ *Auto-detected flag:*\n"
                        f"Tap to use it, or pick a different one.",
                        auto_kb)
                else:
                    send_message(chat_id,
                        f"✅ Name: *{cname_resolved}*\n\n"
                        f"✨ *Pick a flag emoji* for this country\n"
                        f"_(tap one below, or skip)_",
                        get_country_emoji_picker_inline())
                return

        # ── S3/S4 ADD OPERATOR flow ───────────────────────────────
        if action == "s3_add_operator":
            api = state.get("api", "s3")
            cid = state.get("country_id")
            cfg = APIS.get(api, APIS["s1"])
            if state.get("step") == "name":
                state["svc_name"] = text.strip()
                if api == "s3":
                    # S3: each operator has a unique service ID — ask for it next
                    state["step"] = "svc_id"
                    send_message(chat_id,
                        f"✅ Name: *{text.strip()}*\n\n"
                        f"2️⃣ *Send the Service ID for this operator:*\n"
                        f"Example: `amazon`, `amz2`\n\n_/cancel to abort_")
                else:
                    # S4: all operators share the same service ID — go to price
                    state["step"] = "price"
                    send_message(chat_id,
                        f"✅ Name: *{text.strip()}*\n\n"
                        f"2️⃣ *Send the price (₹):*\nExample: `20`\n\n_/cancel to abort_")
                return
            elif state.get("step") == "svc_id":
                # S3 only — capture the unique service ID for this operator
                state["svc_id"] = text.strip()
                state["step"]   = "price"
                send_message(chat_id,
                    f"✅ Service ID: `{text.strip()}`\n\n"
                    f"3️⃣ *Send the price (₹):*\nExample: `20`\n\n_/cancel to abort_")
                return
            elif state.get("step") == "price":
                try:
                    price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ Invalid price. Send a number like `20`:")
                    return
                state["price"] = price
                state["step"]  = "operator"
                send_message(chat_id,
                    f"✅ Price: ₹{price}\n\n"
                    f"4️⃣ *Send the Operator ID:*\nExample: `1`, `2`, `any`\n\n_/cancel to abort_")
                return
            elif state.get("step") == "operator":
                state["operator"] = text.strip()
                state["step"]     = "max_price"
                send_message(chat_id,
                    f"✅ Operator: `{text.strip()}`\n\n"
                    f"🔸 *Send maxPrice ({cfg['name']} highest accepted price):*\n"
                    f"Example: `0.50`\n\n"
                    f"_Numbers above this price will NOT be assigned._\n"
                    f"_Send `0` to skip (no upper limit)._\n"
                    f"_/cancel to abort_")
                return
            elif state.get("step") == "max_price":
                try:
                    max_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ Invalid value. Send a decimal number like `0.50`:")
                    return
                svc_name  = state.get("svc_name", "")
                price     = state.get("price", 0)
                operator  = state.get("operator", "")
                # Use pre-filled svc_id if available (from group-view add or svc_id step), else derive from name
                svc_id    = state.get("svc_id") or (re.sub(r'[^a-z0-9]', '', svc_name.lower())[:10] or "svc")
                fresh_data    = load_data()
                api_countries = get_api_countries(fresh_data, api)
                country       = next((c for c in api_countries if c["id"] == cid), None)
                if not country:
                    send_message(chat_id, "❌ Country not found.")
                    user_states.pop(chat_id, None)
                    return
                entry = {
                    "id": svc_id, "name": svc_name, "price": price,
                    "api": api, "operator": operator, "style": "primary"
                }
                if max_price > 0:
                    entry["max_price"] = max_price
                country["services"].append(entry)
                set_api_countries(fresh_data, api, api_countries)
                save_data(fresh_data)
                user_states.pop(chat_id, None)
                op_line  = f"🔢 Operator: `{operator}`\n" if operator else ""
                mp_line  = f"🔸 maxPrice: `{max_price}`\n" if max_price > 0 else ""
                svcs     = [s for s in country.get("services", []) if s.get("api") == api]
                tgt_norm = svc_name.strip().lower()
                send_message(chat_id,
                    f"✅ *Operator Added!*\n\n"
                    f"`{svc_id}` — {svc_name} — ₹{price}\n"
                    f"{op_line}{mp_line}{cfg['icon']} {cfg['name']}\n\n"
                    f"Operators for *{svc_name}*:",
                    get_s3_svc_group_inline(api, cid, svcs, tgt_norm, svc_id))
                return

        # ── S3/S4 UPDATE PRICE flow ────────────────────────────────
        if action == "s3_update_price":
            api     = state.get("api", "s3")
            cid     = state.get("country_id")
            svc_idx = state.get("svc_idx", 0)
            svc_id  = state.get("svc_id", "")
            cfg     = APIS.get(api, APIS["s1"])
            if state.get("step") == "price":
                try:
                    new_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ Invalid price. Send a number like `25`:")
                    return
                fresh_data    = load_data()
                api_countries = get_api_countries(fresh_data, api)
                country       = next((c for c in api_countries if c["id"] == cid), None)
                if not country:
                    send_message(chat_id, "❌ Country not found.")
                    user_states.pop(chat_id, None)
                    return
                api_svc_count = 0
                old_price     = 0
                svc_name      = ""
                for s in country.get("services", []):
                    if s.get("api") == api:
                        if api_svc_count == svc_idx:
                            old_price  = s.get("price", 0)
                            svc_name   = s["name"]
                            s["price"] = new_price
                            break
                        api_svc_count += 1
                set_api_countries(fresh_data, api, api_countries)
                save_data(fresh_data)
                user_states.pop(chat_id, None)
                svcs     = [s for s in country.get("services", []) if s.get("api") == api]
                tgt_norm = svc_name.strip().lower()
                send_message(chat_id,
                    f"✅ *Price Updated!*\n\n"
                    f"Service: *{svc_name}*\n"
                    f"Old: ₹{old_price}  →  New: ₹{new_price}\n"
                    f"🌍 {country['name']}\n{cfg['icon']} {cfg['name']}\n\n"
                    f"Operators for *{svc_name}*:",
                    get_s3_svc_group_inline(api, cid, svcs, tgt_norm, svc_id))
                return

        if action == "add_service":
            api = state.get("api", "s1")
            cid = state.get("country_id")
            needs_operator = APIS.get(api, {}).get("operator", False)
            if state.get("step") == "id":
                state["service_id"] = text.strip()
                state["step"]       = "name"
                send_message(chat_id, f"✅ ID: `{text.strip()}`\n\nNow send the service name:\nExample: `📱 WhatsApp`")
                return
            elif state.get("step") == "name":
                state["service_name"] = text.strip()
                state["step"]         = "emoji"
                send_message(chat_id,
                    f"✅ Name: *{text.strip()}*\n\n"
                    f"✨ *Pick a premium emoji* for this service\n"
                    f"_(tap one below, or skip)_",
                    get_emoji_picker_inline())
                return
            elif state.get("step") == "price":
                try:
                    price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid price.* Send a number like `20`:")
                    return
                state["service_price"] = price
                if needs_operator:
                    state["step"] = "operator"
                    send_message(chat_id, f"✅ Price: ₹{price}\n\n🔢 *Send the Operator ID:*\nExample: `1`, `2`, `any`\n\n_/cancel to abort_")
                    return
                if api == "s1":
                    # SmsBower supports minPrice/maxPrice range
                    state["step"] = "min_price"
                    send_message(chat_id,
                        f"✅ Price: ₹{price}\n\n"
                        f"🔹 *Send minPrice (SMSBower lowest accepted price):*\n"
                        f"Example: `0.11`\n\n"
                        f"_This is the minimum cost range sent to SMSBower API._\n"
                        f"_/cancel to abort_")
                    return
                if api == "s2":
                    # HeroSMS supports maxPrice via getNumberV2
                    state["step"] = "s2_max_price"
                    send_message(chat_id,
                        f"✅ Price: ₹{price}\n\n"
                        f"🔸 *Send maxPrice (HeroSMS highest accepted price):*\n"
                        f"Example: `0.50`\n\n"
                        f"_Numbers above this cost will NOT be assigned._\n"
                        f"_Send `0` to skip (no price limit)._\n"
                        f"_/cancel to abort_")
                    return
                if api in ("s3", "s4"):
                    # S3/S4 support maxPrice via getNumber (no minPrice in API)
                    state["step"] = "s3_max_price"
                    sname = APIS[api]["name"]
                    send_message(chat_id,
                        f"✅ Price: ₹{price}\n\n"
                        f"🔸 *Send maxPrice ({sname} highest accepted price):*\n"
                        f"Example: `0.50`\n\n"
                        f"_Numbers above this price will NOT be assigned._\n"
                        f"_Send `0` to skip (no upper limit)._\n"
                        f"_/cancel to abort_")
                    return
                # other APIs — save directly
                svc_id        = state.get("service_id")
                raw_name      = state.get("service_name", "")
                emoji         = state.get("emoji", "")
                eid           = state.get("emoji_id", "")
                svc_name      = f"{emoji} {raw_name}".strip() if emoji else raw_name
                cfg           = APIS.get(api, APIS["s1"])
                api_countries = get_api_countries(data, api)
                for c in api_countries:
                    if c["id"] == cid:
                        c["services"].append({"id": svc_id, "name": svc_name, "price": price, "api": api, "emoji_id": eid, "style": state.get("style", "primary")})
                        break
                save_data(data)
                user_states.pop(chat_id, None)
                send_message(chat_id,
                    f"✅ *Service Added!*\n\n"
                    f"`{svc_id}` — {svc_name} — ₹{price}\n{cfg['icon']} {cfg['name']}",
                    get_srv_country_inline(api, cid))
                return
            elif state.get("step") == "s2_max_price":
                try:
                    max_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.50`:")
                    return
                svc_id   = state.get("service_id")
                raw_name = state.get("service_name", "")
                emoji    = state.get("emoji", "")
                eid      = state.get("emoji_id", "")
                svc_name = f"{emoji} {raw_name}".strip() if emoji else raw_name
                price    = state.get("service_price", 0)
                cfg      = APIS.get(api, APIS["s1"])
                api_countries = get_api_countries(data, api)
                for c in api_countries:
                    if c["id"] == cid:
                        entry = {
                            "id": svc_id, "name": svc_name,
                            "price": price, "api": api,
                            "emoji_id": eid, "style": state.get("style", "primary"),
                        }
                        if max_price > 0:
                            entry["max_price"] = max_price
                        c["services"].append(entry)
                        break
                save_data(data)
                user_states.pop(chat_id, None)
                price_line = f"🔸 maxPrice: `{max_price}`\n" if max_price > 0 else ""
                send_message(chat_id,
                    f"✅ *Service Added!*\n\n"
                    f"`{svc_id}` — {svc_name} — ₹{price}\n"
                    f"{price_line}{cfg['icon']} {cfg['name']}",
                    get_srv_country_inline(api, cid))
                return
            elif state.get("step") == "s3_max_price":
                try:
                    max_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.50`:")
                    return
                svc_id    = state.get("service_id")
                raw_name  = state.get("service_name", "")
                emoji     = state.get("emoji", "")
                eid       = state.get("emoji_id", "")
                svc_name  = f"{emoji} {raw_name}".strip() if emoji else raw_name
                price     = state.get("service_price", 0)
                operator  = state.get("operator")
                cfg       = APIS.get(api, APIS["s1"])
                api_countries = get_api_countries(data, api)
                for c in api_countries:
                    if c["id"] == cid:
                        entry = {
                            "id": svc_id, "name": svc_name,
                            "price": price, "api": api,
                            "emoji_id": eid, "style": state.get("style", "primary"),
                        }
                        if operator:
                            entry["operator"] = operator
                        if max_price > 0:
                            entry["max_price"] = max_price
                        c["services"].append(entry)
                        break
                save_data(data)
                user_states.pop(chat_id, None)
                op_line    = f"🔢 Operator: `{operator}`\n" if operator else ""
                price_line = f"🔸 maxPrice: `{max_price}`" if max_price > 0 else ""
                send_message(chat_id,
                    f"✅ *Service Added!*\n\n"
                    f"`{svc_id}` — {svc_name} — ₹{price}\n"
                    + op_line
                    + (f"{price_line}\n" if price_line else "")
                    + f"{cfg['icon']} {cfg['name']}",
                    get_srv_country_inline(api, cid))
                return
            elif state.get("step") == "min_price":
                try:
                    min_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.11`:")
                    return
                state["min_price"] = min_price
                state["step"]      = "max_price"
                send_message(chat_id,
                    f"✅ minPrice: `{min_price}`\n\n"
                    f"🔸 *Send maxPrice (SMSBower highest accepted price):*\n"
                    f"Example: `0.163`\n\n"
                    f"_Numbers above this price will NOT be assigned._\n"
                    f"_/cancel to abort_")
                return
            elif state.get("step") == "max_price":
                try:
                    max_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.163`:")
                    return
                svc_id   = state.get("service_id")
                raw_name = state.get("service_name", "")
                emoji    = state.get("emoji", "")
                eid      = state.get("emoji_id", "")
                svc_name = f"{emoji} {raw_name}".strip() if emoji else raw_name
                price    = state.get("service_price", 0)
                min_price = state.get("min_price", 0)
                cfg           = APIS.get(api, APIS["s1"])
                api_countries = get_api_countries(data, api)
                for c in api_countries:
                    if c["id"] == cid:
                        c["services"].append({
                            "id": svc_id, "name": svc_name,
                            "price": price, "api": api,
                            "emoji_id": eid, "style": state.get("style", "primary"),
                            "min_price": min_price, "max_price": max_price
                        })
                        break
                save_data(data)
                user_states.pop(chat_id, None)
                send_message(chat_id,
                    f"✅ *Service Added!*\n\n"
                    f"`{svc_id}` — {svc_name} — ₹{price}\n"
                    f"🔹 minPrice: `{min_price}`  |  🔸 maxPrice: `{max_price}`\n"
                    f"{cfg['icon']} {cfg['name']}",
                    get_srv_country_inline(api, cid))
                return
            elif state.get("step") == "operator" and needs_operator:
                operator = text.strip()
                state["operator"] = operator
                if api in ("s3", "s4"):
                    state["step"] = "s3_max_price"
                    sname = APIS[api]["name"]
                    send_message(chat_id,
                        f"✅ Operator: `{operator}`\n\n"
                        f"🔸 *Send maxPrice ({sname} highest accepted price):*\n"
                        f"Example: `0.50`\n\n"
                        f"_Numbers above this price will NOT be assigned._\n"
                        f"_Send `0` to skip (no upper limit)._\n"
                        f"_/cancel to abort_")
                    return
                svc_id        = state.get("service_id")
                raw_name      = state.get("service_name", "")
                emoji         = state.get("emoji", "")
                eid           = state.get("emoji_id", "")
                svc_name      = f"{emoji} {raw_name}".strip() if emoji else raw_name
                price         = state.get("service_price", 0)
                cfg           = APIS.get(api, APIS["s1"])
                api_countries = get_api_countries(data, api)
                for c in api_countries:
                    if c["id"] == cid:
                        c["services"].append({
                            "id": svc_id, "name": svc_name,
                            "price": price, "api": api, "operator": operator, "emoji_id": eid,
                            "style": state.get("style", "primary")
                        })
                        break
                save_data(data)
                user_states.pop(chat_id, None)
                send_message(chat_id,
                    f"✅ *Service Added!*\n\n"
                    f"`{svc_id}` — {svc_name} — ₹{price}\n"
                    f"🔢 Operator: `{operator}`\n{cfg['icon']} {cfg['name']}",
                    get_srv_country_inline(api, cid))
                return

        if action in ("add_service_flow", "add_country_to_service"):
            api            = state.get("api", "s1")
            phase          = state.get("phase", 2)
            needs_operator = APIS.get(api, {}).get("operator", False)

            # ── Phase 1: collect service info (only add_service_flow) ──
            if action == "add_service_flow" and phase == 1:
                if state.get("step") == "svc_id":
                    state["svc_id"] = text.strip()
                    state["step"]   = "svc_name"
                    send_message(chat_id, f"✅ ID: `{text.strip()}`\n\nNow send the service name:\nExample: `📱 WhatsApp`")
                    return
                elif state.get("step") == "svc_name":
                    state["svc_name"] = text.strip()
                    state["step"]     = "emoji"
                    send_message(chat_id,
                        f"✅ Name: *{text.strip()}*\n\n"
                        f"✨ *Pick a premium emoji* for this service\n_(tap one below, or skip)_",
                        get_emoji_picker_inline())
                    return

            # ── Phase 2: collect country info ──────────────────────────
            if state.get("step") == "cid":
                state["cid"]  = text.strip()
                state["step"] = "cname"
                send_message(chat_id, f"✅ ID: `{text.strip()}`\n\nNow send the country name:\nExample: `USA` or `India`")
                return
            elif state.get("step") == "cname":
                cname_resolved = _expand_country_name(text.strip())
                state["cname"] = cname_resolved
                state["step"]  = "emoji"
                auto_kb = get_auto_flag_keyboard(cname_resolved)
                if auto_kb:
                    send_message(chat_id,
                        f"✅ Name: *{cname_resolved}*\n\n"
                        f"🏳️ *Auto-detected flag:*\n"
                        f"Tap to use it, or pick a different one.",
                        auto_kb)
                else:
                    send_message(chat_id,
                        f"✅ Name: *{cname_resolved}*\n\n"
                        f"🏳️ *Pick a flag emoji* for this country\n_(tap one below, or skip)_",
                        get_country_emoji_picker_inline())
                return

            # ── Shared: build entry fields ─────────────────────────────
            if action == "add_service_flow":
                svc_id        = state.get("svc_id", "")
                svc_raw_name  = state.get("svc_name", "")
                svc_emoji     = state.get("svc_emoji", "")
                svc_emoji_id  = state.get("svc_emoji_id", "")
                svc_style     = state.get("svc_style", "primary")
                full_svc_name = f"{svc_emoji} {svc_raw_name}".strip() if svc_emoji else svc_raw_name
            else:
                svc_id        = state.get("svc_id", "")
                full_svc_name = state.get("svc_name", "")
                svc_emoji_id  = state.get("svc_emoji_id", "")
                svc_style     = state.get("svc_style", "primary")

            cid        = state.get("cid", "")
            cname_raw  = state.get("cname", "")
            cemoji     = state.get("cemoji", "")
            cemoji_id  = state.get("cemoji_id", "")
            cstyle     = state.get("cstyle", "primary")
            full_cname = f"{cemoji} {cname_raw}".strip() if cemoji else cname_raw
            cfg        = APIS.get(api, APIS["s1"])

            def _do_save(extra_fields=None, price_note=""):
                entry = {
                    "id": svc_id, "name": full_svc_name,
                    "price": state.get("service_price", 0),
                    "api": api, "emoji_id": svc_emoji_id, "style": svc_style,
                    **(extra_fields or {})
                }
                fresh = load_data()
                # S3/S4 services must go into their own country lists
                target_list_key = "s3_countries" if api == "s3" else ("s4_countries" if api == "s4" else "countries")
                target_list = fresh.setdefault(target_list_key, [])
                for c in target_list:
                    if c["id"] == cid:
                        c["services"].append(entry)
                        break
                else:
                    target_list.append({
                        "id": cid, "name": full_cname, "services": [entry],
                        "emoji_id": cemoji_id, "style": cstyle
                    })
                save_data(fresh)
                user_states.pop(chat_id, None)
                price = state.get("service_price", 0)
                send_message(chat_id,
                    f"✅ *Service Added!*\n\n"
                    f"`{svc_id}` — {full_svc_name} — ₹{price}\n"
                    f"🌍 {full_cname}\n"
                    + (f"{price_note}\n" if price_note else "")
                    + f"{cfg['icon']} {cfg['name']}",
                    get_srv_service_countries_inline(fresh, api, svc_id))

            if state.get("step") == "price":
                try:
                    price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid price.* Send a number like `20`:")
                    return
                state["service_price"] = price
                if needs_operator:
                    state["step"] = "operator"
                    send_message(chat_id, f"✅ Price: ₹{price}\n\n🔢 *Send the Operator ID:*\nExample: `1`, `2`, `any`\n\n_/cancel to abort_")
                    return
                if api == "s1":
                    state["step"] = "min_price"
                    send_message(chat_id,
                        f"✅ Price: ₹{price}\n\n"
                        f"🔹 *Send minPrice (SMSBower lowest accepted price):*\n"
                        f"Example: `0.11`\n\n_/cancel to abort_")
                    return
                if api == "s2":
                    state["step"] = "s2_max_price"
                    send_message(chat_id,
                        f"✅ Price: ₹{price}\n\n"
                        f"🔸 *Send maxPrice (HeroSMS highest accepted price):*\n"
                        f"Example: `0.50`\n\nSend `0` to skip.\n_/cancel to abort_")
                    return
                if api in ("s3", "s4"):
                    # S3/S4 supports only maxPrice (no minPrice in API)
                    state["step"] = "s3_max_price"
                    sname = APIS[api]["name"]
                    send_message(chat_id,
                        f"✅ Price: ₹{price}\n\n"
                        f"🔸 *Send maxPrice ({sname} highest accepted price):*\n"
                        f"Example: `0.50`\n\n"
                        f"_Numbers above this price will NOT be assigned._\n"
                        f"_Send `0` to skip (no upper limit)._\n"
                        f"_/cancel to abort_")
                    return
                _do_save()
                return
            elif state.get("step") == "s2_max_price":
                try:
                    max_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.50`:")
                    return
                extra = {"max_price": max_price} if max_price > 0 else None
                note  = f"🔸 maxPrice: `{max_price}`" if max_price > 0 else ""
                _do_save(extra, note)
                return
            elif state.get("step") == "s3_max_price":
                try:
                    max_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.50`:")
                    return
                operator = state.get("operator")
                extra = {}
                if operator:
                    extra["operator"] = operator
                if max_price > 0:
                    extra["max_price"] = max_price
                note_parts = []
                if operator:
                    note_parts.append(f"🔢 Operator: `{operator}`")
                if max_price > 0:
                    note_parts.append(f"🔸 maxPrice: `{max_price}`")
                _do_save(extra or None, "  ".join(note_parts))
                return
            elif state.get("step") == "min_price":
                try:
                    min_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.11`:")
                    return
                state["min_price"] = min_price
                state["step"]      = "max_price"
                send_message(chat_id,
                    f"✅ minPrice: `{min_price}`\n\n"
                    f"🔸 *Send maxPrice (SMSBower highest accepted price):*\n"
                    f"Example: `0.163`\n\n_/cancel to abort_")
                return
            elif state.get("step") == "max_price":
                try:
                    max_price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ *Invalid value.* Send a decimal number like `0.163`:")
                    return
                min_price = state.get("min_price", 0)
                _do_save(
                    {"min_price": min_price, "max_price": max_price},
                    f"🔹 minPrice: `{min_price}`  |  🔸 maxPrice: `{max_price}`"
                )
                return
            elif state.get("step") == "operator" and needs_operator:
                operator = text.strip()
                state["operator"] = operator
                if api in ("s3", "s4"):
                    state["step"] = "s3_max_price"
                    sname = APIS[api]["name"]
                    send_message(chat_id,
                        f"✅ Operator: `{operator}`\n\n"
                        f"🔸 *Send maxPrice ({sname} highest accepted price):*\n"
                        f"Example: `0.50`\n\n"
                        f"_Numbers above this price will NOT be assigned._\n"
                        f"_Send `0` to skip (no upper limit)._\n"
                        f"_/cancel to abort_")
                    return
                _do_save({"operator": operator}, f"🔢 Operator: `{operator}`")
                return

        if action == "add_mail":
            step = state.get("step")
            if step == "id":
                state["mail_id"]   = text.strip()
                state["step"]      = "name"
                send_message(chat_id, f"✅ Service ID: `{text.strip()}`\n\nNow send the *Service Name*:\nExample: `Gmail`")
                return
            elif step == "name":
                state["mail_name"] = text.strip()
                state["step"]      = "emoji"
                send_message(
                    chat_id,
                    f"✅ Name: *{text.strip()}*\n\n"
                    f"✨ *Pick an emoji* for this mail service\n_(tap one below, or skip)_",
                    get_mail_emoji_picker_inline()
                )
                return
            elif step == "domain":
                domain = text.strip()
                state["mail_domain"] = "" if domain == "-" else domain
                state["step"]        = "price"
                send_message(chat_id, f"✅ Domain: `{domain}`\n\n💰 Now send the *Price* in ₹:\nExample: `0`")
                return
            elif step == "price":
                try:
                    price = float(text.strip())
                except ValueError:
                    send_message(chat_id, "⚠️ Invalid price. Send a number like `0`:")
                    return
                ms = {
                    "id":       state["mail_id"],
                    "name":     state["mail_name"],
                    "domain":   state.get("mail_domain", ""),
                    "price":    price,
                    "emoji_id": state.get("emoji_id", ""),
                    "style":    "primary",
                }
                data.setdefault("mail_services", []).append(ms)
                save_data(data)
                user_states.pop(chat_id, None)
                send_message(chat_id,
                    f"✅ *Mail Service Added!*\n━━━━━━━━\n\n"
                    f"ID: `{ms['id']}` — *{ms['name']}*\n"
                    f"Domain: `{ms['domain'] or '-'}`  |  Price: ₹{price}",
                    {"inline_keyboard": [[_btn("📧 Mail Services", callback_data="admin_mail")]]})
                return

        # ── NEXA ADMIN STATES ─────────────────────────────────────

        # ── CUSTOM EMOJI ID INPUT ──────────────────────────────
        if state.get("step") == "epick_custom_input":
            if not is_admin:
                user_states.pop(chat_id, None)
                return
            typed_id = text.strip()
            if not typed_id.isdigit():
                send_message(chat_id,
                    "⚠️ শুধু numeric emoji ID দিন:\n"
                    "Example: `5267598011867491533`")
                return
            # Treat exactly like clicking an epick button with this ID
            emoji_id = typed_id
            chosen   = PICKER_ID_TO_CHAR.get(typed_id, "")
            # Save new custom emoji ID to bot_data so it appears in future picker
            if emoji_id not in PICKER_ID_TO_CHAR:
                custom_list = data.setdefault("custom_picker_emojis", [])
                if emoji_id not in custom_list:
                    custom_list.append(emoji_id)
                    save_data(data)
                PICKER_EMOJIS_DATA.append(("●", emoji_id))
                PICKER_ID_TO_CHAR[emoji_id] = "●"
            orig_action = action  # action is the original flow action
            state["step"] = "emoji"  # temporarily restore for logic below
            if orig_action == "add_country":
                state["emoji"]    = chosen
                state["emoji_id"] = emoji_id
                state["step"]     = "style"
                cname = state.get("country_name", "")
                send_message(chat_id,
                    f"✅ Emoji ID: `{emoji_id}`\nName: *{cname}*\n\n"
                    f"🎨 *Pick a button style:*",
                    get_style_picker_inline())
            elif orig_action == "add_service":
                state["emoji"]    = chosen
                state["emoji_id"] = emoji_id
                state["step"]     = "style"
                svc_name = state.get("service_name", "")
                send_message(chat_id,
                    f"✅ Emoji ID: `{emoji_id}`\nName: *{svc_name}*\n\n"
                    f"🎨 *Pick a button style:*",
                    get_style_picker_inline())
            elif orig_action == "add_service_flow" and state.get("phase", 1) == 1:
                state["svc_emoji"]    = chosen
                state["svc_emoji_id"] = emoji_id
                state["step"]         = "style"
                svc_name = state.get("svc_name", "")
                send_message(chat_id,
                    f"✅ Emoji ID: `{emoji_id}`\nName: *{svc_name}*\n\n"
                    f"🎨 *Pick a service button style:*",
                    get_style_picker_inline())
            elif orig_action == "add_service_flow" and state.get("phase", 1) == 2:
                state["cemoji"]    = chosen
                state["cemoji_id"] = emoji_id
                state["step"]      = "style"
                cname = state.get("cname", "")
                send_message(chat_id,
                    f"✅ Emoji ID: `{emoji_id}`\nCountry: *{cname}*\n\n"
                    f"🎨 *Pick a country button style:*",
                    get_style_picker_inline())
            elif orig_action == "add_country_to_service":
                state["cemoji"]    = chosen
                state["cemoji_id"] = emoji_id
                state["step"]      = "style"
                cname = state.get("cname", "")
                send_message(chat_id,
                    f"✅ Emoji ID: `{emoji_id}`\nCountry: *{cname}*\n\n"
                    f"🎨 *Pick a country button style:*",
                    get_style_picker_inline())
            elif orig_action == "add_mail":
                state["emoji_id"] = emoji_id
                state["step"]     = "domain"
                send_message(chat_id,
                    f"✅ Emoji ID: `{emoji_id}` ✨\n\n"
                    f"Now send the *Domain* (or send `-` for none):\n"
                    f"Example: `gmail.com`")
            else:
                send_message(chat_id, "⚠️ Unknown flow. /cancel করুন।")
            return

        if action == "sct_input":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            api_id = state.get("api_id", "s1")
            try:
                num = int(text.strip())
                if num < 0:
                    raise ValueError
            except (ValueError, TypeError):
                send_message(chat_id, "⚠️ Valid number দিন (e.g. `3`). 0 বা তার বেশি হতে হবে:")
                return
            cfg_api = APIS.get(api_id, APIS["s1"])
            user_states[chat_id] = {"action": "sct_confirm", "api_id": api_id, "num": num}
            send_message(
                chat_id,
                f"⏱ *{cfg_api['icon']} {cfg_api['name']} — Cancel Time*\n━━━━━━━━\n\n"
                f"আপনি *{num}* দিয়েছেন। এটা কি Minutes নাকি Seconds?",
                {
                    "inline_keyboard": [
                        [
                            _btn(f"⏱ {num} Minutes", callback_data="sct_min"),
                            _btn(f"⚡ {num} Seconds", callback_data="sct_sec"),
                        ],
                        [_btn("❌ Cancel",  callback_data="set_cancel_time")],
                    ]
                }
            )
            return

        if action == "sct_confirm":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            send_message(chat_id, "⬆️ উপরের *Minutes* বা *Seconds* বাটনে click করুন।")
            return

        if action == "change_api_key":
            if not is_admin:
                user_states.pop(chat_id, None)
                return
            api_id = state.get("api_id")
            new_key = text.strip()
            if not new_key:
                send_message(chat_id, "⚠️ Key খালি রাখা যাবে না। আবার পাঠান:")
                return
            if api_id not in APIS:
                user_states.pop(chat_id, None)
                send_message(chat_id, "❌ Invalid server.")
                return
            # Update in-memory
            APIS[api_id]["key"] = new_key
            # Persist to bot_data.json
            d = load_data()
            d.setdefault("api_keys", {})[api_id] = new_key
            save_data(d)
            user_states.pop(chat_id, None)
            cfg = APIS[api_id]
            disp = f"`{new_key[:6]}...{new_key[-4:]}`"
            send_message(
                chat_id,
                f"✅ *API Key Updated!*\n━━━━━━━━\n\n"
                f"{cfg['icon']} *{cfg['name']}*\n"
                f"New Key: {disp}",
                get_api_keys_panel_inline()
            )
            return

        if action == "nsrv_set":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            field = state.get("field")
            val   = text.strip()
            if field == "cancel_wait_mins":
                try:
                    val = int(val)
                    if val < 0:
                        raise ValueError
                except (ValueError, TypeError):
                    send_message(chat_id, "⚠️ Send a valid number of minutes (e.g. `3`). Must be 0 or more:")
                    return
                d = load_data()
                d["cancel_wait_mins"] = val
                save_data(d)
                user_states.pop(chat_id, None)
                _disp = f"{val}m" if val > 0 else "0 (instant cancel allowed)"
                send_message(
                    chat_id,
                    f"✅ *Cancel Wait Time Updated*\n━━━━━━━━\n\nNew time: *{_disp}*",
                    get_service_manage_more_inline(d)
                )
                return
            if field == "max_numbers":
                try:
                    val = int(val)
                    if val < 1 or val > 10:
                        raise ValueError
                except (ValueError, TypeError):
                    send_message(chat_id, "⚠️ Send a number between 1 and 10 (e.g. `3`):")
                    return
            _ch_fields = ("sms_channel", "sms_channel_all", "sms_channel_ig", "sms_channel_fb", "sms_channel_wa", "sms_channel_user", "sms_channel_user_buy")
            if field in _ch_fields:
                if val == "-":
                    val = ""
            if not val and field not in ("api_key",) + _ch_fields:
                send_message(chat_id, "⚠️ Value cannot be empty. Send again:")
                return
            d = load_data()
            d.setdefault("tnevs_settings", {})[field] = val
            save_data(d)
            user_states.pop(chat_id, None)
            label_map = {
                "api_key":              "🔑 API Key",
                "sms_channel":          "📡 Live SMS Channel",
                "sms_channel_all":      "📡 All OTP Channel",
                "sms_channel_ig":       "🟠 Instagram OTP Channel",
                "sms_channel_fb":       "🟦 Facebook OTP Channel",
                "sms_channel_wa":       "🟩 WhatsApp OTP Channel",
                "sms_channel_user":     "👤 User SMS Channel",
                "sms_channel_user_buy": "👤 User SMS Buy Channel",
            }
            label = label_map.get(field, field)
            disp  = str(val)[:40] + ("…" if len(str(val)) > 40 else "")
            if field == "sms_channel_user_buy":
                back_markup = get_service_manage_more_inline(d)
                saved_msg   = f"✅ *Saved!*\n━━━━━━━━\n\n{label}: `{disp}`" if val else f"✅ *Cleared!*\n━━━━━━━━\n\n{label} removed."
            elif field in _ch_fields:
                back_markup = get_sms_channel_panel_inline(d)
                saved_msg   = f"✅ *Saved!*\n━━━━━━━━\n\n{label}: `{disp}`" if val else f"✅ *Cleared!*\n━━━━━━━━\n\n{label} removed."
            else:
                back_markup = get_nsrv_api_inline(d)
                saved_msg   = f"✅ *Saved!*\n━━━━━━━━\n\n{label}: `{disp}`"
            send_message(chat_id, saved_msg, back_markup)
            return

        if action == "nsrv_add_srv":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            name = text.strip()
            if not name:
                send_message(chat_id, "⚠️ Send a valid service name (e.g. `WhatsApp`):")
                return
            srv_id = _nsrv_gen_id("srv")
            d = load_data()
            d.setdefault("services_data", {})[srv_id] = {"name": name, "countries": {}}
            save_data(d)
            user_states.pop(chat_id, None)
            send_message(chat_id,
                f"✅ *Service Added!*\n━━━━━━━━\n\n📁 *{name}*\nID: `{srv_id}`\n\n_Now add countries to this service._",
                get_nsrv_services_inline(d))
            return

        if action == "nsrv_add_cnt":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            name   = _expand_country_name(text.strip())
            srv_id = state.get("srv_id")
            if not name:
                send_message(chat_id, "⚠️ Send a valid country name (e.g. `Pakistan`):")
                return
            cnt_id = _nsrv_gen_id("cnt")
            d      = load_data()
            srv    = d.get("services_data", {}).get(srv_id)
            if not srv:
                send_message(chat_id,
                    f"❌ *Service not found in database.*\n━━━━━━━━\n\n"
                    f"_Service ID: `{srv_id}`_\n\n"
                    f"_Please go back and add the service again._",
                    get_nsrv_services_inline(d))
                user_states.pop(chat_id, None)
                return
            srv.setdefault("countries", {})[cnt_id] = {"name": name, "ranges": {}}
            ok = save_data(d)
            user_states.pop(chat_id, None)
            if not ok:
                send_message(chat_id, "❌ *Save failed.* Try again.")
                return
            flag      = get_country_flag(name)
            empty_cnt = {"name": name, "ranges": {}}
            send_message(chat_id,
                f"✅ *Country Added!*\n━━━━━━━━\n\n"
                f"{flag} *{name}*\n\n"
                f"Now tap ➕ *Add Range* to add a number prefix:",
                get_nsrv_ranges_inline(srv_id, cnt_id, empty_cnt))
            return

        if action == "nsrv_add_rng":
            if chat_id != ADMIN_ID:
                user_states.pop(chat_id, None)
                return
            rng_val = text.strip().lstrip("+").replace(" ", "")
            srv_id  = state.get("srv_id")
            cnt_id  = state.get("cnt_id")
            if not rng_val or not rng_val.isdigit():
                send_message(chat_id,
                    "⚠️ *Invalid range* — send digits only (no XXX).\n\n"
                    "Example: `22501`  `8801`  `7404`\n\nTry again:")
                return
            rng_id = _nsrv_gen_id("rng")
            d      = load_data()
            srv    = d.get("services_data", {}).get(srv_id)
            if srv:
                cnt = srv.get("countries", {}).get(cnt_id)
                if cnt:
                    cnt.setdefault("ranges", {})[rng_id] = rng_val
                    save_data(d)
            user_states.pop(chat_id, None)
            srv_f = d.get("services_data", {}).get(srv_id, {})
            cnt_f = srv_f.get("countries", {}).get(cnt_id, {})
            send_message(chat_id,
                f"✅ *Range Added!*\n━━━━━━━━\n\n"
                f"📱 `{rng_val}XXX`\nRID: `{rng_val}`\nID: `{rng_id}`",
                get_nsrv_ranges_inline(srv_id, cnt_id, cnt_f))
            return

    # ── REGULAR COMMANDS ─────────────────────────────────────
    if text.startswith("/start"):
        parts = text.split()
        if len(parts) > 1 and parts[1].startswith("ref_"):
            referrer_id = parts[1][4:]
            uid_str     = str(chat_id)
            if referrer_id != uid_str and uid_str not in data.get("referrals", {}):
                data.setdefault("referrals", {})[uid_str] = referrer_id
                save_data(data)
        send_message(chat_id,
            f"👋 *Welcome to*\n"
            f"🤖 *Virtual Numbers Bot*\n━━━━━━━━",
            get_main_keyboard(is_admin))

    elif text == "Buy Number":
        send_message(chat_id,
            "📱 *Buy Number*\n━━━━━━━━",
            {"inline_keyboard": [
                [
                    {"text": "All Services", "callback_data": "buy_all",    "style": "success", "icon_custom_emoji_id": "5431499171045581032"},
                    {"text": "Search",       "callback_data": "buy_search", "style": "primary", "icon_custom_emoji_id": "5429571366384842791"},
                ]
            ]})

    elif text == "Get Mail":
        mail_services = data.get("mail_services", [])
        if not mail_services:
            send_message(chat_id,
                "❌ *No Mail Services*\n━━━━━━━━\n\n"
                "_Mail services not configured yet. Contact admin._")
            return
        send_message(chat_id,
            "📧 *Get Mail*\n━━━━━━━━\n\n"
            "Choose the mail service you need:",
            get_mail_services_inline(mail_services))

    elif text == "Refer":
        ref_link = f"https://t.me/Akg_numberbot?start=ref_{chat_id}"
        send_message(chat_id,
            f"🔗 *Refer & Earn*\n━━━━━━━━\n\n"
            f"Share your link and earn when friends join!\n"
            f"💰 Earn *3%* of every deposit your referral makes",
            {"inline_keyboard": [[
                {"text": "Invite Now", "copy_text": {"text": ref_link}, "style": "success", "icon_custom_emoji_id": "5323442290708985472"}
            ]]},
            emoji_overrides={"🔗": "6278397714577231283", "💰": "5332275187305958945"})

    elif text == "Deposit":
        _ds     = data.get("deposit_settings", {})
        _upi    = _ds.get("upi_id",      UPI_ID)
        _qr     = _ds.get("qr_url",      "https://t.me/KP_MODZ_2/55")
        _mn     = _ds.get("min_deposit", MIN_DEPOSIT)
        _en     = _ds.get("enabled",     True)
        if not _en:
            send_message(chat_id,
                "🚫 *Deposit is currently disabled.*\n━━━━━━━━\n\n"
                "_Please contact admin for assistance._",
                {"inline_keyboard": [[{
                    "text": "@ITZ_EKBALL", "url": "https://t.me/ITZ_EKBALL",
                    "style": "danger", "icon_custom_emoji_id": "5314504236132747481",
                }]]})
            return
        caption = (
            f"💰 *Add Funds to Wallet*\n━━━━━━━━\n\n"
            f"📲 *Pay via UPI*\n\n"
            f"▸ UPI ID : `{_upi}`\n\n"
            f"⚠️ Minimum deposit : ₹{_mn}\n\n"
            f"━━━━━━━━\nAfter paying tap 👇 to submit proof"
        )
        depo_ovr = {"💰": "6199349120667226859"}
        ok = send_photo(chat_id, _qr, caption, get_submit_payment_inline(), emoji_overrides=depo_ovr)
        if not ok:
            send_message(chat_id, caption, get_submit_payment_inline(), emoji_overrides=depo_ovr)

    elif text == "Account":
        show_account(chat_id, data)

    elif text == "Active Numbers":
        user    = get_user(data, chat_id)
        history = user.get("history", [])
        active  = [h for h in history if h.get("status") == "active"]
        if not active:
            send_message(chat_id,
                "📋 *Active Numbers*\n━━━━━━━━\n\n"
                "✅ No active numbers right now.\n\n"
                "_All completed or cancelled._")
            return
        send_message(chat_id,
            f"📋 *Active Numbers*  ({len(active)} total)\n━━━━━━━━")
        for h in active:
            api      = h.get("api", "s1")
            req_id   = h.get("req_id", "")
            svc      = h.get("service", "")
            ctr      = h.get("country", "")
            svc_icon = _free_svc_icon_char(svc)
            ctr_flag = get_country_flag(ctr)
            prefix   = f"{ctr_flag} {svc_icon}"
            info_p   = pending_sms.get(req_id, {})
            time_str = ""
            if info_p.get("started_at"):
                rem  = max(0, AUTO_CANCEL_SECS - (time.time() - info_p["started_at"]))
                m, s = int(rem // 60), int(rem % 60)
                time_str = f"  •  ⏱ {m}:{s:02d}"
            send_message(chat_id,
                f"{prefix} `{h.get('number', 'N/A')}`\n"
                f"Service: *{svc}*  •  {ctr}\n"
                f"Price ₹{h.get('price', 0)}{time_str}",
                get_number_action_inline(req_id, h.get("cid"), h.get("svc_id"), api))

    elif text == "More":
        send_message(chat_id,
            "➕ *More Options*\n━━━━━━━━",
            get_more_keyboard())

    elif text == "Gift Code":
        markup = {
            "inline_keyboard": [
                [{"text": "Claim Gift Code", "callback_data": "gift_claim", "style": "success", "icon_custom_emoji_id": "5251227707026470504"}],
                [{"text": "Create Gift Code", "callback_data": "gift_create", "style": "primary", "icon_custom_emoji_id": "5251219649667820949"}],
            ]
        }
        send_message(chat_id,
            "🎁 *Gift Code*\n━━━━━━━━",
            markup)

    elif text == "Free Number":
        services_data = data.get("services_data", {})
        has_nexa = any(
            any(len(cnt.get("ranges", {})) > 0 for cnt in srv.get("countries", {}).values())
            for srv in services_data.values()
        )
        if has_nexa:
            send_message(chat_id,
                "📱 *Free Numbers*\n━━━━━━━━",
                get_free_nexa_services_inline(data))
        else:
            send_message(chat_id,
                "❌ *No Free Number Services Available*\n━━━━━━━━\n\n"
                "_Contact admin to configure free number services._")

    elif text == "Temp Mail":
        pending_tmail.pop(chat_id, None)
        creating_id = send_message(chat_id, "⏳ *Creating temp email...*", emoji_overrides={"⏳": "5350719881408966477"})
        mail = create_temp_mail()
        if creating_id:
            delete_message(chat_id, creating_id)
        if not mail:
            send_message(chat_id, "❌ *Failed to create temp email.* Try again later.")
            return
        temp_mails[chat_id] = mail
        msg_id = send_message(chat_id,
            f"📧 *Temp Mail Ready!*\n━━━━━━━━\n\n"
            f"✉️ `{mail['address']}`\n\n"
            f"⏳ *Waiting for SMS...*",
            {"inline_keyboard": [[{"text": "🔄 New Email", "callback_data": f"tmail_new_{chat_id}", "style": "danger", "icon_custom_emoji_id": "5384466886857614542"}]]},
            emoji_overrides={"⏳": "5208619406657082341"}
        )
        with _pending_tmail_lock:
            pending_tmail[chat_id] = {
                "address":    mail["address"],
                "token":      mail["token"],
                "msg_id":     msg_id,
                "started_at": time.time(),
                "known_ids":  set(),
            }

    elif text in ("Back", "⬅️ Back"):
        send_message(chat_id,
            "🏠 *Main Menu*\n━━━━━━━━",
            get_main_keyboard(is_admin))

    elif text == "Admin Panel" and is_admin:
        send_message(chat_id,
            "⚙️ *Admin Panel*\n━━━━━━━━",
            get_admin_main_inline())

    elif photo:
        state = user_states.get(chat_id, {})
        if state.get("action") == "deposit" and state.get("step") == "screenshot":
            # treat as deposit screenshot
            amount  = state.get("amount")
            utr     = state.get("utr")
            file_id = photo[-1]["file_id"]
            caption = (
                f"🔔 *New Deposit Request*\n━━━━━━━━\n\n"
                f"👤 User: `{chat_id}`\n"
                f"💰 Amount: ₹{amount}\n"
                f"🧾 UTR: `{utr}`"
            )
            send_photo_file(ADMIN_ID, file_id, caption, get_admin_depo_inline(chat_id, amount, utr))
            user_states.pop(chat_id, None)
            send_message(chat_id,
                "✅ *Request Submitted!*\n━━━━━━━━\n\n"
                "⏳ Waiting for admin approval...\n_You'll be notified once approved._",
                emoji_overrides={"✅": "6307741669877880674", "⏳": "5350427505805238170"})



# ============= MAIN LOOP =============

def main():
    _startup_data = load_data()
    _get_search_index(_startup_data)   # pre-warm search index at startup
    print("=" * 50)
    print("🤖 VIRTUAL NUMBERS BOT v7.0 — Dual Server")
    print("=" * 50)
    print(f"✅ Admin: {ADMIN_ID}")
    for api_id, cfg in APIS.items():
        print(f"   {cfg['icon']} {cfg['name']} ready")
    print("=" * 50)

    threading.Thread(target=auto_sms_checker, daemon=True).start()
    print("📨 Auto SMS checker started (every 5s)")
    threading.Thread(target=auto_tmail_checker, daemon=True).start()
    print("📧 Auto Temp Mail checker started (every 5s)")
    threading.Thread(target=auto_console_forwarder, daemon=True).start()
    print("📡 Auto console forwarder started (every 10s)")
    threading.Thread(target=auto_free_num_checker, daemon=True).start()
    print("📱 Auto Free Number OTP checker started (every 5s)")
    threading.Thread(target=auto_nsrv_num_checker, daemon=True).start()
    print("🔢 Auto Number Server OTP checker started (every 5s)")

    # Drain ALL pending updates so old messages aren't re-processed on restart
    pending = get_updates(timeout=0)
    if pending:
        last_update_id = pending[-1]["update_id"]
        # Explicitly acknowledge by calling with next offset (two-step drain)
        get_updates(last_update_id + 1, timeout=0)
        print(f"⏩ Skipped {len(pending)} pending update(s) from before startup")
    else:
        last_update_id = 0

    _update_executor = concurrent.futures.ThreadPoolExecutor(max_workers=10, thread_name_prefix="upd")

    def _handle_update(update):
        try:
            if "callback_query" in update:
                cb       = update["callback_query"]
                chat_id  = cb["message"]["chat"]["id"]
                msg_id   = cb["message"]["message_id"]
                cb_fn    = cb.get("from", {})
                with _get_user_lock(chat_id):
                    process_callback(chat_id, cb["data"], cb["id"], msg_id,
                                     first_name=cb_fn.get("first_name", ""),
                                     username=cb_fn.get("username", ""))
            elif "message" in update:
                msg       = update["message"]
                chat_type = msg["chat"].get("type", "private")
                # Only handle private chats — ignore group/channel/supergroup messages
                if chat_type != "private":
                    return
                chat_id    = msg["chat"]["id"]
                text       = msg.get("text", "")
                photo      = msg.get("photo")
                # Always get user info from "from" (the actual sender), not "chat"
                _sender    = msg.get("from", {})
                first_name = _sender.get("first_name", "") or msg["chat"].get("first_name", "")
                username   = _sender.get("username", "") or msg["chat"].get("username", "")
                name       = first_name or "User"
                if text:
                    print(f"📩 {name}: {text[:40]}")
                elif photo:
                    print(f"📸 {name}: [photo]")
                with _get_user_lock(chat_id):
                    process_message(chat_id, text or "", photo=photo,
                                    first_name=first_name, username=username,
                                    message_id=msg.get("message_id"))
        except Exception as e:
            import traceback
            print(f"⚠️ Update processing error: {e}\n{traceback.format_exc()}")

    while True:
        try:
            updates = get_updates(last_update_id + 1)
            for update in updates:
                uid = update["update_id"]
                if uid in _seen_updates:
                    last_update_id = max(last_update_id, uid)
                    continue
                _seen_updates.add(uid)
                # Cap _seen_updates to avoid unbounded memory growth
                if len(_seen_updates) > 5000:
                    _seen_updates.clear()
                    _seen_updates.add(uid)
                last_update_id = max(last_update_id, uid)
                _update_executor.submit(_handle_update, update)
        except KeyboardInterrupt:
            print("\n🛑 Bot stopped")
            _update_executor.shutdown(wait=False)
            break
        except Exception as e:
            print(f"⚠️ Poll error: {e}")
            time.sleep(5)


if __name__ == "__main__":
    main()
